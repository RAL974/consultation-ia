"""
Tests du référentiel articles (moteur/referentiel.py).

Fixture CSV synthétique minimale reproduisant le MOTIF réel observé dans
base/BDD_articles.csv (Legrand -> préfixe "LEG" très majoritaire, 1 ligne
minoritaire contradictoire à ignorer ; d'autres fabricants sans préfixe),
plus une assertion qualitative sur le vrai fichier pour ancrer la
déduction dans la donnée réelle.
"""

import csv as csv_module

import pytest

from moteur.referentiel import Referentiel, deduire_prefixes, cle_designation, _etiquette_type
from moteur.base import BaseArticles
from moteur.comparateur import comparer
from moteur.modele import Article, LigneBesoin

from conftest import ROOT

ENTETES = [
    "Référence", "Désignation", "Fournisseur", "Fabricant", "Catégorie",
    "Tarif approximatif", "CONCAT", "Clé_Réf",
]

LIGNES_BDD = [
    # 5 Legrand préfixées LEG (majorité) + 1 minoritaire EBE (incohérence
    # de saisie, doit être ignorée par la règle déduite)
    ["LEG411651", "Disjoncteur DX3 ID 2P 63A A 30MA", "COMINTER", "Legrand", "Disjoncteurs", "78,14", "LEG411651", "411651"],
    ["LEG406774", "DNX3 4500 1P+N C16", "COMINTER", "Legrand", "Disjoncteurs", "12,50", "LEG406774", "406774"],
    ["LEG100001", "Prise de courant 2P+T", "COMINTER", "Legrand", "Appareillage", "5,20", "LEG100001", "100001"],
    ["LEG100002", "Interrupteur va-et-vient", "COMINTER", "Legrand", "Appareillage", "4,10", "LEG100002", "100002"],
    ["LEG100003", "Va-et-vient double", "COMINTER", "Legrand", "Appareillage", "6,30", "LEG100003", "100003"],
    ["EBE664702", "Interrupteur double va et vient Niloé", "COMINTER", "Legrand", "Appareillage", "7,00", "EBE664702", "664702"],
    # L'Ebénoïd : préfixe EBE (son propre préfixe, cohérent)
    ["EBE500001", "Interrupteur simple", "COMINTER", "L'Ebénoïd", "Appareillage", "3,00", "EBE500001", "500001"],
    # Schneider Electric : pas de préfixe (CONCAT == Clé_Réf == Référence)
    ["MELV429338T", "Commande rotative prolongée NSX250", "COMINTER", "Schneider Electric", "Accessoires", "173,79", "MELV429338T", "MELV429338T"],
    # BLM : pas de préfixe
    ["615250", "Douille DCL E27", "109 DISTRIBUTION", "BLM", "Luminaires", "2,10", "615250", "615250"],
]


@pytest.fixture
def bdd_csv(tmp_path):
    fichier = tmp_path / "BDD_articles.csv"
    with open(fichier, "w", encoding="utf-8-sig", newline="") as f:
        w = csv_module.writer(f, delimiter=";")
        w.writerow(ENTETES)
        w.writerows(LIGNES_BDD)
    return fichier


@pytest.fixture
def referentiel(tmp_path, bdd_csv):
    ref = Referentiel(tmp_path / "moteur")
    ref.importer_bdd(bdd_csv)
    yield ref
    ref.fermer()


def _lignes_dict():
    return [dict(zip(ENTETES, ligne)) for ligne in LIGNES_BDD]


# ------------------------------------------------------------------
# Déduction des préfixes
# ------------------------------------------------------------------
def test_deduire_prefixes_majorite_legrand():
    prefixes = deduire_prefixes(_lignes_dict())
    assert prefixes["Legrand"] == "LEG"


def test_deduire_prefixes_par_fabricant():
    prefixes = deduire_prefixes(_lignes_dict())
    assert prefixes["L'Ebénoïd"] == "EBE"
    # Schneider Electric et BLM n'ont pas de préfixe -> absents du résultat
    assert "Schneider Electric" not in prefixes
    assert "BLM" not in prefixes


def test_deduire_prefixes_minorite_ignoree():
    # La ligne EBE664702 (Fabricant=Legrand) est minoritaire face aux 5
    # lignes LEG -> n'écrase pas la règle Legrand -> LEG.
    prefixes = deduire_prefixes(_lignes_dict())
    assert prefixes["Legrand"] == "LEG"


def test_deduction_sur_vraie_bdd():
    """Ancrage dans la donnée réelle : reste qualitatif (pas de compte
    exact) pour ne pas devenir fragile si base/BDD_articles.csv évolue."""
    with open(ROOT / "base" / "BDD_articles.csv", encoding="utf-8-sig") as f:
        lignes = list(csv_module.DictReader(f, delimiter=";"))
    assert len(lignes) > 5000
    prefixes = deduire_prefixes(lignes)
    assert prefixes.get("Legrand") == "LEG"


# ------------------------------------------------------------------
# Import
# ------------------------------------------------------------------
def test_import_idempotent(tmp_path, bdd_csv):
    ref = Referentiel(tmp_path / "moteur")
    n1 = ref.importer_bdd(bdd_csv)
    n2 = ref.importer_bdd(bdd_csv)
    assert n1 == n2 == len(LIGNES_BDD)
    total = ref.cx.execute("SELECT COUNT(*) FROM articles").fetchone()[0]
    assert total == len(LIGNES_BDD)
    ref.fermer()


# ------------------------------------------------------------------
# Résolution
# ------------------------------------------------------------------
def test_resoudre_connu_via_alias_import(referentiel):
    # La référence telle qu'écrite dans la BDD (=CONCAT) est un alias direct
    cle, statut = referentiel.resoudre("LEG411651")
    assert statut == "connu"
    assert cle == "411651"


def test_resoudre_connu_reference_nue(referentiel):
    cle, statut = referentiel.resoudre("411651")
    assert statut == "connu"
    assert cle == "411651"


def test_resoudre_connu_reference_alphanumerique(referentiel):
    # Cas où coeur_numerique perdrait de l'information (pas de préfixe à
    # retirer, mais alias exact grâce à l'import) : la vraie valeur ajoutée
    # du référentiel par rapport à coeur_numerique() seul.
    cle, statut = referentiel.resoudre("MELV429338T")
    assert statut == "connu"
    assert cle == "MELV429338T"


def test_resoudre_propose_designation_proche(referentiel):
    # "406774-LEG" n'est un alias exact de rien, mais son cœur numérique
    # (406774) correspond à un article connu dont la désignation est proche.
    cle, statut = referentiel.resoudre(
        "406774-LEG", designation="Disjoncteur DNX3 4500 1P+N C16"
    )
    assert statut == "propose"
    assert cle == "406774"


def test_resoudre_nouveau_sans_correspondance(referentiel):
    cle, statut = referentiel.resoudre("ZZZ999999XYZ", designation="Article totalement inconnu")
    assert statut == "nouveau"
    assert cle == "ZZZ999999XYZ"
    # La fois suivante (même exécution ou une prochaine), l'auto-alias créé
    # ci-dessus évite de relancer la recherche de candidats MAIS reste
    # rapporté comme "nouveau" (même clé) : ce n'est pas une correspondance
    # confirmée. Sinon, la MÊME référence obtiendrait un statut différent
    # selon qu'elle est résolue pour la 1ère ou la 2e fois dans une même
    # exécution (ex. l'article d'un devis, puis la ligne de besoin
    # correspondante) -> comparateur.py leur donnerait des clés
    # différentes ("REF:" ou pas) et la ligne ne matcherait plus jamais.
    cle2, statut2 = referentiel.resoudre("ZZZ999999XYZ")
    assert statut2 == "nouveau"
    assert cle2 == cle


def test_resoudre_nouveau_pas_de_doublon(referentiel):
    referentiel.resoudre("NOUVEAU123456", designation="Article X")
    referentiel.resoudre("NOUVEAU123456", designation="Article X")
    n = referentiel.cx.execute(
        "SELECT COUNT(*) FROM articles WHERE cle_normalisee='NOUVEAU123456'"
    ).fetchone()[0]
    assert n == 1


def test_comparateur_reference_inconnue_sans_coeur_numerique_matche_quand_meme(tmp_path, referentiel):
    """
    Régression : une référence absente de la BDD, avec moins de 4 chiffres
    (donc pas de filet de sécurité par cœur numérique dans comparer()), doit
    quand même rapprocher l'offre du devis et la ligne de besoin qui portent
    EXACTEMENT la même référence. L'article est résolu en premier (boucle
    articles de comparer()), la ligne de besoin ensuite : le référentiel ne
    doit pas leur attribuer des clés différentes selon l'ordre de résolution
    (l'auto-alias créé au 1er passage ne doit pas se faire passer pour une
    correspondance confirmée au 2e).
    """
    base = BaseArticles(tmp_path / "base")

    articles = [
        Article(fournisseur="RAVATE", devis="DV1", reference_fournisseur="TUBFFPTTGRD3",
                reference_distributeur="", designation="ICT 20 PREF PTT GRADE 3 S+ VER",
                quantite=6, unite="ROU", prix_brut=136.34, prix_net=136.34, montant=818.04),
    ]
    besoin = [LigneBesoin(reference="TUBFFPTTGRD3", quantite=4, designation="Fourreau grade 3 PTT")]

    resultat = comparer(articles, besoin, base, referentiel)

    assert len(resultat["lignes"]) == 1
    assert resultat["lignes"][0]["offres"].get("RAVATE") is not None
    assert resultat["hors_besoin"] == []


# ------------------------------------------------------------------
# Rapprochement par désignation (besoin SANS référence, ex. bordereau
# architecte "Type A1 - Suspension circulaire..." — chantier Cosinus)
# ------------------------------------------------------------------
def test_cle_designation():
    assert cle_designation("Type A1 - Suspension circulaire Open space") \
        == "DESC:TYPE_A1_SUSPENSION_CIRCULAIRE_OPEN_SPACE"
    # Insensible aux accents/casse : deux désignations "identiques" au
    # sens humain doivent donner la MÊME clé.
    assert cle_designation("type a1 suspension circulaire open space  ") \
        == "DESC:TYPE_A1_SUSPENSION_CIRCULAIRE_OPEN_SPACE"
    assert cle_designation("") == ""
    assert cle_designation("   ") == ""


def test_etiquette_type():
    assert _etiquette_type("Type A1 - Suspension circulaire Open space") == "A1"
    assert _etiquette_type("TYPE A1 - SUSPENSION CIRCULAIRE : SD-WOOD RING...") == "A1"
    assert _etiquette_type("Type L1 - Module de jonction T et L") == "L1"
    assert _etiquette_type("Coffret XL³ 160 classe II 2 rangées") is None


def test_proposer_correspondances_designation_etiquette_type_rattrape_score_faible(referentiel):
    # Régression réelle (chantier Cosinus) : le texte technique environnant
    # dilue le score Jaccard sous le seuil (~0.24 observé) malgré une
    # étiquette "Type A1" identique des deux côtés -> doit quand même être
    # proposé.
    candidats = [
        {"fournisseur": "STAND 64", "reference": "KUBIA-ART00031180",
         "designation": (
             "TYPE A1 - SUSPENSION CIRCULAIRE : SD-WOOD RING SUSPENSION "
             "Ø700MM 70W 8000LM 4000°K IP20 BOIS (KUBIA-ART00031180)"
         ), "quantite": 3},
    ]
    retenus = referentiel.proposer_correspondances_designation(
        "Type A1 - Suspension circulaire Open space", 3, candidats,
    )
    assert len(retenus) == 1
    assert retenus[0]["reference"] == "KUBIA-ART00031180"
    assert retenus[0]["score"] >= 0.9


def test_proposer_correspondances_designation_meilleur_par_fournisseur(referentiel):
    candidats = [
        {"fournisseur": "CLAREO", "reference": "DOW.105942",
         "designation": "Downlight WC Sanitaire Douche CLAREO 19W Diffuseur Opale", "quantite": 9},
        {"fournisseur": "STAND 64", "reference": "ACB-A2033070B",
         "designation": "KOWA APPLIQUE LED applique murale 136x100mm", "quantite": 4},
        # Bruit total (aucun rapport) : ne doit jamais être proposé.
        {"fournisseur": "DEM", "reference": "TUBFF3G6",
         "designation": "FF 3G6 RGE B V/J C50", "quantite": 200},
    ]

    retenus = referentiel.proposer_correspondances_designation(
        "Type A4 - Downlight - WC, Sanitaire, Douche", 9, candidats,
    )

    fournisseurs_retenus = {c["fournisseur"] for c in retenus}
    assert fournisseurs_retenus == {"CLAREO"}
    assert "DOW105942" in referentiel._propositions
    p = referentiel._propositions["DOW105942"]
    assert p["fournisseur"] == "CLAREO"
    assert p["cle"] == "DESC:TYPE_A4_DOWNLIGHT_WC_SANITAIRE_DOUCHE"
    assert p["designation_candidat"] == "Type A4 - Downlight - WC, Sanitaire, Douche"
    assert p["qte_devis"] == 9
    assert p["qte_besoin"] == 9
    assert "TUBFF3G6" not in referentiel._propositions


def test_proposer_correspondances_designation_ne_fusionne_jamais(referentiel):
    """Une proposition par désignation ne doit JAMAIS apparaître comme une
    offre dans comparer() tant qu'elle n'est pas confirmée — même avec un
    score élevé."""
    base = BaseArticles(referentiel.dossier.parent / "base")

    articles = [
        Article(fournisseur="CLAREO", devis="D1", reference_fournisseur="DOW.105942",
                reference_distributeur="", designation="Downlight WC Sanitaire Douche",
                quantite=9, unite="UN", prix_brut=45.0, prix_net=45.0, montant=405.0),
    ]
    besoin = [LigneBesoin(
        reference="", quantite=9, designation="Type A4 - Downlight - WC, Sanitaire, Douche")]

    resultat = comparer(articles, besoin, base, referentiel)

    assert resultat["lignes"][0]["offres"] == {}
    assert "DOW105942" in referentiel._propositions


def test_comparateur_besoin_sans_reference_confirme_ensuite_matche(tmp_path, referentiel):
    """Cycle complet : proposition par désignation -> confirmation acheteur
    (A_confirmer.xlsx) -> la ligne de besoin se rapproche normalement, EXACTEMENT
    comme une proposition par référence (aucun traitement spécial ensuite)."""
    dossier_referentiel = tmp_path / "referentiel"
    base = BaseArticles(tmp_path / "base_besoin_sans_ref")

    articles = [
        Article(fournisseur="CLAREO", devis="D1", reference_fournisseur="DOW.105942",
                reference_distributeur="", designation="Downlight WC Sanitaire Douche",
                quantite=9, unite="UN", prix_brut=45.0, prix_net=45.0, montant=405.0),
    ]
    besoin = [LigneBesoin(
        reference="", quantite=9, designation="Type A4 - Downlight - WC, Sanitaire, Douche")]

    resultat = comparer(articles, besoin, base, referentiel)
    assert resultat["lignes"][0]["offres"] == {}

    fichier = referentiel.ecrire_a_confirmer(dossier_referentiel)
    assert fichier is not None

    from openpyxl import load_workbook
    wb = load_workbook(fichier)
    ws = wb.active
    entetes = [c.value for c in ws[1]]
    col_ref = entetes.index("Référence détectée")
    col_decision = entetes.index("Décision")
    for row in ws.iter_rows(min_row=2):
        if row[col_ref].value == "DOW.105942":
            row[col_decision].value = "OUI"
    wb.save(fichier)

    referentiel.appliquer_confirmations(dossier_referentiel / "A_confirmer.xlsx")

    # Prochaine exécution (même processus ou un nouveau) : la ligne de
    # besoin matche normalement, sans repasser par la désignation.
    resultat2 = comparer(articles, besoin, base, referentiel)
    assert resultat2["lignes"][0]["offres"].get("CLAREO") is not None


# ------------------------------------------------------------------
# Cycle de confirmation (fichier Excel aller-retour)
# ------------------------------------------------------------------
def test_cycle_confirmation(tmp_path, referentiel):
    dossier_referentiel = tmp_path / "referentiel"

    cle, statut = referentiel.resoudre(
        "406774-LEG", designation="Disjoncteur DNX3 4500 1P+N C16"
    )
    assert statut == "propose"

    fichier = referentiel.ecrire_a_confirmer(dossier_referentiel)
    assert fichier is not None and fichier.exists()

    from openpyxl import load_workbook
    wb = load_workbook(fichier)
    ws = wb.active
    entetes = [c.value for c in ws[1]]
    col_ref = entetes.index("Référence détectée")
    col_decision = entetes.index("Décision")
    for row in ws.iter_rows(min_row=2):
        if row[col_ref].value == "406774-LEG":
            row[col_decision].value = "OUI"
    wb.save(fichier)

    referentiel.appliquer_confirmations(dossier_referentiel / "A_confirmer.xlsx")

    cle2, statut2 = referentiel.resoudre("406774-LEG")
    assert statut2 == "connu"
    assert cle2 == "406774"


def test_confirmation_non_devient_sa_propre_cle(tmp_path, referentiel):
    dossier_referentiel = tmp_path / "referentiel"
    referentiel.resoudre("406774-LEG", designation="Disjoncteur DNX3 4500 1P+N C16")
    referentiel.ecrire_a_confirmer(dossier_referentiel)

    from openpyxl import load_workbook
    fichier = dossier_referentiel / "A_confirmer.xlsx"
    wb = load_workbook(fichier)
    ws = wb.active
    entetes = [c.value for c in ws[1]]
    col_ref = entetes.index("Référence détectée")
    col_decision = entetes.index("Décision")
    for row in ws.iter_rows(min_row=2):
        if row[col_ref].value == "406774-LEG":
            row[col_decision].value = "NON"
    wb.save(fichier)

    referentiel.appliquer_confirmations(fichier)

    cle, statut = referentiel.resoudre("406774-LEG")
    assert statut == "connu"
    assert cle == "406774LEG"  # devenue sa propre clé (ref normalisée)


# ------------------------------------------------------------------
# Composés
# ------------------------------------------------------------------
def test_composants_vide_par_defaut(referentiel):
    assert referentiel.composants("411651") == []


def test_comparateur_expansion_composes(tmp_path, referentiel):
    referentiel.cx.execute(
        "INSERT INTO composes VALUES (?,?,?,?,?)",
        ("COFFRET1", "411651", 1, 0, "manuel"),
    )
    referentiel.cx.execute(
        "INSERT INTO composes VALUES (?,?,?,?,?)",
        ("COFFRET1", "406774", 2, 1, "manuel"),
    )
    referentiel.cx.commit()

    base = BaseArticles(tmp_path / "base")

    articles = [
        Article(fournisseur="COMINTER", devis="D1", reference_fournisseur="LEG411651",
                reference_distributeur="", designation="Disjoncteur DX3 ID 2P 63A A 30MA",
                quantite=1, unite="UN", prix_brut=78.14, prix_net=78.14, montant=78.14),
        Article(fournisseur="COMINTER", devis="D1", reference_fournisseur="LEG406774",
                reference_distributeur="", designation="DNX3 4500 1P+N C16",
                quantite=2, unite="UN", prix_brut=12.50, prix_net=12.50, montant=25.00),
    ]
    besoin = [LigneBesoin(reference="COFFRET1", quantite=1, designation="Coffret complet")]

    resultat = comparer(articles, besoin, base, referentiel)

    assert len(resultat["lignes"]) == 2
    refs = sorted(l["reference"] for l in resultat["lignes"])
    assert refs == ["406774", "411651"]
    for ligne in resultat["lignes"]:
        if ligne["reference"] == "406774":
            assert ligne["qte_besoin"] == 2  # 1 (besoin) x 2 (quantite_membre)
        else:
            assert ligne["qte_besoin"] == 1

    base.fermer()
