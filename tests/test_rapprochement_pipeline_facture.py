"""
Logique de moteur/rapprochement/pipeline_facture.py — même esprit que
tests/test_rapprochement_pipeline_bl.py (BL) : objets Facture/
CorrespondanceFacture construits à la main pour la logique pure, un vrai
classeur xlsx (tmp_path, jamais le vrai Suivi) pour ce qui touche à
l'écriture réelle.

Les 5 colonnes facture (N° facture / Date facture / Qté facturée /
PU facturé / Montant facturé HT — voir moteur.rapprochement.ecriture.
ENTETES_FACTURE) ont été créées pour de vrai dans le VRAI Suivi commandes
le 2026-09-01 (voir CLAUDE.md, "colonnes créées dans le Suivi vivant") :
les tests d'écriture réelle ci-dessous utilisent un classeur SYNTHÉTIQUE
(mécanique générique, rapide) ET un test dédié sur une COPIE du vrai
classeur vivant (skipif absent du poste), même esprit que
tests/test_rapprochement_ecriture.py. Un test dédié vérifie aussi le
comportement attendu en l'absence des colonnes (échec propre, pas un
plantage) — utile pour un classeur d'un autre poste pas encore à jour."""

import shutil
import zipfile
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from moteur.rapprochement.ecriture import ColonneNonModifiable, appliquer, lire_entetes
from moteur.rapprochement.pieces import (
    COLONNES_FACTURE_CALCULEES,
    FeuillePiecesAbsente,
    IndexPieces,
    ecrire_pieces,
    feuille_pieces_presente,
    installer_feuille_pieces,
    lire_pieces,
)
from moteur.rapprochement.matching_facture import CauseFacture, CorrespondanceFacture, LigneSuiviFacture, StatutFacture
from moteur.rapprochement.modele_facture import Facture, LigneFacture
from moteur.rapprochement import pipeline_facture
from moteur.rapprochement.pipeline_bl import trouver_fichier_suivi_vivant
from moteur.rapprochement.pipeline_facture import (
    NOM_FEUILLE_SUBSTITUTIONS,
    RapportRapprochementFacture,
    _appliquer_confirmations_substitutions,
    _ecrire_substitutions_probables,
    _est_resolu_facture,
    _resoudre_commandes_facture,
    _verifier_total_ht_facture,
    appliquer_et_archiver_factures,
    archiver_facture,
    classifier_cause_anomalie,
    compter_lignes_a_facturer,
    _chemin_archive_facture,
    _doublon_facture,
    pieces_pour_facture,
    regrouper_par_facture,
    FOURNISSEURS_TARIF_BL_DEPUIS_FACTURE,
)
from moteur.referentiel import Referentiel

from conftest import ROOT


def _facture(fichier, numero_facture="360311", numeros_commande=None, numeros_bl=None,
             date_facture="15/07/2026", lignes=None, type_document="FACTURE"):
    return Facture(
        fournisseur="109 DISTRIBUTION", fichier=fichier, numero_facture=numero_facture,
        date_facture=date_facture, numeros_commande=numeros_commande or [],
        numeros_bl=numeros_bl or [], lignes=lignes or [], type_document=type_document,
    )


def _ligne_facture(**kwargs):
    # montant_ht par défaut = qté × PU (comme un montant IMPRIMÉ cohérent,
    # le cas réel chez 109 Distribution — voir CLAUDE.md) : les tests qui ne
    # s'intéressent pas spécifiquement au recalcul n'ont pas à y penser.
    defaut = dict(reference_fournisseur="REF1", designation="", quantite_facturee=10.0, prix_unitaire_ht=1.0)
    defaut.update(kwargs)
    defaut.setdefault("montant_ht", defaut["quantite_facturee"] * (defaut["prix_unitaire_ht"] or 0))
    return LigneFacture(**defaut)


def _ligne_suivi_facture(ligne_excel, **kwargs):
    defaut = dict(
        reference="REF1", designation="", qte_commandee=10.0, qte_livree=10.0,
        tarif_bl=1.0, tarif_convenu=None,
    )
    defaut.update(kwargs)
    return LigneSuiviFacture(ligne_excel=ligne_excel, **defaut)


def _classeur_avec_colonnes_facture(chemin, lignes=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append([
        "Référence", "Désignation", "Qté commandée", "N° de commande", "Fournisseur",
        "Qté livrée", "Tarif BL", "Tarif convenu",
        *COLONNES_FACTURE_CALCULEES,  # N° facture, Date facture, Qté facturée, PU facturé, Montant facturé HT
    ])
    for ligne in lignes:
        ws.append(ligne)
    wb.save(chemin)
    installer_feuille_pieces(chemin, chemin.parent / "backups")  # P1 : feuille Pièces


def _classeur_sans_colonnes_facture(chemin, lignes=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append([
        "Référence", "Désignation", "Qté commandée", "N° de commande", "Fournisseur",
        "Qté livrée", "Tarif BL", "Tarif convenu",
    ])
    for ligne in lignes:
        ws.append(ligne)
    wb.save(chemin)


# --- regroupement / logique pure --------------------------------------------


def test_regrouper_par_facture_separe_les_statuts():

    f1 = _facture("f1.pdf")
    f2 = _facture("f2.pdf")

    c_sur = CorrespondanceFacture(_ligne_facture(), _ligne_suivi_facture(5), StatutFacture.SUR)
    c_inconnu = CorrespondanceFacture(_ligne_facture(), None, StatutFacture.INCONNU, ["ambigu"])

    rapport = RapportRapprochementFacture(surs=[(f1, c_sur)], inconnus=[(f2, c_inconnu)])

    groupes = regrouper_par_facture(rapport)

    assert set(groupes) == {id(f1), id(f2)}
    assert groupes[id(f1)]["sur"] == [c_sur]
    assert groupes[id(f2)]["inconnu"] == [c_inconnu]


def test_regrouper_par_facture_range_le_bucket_frais_a_part():
    f = _facture("f1.pdf")
    c_frais = CorrespondanceFacture(
        _ligne_facture(reference_fournisseur="ECO-23"), None, StatutFacture.FRAIS,
        ["Frais connu"], CauseFacture.FRAIS,
    )
    rapport = RapportRapprochementFacture(frais=[(f, c_frais)])

    groupes = regrouper_par_facture(rapport)

    assert groupes[id(f)]["frais"] == [c_frais]


def test_est_resolu_facture_le_bucket_frais_nest_jamais_bloquant():
    """Une ligne "frais" (jamais rapprochée, voir charger_frais_fournisseurs)
    ne doit ni bloquer ni être requise pour considérer la facture résolue —
    ni au numérateur ni au dénominateur de _est_resolu_facture."""

    c_sur = CorrespondanceFacture(_ligne_facture(), _ligne_suivi_facture(5), StatutFacture.SUR)
    c_frais = CorrespondanceFacture(
        _ligne_facture(reference_fournisseur="ECO-23"), None, StatutFacture.FRAIS,
        ["Frais connu"], CauseFacture.FRAIS,
    )
    g = {"sur": [c_sur], "a_confirmer": [], "deja_a_jour": [], "inconnu": [], "frais": [c_frais], "anomalies": []}

    assert _est_resolu_facture(g, {id(c_sur)}) is True


def test_est_resolu_facture_toutes_resolues():
    c = CorrespondanceFacture(_ligne_facture(), _ligne_suivi_facture(5), StatutFacture.SUR)
    g = {"sur": [c], "a_confirmer": [], "deja_a_jour": [], "inconnu": [], "anomalies": []}

    assert _est_resolu_facture(g, {id(c)}) is True
    assert _est_resolu_facture(g, set()) is False


def test_est_resolu_facture_avec_anomalie_jamais_resolue():
    g = {"sur": [], "a_confirmer": [], "deja_a_jour": [], "inconnu": [], "anomalies": ["commande introuvable"]}

    assert _est_resolu_facture(g, set()) is False


def test_pieces_pour_facture_construit_une_ligne_facture_montant_imprime():
    """P1 : une ligne Pièces de type Facture par correspondance écrite —
    « Montant HT » reprend le montant IMPRIMÉ (LigneFacture.montant_ht),
    jamais recalculé quand il est disponible ; Chantier/Référence Suivi
    copiés de la ligne Commandes ; Fichier = chemin d'archive."""

    f = _facture("f1.pdf", numero_facture="360311", date_facture="15/07/2026")
    lf = _ligne_facture(quantite_facturee=200.0, prix_unitaire_ht=0.6, montant_ht=120.0)
    lf.numero_commande = "129.034"
    lf.numero_bl = "OBL1"
    ls = _ligne_suivi_facture(5, chantier="Chantier Z", sous_chantier="SC9")
    c = CorrespondanceFacture(lf, ls, StatutFacture.SUR)

    pieces, ecritures, montants_recalcules, tarif_bl_ecrit = pieces_pour_facture([(f, c)], "Traités")

    [p] = pieces
    assert p["Type"] == "Facture" and p["Fournisseur"] == "109 Distribution"
    assert p["N° pièce"] == "360311" and p["Date pièce"] == date(2026, 7, 15)
    assert p["N° de commande"] == "129.034" and p["Chantier"] == "Chantier Z" and p["Sous-Chantier"] == "SC9"
    assert p["Référence Suivi"] == "REF1" and p["Référence fournisseur"] == "REF1"
    assert p["Qté"] == 200.0 and p["PU HT"] == 0.6 and p["Montant HT"] == 120.0
    assert p["N° BL lié"] == "OBL1" and p["Mode de rapprochement"] == "Auto"
    assert p["Fichier"].cache == "2026-07-15 - 109 DISTRIBUTION - Facture 360311 - BC 129.034.pdf"
    assert p["ID pièce"] == "109 Distribution|Facture|360311|129.034|REF1|OBL1"
    assert p["Commentaire"] is None
    assert ecritures == [] and montants_recalcules == [] and tarif_bl_ecrit == []


def test_pieces_pour_facture_recalcule_montant_absent_et_le_signale():
    """Facture sans montant par ligne : Montant HT = Qté × PU, Commentaire
    « montant recalculé » ET entrée dans montants_recalcules — JAMAIS
    silencieusement (voir CLAUDE.md)."""

    f = _facture("f1.pdf", numero_facture="360311")
    lf = _ligne_facture(reference_fournisseur="REFX", quantite_facturee=10.0, prix_unitaire_ht=2.5, montant_ht=None)
    c = CorrespondanceFacture(lf, _ligne_suivi_facture(5), StatutFacture.SUR)

    pieces, _, montants_recalcules, _ = pieces_pour_facture([(f, c)], "Traités")

    [p] = pieces
    assert p["Montant HT"] == 25.0 and p["Commentaire"] == "montant recalculé"
    [m] = montants_recalcules
    assert m["fichier"] == "f1.pdf" and m["facture"] == "360311" and m["reference"] == "REFX"
    assert m["ligne_excel"] == 5 and m["montant"] == 25.0


def test_pieces_pour_facture_ignore_deja_a_jour_et_modes():
    f = _facture("f1.pdf")
    c_sur = CorrespondanceFacture(_ligne_facture(), _ligne_suivi_facture(5), StatutFacture.SUR)
    c_deja = CorrespondanceFacture(_ligne_facture(), _ligne_suivi_facture(6), StatutFacture.DEJA_A_JOUR)
    c_confirme = CorrespondanceFacture(_ligne_facture(), _ligne_suivi_facture(7), StatutFacture.A_CONFIRMER, ["qté"], CauseFacture.QTE_PARTIELLE)
    c_repli = CorrespondanceFacture(_ligne_facture(), _ligne_suivi_facture(8), StatutFacture.A_CONFIRMER, ["proche"], CauseFacture.CLE_PARTIELLE)

    pieces, _, _, _ = pieces_pour_facture([(f, c_sur), (f, c_deja), (f, c_confirme), (f, c_repli)], "Traités")

    assert [p["Mode de rapprochement"] for p in pieces] == ["Auto", "Confirmé", "Équivalence"]


def test_pieces_pour_facture_ajoute_les_frais_des_factures_ecrites():
    """Les frais connus (StatutFacture.FRAIS) de la facture écrite deviennent
    des lignes Pièces de type Frais (sans Référence Suivi, chantier de la
    commande) ; ceux d'une facture NON écrite dans ce passage attendent."""

    f = _facture("f1.pdf", numero_facture="6107800")
    f.fournisseur = "COREDIME"
    lf = _ligne_facture(montant_ht=35.0)
    lf.numero_commande = "102.061"
    c = CorrespondanceFacture(lf, _ligne_suivi_facture(5, chantier="Chantier Q"), StatutFacture.SUR)
    lf_frais = _ligne_facture(reference_fournisseur="ECO-23", designation="Éco-participation", quantite_facturee=1.0, prix_unitaire_ht=0.8, montant_ht=0.8)
    lf_frais.numero_commande = "102.061"
    c_frais = CorrespondanceFacture(lf_frais, None, StatutFacture.FRAIS, ["Frais connu (Éco-part)"], CauseFacture.FRAIS)
    f2 = _facture("f2.pdf", numero_facture="9999")
    c_frais2 = CorrespondanceFacture(_ligne_facture(reference_fournisseur="PORT"), None, StatutFacture.FRAIS, [], CauseFacture.FRAIS)

    pieces, _, _, _ = pieces_pour_facture([(f, c)], "Traités", frais=[(f, c_frais), (f2, c_frais2)])

    assert [p["Type"] for p in pieces] == ["Facture", "Frais"]
    frais = pieces[1]
    assert frais["Référence Suivi"] is None and frais["Référence fournisseur"] == "ECO-23"
    assert frais["Chantier"] == "Chantier Q" and frais["N° de commande"] == "102.061"
    assert frais["Montant HT"] == 0.8 and frais["Commentaire"] == "Frais connu (Éco-part)"
    assert frais["ID pièce"] == "COREDIME|Frais|6107800|102.061|F:ECO-23|"


# --- exception Tarif BL depuis facture (liste blanche, session F4) -------


def test_pieces_pour_facture_ecrit_tarif_bl_pour_fournisseur_en_liste_blanche():
    """COREDIME : ses BL n'affichent jamais de prix (voir
    moteur.fournisseurs.coredime) — Tarif BL doit être renseigné depuis la
    facture quand il est encore vide, sinon "Statut commande"/⚠️
    Surfacturation (formule Excel qui LIT Tarif BL directement) resterait
    aveugle pour ce fournisseur. Proposition validée par l'acheteur en une
    phrase (voir CLAUDE.md, cadrage F4). Seule colonne de Commandes encore
    écrite par ce flux depuis P1."""

    assert "COREDIME" in FOURNISSEURS_TARIF_BL_DEPUIS_FACTURE

    f = _facture("f1.pdf", numero_facture="6107293")
    f.fournisseur = "COREDIME"
    lf = _ligne_facture(quantite_facturee=500.0, prix_unitaire_ht=0.35, montant_ht=175.0)
    ls = _ligne_suivi_facture(5, tarif_bl=None)
    c = CorrespondanceFacture(lf, ls, StatutFacture.SUR)

    pieces, ecritures, _, tarif_bl_ecrit = pieces_pour_facture([(f, c)], "Traités")

    assert len(pieces) == 1
    [e] = ecritures
    assert (e.ligne, e.colonne, e.valeur) == (5, "Tarif BL", 0.35)
    [t] = tarif_bl_ecrit
    assert t["fichier"] == "f1.pdf" and t["reference"] == lf.reference_fournisseur
    assert t["ligne_excel"] == 5 and t["tarif_bl"] == 0.35


def test_pieces_pour_facture_necrase_jamais_un_tarif_bl_deja_present():
    f = _facture("f1.pdf", numero_facture="6107293")
    f.fournisseur = "COREDIME"
    lf = _ligne_facture(quantite_facturee=500.0, prix_unitaire_ht=0.35, montant_ht=175.0)
    c = CorrespondanceFacture(lf, _ligne_suivi_facture(5, tarif_bl=0.33), StatutFacture.SUR)

    _, ecritures, _, tarif_bl_ecrit = pieces_pour_facture([(f, c)], "Traités")

    assert ecritures == [] and tarif_bl_ecrit == []


def test_pieces_pour_facture_najoute_pas_tarif_bl_hors_liste_blanche():
    f = _facture("f1.pdf")  # 109 DISTRIBUTION
    lf = _ligne_facture(quantite_facturee=500.0, prix_unitaire_ht=0.35, montant_ht=175.0)
    c = CorrespondanceFacture(lf, _ligne_suivi_facture(5, tarif_bl=None), StatutFacture.SUR)

    _, ecritures, _, tarif_bl_ecrit = pieces_pour_facture([(f, c)], "Traités")

    assert ecritures == [] and tarif_bl_ecrit == []


# --- garde-fou doublon de facture (niveau document, P1) -------------------


def _pieces_existantes(fichier_libelle, numero="360311", fournisseur="109 Distribution"):
    return IndexPieces([{
        "ID pièce": "x", "Type": "Facture", "Fournisseur": fournisseur, "N° pièce": numero,
        "N° de commande": "129.034", "Référence Suivi": "REF1", "Fichier": fichier_libelle,
    }])


def test_doublon_facture_meme_numero_autre_document():
    f = _facture("Facture_360311.pdf", numero_facture="360311", date_facture="15/07/2026")
    autre = _pieces_existantes("2026-06-01 - 109 DISTRIBUTION - Facture 360311 - BC 129.034.pdf")
    raison = _doublon_facture(f, autre)
    assert raison and "DOUBLON" in raison and "360311" in raison
    assert pipeline_facture.classifier_cause_anomalie(raison) is CauseFacture.DOUBLON_FACTURE


def test_doublon_facture_meme_document_redepose_nest_pas_un_doublon():
    f = _facture("Facture_360311.pdf", numero_facture="360311", date_facture="15/07/2026")
    meme = _pieces_existantes("2026-07-15 - 109 DISTRIBUTION - Facture 360311 - BC 129.034.pdf")
    assert _doublon_facture(f, meme) is None
    autre_commande = _pieces_existantes("2026-07-15 - 109 DISTRIBUTION - Facture 360311 - BC 130.001.pdf")
    assert _doublon_facture(f, autre_commande) is None  # facture multi-BC : même document
    assert _doublon_facture(f, IndexPieces([])) is None
    assert _doublon_facture(f, IndexPieces(None)) is None  # pas de feuille Pièces : pas de garde-fou possible


def test_chemin_archive_facture_est_le_nom_de_larchive():
    f = _facture("f1.pdf", numero_facture="360311", date_facture="15/07/2026")
    chemin = _chemin_archive_facture(f, "129.034", "Traités")
    assert chemin.as_posix().endswith("Traités/129.034/2026-07-15 - 109 DISTRIBUTION - Facture 360311 - BC 129.034.pdf")


def test_verifier_total_ht_facture_aucun_ecart_rien_a_signaler():
    lf = _ligne_facture(quantite_facturee=10.0, prix_unitaire_ht=5.0, montant_ht=50.0)
    f = _facture("f1.pdf", lignes=[lf])
    f.total_ht_affiche = 50.0

    assert _verifier_total_ht_facture(f) is None


def test_verifier_total_ht_facture_total_absent_rien_a_signaler():
    f = _facture("f1.pdf", lignes=[_ligne_facture()])
    f.total_ht_affiche = None

    assert _verifier_total_ht_facture(f) is None


def test_verifier_total_ht_facture_zero_ligne_signale_le_montant_manquant():
    """Cas réel qui a motivé ce contrôle (session S0,
    facture_coredime_6108846_remise_double_x3.pdf) : 0 ligne extraite pour
    un Total HT affiché non nul — sans ce contrôle, rien ne signalait les
    196,92€ manquants, la facture pouvait être archivée en silence."""

    f = _facture("f1.pdf", lignes=[])
    f.total_ht_affiche = 196.92

    raison = _verifier_total_ht_facture(f)

    assert raison is not None
    assert "196.92" in raison
    assert "0.00" in raison


def test_verifier_total_ht_facture_petit_ecart_sous_le_seuil_tolere():
    lf = _ligne_facture(quantite_facturee=1.0, prix_unitaire_ht=10.0, montant_ht=10.0)
    f = _facture("f1.pdf", lignes=[lf])
    f.total_ht_affiche = 10.01  # 1 centime, sous SEUIL_TOTAL_ECART_FACTURE

    assert _verifier_total_ht_facture(f) is None


# --- classification des causes (session S0, correction 1e) -----------------


def test_classifier_cause_anomalie_reconnait_les_motifs_deja_en_clair():
    assert classifier_cause_anomalie("Facture d'AVOIR — jamais rapprochée automatiquement") is CauseFacture.AVOIR
    assert classifier_cause_anomalie(
        "BL 123 : commande passée sur un bon manuel (« BC 241766 »)"
    ) is CauseFacture.BDC_MANUEL_24X
    assert classifier_cause_anomalie("Écart de Total HT : +5.00€ (...)") is CauseFacture.TOTAL_ECART
    assert classifier_cause_anomalie("Commande 123.089 introuvable dans le Suivi pour « X »") is CauseFacture.COMMANDE_ABSENTE
    assert classifier_cause_anomalie("BL 1 : n° de commande introuvable (...)") is CauseFacture.COMMANDE_ABSENTE
    assert classifier_cause_anomalie("Fournisseur non reconnu") is CauseFacture.FOURNISSEUR_INCONNU
    assert classifier_cause_anomalie("Fournisseur RAVATE reconnu mais pas encore de parser facture") is CauseFacture.PARSER_ABSENT
    assert classifier_cause_anomalie("Aucune ligne extraite (facture 123)") is CauseFacture.ZERO_LIGNE
    assert classifier_cause_anomalie("PDF illisible (corrompu)") is CauseFacture.ANNEXE_SANS_TEXTE
    assert classifier_cause_anomalie("Erreur de lecture (X)") is CauseFacture.ANNEXE_SANS_TEXTE


def test_classifier_cause_anomalie_aucun_motif_reconnu_retourne_none():
    assert classifier_cause_anomalie("Suivi commandes introuvable") is None


# --- résolution de commande par bloc de BL ----------------------------------


def test_resoudre_commandes_facture_entete_direct_applique_a_tous_les_blocs():
    """En-tête clair (N°Réf.Client au format Suivi) -> appliqué à TOUS les
    blocs de BL de la facture, même s'il y en a plusieurs (cas de loin le
    plus courant sur le lot de cadrage réel, voir CLAUDE.md)."""

    f = _facture(
        "f1.pdf", numeros_commande=["129.034"],
        lignes=[
            _ligne_facture(reference_fournisseur="A", numero_bl="BL1"),
            _ligne_facture(reference_fournisseur="B", numero_bl="BL2"),
        ],
    )

    resolutions = _resoudre_commandes_facture(f, fichier_suivi=None)  # jamais lu si l'en-tête suffit

    assert resolutions["BL1"] == ("129.034", False, None)
    assert resolutions["BL2"] == ("129.034", False, None)


def test_resoudre_commandes_facture_deduction_par_contenu_en_repli(tmp_path):
    """N°Réf.Client vide/non exploitable -> repli sur la déduction par
    contenu (même mécanisme que côté BL, voir
    matching.deduire_commande_par_contenu) : au moins 2 références/quantités
    du bloc concordent EXACTEMENT avec une seule commande du Suivi."""

    # Références avec un cœur NUMÉRIQUE (>= 4 chiffres) — deduire_commande_
    # par_contenu s'appuie sur moteur.base.coeur_numerique, qui ignore toute
    # référence sans chiffres exploitables (voir son seuil de fiabilité).
    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi, lignes=[
        ["REF4501", "", 5, "129.099", "109 DISTRIBUTION", 5, 1.0, None, None, None, None, None, None],
        ["REF7802", "", 3, "129.099", "109 DISTRIBUTION", 3, 1.0, None, None, None, None, None, None],
    ])

    f = _facture(
        "f1.pdf", numeros_commande=[],  # en-tête illisible/non exploitable
        lignes=[
            _ligne_facture(reference_fournisseur="REF4501", quantite_facturee=5.0, numero_bl="BL1"),
            _ligne_facture(reference_fournisseur="REF7802", quantite_facturee=3.0, numero_bl="BL1"),
        ],
    )

    resolutions = _resoudre_commandes_facture(f, chemin_suivi)

    numero_commande, deduit, raison = resolutions["BL1"]
    assert numero_commande == "129.099"
    assert deduit is True
    assert "déduit du contenu" in raison


def test_resoudre_commandes_facture_rien_de_deductible_retourne_none(tmp_path):

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi, lignes=[
        ["AUTREREF", "", 5, "999.999", "109 DISTRIBUTION", 5, 1.0, None, None, None, None, None, None],
    ])

    f = _facture(
        "f1.pdf", numeros_commande=[],
        lignes=[_ligne_facture(reference_fournisseur="INCONNUE", numero_bl="BL1")],
    )

    resolutions = _resoudre_commandes_facture(f, chemin_suivi)

    assert resolutions["BL1"] == (None, False, None)


# --- archivage (copie multi-commande, jamais un déplacement) ---------------


def test_archiver_facture_copie_dans_chaque_commande_concernee_et_supprime_la_source(tmp_path):
    """Une facture peut couvrir PLUSIEURS commandes (cas réel,
    Facture_365533.pdf) — copiée dans CHAQUE dossier de commande concerné,
    jamais découpée (contrairement à un BL Cominter multi-BL, voir bandeau
    du module) ; la source n'est supprimée qu'une fois toutes les copies
    faites."""

    source = tmp_path / "a_traiter" / "Factures" / "Facture_365533.pdf"
    source.parent.mkdir(parents=True)
    source.write_bytes(b"%PDF-1.4 factice")

    traites = tmp_path / "a_traiter" / "BL" / "Traités"
    f = _facture("Facture_365533.pdf", numero_facture="365533", date_facture="17/08/2026")

    cibles = archiver_facture(source, f, ["132.008", "132.033"], traites)

    assert not source.exists()
    assert len(cibles) == 2
    assert {c.parent.name for c in cibles} == {"132.008", "132.033"}
    for c in cibles:
        assert c.exists()
        assert "365533" in c.name
        assert "109 DISTRIBUTION" in c.name


def test_archiver_facture_evite_lecrasement(tmp_path):

    traites = tmp_path / "Traités"
    f = _facture("f.pdf", numero_facture="1", date_facture="01/01/2026")

    source1 = tmp_path / "f1.pdf"
    source1.write_bytes(b"1")
    [cible1] = archiver_facture(source1, f, ["C1"], traites)

    source2 = tmp_path / "f2.pdf"
    source2.write_bytes(b"2")
    [cible2] = archiver_facture(source2, f, ["C1"], traites)

    assert cible1 != cible2
    assert cible1.exists() and cible2.exists()


# --- résorption --------------------------------------------------------------


def test_compter_lignes_a_facturer(tmp_path):

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi, lignes=[
        ["REF1", "", 10, "C1", "109 DISTRIBUTION", 10, 1.0, None, None, None, None, None, None],  # déjà facturée (Pièces)
        ["REF2", "", 5, "C2", "109 DISTRIBUTION", 5, 1.0, None, None, None, None, None, None],  # livrée, pas facturée
        ["REF3", "", 5, "C3", "109 DISTRIBUTION", 0, None, None, None, None, None, None, None],  # pas livrée du tout
    ])
    # P1 : « déjà facturée » se lit dans la feuille Pièces (source de
    # vérité), pas dans les colonnes de Commandes (formules, cache).
    from moteur.rapprochement.pieces import nouvelle_piece
    ecrire_pieces(chemin_suivi, [nouvelle_piece(
        "Facture", "109 DISTRIBUTION", "F1", date(2026, 7, 15), "C1", None, None, "REF1", "REF1", "", 10, 1.0, 10.0,
    )], tmp_path / "backups")

    r = compter_lignes_a_facturer(chemin_suivi, "109 DISTRIBUTION")

    assert r == {"livrees": 2, "a_facturer": 1, "deja_facturees": 1}


def test_compter_lignes_a_facturer_colonnes_absentes_reste_honnete(tmp_path):
    """Sans les colonnes facture (état actuel du vrai Suivi, voir CLAUDE.md),
    toute ligne livrée ressort "à facturer" — signal honnête, pas un 0
    trompeur qui masquerait leur absence."""

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_sans_colonnes_facture(chemin_suivi, lignes=[
        ["REF1", "", 10, "C1", "109 DISTRIBUTION", 10, 1.0, None],
    ])

    r = compter_lignes_a_facturer(chemin_suivi, "109 DISTRIBUTION")

    assert r == {"livrees": 1, "a_facturer": 1, "deja_facturees": 0}


# --- écriture réelle de bout en bout (classeur synthétique, jamais le vrai) -


def test_appliquer_et_archiver_factures_ecrit_et_archive(tmp_path):
    """Preuve que le mécanisme d'écriture fonctionne de bout en bout — sur
    un classeur SYNTHÉTIQUE qui a les colonnes facture (le vrai Suivi ne
    les a pas encore, voir bandeau du module)."""

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi, lignes=[
        ["REF1", "Article test", 200, "129.034", "109 DISTRIBUTION", 200, 0.6, None, None, None, None, None, None],
    ])

    dossier_a_traiter = tmp_path / "a_traiter" / "Factures"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "f1.pdf").write_bytes(b"1")

    lf = _ligne_facture(reference_fournisseur="REF1", quantite_facturee=200.0, prix_unitaire_ht=0.6)
    lf.numero_commande = "129.034"
    f = _facture("f1.pdf", numero_facture="360311", date_facture="15/07/2026", lignes=[lf])

    ls = _ligne_suivi_facture(2, reference="REF1", qte_commandee=200.0, qte_livree=200.0, tarif_bl=0.6)
    c = CorrespondanceFacture(lf, ls, StatutFacture.SUR)

    rapport = RapportRapprochementFacture(surs=[(f, c)], fichier_suivi=chemin_suivi)

    resume = appliquer_et_archiver_factures(tmp_path, dossier_a_traiter, rapport, rapport.surs)

    assert resume["lignes_ecrites"] == 1
    assert resume["pieces_ecrites"] == 1
    assert resume["sauvegarde"].exists()
    assert resume["montants_recalcules"] == []  # montant_ht imprimé (défaut du helper), rien à recalculer

    # P1 : la ligne de document est dans Pièces, plus dans les 5 colonnes
    # de Commandes (devenues des formules — jamais écrites par l'outil).
    [p] = lire_pieces(chemin_suivi)
    assert p["Type"] == "Facture" and p["N° pièce"] == "360311"
    assert p["N° de commande"] == "129.034" and p["Référence Suivi"] == "REF1"
    assert p["Qté"] == 200.0 and p["PU HT"] == 0.6 and p["Montant HT"] == 120.0
    assert p["Fichier"] == "2026-07-15 - 109 DISTRIBUTION - Facture 360311 - BC 129.034.pdf"
    from openpyxl import load_workbook
    wb = load_workbook(chemin_suivi, data_only=True)
    ws = wb["Commandes"]
    assert ws["I2"].value is None and ws["K2"].value is None  # rien écrit dans les colonnes facture

    [(fichier, cibles)] = resume["factures_archivees"]
    assert fichier == "f1.pdf"
    [cible] = cibles
    assert cible.exists()
    assert cible.parent.name == "129.034"
    assert not (dossier_a_traiter / "f1.pdf").exists()

    assert resume["resorption"]["109 DISTRIBUTION"] is not None
    assert resume["tarif_bl_ecrit_depuis_facture"] == []  # 109 DISTRIBUTION hors liste blanche (voir bandeau)
    assert resume["chemin_rapport"].exists()


def test_appliquer_et_archiver_factures_rapporte_frais_et_causes(tmp_path):
    """Le rapport écrit (resume["causes"] + le fichier texte) reflète les
    lignes "frais" (jamais écrites, jamais bloquantes) et un compte rendu
    chiffré par cause (session S0, corrections 1c/1e) — "Pas de résiduel
    unique" : une ligne à confirmer non résolue doit apparaître dans
    resume["causes"], jamais silencieusement absente."""

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi)

    dossier_a_traiter = tmp_path / "a_traiter" / "Factures"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "f1.pdf").write_bytes(b"1")

    f = _facture("f1.pdf")
    c_frais = CorrespondanceFacture(
        _ligne_facture(reference_fournisseur="ECO-23", montant_ht=0.8), None, StatutFacture.FRAIS,
        ["Frais connu (Éco-participation)"], CauseFacture.FRAIS,
    )
    c_confirmer = CorrespondanceFacture(
        _ligne_facture(reference_fournisseur="REFX"), _ligne_suivi_facture(5), StatutFacture.A_CONFIRMER,
        ["Qté facturée différente"], CauseFacture.QTE_PARTIELLE,
    )

    rapport = RapportRapprochementFacture(
        frais=[(f, c_frais)], a_confirmer=[(f, c_confirmer)], fichier_suivi=chemin_suivi,
    )

    resume = appliquer_et_archiver_factures(tmp_path, dossier_a_traiter, rapport, [])

    assert resume["causes"] == {"frais": 1, "qte_partielle": 1}

    texte_rapport = resume["chemin_rapport"].read_text(encoding="utf-8")
    assert "1 ligne(s) de frais connus" in texte_rapport
    assert "0.80" in texte_rapport
    assert "Répartition par cause" in texte_rapport
    assert "frais : 1" in texte_rapport
    assert "qte_partielle : 1" in texte_rapport


def test_appliquer_et_archiver_factures_bloc_non_resolu_va_en_a_verifier(tmp_path):

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi)

    dossier_a_traiter = tmp_path / "a_traiter" / "Factures"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "f1.pdf").write_bytes(b"1")

    f = _facture("f1.pdf")
    c_inconnu = CorrespondanceFacture(_ligne_facture(), None, StatutFacture.INCONNU, ["référence inconnue"])

    rapport = RapportRapprochementFacture(inconnus=[(f, c_inconnu)], fichier_suivi=chemin_suivi)

    resume = appliquer_et_archiver_factures(tmp_path, dossier_a_traiter, rapport, [])

    assert resume["factures_archivees"] == []
    [(fichier, cible, raisons)] = resume["factures_a_verifier"]
    assert fichier == "f1.pdf"
    assert cible == dossier_a_traiter / "À vérifier" / "f1.pdf"
    assert cible.exists()
    assert "référence inconnue" in raisons[0]


def test_appliquer_et_archiver_factures_sans_parser_reste_en_place():
    """Un fournisseur RECONNU mais sans parser facture n'est pas déplacé
    vers À vérifier/ (ce n'est pas une décision humaine en attente, juste
    un fournisseur pas encore couvert) — laissé tel quel dans
    a_traiter/Factures/, listé dans resume["factures_sans_parser"]."""

    from moteur.rapprochement.pipeline_facture import _est_anomalie_sans_parser
    assert _est_anomalie_sans_parser("Fournisseur RAVATE reconnu mais pas encore de parser facture")
    assert not _est_anomalie_sans_parser("Fournisseur non reconnu")
    assert not _est_anomalie_sans_parser("PDF illisible (corrompu)")


def test_appliquer_et_archiver_factures_sans_parser_vs_vraie_anomalie(tmp_path):
    """Confronte les deux cas dans le MÊME lot : le fichier "sans parser"
    reste en place et est reporté à part ; le fichier avec une vraie
    anomalie de lecture (fournisseur non reconnu) part, lui, vers
    À vérifier/ comme avant — comportement inchangé pour ce cas."""

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi)

    dossier_a_traiter = tmp_path / "a_traiter" / "Factures"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "ravate.pdf").write_bytes(b"1")
    (dossier_a_traiter / "illisible.pdf").write_bytes(b"2")

    rapport = RapportRapprochementFacture(
        anomalies_lecture=[
            ("ravate.pdf", "Fournisseur RAVATE reconnu mais pas encore de parser facture"),
            ("illisible.pdf", "Fournisseur non reconnu"),
        ],
        fichier_suivi=chemin_suivi,
    )

    resume = appliquer_et_archiver_factures(tmp_path, dossier_a_traiter, rapport, [])

    assert resume["factures_sans_parser"] == [
        ("ravate.pdf", "Fournisseur RAVATE reconnu mais pas encore de parser facture"),
    ]
    assert (dossier_a_traiter / "ravate.pdf").exists()  # jamais déplacé

    [(fichier, cible, _)] = resume["factures_a_verifier"]
    assert fichier == "illisible.pdf"
    assert cible == dossier_a_traiter / "À vérifier" / "illisible.pdf"
    assert not (dossier_a_traiter / "illisible.pdf").exists()  # bien déplacé, comportement inchangé


def test_ecriture_echoue_proprement_sans_feuille_pieces(tmp_path):
    """P1 : un classeur sans feuille Pièces (export périmé, poste pas à jour)
    doit échouer proprement -> FeuillePiecesAbsente, message clair, jamais
    un plantage générique ni une écriture partielle ; rien n'est déplacé."""

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_sans_colonnes_facture(chemin_suivi, lignes=[
        ["REF1", "", 10, "C1", "109 DISTRIBUTION", 10, 1.0, None],
    ])
    assert not feuille_pieces_presente(chemin_suivi)

    dossier_a_traiter = tmp_path / "a_traiter" / "Factures"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "f1.pdf").write_bytes(b"1")

    lf = _ligne_facture(reference_fournisseur="REF1", quantite_facturee=10.0, prix_unitaire_ht=1.0)
    f = _facture("f1.pdf", lignes=[lf])
    ls = _ligne_suivi_facture(2, reference="REF1", qte_livree=10.0, tarif_bl=1.0)
    c = CorrespondanceFacture(lf, ls, StatutFacture.SUR)

    rapport = RapportRapprochementFacture(surs=[(f, c)], fichier_suivi=chemin_suivi)

    with pytest.raises(FeuillePiecesAbsente):
        appliquer_et_archiver_factures(tmp_path, dossier_a_traiter, rapport, rapport.surs)

    assert (dossier_a_traiter / "f1.pdf").exists()


def test_appliquer_et_archiver_factures_est_idempotent_par_id_piece(tmp_path):
    """Redéposer la même facture : les lignes Pièces ne sont jamais
    réécrites (ID pièce déjà présent), comptées dans pieces_ignorees."""

    chemin_suivi = tmp_path / "suivi.xlsx"
    _classeur_avec_colonnes_facture(chemin_suivi, lignes=[
        ["REF1", "Article test", 200, "129.034", "109 DISTRIBUTION", 200, 0.6, None, None, None, None, None, None],
    ])
    dossier_a_traiter = tmp_path / "a_traiter" / "Factures"
    dossier_a_traiter.mkdir(parents=True)

    def _passage():
        (dossier_a_traiter / "f1.pdf").write_bytes(b"1")
        lf = _ligne_facture(reference_fournisseur="REF1", quantite_facturee=200.0, prix_unitaire_ht=0.6)
        lf.numero_commande = "129.034"
        f = _facture("f1.pdf", numero_facture="360311", date_facture="15/07/2026", lignes=[lf])
        c = CorrespondanceFacture(lf, _ligne_suivi_facture(2, reference="REF1", qte_commandee=200.0, qte_livree=200.0, tarif_bl=0.6), StatutFacture.SUR)
        rapport = RapportRapprochementFacture(surs=[(f, c)], fichier_suivi=chemin_suivi)
        return appliquer_et_archiver_factures(tmp_path, dossier_a_traiter, rapport, rapport.surs)

    assert _passage()["pieces_ecrites"] == 1
    resume = _passage()
    assert resume["pieces_ecrites"] == 0
    assert resume["pieces_ignorees"] == ["109 Distribution|Facture|360311|129.034|REF1|"]
    assert len(lire_pieces(chemin_suivi)) == 1


# --- écriture réelle sur une COPIE du vrai Suivi commandes vivant ----------


@pytest.mark.skipif(
    trouver_fichier_suivi_vivant(ROOT) is None,
    reason="Classeur Suivi commandes VIVANT introuvable depuis ce poste",
)
def test_pieces_pour_facture_sur_le_vrai_suivi_vivant(tmp_path):
    """Preuve au niveau PIPELINE sur une COPIE du vrai classeur vivant
    (jamais l'original) : la ligne Pièces construite par
    pieces_pour_facture() s'écrit via ecrire_pieces() en ne modifiant que
    la feuille Pièces, sa définition de tableau (et styles.xml si un style
    manquait) — tout le reste du zip (Commandes, Dashboard, Analyses,
    Calculs, validations, sharedStrings...) reste identique octet pour
    octet. La feuille Pièces est installée sur la copie si le vivant ne l'a
    pas encore (avant l'étape 7 de P1)."""

    fichier_reel = trouver_fichier_suivi_vivant(ROOT)
    copie = tmp_path / fichier_reel.name
    shutil.copy2(fichier_reel, copie)
    if not feuille_pieces_presente(copie):
        installer_feuille_pieces(copie, tmp_path / "backups")

    with zipfile.ZipFile(copie) as z:
        contenu_avant = {n: z.read(n) for n in z.namelist()}

    f = _facture("f1.pdf", numero_facture="TEST-360999", date_facture="01/09/2026")
    lf = _ligne_facture(reference_fournisseur="TESTREF", quantite_facturee=7.0, prix_unitaire_ht=12.5, montant_ht=87.5)
    lf.numero_commande = "TEST.001"
    ls = _ligne_suivi_facture(2, reference="TESTREF", qte_livree=7.0, tarif_bl=12.5, chantier="Test")
    c = CorrespondanceFacture(lf, ls, StatutFacture.SUR)

    pieces, ecritures, montants_recalcules, _ = pieces_pour_facture([(f, c)], tmp_path / "Traités")
    assert montants_recalcules == [] and ecritures == []
    resultat = ecrire_pieces(copie, pieces, tmp_path / "backups")
    assert resultat["ajoutees"] == 1

    with zipfile.ZipFile(copie) as z:
        contenu_apres = {n: z.read(n) for n in z.namelist()}
    modifiees = {n for n in contenu_avant if contenu_avant[n] != contenu_apres[n]}
    assert "xl/worksheets/sheet1.xml" not in modifiees  # Commandes jamais touchée
    assert all(n.startswith("xl/worksheets/sheet") or n.startswith("xl/tables/table") or n == "xl/styles.xml" for n in modifiees)
    assert len([n for n in modifiees if n.startswith("xl/worksheets/")]) == 1

    derniere = lire_pieces(copie)[-1]
    assert derniere["N° pièce"] == "TEST-360999" and derniere["Qté"] == 7.0 and derniere["Montant HT"] == 87.5
    assert derniere["Chantier"] == "Test"


# --- "Substitutions probables" : feuille dédiée + apprentissage (étape 1) --


def _rapport_avec_substitution(fichier="f1.pdf", commande="M3.14.342",
                                ref_facturee="LEG06620", ref_suivi="5120"):
    f = _facture(fichier, numero_facture="6108234", numeros_commande=[commande])
    lf = _ligne_facture(reference_fournisseur=ref_facturee, designation="ICTA 3422 20 ATF", quantite_facturee=100.0, prix_unitaire_ht=0.37)
    lf.numero_commande = commande
    ls = _ligne_suivi_facture(5353, reference=ref_suivi, designation="ICT 20 BLEU TURBO G-ROUL 100M", qte_livree=100.0, tarif_bl=0.37)
    c = CorrespondanceFacture(lf, ls, StatutFacture.A_CONFIRMER, ["Substitution probable"], CauseFacture.SUBSTITUTION_PROBABLE)

    rapport = RapportRapprochementFacture()
    rapport.a_confirmer.append((f, c))
    return rapport


def test_ecrire_substitutions_probables_cree_la_feuille(tmp_path):
    dossier = tmp_path / "referentiel"
    dossier.mkdir()
    rapport = _rapport_avec_substitution()

    _ecrire_substitutions_probables(dossier, "A_confirmer_Facture.xlsx", rapport)

    fichier = dossier / "A_confirmer_Facture.xlsx"
    assert fichier.exists()

    wb = load_workbook(fichier)
    assert NOM_FEUILLE_SUBSTITUTIONS in wb.sheetnames
    ws = wb[NOM_FEUILLE_SUBSTITUTIONS]
    lignes = list(ws.iter_rows(values_only=True))
    assert len(lignes) == 2  # en-tête + 1 ligne
    entetes = lignes[0]
    ligne = dict(zip(entetes, lignes[1]))
    assert ligne["Référence facturée"] == "LEG06620"
    assert ligne["Référence Suivi (connue)"] == "5120"
    assert ligne["Qté facturée"] == 100.0
    assert ligne["Commande"] == "M3.14.342"


def test_ecrire_substitutions_probables_ne_touche_pas_la_feuille_primaire(tmp_path):
    """La feuille "À confirmer" déjà écrite par referentiel.ecrire_a_confirmer()
    (round-trip openpyxl) doit rester intacte quand on ajoute la feuille
    "Substitutions probables" au même fichier."""

    dossier = tmp_path / "referentiel"
    dossier.mkdir()
    fichier = dossier / "A_confirmer_Facture.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "À confirmer"
    ws.append(["Fournisseur", "Décision"])
    ws.append(["COREDIME", ""])
    wb.save(fichier)

    _ecrire_substitutions_probables(dossier, "A_confirmer_Facture.xlsx", _rapport_avec_substitution())

    wb2 = load_workbook(fichier)
    assert set(wb2.sheetnames) == {"À confirmer", NOM_FEUILLE_SUBSTITUTIONS}
    ws2 = wb2["À confirmer"]
    # Round-trip openpyxl : une chaîne vide écrite se relit None (quirk
    # déjà connu), pas une régression de ce test.
    assert [c.value for c in ws2[2]] == ["COREDIME", None]


def test_ecrire_substitutions_probables_supprime_la_feuille_si_plus_aucune(tmp_path):
    dossier = tmp_path / "referentiel"
    dossier.mkdir()
    fichier = dossier / "A_confirmer_Facture.xlsx"

    _ecrire_substitutions_probables(dossier, "A_confirmer_Facture.xlsx", _rapport_avec_substitution())
    assert fichier.exists()

    rapport_vide = RapportRapprochementFacture()
    _ecrire_substitutions_probables(dossier, "A_confirmer_Facture.xlsx", rapport_vide)

    # Plus aucune proposition et rien d'autre dans le classeur -> fichier supprimé.
    assert not fichier.exists()


def test_ecrire_substitutions_probables_rien_a_ecrire_et_fichier_absent(tmp_path):
    dossier = tmp_path / "referentiel"
    dossier.mkdir()
    _ecrire_substitutions_probables(dossier, "A_confirmer_Facture.xlsx", RapportRapprochementFacture())
    assert not (dossier / "A_confirmer_Facture.xlsx").exists()


def test_appliquer_confirmations_substitutions_apprend_dans_equivalences_bl(tmp_path):
    dossier = tmp_path / "referentiel"
    dossier.mkdir()
    rapport = _rapport_avec_substitution()
    _ecrire_substitutions_probables(dossier, "A_confirmer_Facture.xlsx", rapport)

    # L'acheteur coche "OUI" dans le fichier régénéré.
    fichier = dossier / "A_confirmer_Facture.xlsx"
    wb = load_workbook(fichier)
    ws = wb[NOM_FEUILLE_SUBSTITUTIONS]
    entetes = [c.value for c in ws[1]]
    ws.cell(row=2, column=entetes.index("Décision") + 1, value="OUI")
    wb.save(fichier)

    ref = Referentiel(tmp_path / "moteur")
    try:
        n = _appliquer_confirmations_substitutions(dossier, "A_confirmer_Facture.xlsx", ref)
        assert n == 1
    finally:
        ref.fermer()

    contenu = (dossier / "equivalences_bl.csv").read_text(encoding="utf-8-sig")
    assert "5120;LEG06620" in contenu
    assert "6108234" in contenu  # n° de facture tracé dans la Note
    assert "M3.14.342" in contenu  # n° de commande tracé dans la Note


def test_appliquer_confirmations_substitutions_ignore_sans_decision(tmp_path):
    dossier = tmp_path / "referentiel"
    dossier.mkdir()
    _ecrire_substitutions_probables(dossier, "A_confirmer_Facture.xlsx", _rapport_avec_substitution())

    ref = Referentiel(tmp_path / "moteur")
    try:
        assert _appliquer_confirmations_substitutions(dossier, "A_confirmer_Facture.xlsx", ref) == 0
    finally:
        ref.fermer()

    assert not (dossier / "equivalences_bl.csv").exists()


def test_appliquer_confirmations_substitutions_fichier_absent(tmp_path):
    dossier = tmp_path / "referentiel"
    dossier.mkdir()
    ref = Referentiel(tmp_path / "moteur")
    try:
        assert _appliquer_confirmations_substitutions(dossier, "A_confirmer_Facture.xlsx", ref) == 0
    finally:
        ref.fermer()
