"""
Bout en bout : Besoin.txt -> devis PDF réels -> Comparatif -> décision de
l'acheteur -> Panier.

Utilise 3 VRAIS PDF de devis (tests/fixtures/ravate.pdf, coredime.pdf,
dem.pdf — règle d'or du projet, jamais de PDF synthétique) qui partagent
un même article sous 3 écritures différentes ("CAP856574" / "CAPCAP856574"
/ "CAP856574") : de quoi exercer un vrai rapprochement 3 fournisseurs sans
dépendre de la BDD achats (cœur numérique commun "856574").
"""

import shutil

import pytest
from openpyxl import load_workbook

from moteur.pipeline import generer_comparatif
from moteur.panier import (
    generer_panier, MAPPING_FOURNISSEURS, nom_fournisseur_canonique,
    _resoudre_chantier, trouver_fichier_suivi,
)
from moteur.detecteur import _FOURNISSEURS

from conftest import ROOT, FIXTURES


def _preparer_projet(tmp_path):

    dossier_devis = tmp_path / "devis"
    dossier_devis.mkdir()
    for f in ("ravate.pdf", "coredime.pdf", "dem.pdf"):
        shutil.copy(FIXTURES / f, dossier_devis / f)

    dossier_referentiel = tmp_path / "referentiel"
    dossier_referentiel.mkdir()
    (dossier_referentiel / "composes.csv").write_text(
        "Cle_besoin;Membre;Quantite\n"
        "KITSIBTOP;SIBP02000;1\n"
        "KITSIBTOP;SIBP03200;1\n",
        encoding="utf-8",
    )

    fichier_besoin = tmp_path / "Besoin Test.txt"
    fichier_besoin.write_text(
        "Désignation\tRéférence\tQté\n"
        "Manchon prédalle HT65 axe74;856574;50\n"
        "Kit boîtier + couvercle SIBTOP;KITSIBTOP;3\n"
        "Fourreau grade 3 PTT;TUBFFPTTGRD3;4\n"
        "Pot simple air'métic;EUR52097;20\n"
        "Article totalement introuvable chez qui que ce soit;999999NOPE;10\n",
        encoding="utf-8",
    )

    return fichier_besoin, dossier_devis


def _valeurs_par_entete(ws):
    entetes = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    lignes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        if row[0] == "TOTAL (si tout pris chez ce fournisseur)":
            continue
        lignes.append(dict(zip(entetes, row)))
    return lignes


def test_mapping_fournisseurs_couvre_tous_les_fournisseurs_connus():
    # Régression : tout nouveau fournisseur ajouté à detecteur.py doit
    # aussi être mappé ici, sinon le Panier recopie un nom qui ne
    # correspond à rien dans le Suivi commandes sans que personne s'en
    # aperçoive avant le collage.
    noms = {nom for nom, _ in _FOURNISSEURS if nom != "LEGRAND"}
    manquants = noms - set(MAPPING_FOURNISSEURS)
    assert not manquants, f"Fournisseur(s) non mappé(s) pour le Panier : {manquants}"


def test_resoudre_chantier_ambigu_reste_tel_quel():
    # Deux chantiers "Doujani" existent réellement dans le Suivi (137 ZAC
    # Doujani / 155 Ecole Doujani) : jamais de code deviné au hasard.
    chantiers = ["137 ZAC Doujani", "155 Ecole Doujani", "070 Singuizou"]
    nom, ok = _resoudre_chantier("Doujani", chantiers)
    assert not ok
    assert nom == "Doujani"

    nom, ok = _resoudre_chantier("Singuizou", chantiers)
    assert ok
    assert nom == "070 Singuizou"


def test_nom_fournisseur_canonique():
    assert nom_fournisseur_canonique("RAVATE") == ("RAVATE ELEC", True)
    assert nom_fournisseur_canonique("SAGEES") == ("Sagees", True)
    assert nom_fournisseur_canonique("UN FOURNISSEUR INEDIT") == ("UN FOURNISSEUR INEDIT", False)


def test_pipeline_complet_besoin_vers_panier(tmp_path):

    fichier_besoin, dossier_devis = _preparer_projet(tmp_path)

    fichier_comparatif = generer_comparatif(tmp_path, fichier_besoin, dossier_devis)
    assert fichier_comparatif is not None

    wb = load_workbook(fichier_comparatif)
    ws = wb["Comparatif"]
    lignes = _valeurs_par_entete(ws)

    # 5 lignes de besoin -> le composé KITSIBTOP se décompose en 2 lignes :
    # 4 lignes "normales" + 2 lignes du composé = 6, moins la ligne d'en-tête
    # fantôme qu'il ne doit PLUS y avoir (régression corrigée cette session).
    assert len(lignes) == 6

    # La ligne sans offre doit être remontée EN TÊTE (jamais perdue en silence).
    assert lignes[0]["Référence"] == "999999NOPE"
    assert lignes[0]["Meilleur prix"] == "AUCUNE OFFRE"

    ligne_manchon = next(l for l in lignes if l["Référence"] == "856574")
    # 3 fournisseurs sur cette ligne, mieux-disant = COREDIME (1.2 < 1.31 < 1.41)
    assert ligne_manchon["Fournisseur retenu"] == "COREDIME"
    for f in ("COREDIME", "DEM", "RAVATE"):
        assert ligne_manchon[f] is not None

    # Décision de l'acheteur : override du mieux-disant -> DEM (pas le moins
    # cher). C'est exactement ce que doit pouvoir faire l'acheteur dans Excel.
    for row in ws.iter_rows(min_row=2):
        cell_ref = next((c for c in row if c.value == "856574"), None)
        if cell_ref is not None:
            row_idx = cell_ref.row
            entetes = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col_retenu = entetes.index("Fournisseur retenu") + 1
            ws.cell(row=row_idx, column=col_retenu).value = "DEM"
            break
    wb.save(fichier_comparatif)

    fichier_panier = generer_panier(tmp_path, fichier_besoin, dossier_devis, fichier_comparatif)
    assert fichier_panier is not None

    wb2 = load_workbook(fichier_panier)
    ws_panier = wb2["Panier"]
    lignes_panier = _valeurs_par_entete(ws_panier)

    # Colonnes de repli (aucun export du Suivi commandes dans ce dossier
    # de test isolé) : mêmes noms, dans le même ordre.
    entetes_panier = [c.value for c in next(ws_panier.iter_rows(min_row=1, max_row=1))]
    assert entetes_panier == [
        "Référence", "Désignation", "Qté commandée", "Tarif convenu",
        "Devis associé", "Chantier", "Fournisseur", "Type de commande",
    ]

    # La ligne sans offre n'apparaît PAS dans le panier commandable...
    assert not any(l["Référence"] == "999999NOPE" for l in lignes_panier)

    # ... mais n'est pas perdue pour autant : elle est listée à part.
    ws_non_commandees = wb2["Non commandées"]
    non_commandees = _valeurs_par_entete(ws_non_commandees)
    assert any(n["Référence"] == "999999NOPE" for n in non_commandees)

    # Le manchon a bien suivi la décision de l'acheteur (DEM), PAS le
    # mieux-disant par défaut (COREDIME) : référence, devis et tarif RÉELS
    # de DEM, pas ceux de COREDIME.
    ligne_manchon_panier = next(l for l in lignes_panier if l["Référence"] == "CAPCAP856574")
    assert ligne_manchon_panier["Fournisseur"] == "DEM"
    assert ligne_manchon_panier["Devis associé"] == "821035"
    assert round(ligne_manchon_panier["Tarif convenu"], 4) == 1.41
    assert ligne_manchon_panier["Type de commande"] == "Chantier"

    # Le composé (Kit boîtier + couvercle) donne bien 2 lignes de commande,
    # chacune à sa quantité (3 x 1), chez COREDIME (seul à les vendre).
    refs_composant = {"SIBP02000", "SIBP03200"}
    lignes_composant = [l for l in lignes_panier if l["Référence"] in refs_composant]
    assert len(lignes_composant) == 2
    for l in lignes_composant:
        assert l["Qté commandée"] == 3
        assert l["Fournisseur"] == "COREDIME"

    # Fournisseur RAVATE -> nom canonique "RAVATE ELEC" du Suivi commandes,
    # pas le nom interne "RAVATE".
    ligne_ravate = next(l for l in lignes_panier if l["Référence"] == "TUBFFPTTGRD3")
    assert ligne_ravate["Fournisseur"] == "RAVATE ELEC"
    assert ligne_ravate["Qté commandée"] == 4

    # Un bloc de lignes PAR fournisseur retenu : trié, regroupé.
    fournisseurs_ordre = [l["Fournisseur"] for l in lignes_panier]
    assert fournisseurs_ordre == sorted(fournisseurs_ordre)


@pytest.mark.skipif(
    trouver_fichier_suivi(ROOT) is None,
    reason="Export du Suivi commandes absent de ce poste",
)
def test_colonnes_calees_sur_le_vrai_suivi_commandes(tmp_path):

    fichier_besoin, dossier_devis = _preparer_projet(tmp_path)
    fichier_comparatif = generer_comparatif(tmp_path, fichier_besoin, dossier_devis)

    # On dépose une copie du vrai export du Suivi dans le dossier de test
    # pour que generer_panier() la trouve (recherche à la racine du projet).
    fichier_suivi = trouver_fichier_suivi(ROOT)
    shutil.copy(fichier_suivi, tmp_path / fichier_suivi.name)

    fichier_panier = generer_panier(tmp_path, fichier_besoin, dossier_devis, fichier_comparatif)

    wb = load_workbook(fichier_panier)
    entetes_panier = [c.value for c in next(wb["Panier"].iter_rows(min_row=1, max_row=1))]

    wb_suivi = load_workbook(fichier_suivi, read_only=True, data_only=True)
    entetes_suivi = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(wb_suivi["Commandes"].iter_rows(min_row=1, max_row=1))
    ]
    wb_suivi.close()

    assert entetes_panier == entetes_suivi
