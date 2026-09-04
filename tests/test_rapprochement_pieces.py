"""
Modèle de la feuille « Pièces » (moteur/rapprochement/pieces.py) : ID,
dédoublonnage, construction d'une ligne, formules Commandes, lecture et
index — sur des objets construits à la main et un classeur synthétique
(voir tests/test_rapprochement_ecriture.py pour les socles XML eux-mêmes).
"""

from datetime import date, datetime

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from moteur.rapprochement.pieces import (
    COLONNES_FACTURE_CALCULEES,
    COLONNES_PIECES,
    FORMULES_COMMANDES_BASCULE,
    FORMULES_COMMANDES_NOUVELLES,
    IndexPieces,
    TYPE_AVOIR,
    TYPE_FACTURE,
    TYPE_FRAIS,
    ajouter_colonnes_commandes,
    basculer_colonnes_facture,
    cle_reference,
    construire_id_piece,
    dedoublonner_ids,
    ecrire_pieces,
    formule_fichier,
    installer_feuille_pieces,
    lire_pieces,
    nom_fournisseur_suivi,
    nouvelle_piece,
)


def _piece(**kw):
    defaut = dict(
        type_piece=TYPE_FACTURE, fournisseur="COREDIME", numero_piece="6100600", date_piece=date(2026, 1, 28),
        numero_commande="131.082", chantier="Chantier X", sous_chantier=None, reference_suivi="06620",
        reference_fournisseur="06620", designation="ICT 20", qte=800, pu_ht=0.35, montant_ht=280.0,
        numero_bl_lie="B010001", mode="Migré", fichier=r"X:\archive\f.pdf",
        date_ecriture=datetime(2026, 9, 4, 20, 0, 0),
    )
    defaut.update(kw)
    return nouvelle_piece(**defaut)


def test_nom_fournisseur_suivi_normalise_les_saisies_libres():
    assert nom_fournisseur_suivi("Coredime") == "COREDIME"
    assert nom_fournisseur_suivi("109 DISTRIBUTION") == "109 Distribution"
    assert nom_fournisseur_suivi("COMINTER Mayotte") == "COMINTER Mayotte"
    assert nom_fournisseur_suivi("ELECTRIC PLUS") == "GMR"
    assert nom_fournisseur_suivi("Inconnu SARL") == "Inconnu SARL"


def test_cle_reference_unifie_nombre_et_texte():
    assert cle_reference(5120) == cle_reference("5120") == cle_reference(5120.0) == "5120"
    assert cle_reference(" 06620 ") == "06620"
    assert cle_reference(None) == ""


def test_construire_id_piece_et_dedoublonnage():
    id1 = construire_id_piece("COREDIME", "Facture", "6100600", "131.082", "06620", "B1")
    assert id1 == "COREDIME|Facture|6100600|131.082|06620|B1"
    pieces = [{"ID pièce": id1}, {"ID pièce": id1}, {"ID pièce": id1}, {"ID pièce": "autre"}]
    dedoublonner_ids(pieces)
    assert [p["ID pièce"] for p in pieces] == [id1, id1 + "#2", id1 + "#3", "autre"]
    dedoublonner_ids(pieces)  # rejouer le même lot : inchangé (déterministe)
    assert [p["ID pièce"] for p in pieces] == [id1, id1 + "#2", id1 + "#3", "autre"]


def test_nouvelle_piece_remplit_les_26_colonnes():
    p = _piece()
    assert list(p) == list(COLONNES_PIECES)
    assert p["ID pièce"] == "COREDIME|Facture|6100600|131.082|06620|B010001"
    assert p["Fournisseur"] == "COREDIME"
    assert p["Date pièce"] == date(2026, 1, 28)
    assert p["Fichier"].texte == 'HYPERLINK("X:\\archive\\f.pdf","f.pdf")'
    assert p["Fichier"].cache == "f.pdf"
    assert p["Prix de référence"] is None and p["Demande d'avoir"] is None  # P2/P4
    assert p["Commentaire"] is None
    with pytest.raises(ValueError):
        _piece(type_piece="Ticket")


def test_nouvelle_piece_frais_sans_ligne_suivi_a_un_id_distinct():
    p = _piece(type_piece=TYPE_FRAIS, reference_suivi=None, reference_fournisseur="ECO-23")
    assert p["ID pièce"] == "COREDIME|Frais|6100600|131.082|F:ECO-23|B010001"
    assert p["Référence Suivi"] is None


def test_formules_commandes_sont_en_anglais_avec_prefixes_xlfn():
    assert set(FORMULES_COMMANDES_BASCULE) == set(COLONNES_FACTURE_CALCULEES)
    assert FORMULES_COMMANDES_BASCULE["N° facture"].array is True
    assert "_xlfn.TEXTJOIN(" in FORMULES_COMMANDES_BASCULE["N° facture"].texte
    assert "_xlfn._xlws.FILTER(" in FORMULES_COMMANDES_BASCULE["N° facture"].texte
    assert "_xlfn.UNIQUE(" in FORMULES_COMMANDES_BASCULE["N° facture"].texte
    assert FORMULES_COMMANDES_BASCULE["Date facture"].texte.count("_xlfn.MAXIFS(") == 2
    assert FORMULES_COMMANDES_BASCULE["Qté facturée"].texte.startswith("SUMIFS(Pieces[Qté],")
    assert '"Avoir"' in FORMULES_COMMANDES_BASCULE["Qté facturée"].texte
    for f in list(FORMULES_COMMANDES_BASCULE.values()) + list(FORMULES_COMMANDES_NOUVELLES.values()):
        assert not f.texte.startswith("=")
        assert "[@" not in f.texte  # forme longue [[#This Row],[...]] uniquement
    assert "⛔ Sur-facturée" in FORMULES_COMMANDES_NOUVELLES["Statut facture"].texte


def test_index_pieces_cumul_par_ligne_de_commande_et_doublons_de_facture():
    pieces = [
        _piece(numero_piece="F1", qte=10, montant_ht=10.0),
        _piece(numero_piece="F2", qte=5, montant_ht=5.0, numero_bl_lie="B2"),
        _piece(type_piece=TYPE_AVOIR, numero_piece="AV1", qte=-2, montant_ht=-2.0),
        _piece(numero_piece="F1", reference_suivi=5120, qte=1, montant_ht=1.0),
        _piece(type_piece="BL", numero_piece="B1", qte=99, montant_ht=None),
    ]
    for p in pieces:
        p["Fichier"] = p["Fichier"].cache  # comme relu par openpyxl (valeur en cache)
    index = IndexPieces(pieces)
    assert index.disponible
    assert index.qte_facturee("131.082", "06620") == 13  # 10 + 5 - 2, le BL ignoré
    assert index.montant_facture("131.082", "06620") == 13.0
    assert index.numeros_factures("131.082", "06620") == ["F1", "F2"]
    assert index.qte_facturee("131.082", "5120") == 1 and index.qte_facturee("131.082", 5120.0) == 1
    assert index.qte_facturee("999", "x") == 0
    assert index.fichiers_facture("Coredime", "F1") == {"f.pdf"}
    assert index.fichiers_facture("COREDIME", "inconnue") == set()
    assert pieces[0]["ID pièce"] in index.ids

    vide = IndexPieces(None)
    assert not vide.disponible and vide.qte_facturee("131.082", "06620") == 0


def _classeur_commandes(chemin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "N° de commande", "Fournisseur", "Qté livrée", "Facturé BL",
               *COLONNES_FACTURE_CALCULEES])
    ws.append(["06620", 800, "131.082", "COREDIME", 800, 280.0, "6100600", date(2026, 1, 28), 800, 0.35, 280.0])
    ws.append([5120, 10, "131.082", "COREDIME", 10, 35.0, None, None, None, None, None])
    table = Table(displayName="Commandes", ref="A1:K3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.save(chemin)


def test_installer_ecrire_et_relire_les_pieces(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin)
    assert lire_pieces(chemin) is None  # feuille absente : None, jamais une liste vide trompeuse

    with pytest.raises(ValueError):
        ecrire_pieces(chemin, [_piece()], tmp_path / "backups")  # jamais d'écriture perdue en silence

    installer_feuille_pieces(chemin, tmp_path / "backups")
    assert lire_pieces(chemin) == []

    pieces = [_piece(), _piece(numero_piece="F2", reference_suivi=5120, qte=3, montant_ht=1.05)]
    resultat = ecrire_pieces(chemin, pieces, tmp_path / "backups")
    assert resultat["ajoutees"] == 2
    assert ecrire_pieces(chemin, [_piece()], tmp_path / "backups")["ajoutees"] == 0  # idempotent

    relues = lire_pieces(chemin)
    assert len(relues) == 2
    assert relues[0]["ID pièce"] == pieces[0]["ID pièce"]
    assert relues[0]["Fichier"] == "f.pdf"  # cache de la formule HYPERLINK
    assert relues[0]["Date pièce"].date() == date(2026, 1, 28)
    assert relues[0]["Date d'écriture"] == datetime(2026, 9, 4, 20, 0, 0)
    assert relues[1]["Référence Suivi"] == 5120  # type conservé (nombre), comme dans Commandes
    assert relues[1]["_ligne_excel"] == 3

    wb = load_workbook(chemin)
    assert wb.sheetnames == ["Commandes", "Pièces"]
    assert wb["Pièces"].tables["Pieces"].ref == f"A1:{chr(ord('A') + len(COLONNES_PIECES) - 1)}3"


def test_basculer_et_ajouter_les_colonnes_commandes(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin)
    installer_feuille_pieces(chemin, tmp_path / "backups")
    basculer_colonnes_facture(chemin, tmp_path / "backups")
    ajouter_colonnes_commandes(chemin, tmp_path / "backups")

    wb = load_workbook(chemin)
    ws = wb["Commandes"]
    entetes = [c.value for c in ws[1]]
    assert entetes[-4:] == ["Reste à facturer HT", "Écart facture €", "Qté retournée", "Statut facture"]
    col_qte = entetes.index("Qté facturée") + 1
    assert ws.cell(row=2, column=col_qte).value == "=" + FORMULES_COMMANDES_BASCULE["Qté facturée"].texte
    assert ws.cell(row=3, column=col_qte).value == "=" + FORMULES_COMMANDES_BASCULE["Qté facturée"].texte
    col_num = entetes.index("N° facture") + 1
    assert ws.cell(row=2, column=col_num).value.text == "=" + FORMULES_COMMANDES_BASCULE["N° facture"].texte
    cols = {c.name: c for c in ws.tables["Commandes"].tableColumns}
    assert cols["N° facture"].calculatedColumnFormula.array is True
    assert cols["Statut facture"].calculatedColumnFormula.attr_text == FORMULES_COMMANDES_NOUVELLES["Statut facture"].texte
    assert ws.tables["Commandes"].ref == "A1:O3"
