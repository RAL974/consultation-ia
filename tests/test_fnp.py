# -*- coding: utf-8 -*-
"""
Tests de moteur/fnp.py — état mensuel des Factures Non Parvenues (clôture
comptable, demande de la DAF). Classeurs SYNTHÉTIQUES (tmp_path, jamais le
vrai Suivi commandes pour les cas unitaires) reproduisant la structure de
dossiers attendue par trouver_fichier_suivi_vivant() :

  <racine_tmp>/Consultation AI/                          <- dossier_projet
  <racine_tmp>/1.3.0.1. Commandes courantes/1.3.0.1. Suivi commandes - 2026.xlsx
  <racine_tmp>/1.3.0. Suivi commandes spéciales.xlsm

Un test dédié (test_calculer_rapport_fnp_sur_le_vrai_suivi) tourne en plus
sur le VRAI classeur vivant, en LECTURE SEULE, ignoré si absent du poste —
même esprit que tests/test_rapprochement_ecriture.py : verrouille que les
en-têtes réels n'ont pas dérivé, sans figer de valeur métier (qui change à
chaque commande passée)."""

from datetime import date, datetime

import pytest
from openpyxl import Workbook

from moteur.fnp import (
    DATE_CREATION_COLONNES_FACTURE,
    SuiviIntrouvable,
    _controler_couverture_transitaires,
    _repli_transitaires_suivi_principal,
    calculer_rapport_fnp,
    dernier_jour_mois,
    ecrire_classeur_fnp,
    generer_etat_fnp,
    lire_dossiers_transitaires_non_factures,
    lire_lignes_bl_non_facturees,
    mois_en_lettres,
    mois_precedent_complet,
)
from moteur.rapprochement.pipeline_bl import trouver_fichier_suivi_vivant

from conftest import ROOT

ENTETES_COMMANDES = [
    "Référence", "Désignation", "Qté commandée", "Tarif convenu", "N° de commande",
    "Fournisseur", "Chantier", "Date de livraison", "Qté livrée", "Tarif BL",
    "Note", "N° facture", "Date facture", "Facturé BL", "Transitaire",
    "N° dossier revient", "Montant total commande",
]
_IDX = {nom: i for i, nom in enumerate(ENTETES_COMMANDES)}


def _ligne_commande(**kwargs):
    """Une ligne de la feuille "Commandes", valeurs par défaut = ligne FNP
    typique (livrée en août, jamais facturée, tarif BL connu)."""

    defaut = {
        "Référence": "REF1", "Désignation": "Article", "Qté commandée": 10,
        "Tarif convenu": 0, "N° de commande": "M1.00.001", "Fournisseur": "COREDIME",
        "Chantier": "100 Chantier Test", "Date de livraison": datetime(2026, 8, 15),
        "Qté livrée": 10, "Tarif BL": 5.0, "Note": "", "N° facture": "", "Date facture": None,
        "Facturé BL": 50.0, "Transitaire": "", "N° dossier revient": "", "Montant total commande": 50.0,
    }
    defaut.update(kwargs)
    return [defaut[nom] for nom in ENTETES_COMMANDES]


def _classeur_commandes(chemin, lignes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(ENTETES_COMMANDES)
    for l in lignes:
        ws.append(l)
    wb.save(chemin)


ENTETES_SPECIALES = [
    "Désignation", "Commande", "Chantier", "Fournisseur", "N° dossier revient",
    "Montant commande", "Transitaire", "Réf trsprt", "Date de départ",
    "ETA ou arrivée réelle", "Expédition facturée", "Coût estimé",
]
_IDX_SPE = {nom: i for i, nom in enumerate(ENTETES_SPECIALES)}


def _ligne_speciale(**kwargs):
    defaut = {
        "Désignation": "Marchandise", "Commande": "M1.00.001", "Chantier": "100 Chantier Test",
        "Fournisseur": "Clareo", "N° dossier revient": "R26.001", "Montant commande": 1000.0,
        "Transitaire": "Steinweg", "Réf trsprt": "REF-T1", "Date de départ": datetime(2026, 7, 1),
        "ETA ou arrivée réelle": datetime(2026, 8, 10), "Expédition facturée": 0, "Coût estimé": 300.0,
    }
    defaut.update(kwargs)
    return [defaut[nom] for nom in ENTETES_SPECIALES]


def _classeur_speciales(chemin, lignes, feuille="Suivi"):
    wb = Workbook()
    ws = wb.active
    ws.title = feuille
    ws.append(ENTETES_SPECIALES)
    for l in lignes:
        ws.append(l)
    wb.save(chemin)


def _arborescence(tmp_path):
    """Construit <tmp_path>/Consultation AI (dossier_projet) + les deux
    fichiers frères attendus, retourne (dossier_projet, chemin_suivi,
    chemin_speciales) — les DEUX derniers chemins ne sont pas encore créés,
    à l'appelant de le faire via _classeur_commandes/_classeur_speciales."""

    dossier_projet = tmp_path / "Consultation AI"
    dossier_projet.mkdir()
    dossier_courantes = tmp_path / "1.3.0.1. Commandes courantes"
    dossier_courantes.mkdir()
    chemin_suivi = dossier_courantes / "1.3.0.1. Suivi commandes - 2026.xlsx"
    chemin_speciales = tmp_path / "1.3.0. Suivi commandes spéciales.xlsm"
    return dossier_projet, chemin_suivi, chemin_speciales


# --- dates -------------------------------------------------------------


def test_dernier_jour_mois():
    assert dernier_jour_mois("2026-08") == date(2026, 8, 31)
    assert dernier_jour_mois("2026-02") == date(2026, 2, 28)
    assert dernier_jour_mois("2028-02") == date(2028, 2, 29)  # bissextile
    assert dernier_jour_mois("2026-04") == date(2026, 4, 30)


def test_mois_precedent_complet():
    assert mois_precedent_complet(date(2026, 9, 1)) == "2026-08"
    assert mois_precedent_complet(date(2026, 9, 30)) == "2026-08"
    assert mois_precedent_complet(date(2026, 1, 15)) == "2025-12"


def test_mois_en_lettres():
    assert mois_en_lettres("2026-08") == "août 2026"
    assert mois_en_lettres("2026-01") == "janvier 2026"


# --- volet (a) : lire_lignes_bl_non_facturees ---------------------------


def test_ligne_simple_valorisee_par_tarif_bl(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande()])

    valorisees, sans_prix, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))

    assert len(valorisees) == 1
    assert not sans_prix
    l = valorisees[0]
    assert l.montant_ht == 50.0
    assert l.source_prix == "Tarif BL"
    assert l.anciennete_jours == (date(2026, 8, 31) - date(2026, 8, 15)).days


def test_repli_sur_tarif_convenu_si_pas_de_tarif_bl(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{"Tarif BL": 0, "Tarif convenu": 4.0, "Facturé BL": 40.0})])

    valorisees, sans_prix, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))

    assert len(valorisees) == 1
    assert valorisees[0].source_prix == "Tarif convenu"
    assert valorisees[0].montant_ht == 40.0


def test_ligne_livree_sans_aucun_prix_va_dans_sans_prix(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{"Tarif BL": 0, "Tarif convenu": 0, "Facturé BL": 0})])

    valorisees, sans_prix, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))

    assert not valorisees
    assert len(sans_prix) == 1
    assert sans_prix[0].source_prix == "Aucune"


def test_qte_livree_nulle_exclue(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{"Qté livrée": 0})])

    valorisees, sans_prix, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert not valorisees and not sans_prix


def test_livraison_apres_fin_de_mois_exclue(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{"Date de livraison": datetime(2026, 9, 3)})])

    valorisees, _, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert not valorisees


def test_commande_annulee_exclue(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{"Note": "Commande annulée"})])

    valorisees, sans_prix, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert not valorisees and not sans_prix


def test_rupture_fournisseur_nest_pas_exclue(tmp_path):
    """Seule "Commande annulée" exclut une ligne — les autres valeurs
    "magiques" de Note (Rupture fournisseur, Reliquat soldé) sont juste
    reportées telles quelles (voir périmètre donné par la DAF)."""
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{"Note": "Rupture fournisseur"})])

    valorisees, _, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert len(valorisees) == 1
    assert valorisees[0].note == "Rupture fournisseur"


def test_facture_deja_recue_avant_cloture_exclue(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{
        "N° facture": "F123", "Date facture": datetime(2026, 8, 20),
    })])

    valorisees, sans_prix, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert not valorisees and not sans_prix


def test_facture_recue_apres_cloture_reste_une_fnp(tmp_path):
    """Facturée, mais après la fin du mois M -> à la clôture de M, la
    facture n'était pas encore là : toujours une FNP pour M."""
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{
        "N° facture": "F123", "Date facture": datetime(2026, 9, 5),
    })])

    valorisees, _, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert len(valorisees) == 1


def test_numero_facture_sans_date_facture_traite_comme_deja_facturee(tmp_path):
    """Cas limite : un N° facture peut être écrit sans date parseable (voir
    pipeline_facture.ecritures_pour_facture, qui n'écrit la date que si elle
    a pu être parsée) — traité prudemment comme "déjà facturée" plutôt que
    risquer de compter une ligne deux fois."""
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{"N° facture": "F123", "Date facture": None})])

    valorisees, sans_prix, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert not valorisees and not sans_prix


def test_filtre_depuis_exclut_les_livraisons_plus_anciennes(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [
        _ligne_commande(**{"Référence": "VIEUX", "Date de livraison": datetime(2026, 6, 1)}),
        _ligne_commande(**{"Référence": "RECENT", "Date de livraison": datetime(2026, 8, 20)}),
    ])

    valorisees, _, _ = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31), depuis=date(2026, 8, 1))
    assert [l.reference for l in valorisees] == ["RECENT"]


def test_transitaire_renseigne_collecte_pour_le_controle_de_couverture(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [_ligne_commande(**{
        "Transitaire": "Steinweg", "N° dossier revient": "R26.099",
        "N° de commande": "M9.00.009", "Fournisseur": "Clareo", "Chantier": "999 Chantier",
    })])

    _, _, transitaires_vus = lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))
    assert transitaires_vus == [("R26.099", "M9.00.009", "Clareo", "999 Chantier")]


def test_colonne_manquante_leve_keyerror(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Désignation"])  # loin d'être complet
    wb.save(chemin)

    with pytest.raises(KeyError):
        lire_lignes_bl_non_facturees(chemin, date(2026, 8, 31))


# --- volet (b) : lire_dossiers_transitaires_non_factures -----------------


def test_dossier_non_facture_et_arrive_a_temps_inclus(tmp_path):
    chemin = tmp_path / "speciales.xlsm"
    _classeur_speciales(chemin, [_ligne_speciale()])

    dossiers, avert, connus = lire_dossiers_transitaires_non_factures(chemin, date(2026, 8, 31))

    assert not avert
    assert len(dossiers) == 1
    assert dossiers[0].cout_estime == 300.0
    assert connus == {"R26.001"}


def test_dossier_deja_facture_exclu(tmp_path):
    chemin = tmp_path / "speciales.xlsm"
    _classeur_speciales(chemin, [_ligne_speciale(**{"Expédition facturée": 1})])

    dossiers, _, connus = lire_dossiers_transitaires_non_factures(chemin, date(2026, 8, 31))
    assert not dossiers
    assert connus == {"R26.001"}  # toujours compté comme "connu", même déjà facturé


def test_dossier_arrive_apres_fin_de_mois_exclu(tmp_path):
    chemin = tmp_path / "speciales.xlsm"
    _classeur_speciales(chemin, [_ligne_speciale(**{"ETA ou arrivée réelle": datetime(2026, 9, 2)})])

    dossiers, _, _ = lire_dossiers_transitaires_non_factures(chemin, date(2026, 8, 31))
    assert not dossiers


def test_dossier_sans_date_arrivee_exclu(tmp_path):
    chemin = tmp_path / "speciales.xlsm"
    _classeur_speciales(chemin, [_ligne_speciale(**{"ETA ou arrivée réelle": None})])

    dossiers, _, _ = lire_dossiers_transitaires_non_factures(chemin, date(2026, 8, 31))
    assert not dossiers


def test_dossier_sans_commande_inclus_quand_meme(tmp_path):
    """Confirmé par l'acheteur (cadrage) : le lien à un N° de commande
    n'est jamais un critère d'inclusion/exclusion pour ce volet."""
    chemin = tmp_path / "speciales.xlsm"
    _classeur_speciales(chemin, [_ligne_speciale(**{"Commande": ""})])

    dossiers, _, _ = lire_dossiers_transitaires_non_factures(chemin, date(2026, 8, 31))
    assert len(dossiers) == 1
    assert dossiers[0].numero_commande == ""


def test_classeur_speciales_introuvable_renvoie_avertissement():
    dossiers, avert, connus = lire_dossiers_transitaires_non_factures(
        "Z:/chemin/inexistant.xlsm", date(2026, 8, 31),
    )
    assert dossiers == [] and connus == set()
    assert "introuvable" in avert


def test_feuille_suivi_absente_renvoie_avertissement(tmp_path):
    chemin = tmp_path / "speciales.xlsm"
    wb = Workbook()
    wb.active.title = "AutreFeuille"
    wb.save(chemin)

    dossiers, avert, _ = lire_dossiers_transitaires_non_factures(chemin, date(2026, 8, 31))
    assert not dossiers
    assert "Suivi" in avert


# --- contrôle de couverture / repli --------------------------------------


def test_controler_couverture_signale_seulement_les_vrais_trous():
    vus = [
        ("R26.001", "M1", "F1", "C1"),   # connu -> pas signalé
        ("R26.099", "M9", "F9", "C9"),   # inconnu -> signalé
        ("", "M2", "F2", "C2"),          # pas de N° dossier -> pas vérifiable, pas signalé
    ]
    manquants = _controler_couverture_transitaires(vus, {"R26.001"})
    assert manquants == [("R26.099", "M9", "F9", "C9")]


def test_repli_transitaires_suivi_principal(tmp_path):
    chemin = tmp_path / "suivi.xlsx"
    _classeur_commandes(chemin, [
        _ligne_commande(**{"Transitaire": "DHL", "Montant total commande": 999.0}),
        _ligne_commande(**{"Référence": "SANS_TRANSITAIRE", "Transitaire": ""}),
    ])

    dossiers = _repli_transitaires_suivi_principal(chemin, date(2026, 8, 31))

    assert len(dossiers) == 1
    assert dossiers[0].transitaire == "DHL"
    assert dossiers[0].cout_estime is None
    assert dossiers[0].montant_marchandise == 999.0


# --- bout en bout ---------------------------------------------------------


def test_generer_etat_fnp_bout_en_bout(tmp_path):
    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [
        _ligne_commande(),
        _ligne_commande(**{"Référence": "SANS_PRIX", "Tarif BL": 0, "Tarif convenu": 0, "Facturé BL": 0}),
    ])
    _classeur_speciales(chemin_speciales, [_ligne_speciale()])

    chemin_sortie = generer_etat_fnp(dossier_projet, "2026-08")

    assert chemin_sortie.exists()
    assert chemin_sortie.parent.name == "rapports"

    from openpyxl import load_workbook
    wb = load_workbook(chemin_sortie)
    # "Factures reçues" et "Déclaré (hors outil)" : v1.1, session S0 (étapes
    # 4a/4b) — voir test_generer_etat_fnp_bout_en_bout_sans_exclusion pour le
    # cas --sans-exclusion (pas d'onglet "Factures reçues" dans ce cas).
    assert wb.sheetnames == [
        "Synthèse", "BL non facturés", "Transitaires", "Factures reçues", "Déclaré (hors outil)",
    ]
    # 1 ligne d'en-tête + 1 ligne valorisée + ligne vide + titre bloc + en-tête bloc + 1 ligne sans prix
    assert wb["BL non facturés"].max_row == 6
    assert wb["Transitaires"].max_row == 2  # en-tête + 1 dossier
    assert wb["Factures reçues"].max_row == 1  # en-tête seulement, aucune facture PDF déposée dans ce test
    assert wb["Déclaré (hors outil)"].max_row == 1  # en-tête seulement, aucun fichier fnp_ajustements_*.csv dans ce test


def test_generer_etat_fnp_bout_en_bout_sans_exclusion(tmp_path):
    """--sans-exclusion (repli 16h, voir CLAUDE.md) : pas de scan de
    a_traiter/Factures/, pas d'onglet "Factures reçues" — les étapes 4b/4c
    restent actives."""

    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [_ligne_commande()])
    _classeur_speciales(chemin_speciales, [_ligne_speciale()])

    chemin_sortie = generer_etat_fnp(dossier_projet, "2026-08", appliquer_exclusion=False)

    from openpyxl import load_workbook
    wb = load_workbook(chemin_sortie)
    assert wb.sheetnames == ["Synthèse", "BL non facturés", "Transitaires", "Déclaré (hors outil)"]


def test_generer_etat_fnp_repli_si_commandes_speciales_absent(tmp_path):
    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [_ligne_commande(**{"Transitaire": "DHL"})])
    # chemin_speciales volontairement jamais créé

    rapport = calculer_rapport_fnp(dossier_projet, "2026-08")

    assert rapport.transitaire_repli_utilise
    assert "introuvable" in rapport.transitaire_avertissement
    assert len(rapport.dossiers_transitaires) == 1
    assert rapport.dossiers_transitaires[0].cout_estime is None


def test_generer_etat_fnp_suivi_introuvable_leve(tmp_path):
    dossier_projet = tmp_path / "Consultation AI"
    dossier_projet.mkdir()
    # Aucun "1.3.0.1. Commandes courantes/" à côté.

    with pytest.raises(SuiviIntrouvable):
        generer_etat_fnp(dossier_projet, "2026-08")


# --- v1.1 (session S0), étape 4a : exclusion "facture reçue non rapprochée" -


def _ligne_fnp(**kwargs):
    from moteur.fnp import LigneFNP
    defaut = dict(
        ligne_excel=5, fournisseur="COREDIME", numero_commande="M1.00.001", chantier="100 Chantier",
        reference="REF1", designation="Article", qte_livree=10.0, montant_ht=50.0, source_prix="Tarif BL",
        date_livraison=date(2026, 8, 15), anciennete_jours=16, numero_facture="", date_facture=None, note="",
    )
    defaut.update(kwargs)
    return LigneFNP(**defaut)


def test_appliquer_exclusion_factures_recues_retire_la_ligne_matchee():
    from moteur.fnp import _appliquer_exclusion_factures_recues

    l = _ligne_fnp(ligne_excel=5, montant_ht=123.45)
    lignes_excel_facturees = {5: {
        "numero_facture": "F999", "date_facture": date(2026, 8, 20),
        "reference": "REF1", "fournisseur": "COREDIME", "numero_commande": "M1.00.001",
    }}

    restantes, sans_prix_restantes, recues = _appliquer_exclusion_factures_recues([l], [], lignes_excel_facturees)

    assert restantes == []
    assert sans_prix_restantes == []
    [r] = recues
    assert r.ligne_excel == 5
    assert r.numero_facture == "F999"
    assert r.date_facture == date(2026, 8, 20)
    assert r.montant_facture_bl == 123.45


def test_appliquer_exclusion_factures_recues_ligne_non_matchee_reste():
    from moteur.fnp import _appliquer_exclusion_factures_recues

    l = _ligne_fnp(ligne_excel=7)

    restantes, sans_prix_restantes, recues = _appliquer_exclusion_factures_recues([l], [], {5: {}})

    assert restantes == [l]
    assert recues == []


def test_appliquer_exclusion_factures_recues_fonctionne_aussi_sur_sans_prix():
    from moteur.fnp import _appliquer_exclusion_factures_recues

    l = _ligne_fnp(ligne_excel=9, montant_ht=0.0, source_prix="Aucune")
    lignes_excel_facturees = {9: {
        "numero_facture": "F1", "date_facture": None,
        "reference": "REF1", "fournisseur": "COREDIME", "numero_commande": "M1.00.001",
    }}

    restantes_bl, restantes_sans_prix, recues = _appliquer_exclusion_factures_recues([], [l], lignes_excel_facturees)

    assert restantes_sans_prix == []
    [r] = recues
    assert r.ligne_excel == 9
    assert r.date_facture is None


def test_identifier_lignes_excel_facturees_sur_une_vraie_facture(tmp_path):
    """Intégration RÉELLE (voir CLAUDE.md, règle d'or — vrai PDF, pas de
    texte inventé) : facture_dist109_1_simple.pdf (109 DISTRIBUTION,
    commande 123.075, facture 360311 datée 15/07/2026, référence P03101)
    déposée dans a_traiter/Factures/ d'un dossier_projet SYNTHÉTIQUE ->
    _identifier_lignes_excel_facturees() doit la retrouver et associer la
    bonne ligne_excel."""

    from moteur.fnp import _identifier_lignes_excel_facturees

    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [
        _ligne_commande(**{
            "Référence": "P03101", "N° de commande": "123.075", "Fournisseur": "109 DISTRIBUTION",
            "Qté livrée": 200, "N° facture": "", "Date facture": None,
        }),
    ])

    dossier_factures = dossier_projet / "a_traiter" / "Factures"
    dossier_factures.mkdir(parents=True)
    fixture = ROOT / "tests" / "fixtures" / "facture_dist109_1_simple.pdf"
    (dossier_factures / fixture.name).write_bytes(fixture.read_bytes())

    lignes_excel_facturees, n_bdc_manuel_24x = _identifier_lignes_excel_facturees(
        dossier_projet, fin_de_mois=date(2026, 7, 31),
    )

    assert n_bdc_manuel_24x == 0
    [(ligne_excel, info)] = lignes_excel_facturees.items()
    assert ligne_excel == 2  # 1re ligne de données (après l'en-tête)
    assert info["numero_facture"] == "360311"
    assert info["date_facture"] == date(2026, 7, 15)
    assert info["reference"] == "P03101"


def test_identifier_lignes_excel_facturees_facture_posterieure_a_fin_de_mois_exclue(tmp_path):
    """Même pièce réelle, mais fin_de_mois AVANT la date de la facture
    (15/07/2026) -> non retenue (seules les factures datées <= fin_de_mois
    comptent, voir bandeau du module)."""

    from moteur.fnp import _identifier_lignes_excel_facturees

    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [
        _ligne_commande(**{
            "Référence": "P03101", "N° de commande": "123.075", "Fournisseur": "109 DISTRIBUTION",
            "Qté livrée": 200,
        }),
    ])

    dossier_factures = dossier_projet / "a_traiter" / "Factures"
    dossier_factures.mkdir(parents=True)
    fixture = ROOT / "tests" / "fixtures" / "facture_dist109_1_simple.pdf"
    (dossier_factures / fixture.name).write_bytes(fixture.read_bytes())

    lignes_excel_facturees, _ = _identifier_lignes_excel_facturees(dossier_projet, fin_de_mois=date(2026, 6, 30))

    assert lignes_excel_facturees == {}


def test_identifier_lignes_excel_facturees_dossier_absent_retourne_vide(tmp_path):
    from moteur.fnp import _identifier_lignes_excel_facturees

    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [_ligne_commande()])
    # a_traiter/Factures/ jamais créé.

    lignes_excel_facturees, n_bdc = _identifier_lignes_excel_facturees(dossier_projet, fin_de_mois=date(2026, 8, 31))

    assert lignes_excel_facturees == {}
    assert n_bdc == 0


def test_calculer_rapport_fnp_applique_lexclusion_bout_en_bout(tmp_path):
    """La ligne Suivi P03101/123.075 est livrée en juillet, jamais facturée
    dans le Suivi (N° facture vide) -> sans 4a, elle apparaîtrait dans le
    volet (a). Avec la vraie facture PDF déposée, elle doit en sortir et
    se retrouver dans factures_recues_non_rapprochees."""

    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [
        _ligne_commande(**{
            "Référence": "P03101", "N° de commande": "123.075", "Fournisseur": "109 DISTRIBUTION",
            "Chantier": "100 Chantier Test", "Date de livraison": datetime(2026, 7, 10),
            "Qté livrée": 200, "Tarif BL": 0.6, "Facturé BL": 120.0,
            "N° facture": "", "Date facture": None,
        }),
    ])
    _classeur_speciales(chemin_speciales, [])

    dossier_factures = dossier_projet / "a_traiter" / "Factures"
    dossier_factures.mkdir(parents=True)
    fixture = ROOT / "tests" / "fixtures" / "facture_dist109_1_simple.pdf"
    (dossier_factures / fixture.name).write_bytes(fixture.read_bytes())

    rapport = calculer_rapport_fnp(dossier_projet, "2026-07")

    assert rapport.lignes_bl == []
    [f] = rapport.factures_recues_non_rapprochees
    assert f.numero_facture == "360311"
    assert f.montant_facture_bl == 120.0


# --- v1.1 (session S0), étape 4b : ajustements déclarés ---------------------


def test_lire_ajustements_fnp_fichier_absent_liste_vide(tmp_path):
    from moteur.fnp import lire_ajustements_fnp

    assert lire_ajustements_fnp(tmp_path / "inexistant.csv") == []


def test_lire_ajustements_fnp_lit_le_csv(tmp_path):
    from moteur.fnp import lire_ajustements_fnp

    chemin = tmp_path / "fnp_ajustements_2026-08.csv"
    chemin.write_text(
        "type;libelle;fournisseur_ou_transitaire;chantier;piece;date_livraison;montant_ht;source;commentaire\n"
        "BDC_MANUEL;Cable H07;COREDIME;100 Chantier;BC24.0512;12/08/2026;150,50;carnet papier;vu avec le chef\n"
        "TRANSIT;Conteneur;Steinweg;;D26.099;;2000;mail du 20/08;\n",
        encoding="utf-8",
    )

    ajustements = lire_ajustements_fnp(chemin)

    assert len(ajustements) == 2
    a0 = ajustements[0]
    assert a0.type == "BDC_MANUEL"
    assert a0.libelle == "Cable H07"
    assert a0.date_livraison == date(2026, 8, 12)
    assert a0.montant_ht == 150.5
    a1 = ajustements[1]
    assert a1.type == "TRANSIT"
    assert a1.date_livraison is None
    assert a1.montant_ht == 2000.0


def test_lire_ajustements_fnp_ignore_ligne_sans_type(tmp_path):
    from moteur.fnp import lire_ajustements_fnp

    chemin = tmp_path / "fnp_ajustements_2026-08.csv"
    chemin.write_text(
        "type;libelle;fournisseur_ou_transitaire;chantier;piece;date_livraison;montant_ht;source;commentaire\n"
        ";;;;;;;;\n",
        encoding="utf-8",
    )

    assert lire_ajustements_fnp(chemin) == []


def test_lire_ajustements_fnp_ignore_les_lignes_de_commentaire(tmp_path):
    from moteur.fnp import lire_ajustements_fnp

    chemin = tmp_path / "fnp_ajustements_2026-08.csv"
    chemin.write_text(
        "# ceci est un commentaire d'en-tête\n"
        "type;libelle;fournisseur_ou_transitaire;chantier;piece;date_livraison;montant_ht;source;commentaire\n"
        "AUTRE;Test;X;;;;10;;\n",
        encoding="utf-8",
    )

    assert len(lire_ajustements_fnp(chemin)) == 1


def test_nom_fichier_ajustements():
    from moteur.fnp import nom_fichier_ajustements

    assert nom_fichier_ajustements("2026-08") == "fnp_ajustements_2026-08.csv"


def test_calculer_rapport_fnp_lit_les_ajustements_du_bon_mois(tmp_path):
    """referentiel/fnp_ajustements_<mois>.csv, PAS un nom fixe — un fichier
    d'un AUTRE mois ne doit jamais être lu par erreur."""

    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [_ligne_commande()])
    _classeur_speciales(chemin_speciales, [])

    dossier_referentiel = dossier_projet / "referentiel"
    dossier_referentiel.mkdir()
    (dossier_referentiel / "fnp_ajustements_2026-08.csv").write_text(
        "type;libelle;fournisseur_ou_transitaire;chantier;piece;date_livraison;montant_ht;source;commentaire\n"
        "AUTRE;Test aout;X;;;;42;;\n",
        encoding="utf-8",
    )
    (dossier_referentiel / "fnp_ajustements_2026-07.csv").write_text(
        "type;libelle;fournisseur_ou_transitaire;chantier;piece;date_livraison;montant_ht;source;commentaire\n"
        "AUTRE;Test juillet, ne doit PAS apparaitre;X;;;;999;;\n",
        encoding="utf-8",
    )

    rapport = calculer_rapport_fnp(dossier_projet, "2026-08", appliquer_exclusion=False)

    [a] = rapport.ajustements
    assert a.libelle == "Test aout"
    assert a.montant_ht == 42.0


# --- v1.1 (session S0), étape 4c : réserves de périmètre --------------------


def test_compter_dossiers_speciales(tmp_path):
    from moteur.fnp import compter_dossiers_speciales

    chemin = tmp_path / "speciales.xlsm"
    _classeur_speciales(chemin, [_ligne_speciale(), _ligne_speciale(**{"N° dossier revient": "R26.002"})])

    assert compter_dossiers_speciales(chemin) == 2


def test_compter_dossiers_speciales_fichier_absent_zero(tmp_path):
    from moteur.fnp import compter_dossiers_speciales

    assert compter_dossiers_speciales(tmp_path / "inexistant.xlsm") == 0


def test_calculer_rapport_fnp_reserves_bdc_manuel_et_dossiers_speciales(tmp_path):
    """rapport.reserves reflète : le nombre de dossiers Commandes spéciales
    (tout statut, voir compter_dossiers_speciales) et le nombre de
    transitaires sans estimation déjà connu du volet (b)."""

    dossier_projet, chemin_suivi, chemin_speciales = _arborescence(tmp_path)
    _classeur_commandes(chemin_suivi, [_ligne_commande()])
    _classeur_speciales(chemin_speciales, [
        _ligne_speciale(),
        _ligne_speciale(**{"N° dossier revient": "R26.002", "Coût estimé": None}),
    ])

    rapport = calculer_rapport_fnp(dossier_projet, "2026-08", appliquer_exclusion=False)

    assert rapport.reserves.n_bdc_manuel_24x == 0  # exclusion désactivée, jamais scanné
    assert rapport.reserves.n_dossiers_speciales_total == 2
    assert rapport.reserves.n_transitaires_sans_estimation == 1  # le 2e dossier, sans Coût estimé


# --- vrai classeur (lecture seule, ignoré si absent du poste) ------------


def _chemin_vrai_suivi():
    return trouver_fichier_suivi_vivant(ROOT)


@pytest.mark.skipif(_chemin_vrai_suivi() is None, reason="Suivi commandes vivant absent de ce poste")
def test_calculer_rapport_fnp_sur_le_vrai_suivi():
    """Vérifie que les en-têtes réels n'ont pas dérivé — ne fige AUCUNE
    valeur métier (les commandes réelles changent en continu), juste que le
    calcul tourne sans exception et retourne des types cohérents.

    appliquer_exclusion=False (voir _identifier_lignes_excel_facturees) :
    ce test vérifie les en-têtes du SUIVI, pas la 4a — sans ce False, il
    scannerait le VRAI a_traiter/Factures/ (des centaines de PDF réels,
    OCR compris) à CHAQUE exécution de la suite de tests, beaucoup trop
    lent pour un test de non-régression sur les seuls en-têtes."""

    rapport = calculer_rapport_fnp(ROOT, "2026-08", appliquer_exclusion=False)

    assert rapport.chemin_suivi is not None
    assert isinstance(rapport.lignes_bl, list)
    assert isinstance(rapport.dossiers_transitaires, list)
    for l in rapport.lignes_bl:
        assert l.montant_ht >= 0
        assert l.source_prix in ("Tarif BL", "Tarif convenu")
    for l in rapport.lignes_sans_prix:
        assert l.source_prix == "Aucune"
