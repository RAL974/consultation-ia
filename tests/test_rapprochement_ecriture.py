"""
Socle d'écriture sécurisée dans le Suivi commandes (moteur/rapprochement/
ecriture.py, branche Rapprochement AI — voir CLAUDE.md).

Deux familles de tests :
- sur un classeur SYNTHÉTIQUE (mais structurellement représentatif : Excel
  Table, colonne formule, validation de données) pour vérifier la mécanique
  générique du patch XML sans dépendre du fichier personnel de l'acheteur ;
- sur le VRAI "1.3.0.1. Suivi commandes - 2026.xlsx" (jamais l'original —
  toujours une copie dans tmp_path), ignoré si absent de ce poste (comme
  tests/test_panier.py::test_colonnes_calees_sur_le_vrai_suivi_commandes),
  pour la preuve définitive que rien d'autre que la cellule visée ne bouge.
"""

import re
import shutil
import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

from moteur.rapprochement.ecriture import (
    COLONNES_MODIFIABLES,
    Ecriture,
    ClasseurVerrouille,
    ColonneNonModifiable,
    Formule,
    _decaler_references_ligne,
    ajouter_colonnes_calculees,
    ajouter_entetes_saisie,
    ajouter_feuille_tableau,
    ajouter_lignes_tableau,
    appliquer,
    basculer_colonnes_en_formules,
    est_verrouille,
    lire_entetes,
    reparer_formules_ligne,
    sauvegarder,
    simuler,
)
from moteur.rapprochement.pieces import COLONNES_FACTURE_CALCULEES
from moteur.panier import trouver_fichier_suivi
from moteur.rapprochement.pipeline_bl import trouver_fichier_suivi_vivant

from conftest import ROOT


def _classeur_synthetique(chemin):
    """Un classeur minimal mais structurellement représentatif du vrai
    Suivi commandes : une feuille "Commandes" avec un Excel Table, une
    colonne formule (jamais modifiable) et une validation de données."""

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Statut", "Note"])
    ws.append(["ART1", 10, None, None, '=IF(C2="","attente","reçu")', None])
    ws.append(["ART2", 5, 5, 12.5, '=IF(C3="","attente","reçu")', None])

    table = Table(displayName="Commandes", ref="A1:F3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    dv = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    dv.add("F2:F3")
    ws.add_data_validation(dv)

    wb.save(chemin)


@pytest.fixture
def classeur(tmp_path):
    chemin = tmp_path / "Suivi_test.xlsx"
    _classeur_synthetique(chemin)
    return chemin


def test_lire_entetes(classeur):
    entetes = lire_entetes(classeur)
    assert entetes["Qté livrée"] == 3
    assert entetes["Tarif BL"] == 4


def test_colonne_formule_refusee(classeur):
    with pytest.raises(ColonneNonModifiable):
        appliquer(classeur, [Ecriture(2, "Statut", "reçu")], classeur.parent / "backups")


def test_colonne_inconnue_refusee(classeur):
    with pytest.raises(ColonneNonModifiable):
        appliquer(classeur, [Ecriture(2, "Colonne fantôme", 1)], classeur.parent / "backups")


def test_simuler_ne_modifie_rien(classeur):
    avant = classeur.read_bytes()
    rapport = simuler(classeur, [Ecriture(2, "Qté livrée", 10), Ecriture(2, "Tarif BL", 9.9)])
    assert classeur.read_bytes() == avant

    par_colonne = {r["colonne"]: r for r in rapport}
    assert par_colonne["Qté livrée"]["ancienne_valeur"] is None
    assert par_colonne["Qté livrée"]["nouvelle_valeur"] == 10
    assert par_colonne["Tarif BL"]["nouvelle_valeur"] == 9.9


def test_appliquer_ecrit_cellule_vide_et_preserve_le_reste(classeur, tmp_path):
    dossier_backups = tmp_path / "backups"
    sauvegarde = appliquer(
        classeur,
        [Ecriture(2, "Qté livrée", 10), Ecriture(2, "Tarif BL", 19.9), Ecriture(2, "Note", "Rupture fournisseur")],
        dossier_backups,
    )

    assert sauvegarde.exists()
    assert sauvegarde.read_bytes() != classeur.read_bytes()  # la sauvegarde est l'ANCIEN état

    wb = load_workbook(classeur, data_only=False)
    ws = wb["Commandes"]
    assert ws["C2"].value == 10
    assert ws["D2"].value == 19.9
    assert ws["F2"].value == "Rupture fournisseur"
    # la formule voisine, jamais touchée par l'écriture, doit survivre :
    assert str(ws["E2"].value).startswith("=IF(")
    # le Table et la validation de données doivent survivre :
    assert "Commandes" in ws.tables
    assert len(ws.data_validations.dataValidation) == 1


def test_appliquer_ecrit_cellule_deja_remplie(classeur):
    appliquer(classeur, [Ecriture(3, "Qté livrée", 7)], classeur.parent / "backups")
    wb = load_workbook(classeur)
    assert wb["Commandes"]["C3"].value == 7


def test_verrou_bloque_ecriture(classeur):
    verrou = classeur.parent / f"~${classeur.name}"
    verrou.write_text("")
    assert est_verrouille(classeur)
    with pytest.raises(ClasseurVerrouille):
        appliquer(classeur, [Ecriture(2, "Qté livrée", 1)], classeur.parent / "backups")


def test_sauvegarde_purge_les_anciennes(classeur, tmp_path):
    import time

    dossier_backups = tmp_path / "backups"
    ancienne = sauvegarder(classeur, dossier_backups)
    vieux_temps = time.time() - 31 * 86400
    import os

    os.utime(ancienne, (vieux_temps, vieux_temps))

    recente = sauvegarder(classeur, dossier_backups)

    assert not ancienne.exists()
    assert recente.exists()


def test_colonnes_facture_plus_jamais_modifiables():
    """Depuis P1 (feuille Pièces, voir CLAUDE.md) les 5 colonnes facture de
    Commandes sont des colonnes CALCULÉES : l'outil n'y écrit plus jamais
    une valeur — seules les 4 colonnes BL historiques restent saisissables."""

    assert set(COLONNES_MODIFIABLES) == {"Date de livraison", "Qté livrée", "Tarif BL", "Note"}
    assert not set(COLONNES_FACTURE_CALCULEES) & set(COLONNES_MODIFIABLES)


def test_appliquer_refuse_les_colonnes_facture(tmp_path):
    """Même présentes dans la feuille, les 5 colonnes facture sont refusées
    par appliquer() (elles portent une formule depuis P1) — rien n'est
    écrit, aucune sauvegarde créée."""

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté livrée", "Tarif BL", "N° facture", "Date facture", "Qté facturée", "PU facturé"])
    ws.append(["ART1", 10, 5.0, None, None, None, None])
    chemin = tmp_path / "suivi_facture.xlsx"
    wb.save(chemin)
    avant = chemin.read_bytes()

    with pytest.raises(ColonneNonModifiable):
        appliquer(chemin, [Ecriture(2, "N° facture", "360311")], tmp_path / "backups")
    assert chemin.read_bytes() == avant
    assert not (tmp_path / "backups").exists()


def test_ajouter_entetes_saisie_integre_les_colonnes_dans_le_tableau(classeur, tmp_path):
    """Les nouveaux en-têtes vont à la suite de la dernière colonne
    existante ET rejoignent le tableau structuré (ref étendue, autoFilter
    étendu, tableColumns complété) — jamais posés à côté : un tri du
    tableau ne suit pas les colonnes restées hors tableau, les données
    facture se retrouveraient sur les mauvaises lignes (décision explicite
    de l'acheteur, voir CLAUDE.md)."""

    sauvegarde = ajouter_entetes_saisie(
        classeur, ["N° facture", "Date facture"], tmp_path / "backups"
    )
    assert sauvegarde.exists()

    entetes = lire_entetes(classeur)
    assert entetes["N° facture"] == 7  # G, juste après F ("Note")
    assert entetes["Date facture"] == 8  # H

    wb = load_workbook(classeur, data_only=False)
    ws = wb["Commandes"]
    assert ws["G1"].value == "N° facture"
    assert ws["H1"].value == "Date facture"
    assert str(ws["E2"].value).startswith("=IF(")  # formule voisine intacte

    table = ws.tables["Commandes"]
    assert table.ref == "A1:H3"  # le tableau s'étend pour couvrir les 2 colonnes
    noms_colonnes = [c.name for c in table.tableColumns]
    assert len(noms_colonnes) == 8
    assert noms_colonnes[-2:] == ["N° facture", "Date facture"]


def test_ajouter_entetes_saisie_refuse_colonne_deja_dans_le_tableau(classeur, tmp_path):
    """Garde-fou défensif : un nom déjà présent comme colonne du TABLEAU
    (mais pas forcément détecté par lire_entetes(), qui ne lit que la
    ligne 1 de la FEUILLE) doit aussi bloquer — simule une
    désynchronisation entre la feuille et la définition du tableau."""

    with zipfile.ZipFile(classeur) as z:
        contenu = {n: z.read(n) for n in z.namelist()}
    xml = contenu["xl/tables/table1.xml"].decode("utf-8")
    xml = xml.replace(
        "</tableColumns>", '<tableColumn id="7" name="N° facture"/></tableColumns>'
    )
    contenu["xl/tables/table1.xml"] = xml.encode("utf-8")
    with zipfile.ZipFile(classeur, "w", zipfile.ZIP_DEFLATED) as z:
        for nom, donnees in contenu.items():
            z.writestr(nom, donnees)

    avant = classeur.read_bytes()
    with pytest.raises(ValueError):
        ajouter_entetes_saisie(classeur, ["N° facture"], tmp_path / "backups")
    assert classeur.read_bytes() == avant  # refus AVANT toute écriture
    assert not (tmp_path / "backups").exists()


def test_ajouter_entetes_saisie_refuse_nom_deja_present(classeur, tmp_path):
    avant = classeur.read_bytes()
    with pytest.raises(ValueError):
        ajouter_entetes_saisie(classeur, ["Note"], tmp_path / "backups")
    assert classeur.read_bytes() == avant  # refus AVANT toute écriture


def test_ajouter_entetes_saisie_refuse_noms_en_double(classeur, tmp_path):
    with pytest.raises(ValueError):
        ajouter_entetes_saisie(classeur, ["N° facture", "N° facture"], tmp_path / "backups")


def test_ajouter_entetes_saisie_bloque_si_verrouille(classeur, tmp_path):
    verrou = classeur.parent / f"~${classeur.name}"
    verrou.write_text("")
    with pytest.raises(ClasseurVerrouille):
        ajouter_entetes_saisie(classeur, ["N° facture"], tmp_path / "backups")


def test_ajouter_entetes_saisie_refuse_cellule_deja_presente(classeur, tmp_path):
    """Garde-fou défensif : même une cellule présente mais SANS valeur (donc
    invisible de lire_entetes) à l'emplacement visé doit bloquer l'ajout —
    simule un résidu jamais nettoyé à la position G1 (celle que la fonction
    viserait normalement)."""

    with zipfile.ZipFile(classeur) as z:
        contenu = {n: z.read(n) for n in z.namelist()}
    xml = contenu["xl/worksheets/sheet1.xml"].decode("utf-8")
    xml = re.sub(
        r'(<row r="1"[^>]*>.*?)(</row>)',
        r'\1<c r="G1" s="1"/>\2',
        xml,
        count=1,
        flags=re.S,
    )
    contenu["xl/worksheets/sheet1.xml"] = xml.encode("utf-8")
    with zipfile.ZipFile(classeur, "w", zipfile.ZIP_DEFLATED) as z:
        for nom, donnees in contenu.items():
            z.writestr(nom, donnees)

    avant = classeur.read_bytes()
    with pytest.raises(ValueError):
        ajouter_entetes_saisie(classeur, ["N° facture"], tmp_path / "backups")
    assert classeur.read_bytes() == avant  # refus AVANT toute écriture (pas de sauvegarde non plus)
    assert not (tmp_path / "backups").exists()


@pytest.mark.skipif(
    trouver_fichier_suivi_vivant(ROOT) is None,
    reason="Classeur Suivi commandes VIVANT introuvable depuis ce poste (dossier "
    "frère « 1.3.0.1. Commandes courantes » absent ou inaccessible)",
)
def test_ajouter_entetes_saisie_sur_le_vrai_suivi_commandes_vivant(tmp_path):
    """La preuve définitive pour l'étape 1 (voir CLAUDE.md) : sur une COPIE
    du classeur VIVANT réellement utilisé par l'acheteur (refondu — feuilles
    Dashboard/Analyses/Calculs — jamais l'export périmé de la racine du
    dépôt, voir trouver_fichier_suivi_vivant), ajouter des colonnes DANS le
    tableau structuré Commandes (décision explicite de l'acheteur — un tri
    du tableau ne suit pas une colonne restée hors tableau) ne touche
    STRICTEMENT que la feuille (ligne 1 + <dimension>) et la définition du
    tableau (ref, autoFilter, tableColumns) : les formules existantes
    (Statut commande, Facturé BL...) et tout le reste du zip survivent
    identiques.

    L'état "avant" (ref du tableau, nombre de colonnes) est lu dynamiquement
    plutôt que supposé figé : les 5 vraies colonnes facture (ENTETES_FACTURE)
    ont été créées pour de vrai dans le classeur vivant à l'étape 2 de cette
    même session — ce test réutilise donc 2 noms JAMAIS utilisés en pratique,
    uniquement pour continuer à prouver le MÉCANISME sur la vraie structure
    du fichier, indéfiniment (même après l'ajout d'autres colonnes futures)."""

    fichier_reel = trouver_fichier_suivi_vivant(ROOT)
    copie = tmp_path / fichier_reel.name
    shutil.copy2(fichier_reel, copie)
    avant_fichier = copie.read_bytes()

    with zipfile.ZipFile(copie) as z:
        contenu_avant = {n: z.read(n) for n in z.namelist()}

    wb_avant = load_workbook(copie, data_only=False)
    ws_avant = wb_avant["Commandes"]
    table_avant = ws_avant.tables["Commandes"]
    ref_avant = table_avant.ref
    nb_colonnes_avant = len(table_avant.tableColumns)
    derniere_colonne_avant = max(lire_entetes(copie).values())
    formule_statut_avant = ws_avant["U100"].value
    formule_facture_bl_avant = ws_avant["AV100"].value
    wb_avant.close()

    noms_test = ["Colonne test rapprochement A", "Colonne test rapprochement B"]
    sauvegarde = ajouter_entetes_saisie(copie, noms_test, tmp_path / "backups")
    assert sauvegarde.exists()
    assert sauvegarde.read_bytes() == avant_fichier  # sauvegarde = état AVANT patch, intégral
    assert sauvegarde.read_bytes() != copie.read_bytes()

    with zipfile.ZipFile(copie) as z:
        contenu_apres = {n: z.read(n) for n in z.namelist()}

    chemin_feuille = "xl/worksheets/sheet1.xml"
    chemin_table = "xl/tables/table1.xml"  # tableau structuré "Commandes"
    assert set(contenu_avant) == set(contenu_apres)
    parties_modifiees = {chemin_feuille, chemin_table}
    for nom in contenu_avant:
        if nom in parties_modifiees:
            assert contenu_avant[nom] != contenu_apres[nom]
        else:
            assert contenu_avant[nom] == contenu_apres[nom], f"partie modifiée à tort : {nom}"

    # À l'intérieur de sheet1.xml : seules la ligne 1 et <dimension> ont
    # changé, rigoureusement rien d'autre (aucune autre ligne, aucun
    # décalage de calcChain/sharedStrings...).
    def _sans_ligne1_ni_dimension(xml):
        xml = re.sub(r'<row r="1"[^>]*>.*?</row>', "", xml, count=1, flags=re.S)
        xml = re.sub(r'<dimension ref="[^"]*"\s*/>', "", xml, count=1)
        return xml

    xml_feuille_avant = contenu_avant[chemin_feuille].decode("utf-8")
    xml_feuille_apres = contenu_apres[chemin_feuille].decode("utf-8")
    assert _sans_ligne1_ni_dimension(xml_feuille_avant) == _sans_ligne1_ni_dimension(xml_feuille_apres)

    nouvelle_derniere_colonne = get_column_letter(derniere_colonne_avant + len(noms_test))
    m_dim = re.search(r'<dimension ref="[A-Z]+\d+:([A-Z]+)(\d+)"\s*/>', xml_feuille_avant)
    assert m_dim is not None
    ligne_fin = m_dim.group(2)
    assert f'<dimension ref="A1:{nouvelle_derniere_colonne}{ligne_fin}"/>' in xml_feuille_apres

    entetes_apres = lire_entetes(copie)
    for i, nom in enumerate(noms_test):
        assert entetes_apres[nom] == derniere_colonne_avant + 1 + i

    # À l'intérieur de table1.xml : reconstruction INDÉPENDANTE (remplacement
    # de texte brut, pas la même logique regex que _etendre_tableau) de ce
    # que le patch doit produire — seuls le ref du <table>, le ref du
    # <autoFilter> et le count de <tableColumns> changent, plus les
    # <tableColumn> ajoutés en toute fin, rien d'autre.
    xml_table_avant = contenu_avant[chemin_table].decode("utf-8")
    xml_table_apres = contenu_apres[chemin_table].decode("utf-8")
    ancienne_ref_attr = f'ref="{ref_avant}"'
    ancien_count_attr = f'<tableColumns count="{nb_colonnes_avant}">'
    assert ancienne_ref_attr in xml_table_avant
    assert ancien_count_attr in xml_table_avant

    m_ref = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)$", ref_avant)
    debut_col, debut_ligne, fin_col, fin_ligne = m_ref.groups()
    nouvelle_fin_col = get_column_letter(column_index_from_string(fin_col) + len(noms_test))
    nouvelle_ref = f"{debut_col}{debut_ligne}:{nouvelle_fin_col}{fin_ligne}"

    ids_existants = [int(i) for i in re.findall(r'<tableColumn\b[^>]*\bid="(\d+)"', xml_table_avant)]
    id_suivant = max(ids_existants) + 1
    fragments = "".join(
        f'<tableColumn id="{id_suivant + i}" name="{nom}"/>' for i, nom in enumerate(noms_test)
    )
    nouveau_count_attr = f'<tableColumns count="{nb_colonnes_avant + len(noms_test)}">'
    xml_table_attendu = xml_table_avant.replace(ancienne_ref_attr, f'ref="{nouvelle_ref}"')
    xml_table_attendu = xml_table_attendu.replace(ancien_count_attr, nouveau_count_attr)
    xml_table_attendu = xml_table_attendu.replace("</tableColumns>", fragments + "</tableColumns>")
    assert xml_table_apres == xml_table_attendu

    wb_apres = load_workbook(copie, data_only=False)
    ws_apres = wb_apres["Commandes"]
    col1 = get_column_letter(derniere_colonne_avant + 1)
    col2 = get_column_letter(derniere_colonne_avant + 2)
    assert ws_apres[f"{col1}1"].value == noms_test[0]
    assert ws_apres[f"{col2}1"].value == noms_test[1]
    # le tableau structuré Commandes s'étend pour couvrir les nouvelles
    # colonnes (indépendant, via le parseur XML d'openpyxl) :
    table_apres = ws_apres.tables["Commandes"]
    assert table_apres.ref == nouvelle_ref
    noms_colonnes_apres = [c.name for c in table_apres.tableColumns]
    assert len(noms_colonnes_apres) == nb_colonnes_avant + len(noms_test)
    assert noms_colonnes_apres[-len(noms_test):] == noms_test
    # les formules existantes, protégées par COLONNES_MODIFIABLES, sont
    # EXACTEMENT les mêmes qu'avant (aucun calculatedColumnFormula ajouté
    # aux nouvelles colonnes, aucune formule existante retouchée) :
    assert ws_apres["U100"].value == formule_statut_avant
    assert ws_apres["AV100"].value == formule_facture_bl_avant
    wb_apres.close()


@pytest.mark.skipif(
    trouver_fichier_suivi_vivant(ROOT) is None,
    reason="Classeur Suivi commandes VIVANT introuvable depuis ce poste",
)
def test_appliquer_refuse_les_colonnes_facture_sur_le_vrai_suivi_vivant():
    """Les 5 colonnes facture existent bien dans le vivant (créées le
    2026-09-01) mais, depuis P1, appliquer() les REFUSE : ce sont des
    colonnes calculées depuis la feuille Pièces. Lecture seule des en-têtes
    du vivant, refus AVANT toute sauvegarde/écriture."""

    fichier_reel = trouver_fichier_suivi_vivant(ROOT)
    entetes = lire_entetes(fichier_reel)
    assert set(COLONNES_FACTURE_CALCULEES) <= set(entetes)

    mtime_avant = fichier_reel.stat().st_mtime
    with pytest.raises(ColonneNonModifiable):
        appliquer(fichier_reel, [Ecriture(2, "N° facture", "TEST-360999")], ROOT / "backups")
    assert fichier_reel.stat().st_mtime == mtime_avant


@pytest.mark.skipif(
    trouver_fichier_suivi(ROOT) is None,
    reason="Export du Suivi commandes absent de ce poste",
)
def test_ecriture_chirurgicale_sur_le_vrai_suivi_commandes(tmp_path):
    """La preuve définitive : sur une COPIE du vrai classeur, seule la
    cellule visée change — tout le reste du zip est OCTET POUR OCTET
    identique à l'original (tableaux, validations, calcChain, sharedStrings,
    customXml, printerSettings inclus)."""

    fichier_reel = trouver_fichier_suivi(ROOT)
    copie = tmp_path / fichier_reel.name
    shutil.copy2(fichier_reel, copie)

    with zipfile.ZipFile(copie) as z:
        contenu_avant = {n: z.read(n) for n in z.namelist()}

    entetes = lire_entetes(copie)
    assert set(("Date de livraison", "Qté livrée", "Tarif BL", "Note")) <= set(entetes)

    # Ligne 2 = 1ère commande réelle : on y écrit une quantité livrée
    # arbitraire, jamais destinée à être sauvegardée dans le vrai fichier
    # (on travaille sur `copie`, dans tmp_path).
    appliquer(copie, [Ecriture(2, "Qté livrée", 12345)], tmp_path / "backups")

    with zipfile.ZipFile(copie) as z:
        contenu_apres = {n: z.read(n) for n in z.namelist()}

    chemin_feuille = "xl/worksheets/sheet1.xml"  # "Commandes" est le 1er onglet
    assert set(contenu_avant) == set(contenu_apres)
    for nom in contenu_avant:
        if nom == chemin_feuille:
            assert contenu_avant[nom] != contenu_apres[nom]
        else:
            assert contenu_avant[nom] == contenu_apres[nom], f"partie modifiée à tort : {nom}"

    wb = load_workbook(copie, data_only=True)
    assert wb["Commandes"]["P2"].value == 12345
    # les 16 Excel Tables du vrai classeur doivent toutes survivre :
    assert len(wb["Commandes"].tables) >= 1


# ===========================================================================
# Socles P1 (feuille Pièces) — voir CLAUDE.md « Feuille Pièces — modèle,
# socles, migration » : 1. feuille + tableau, 2. lignes en fin de tableau,
# 3. colonnes calculées / bascule ; plus la réparation d'une ligne.
# Sur le classeur SYNTHÉTIQUE d'abord, puis sur une COPIE du vivant.
# ===========================================================================


def _parties(chemin):
    with zipfile.ZipFile(chemin) as z:
        return {n: z.read(n) for n in z.namelist()}


def _xml_bien_forme(contenu: dict):
    import xml.etree.ElementTree as ET

    for nom, donnees in contenu.items():
        if nom.endswith(".xml") or nom.endswith(".rels"):
            ET.fromstring(donnees)  # lève si mal formé


COLONNES_TEST = ["ID", "Type", "Date", "Qté", "Montant", "Fichier", "Écrit le"]


def test_ajouter_feuille_tableau_cree_feuille_et_tableau(classeur, tmp_path):
    avant = _parties(classeur)
    sauvegarde = ajouter_feuille_tableau(
        classeur, "Pièces", "Pieces", COLONNES_TEST, tmp_path / "backups", largeurs={"Fichier": 40},
    )
    assert sauvegarde.exists()
    apres = _parties(classeur)
    _xml_bien_forme(apres)

    nouvelles = set(apres) - set(avant)
    assert nouvelles == {"xl/worksheets/sheet2.xml", "xl/worksheets/_rels/sheet2.xml.rels", "xl/tables/table2.xml"}
    modifiees = {n for n in avant if avant[n] != apres[n]}
    assert modifiees == {"xl/workbook.xml", "xl/_rels/workbook.xml.rels", "[Content_Types].xml"}

    wb = load_workbook(classeur)
    assert wb.sheetnames == ["Commandes", "Pièces"]  # juste après la feuille modèle
    ws = wb["Pièces"]
    assert [c.value for c in ws[1]] == COLONNES_TEST
    table = ws.tables["Pieces"]
    assert table.ref == "A1:G2"  # une ligne de données vide sous l'en-tête
    assert [c.name for c in table.tableColumns] == COLONNES_TEST
    assert table.tableStyleInfo.name == wb["Commandes"].tables["Commandes"].tableStyleInfo.name
    assert wb["Commandes"].tables["Commandes"].ref == "A1:F3"  # tableau modèle intact


def test_ajouter_feuille_tableau_refuse_doublons(classeur, tmp_path):
    ajouter_feuille_tableau(classeur, "Pièces", "Pieces", COLONNES_TEST, tmp_path / "backups")
    avant = classeur.read_bytes()
    with pytest.raises(ValueError):
        ajouter_feuille_tableau(classeur, "Pièces", "Autre", COLONNES_TEST, tmp_path / "backups")
    with pytest.raises(ValueError):
        ajouter_feuille_tableau(classeur, "Autre", "Pieces", COLONNES_TEST, tmp_path / "backups")
    with pytest.raises(ValueError):
        ajouter_feuille_tableau(classeur, "Autre", "Nom invalide", COLONNES_TEST, tmp_path / "backups")
    assert classeur.read_bytes() == avant


def test_ajouter_feuille_tableau_bloque_si_verrouille(classeur, tmp_path):
    (classeur.parent / f"~${classeur.name}").write_text("")
    with pytest.raises(ClasseurVerrouille):
        ajouter_feuille_tableau(classeur, "Pièces", "Pieces", COLONNES_TEST, tmp_path / "backups")


def _trois_lignes():
    from datetime import datetime as _dt

    return [
        {"ID": "A|1", "Type": "Facture", "Date": date(2026, 9, 4), "Qté": 3, "Montant": 12.5,
         "Fichier": Formule('HYPERLINK("X:\\\\dossier\\\\f1.pdf","f1.pdf")', cache="f1.pdf"),
         "Écrit le": _dt(2026, 9, 4, 18, 30, 0)},
        {"ID": "A|2", "Type": "BL", "Date": date(2026, 9, 1), "Qté": -1, "Montant": 0},
        {"ID": "A|3", "Type": "Frais", "Qté": 1, "Montant": 7.2, "Écrit le": _dt(2026, 9, 4, 18, 30, 5)},
    ]


def test_ajouter_lignes_tableau_ajoute_trois_lignes_et_etend_le_tableau(classeur, tmp_path):
    ajouter_feuille_tableau(classeur, "Pièces", "Pieces", COLONNES_TEST, tmp_path / "backups")
    avant = _parties(classeur)

    resultat = ajouter_lignes_tableau(
        classeur, "Pièces", "Pieces", _trois_lignes(), tmp_path / "backups",
        colonne_id="ID", styles_colonnes={"Montant": "monnaie"},
    )
    assert resultat["ajoutees"] == 3
    assert resultat["ignorees"] == []
    assert (resultat["premiere_ligne"], resultat["derniere_ligne"]) == (2, 4)

    apres = _parties(classeur)
    _xml_bien_forme(apres)
    modifiees = {n for n in avant if avant[n] != apres[n]}
    assert modifiees == {"xl/worksheets/sheet2.xml", "xl/tables/table2.xml", "xl/styles.xml"}
    # aucune chaîne partagée ajoutée : textes en inlineStr
    assert avant.get("xl/sharedStrings.xml") == apres.get("xl/sharedStrings.xml")

    wb = load_workbook(classeur)
    ws = wb["Pièces"]
    assert ws.tables["Pieces"].ref == "A1:G4"
    assert ws.tables["Pieces"].autoFilter.ref == "A1:G4"
    assert ws["A2"].value == "A|1" and ws["B2"].value == "Facture"
    assert ws["C2"].value == date(2026, 9, 4) or ws["C2"].value.date() == date(2026, 9, 4)
    assert ws["C2"].number_format != "General"  # style date appliqué
    assert ws["D3"].value == -1
    assert ws["E2"].value == 12.5 and ws["E2"].number_format != "General"
    assert str(ws["F2"].value).startswith("=HYPERLINK(")
    assert ws["G2"].value.hour == 18 and ws["G2"].number_format != "General"  # datetime
    assert ws["A4"].value == "A|3" and ws["C4"].value is None

    wb_v = load_workbook(classeur, data_only=True)
    assert wb_v["Pièces"]["F2"].value == "f1.pdf"  # cache de la formule HYPERLINK


def test_ajouter_lignes_tableau_est_idempotent_par_colonne_cle(classeur, tmp_path):
    ajouter_feuille_tableau(classeur, "Pièces", "Pieces", COLONNES_TEST, tmp_path / "backups")
    ajouter_lignes_tableau(classeur, "Pièces", "Pieces", _trois_lignes(), tmp_path / "backups", colonne_id="ID")
    avant = classeur.read_bytes()

    lignes = _trois_lignes() + [{"ID": "A|4", "Type": "Avoir", "Qté": -2, "Montant": -3.0}, {"ID": "A|4", "Type": "x"}]
    resultat = ajouter_lignes_tableau(classeur, "Pièces", "Pieces", lignes, tmp_path / "backups", colonne_id="ID")
    assert resultat["ajoutees"] == 1
    assert resultat["ignorees"] == ["A|1", "A|2", "A|3", "A|4"]  # 3 déjà là + 1 doublon dans le lot
    wb = load_workbook(classeur)
    assert wb["Pièces"].tables["Pieces"].ref == "A1:G5"
    assert wb["Pièces"]["A5"].value == "A|4"

    # tout déjà présent -> rien n'est écrit, pas de sauvegarde
    resultat = ajouter_lignes_tableau(classeur, "Pièces", "Pieces", _trois_lignes(), tmp_path / "backups", colonne_id="ID")
    assert resultat["ajoutees"] == 0 and resultat["sauvegarde"] is None


def test_ajouter_lignes_tableau_refuse_colonne_inconnue_et_id_vide(classeur, tmp_path):
    ajouter_feuille_tableau(classeur, "Pièces", "Pieces", COLONNES_TEST, tmp_path / "backups")
    avant = classeur.read_bytes()
    with pytest.raises(ValueError):
        ajouter_lignes_tableau(classeur, "Pièces", "Pieces", [{"ID": "x", "Fantôme": 1}], tmp_path / "backups", colonne_id="ID")
    with pytest.raises(ValueError):
        ajouter_lignes_tableau(classeur, "Pièces", "Pieces", [{"Type": "x"}], tmp_path / "backups", colonne_id="ID")
    assert classeur.read_bytes() == avant


def test_ajouter_lignes_tableau_relit_les_id_meme_reecrits_en_chaines_partagees(classeur, tmp_path):
    """Excel, à sa prochaine sauvegarde, transforme les inlineStr en
    chaînes partagées : l'idempotence doit survivre à cette réécriture."""

    ajouter_feuille_tableau(classeur, "Pièces", "Pieces", COLONNES_TEST, tmp_path / "backups")
    ajouter_lignes_tableau(classeur, "Pièces", "Pieces", _trois_lignes(), tmp_path / "backups", colonne_id="ID")
    wb = load_workbook(classeur)
    wb.save(classeur)  # openpyxl réécrit tout en sharedStrings (comme Excel)

    resultat = ajouter_lignes_tableau(classeur, "Pièces", "Pieces", _trois_lignes(), tmp_path / "backups", colonne_id="ID")
    assert resultat["ajoutees"] == 0
    assert resultat["ignorees"] == ["A|1", "A|2", "A|3"]


def _injecter_faux_calc_chain(chemin):
    """Le classeur synthétique (openpyxl) n'a pas de calcChain — on en
    fabrique un, relié et déclaré comme Excel le ferait."""

    contenu = _parties(chemin)
    contenu["xl/calcChain.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><c r="E2" i="1"/></calcChain>'
    )
    rels = contenu["xl/_rels/workbook.xml.rels"].decode("utf-8").replace(
        "</Relationships>",
        '<Relationship Id="rId99" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain" '
        'Target="calcChain.xml"/></Relationships>',
    )
    contenu["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
    ct = contenu["[Content_Types].xml"].decode("utf-8").replace(
        "</Types>",
        '<Override PartName="/xl/calcChain.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/></Types>',
    )
    contenu["[Content_Types].xml"] = ct.encode("utf-8")
    with zipfile.ZipFile(chemin, "w", zipfile.ZIP_DEFLATED) as z:
        for nom, donnees in contenu.items():
            z.writestr(nom, donnees)


def test_ajouter_colonnes_calculees_formule_sur_chaque_ligne_et_calc_chain_supprime(classeur, tmp_path):
    _injecter_faux_calc_chain(classeur)
    avant = _parties(classeur)
    assert "xl/calcChain.xml" in avant

    formules = {
        "Reste": 'IFERROR(N(Commandes[[#This Row],[Qté commandée]])-N(Commandes[[#This Row],[Qté livrée]]),"")',
        "Libellés": Formule(
            '_xlfn.TEXTJOIN("; ",TRUE,_xlfn.UNIQUE(_xlfn._xlws.FILTER(Commandes[Référence],'
            'Commandes[Qté livrée]=Commandes[[#This Row],[Qté livrée]],"")))', array=True,
        ),
    }
    sauvegarde = ajouter_colonnes_calculees(
        classeur, "Commandes", "Commandes", formules, tmp_path / "backups", styles_colonnes={"Reste": "monnaie"},
    )
    assert sauvegarde.exists()

    apres = _parties(classeur)
    _xml_bien_forme(apres)
    assert "xl/calcChain.xml" not in apres
    assert "calcChain" not in apres["xl/_rels/workbook.xml.rels"].decode("utf-8")
    assert "calcChain" not in apres["[Content_Types].xml"].decode("utf-8")
    assert 'fullCalcOnLoad="1"' in apres["xl/workbook.xml"].decode("utf-8")

    wb = load_workbook(classeur)
    ws = wb["Commandes"]
    assert ws["G1"].value == "Reste" and ws["H1"].value == "Libellés"
    assert ws["G2"].value == "=" + formules["Reste"]
    assert ws["G3"].value == "=" + formules["Reste"]  # la MÊME chaîne sur chaque ligne
    assert ws["H2"].value.text == "=" + formules["Libellés"].texte  # ArrayFormula (t="array")
    assert ws["H2"].value.ref == "H2"
    table = ws.tables["Commandes"]
    assert table.ref == "A1:H3"
    cols = {c.name: c for c in table.tableColumns}
    assert cols["Reste"].calculatedColumnFormula.attr_text == formules["Reste"]
    assert cols["Libellés"].calculatedColumnFormula.array is True
    assert str(ws["E2"].value).startswith("=IF(")  # formule voisine intacte
    assert len(ws.data_validations.dataValidation) == 1
    assert ws["G2"].number_format != "General"


def test_ajouter_colonnes_calculees_refuse_colonne_existante(classeur, tmp_path):
    avant = classeur.read_bytes()
    with pytest.raises(ValueError):
        ajouter_colonnes_calculees(classeur, "Commandes", "Commandes", {"Note": "1+1"}, tmp_path / "backups")
    assert classeur.read_bytes() == avant


def test_basculer_colonnes_en_formules_remplace_les_valeurs_saisies(classeur, tmp_path):
    formule = "N(Commandes[[#This Row],[Qté commandée]])*2"
    basculer_colonnes_en_formules(classeur, "Commandes", "Commandes", {"Qté livrée": formule}, tmp_path / "backups")

    wb = load_workbook(classeur)
    ws = wb["Commandes"]
    assert ws["C2"].value == "=" + formule  # cellule absente avant (ART1 sans Qté livrée) : créée
    assert ws["C3"].value == "=" + formule  # valeur 5 remplacée
    cols = {c.name: c for c in ws.tables["Commandes"].tableColumns}
    assert cols["Qté livrée"].calculatedColumnFormula.attr_text == formule
    assert ws.tables["Commandes"].ref == "A1:F3"  # pas de colonne ajoutée
    assert 'fullCalcOnLoad="1"' in _parties(classeur)["xl/workbook.xml"].decode("utf-8")

    with pytest.raises(ValueError):
        basculer_colonnes_en_formules(classeur, "Commandes", "Commandes", {"Fantôme": formule}, tmp_path / "backups")


def test_decaler_references_ligne():
    assert _decaler_references_ligne("SUMIF(Commandes[N° de commande],H3,Commandes[Facturé BL])", 3, 2) == (
        "SUMIF(Commandes[N° de commande],H2,Commandes[Facturé BL])"
    )
    assert _decaler_references_ligne("$A$3+A3+A30+AB3+'Feuille'!C3", 3, 2) == "$A$3+A2+A30+AB2+'Feuille'!C2"
    assert _decaler_references_ligne("Commandes[[#This Row],[Tarif BL]]*1.15", 3, 2) == "Commandes[[#This Row],[Tarif BL]]*1.15"
    assert _decaler_references_ligne("LOG10(3)", 3, 2) == "LOG10(3)"


def test_reparer_formules_ligne_recopie_la_formule_modele(classeur, tmp_path):
    """Ligne 2 : formule décalée (celle du Statut) à la place de la bonne ;
    ligne 3 = modèle. Seules les cellules listées sont touchées."""

    contenu = _parties(classeur)
    xml = contenu["xl/worksheets/sheet1.xml"].decode("utf-8")
    xml = re.sub(
        r'(<c r="E2"[^>]*>)<f>.*?</f>', r'\1<f>IF(D2="","x","y")</f>', xml, count=1, flags=re.S,
    )
    contenu["xl/worksheets/sheet1.xml"] = xml.encode("utf-8")
    with zipfile.ZipFile(classeur, "w", zipfile.ZIP_DEFLATED) as z:
        for nom, donnees in contenu.items():
            z.writestr(nom, donnees)
    assert load_workbook(classeur)["Commandes"]["E2"].value == '=IF(D2="","x","y")'

    ecrit = reparer_formules_ligne(classeur, tmp_path / "backups", colonnes=[5])
    assert ecrit == {5: 'IF(C2="","attente","reçu")'}  # C3 -> C2

    wb = load_workbook(classeur)
    ws = wb["Commandes"]
    assert ws["E2"].value == '=IF(C2="","attente","reçu")'
    assert ws["E3"].value == '=IF(C3="","attente","reçu")'
    assert ws["A2"].value == "ART1" and ws["B2"].value == 10  # voisines intactes

    assert reparer_formules_ligne(classeur, tmp_path / "backups", colonnes=[5]) == {}  # déjà identique
    with pytest.raises(ValueError):
        reparer_formules_ligne(classeur, tmp_path / "backups", colonnes=[1])  # A3 : pas une formule


# --- Sur une COPIE du classeur vivant ----------------------------------------


@pytest.mark.skipif(
    trouver_fichier_suivi_vivant(ROOT) is None,
    reason="Classeur Suivi commandes VIVANT introuvable depuis ce poste",
)
def test_socles_feuille_et_lignes_sur_une_copie_du_vrai_suivi_vivant(tmp_path):
    """Socles 1 et 2 sur une copie du vrai classeur : 16 -> 17 tableaux,
    onglet juste après Commandes, seules les parties attendues changent,
    XML bien formé partout, lignes relues par openpyxl."""

    fichier_reel = trouver_fichier_suivi_vivant(ROOT)
    copie = tmp_path / fichier_reel.name
    shutil.copy2(fichier_reel, copie)
    avant = _parties(copie)
    nb_tables_avant = len([n for n in avant if re.match(r"xl/tables/table\d+\.xml$", n)])
    onglets_avant = re.findall(r'<sheet\b[^>]*\bname="([^"]*)"', avant["xl/workbook.xml"].decode("utf-8"))

    nom_feuille, nom_table = "Test P1 socle", "TestP1Socle"
    ajouter_feuille_tableau(copie, nom_feuille, nom_table, COLONNES_TEST, tmp_path / "backups")

    apres = _parties(copie)
    _xml_bien_forme({n: d for n, d in apres.items() if n != "xl/worksheets/sheet1.xml"})  # sheet1 = 58 Mo, inchangée (vérifié ci-dessous)
    assert len([n for n in apres if re.match(r"xl/tables/table\d+\.xml$", n)]) == nb_tables_avant + 1
    nouvelles = set(apres) - set(avant)
    assert len(nouvelles) == 3
    modifiees = {n for n in avant if avant[n] != apres[n]}
    assert modifiees == {"xl/workbook.xml", "xl/_rels/workbook.xml.rels", "[Content_Types].xml"}
    onglets_apres = re.findall(r'<sheet\b[^>]*\bname="([^"]*)"', apres["xl/workbook.xml"].decode("utf-8"))
    assert onglets_apres == [onglets_avant[0], nom_feuille] + onglets_avant[1:]
    assert avant["xl/calcChain.xml"] == apres["xl/calcChain.xml"]  # socle 1 : aucune formule, calcChain intact

    resultat = ajouter_lignes_tableau(copie, nom_feuille, nom_table, _trois_lignes(), tmp_path / "backups", colonne_id="ID")
    assert resultat["ajoutees"] == 3

    wb = load_workbook(copie, read_only=True)
    assert wb.sheetnames[1] == nom_feuille
    ws = wb[nom_feuille]
    lignes = list(ws.iter_rows(values_only=True))
    assert list(lignes[0]) == COLONNES_TEST
    assert lignes[1][0] == "A|1" and lignes[3][0] == "A|3"
    wb.close()


@pytest.mark.skipif(
    trouver_fichier_suivi_vivant(ROOT) is None,
    reason="Classeur Suivi commandes VIVANT introuvable depuis ce poste",
)
def test_socle_colonnes_calculees_sur_une_copie_du_vrai_suivi_vivant(tmp_path):
    """Socle 3 sur une copie du vrai classeur (feuille de 58 Mo, 6 500+
    lignes) : la formule est posée sur CHAQUE ligne du tableau Commandes,
    calcChain retiré proprement, les formules existantes (Statut commande
    U100, Facturé BL AV100) strictement identiques."""

    fichier_reel = trouver_fichier_suivi_vivant(ROOT)
    copie = tmp_path / fichier_reel.name
    shutil.copy2(fichier_reel, copie)
    avant = _parties(copie)
    xml_avant = avant["xl/worksheets/sheet1.xml"].decode("utf-8")
    ref_avant = re.search(r'<table\b[^>]*\sref="([^"]*)"', avant["xl/tables/table1.xml"].decode("utf-8")).group(1)
    derniere_ligne = int(re.match(r"[A-Z]+\d+:[A-Z]+(\d+)", ref_avant).group(1))
    nb_cols_avant = len(lire_entetes(copie))

    def _formule(xml, ref):
        m = re.search(rf'<c r="{ref}"[^>]*>.*?<f[^>]*>(.*?)</f>', xml, re.S)
        return m.group(1) if m else None

    u100_avant, av100_avant = _formule(xml_avant, "U100"), _formule(xml_avant, "AV100")
    assert u100_avant and av100_avant

    nom = "Test P1 calcul"
    formule = 'IFERROR(N(Commandes[[#This Row],[Qté commandée]])-N(Commandes[[#This Row],[Qté livrée]]),"")'
    ajouter_colonnes_calculees(copie, "Commandes", "Commandes", {nom: formule}, tmp_path / "backups")

    apres = _parties(copie)
    assert "xl/calcChain.xml" not in apres
    assert "calcChain" not in apres["xl/_rels/workbook.xml.rels"].decode("utf-8")
    assert "calcChain" not in apres["[Content_Types].xml"].decode("utf-8")
    modifiees = {n for n in avant if n in apres and avant[n] != apres[n]}
    assert modifiees == {
        "xl/worksheets/sheet1.xml", "xl/tables/table1.xml", "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels", "[Content_Types].xml",
    }
    _xml_bien_forme({n: apres[n] for n in modifiees})

    xml_apres = apres["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert _formule(xml_apres, "U100") == u100_avant
    assert _formule(xml_apres, "AV100") == av100_avant
    lettre = get_column_letter(nb_cols_avant + 1)
    attendu = formule.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    assert _formule(xml_apres, f"{lettre}2") == attendu
    assert _formule(xml_apres, f"{lettre}{derniere_ligne}") == attendu
    assert len(re.findall(rf'<c r="{lettre}\d+"', xml_apres)) == derniere_ligne  # en-tête + une par ligne
    assert lire_entetes(copie)[nom] == nb_cols_avant + 1
    xml_table = apres["xl/tables/table1.xml"].decode("utf-8")
    assert f"<calculatedColumnFormula>{attendu}</calculatedColumnFormula>" in xml_table
    assert re.search(r'<table\b[^>]*\sref="A1:' + lettre + str(derniere_ligne) + '"', xml_table)
