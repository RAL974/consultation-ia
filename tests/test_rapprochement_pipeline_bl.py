"""
Logique pure de moteur/rapprochement/pipeline_bl.py (regroupement par BL,
construction des Ecriture, nommage d'archive) — sur des objets BonLivraison/
Correspondance construits à la main, sans OCR (voir
tests/test_parsers_bl_dist109.py et tests/test_rapprochement_matching.py
pour les parties qui en dépendent). Le découpage PDF par BL individuel a
besoin d'un vrai PDF ouvrable (voir tests/fixtures/) pour les quelques tests
qui l'exercent — un fichier factice ne suffit pas à fitz.
"""

from datetime import date

import pytest
from openpyxl import Workbook

from moteur.rapprochement.ecriture import Ecriture
from moteur.rapprochement.pieces import FeuillePiecesAbsente, installer_feuille_pieces, lire_pieces
from moteur.rapprochement.matching import LigneSuivi, Statut, Correspondance
from moteur.rapprochement.modele_bl import BonLivraison, LigneBL
from moteur.rapprochement import pipeline_bl
from moteur.rapprochement.pipeline_bl import (
    RapportRapprochement,
    _desamorcer_conflits_meme_ligne_suivi,
    _est_resolu,
    archiver_bl,
    appliquer_et_archiver,
    ecritures_pour,
    pieces_pour_bl,
    regrouper_par_bl,
    trouver_bon_de_commande,
)


def _bl(fichier, numero_commande="123.096", numero_bl="735136", date_bl="06/08/2026", lignes=None):
    return BonLivraison(
        fournisseur="109 DISTRIBUTION", fichier=fichier, numero_bl=numero_bl,
        date_bl=date_bl, numero_commande=numero_commande, lignes=lignes or [],
    )


def _ligne_suivi(ligne_excel):
    return LigneSuivi(
        ligne_excel=ligne_excel, reference="81000298", designation="", qte_commandee=2,
        qte_livree=0, tarif_bl=None, date_livraison=None, statut="", note="",
    )


def _ligne_bl():
    return LigneBL(reference_fournisseur="81000298", designation="", quantite_livree=2.0, prix_net=32.0, montant=64.0)


def test_regrouper_par_bl_separe_les_statuts_par_bl():

    bl1 = _bl("bl1.pdf")
    bl2 = _bl("bl2.pdf")

    c_sur = Correspondance(_ligne_bl(), _ligne_suivi(5), Statut.SUR)
    c_inconnu = Correspondance(_ligne_bl(), None, Statut.INCONNU, ["ambigu"])

    rapport = RapportRapprochement(surs=[(bl1, c_sur)], inconnus=[(bl2, c_inconnu)])

    groupes = regrouper_par_bl(rapport)

    assert set(groupes) == {id(bl1), id(bl2)}
    assert groupes[id(bl1)]["sur"] == [c_sur]
    assert groupes[id(bl2)]["inconnu"] == [c_inconnu]
    assert groupes[id(bl1)]["inconnu"] == []


def test_regrouper_par_bl_ne_fusionne_pas_deux_bl_du_meme_fichier():
    """BUG RÉEL CORRIGÉ (session R2 suite) : grouper par NOM DE FICHIER
    fusionnait à tort les lignes de plusieurs BL d'un même fichier
    (Cominter Ouest) dans un seul groupe — la clé doit être l'OBJET BL
    (id()), jamais son fichier d'origine, pour permettre l'archivage
    individuel (voir CLAUDE.md, "archivage par BL individuel")."""

    bl1 = _bl("multi.pdf", numero_bl="A")
    bl2 = _bl("multi.pdf", numero_bl="B")

    c1 = Correspondance(_ligne_bl(), _ligne_suivi(5), Statut.SUR)
    c2 = Correspondance(_ligne_bl(), None, Statut.INCONNU, ["ambigu"])

    rapport = RapportRapprochement(surs=[(bl1, c1)], inconnus=[(bl2, c2)])

    groupes = regrouper_par_bl(rapport)

    assert len(groupes) == 2
    assert groupes[id(bl1)]["sur"] == [c1]
    assert groupes[id(bl1)]["inconnu"] == []
    assert groupes[id(bl2)]["inconnu"] == [c2]
    assert groupes[id(bl2)]["sur"] == []


def test_desamorcer_conflits_meme_ligne_suivi_deux_fichiers_meme_cible():
    """Cas réel (session R2 suite, Cominter Ouest) : le même BL déposé
    deux fois (vieux fichier + nouveau scan) proposait chacun une ligne
    "sûre" vers la MÊME ligne du Suivi -> sans ce garde-fou, les deux
    auraient été écrites comme deux incréments indépendants, doublant la
    quantité livrée réellement enregistrée."""

    bl1 = _bl("BL M3.23.030 MABOOC.pdf", numero_commande="M3.23.030")
    bl2 = _bl("doc_nouveau_scan.pdf", numero_commande="M3.23.030")

    ls = _ligne_suivi(5)
    c1 = Correspondance(_ligne_bl(), ls, Statut.SUR)
    c2 = Correspondance(_ligne_bl(), ls, Statut.SUR)

    rapport = RapportRapprochement(surs=[(bl1, c1), (bl2, c2)])

    _desamorcer_conflits_meme_ligne_suivi(rapport)

    assert rapport.surs == [(bl1, c1)]
    assert len(rapport.a_confirmer) == 1
    bl_confirme, c_confirme = rapport.a_confirmer[0]
    assert bl_confirme is bl2
    assert c_confirme.statut is Statut.A_CONFIRMER
    assert "BL M3.23.030 MABOOC.pdf" in c_confirme.raisons[0]


def test_desamorcer_conflits_meme_ligne_suivi_meme_fichier_plusieurs_lignes_ok():
    """Deux lignes DIFFÉRENTES d'un même BL (même fichier) visant des
    lignes Suivi différentes ne doivent jamais être touchées par ce
    garde-fou — seul un conflit ENTRE fichiers doit déclencher quelque
    chose."""

    bl = _bl("bl1.pdf")
    c1 = Correspondance(_ligne_bl(), _ligne_suivi(5), Statut.SUR)
    c2 = Correspondance(_ligne_bl(), _ligne_suivi(6), Statut.SUR)

    rapport = RapportRapprochement(surs=[(bl, c1), (bl, c2)])

    _desamorcer_conflits_meme_ligne_suivi(rapport)

    assert rapport.surs == [(bl, c1), (bl, c2)]
    assert rapport.a_confirmer == []


def test_ecritures_pour_construit_qte_tarif_date():

    bl = _bl("bl1.pdf")
    c = Correspondance(_ligne_bl(), _ligne_suivi(5), Statut.SUR)

    ecritures = ecritures_pour([(bl, c)])

    assert Ecriture(5, "Qté livrée", 2.0) in ecritures
    assert Ecriture(5, "Tarif BL", 32.0) in ecritures
    assert Ecriture(5, "Date de livraison", date(2026, 8, 6)) in ecritures


def test_ecritures_pour_ignore_les_correspondances_deja_a_jour():
    """Garde-fou (défense en profondeur) : une correspondance "déjà à
    jour" ne doit JAMAIS produire d'Ecriture, même si elle se retrouve par
    erreur dans la liste passée — bug réel rencontré en recette (repli
    référence proche traité à tort comme une livraison fraîche, doublant
    une quantité déjà exacte dans le vrai classeur)."""

    bl = _bl("bl1.pdf")
    c_sur = Correspondance(_ligne_bl(), _ligne_suivi(5), Statut.SUR)
    c_deja = Correspondance(_ligne_bl(), _ligne_suivi(6), Statut.DEJA_A_JOUR)

    ecritures = ecritures_pour([(bl, c_sur), (bl, c_deja)])

    assert all(e.ligne != 6 for e in ecritures)
    assert any(e.ligne == 5 for e in ecritures)


def test_ecritures_pour_date_illisible_repli_sur_aujourdhui():

    bl = _bl("bl1.pdf", date_bl="108/2026")  # garbled OCR date (cas réel BL 132.008)
    c = Correspondance(_ligne_bl(), _ligne_suivi(5), Statut.SUR)

    ecritures = ecritures_pour([(bl, c)])

    dates = [e.valeur for e in ecritures if e.colonne == "Date de livraison"]
    assert dates == [date.today()]


def test_archiver_bl_nomme_et_deplace(tmp_path):

    source = tmp_path / "a_traiter" / "BL 123.096.pdf"
    source.parent.mkdir(parents=True)
    source.write_bytes(b"%PDF-1.4 factice")

    # Dossier "Traités" créé par l'acheteur sous a_traiter/BL/, où elle
    # récupère les BL numérisés déjà rapprochés pour agrafer BdC+BL papier
    # et archiver dans les classeurs physiques (demande explicite, session
    # R2 suite — remplace le "rapproches/<fournisseur>/<AAAA-MM>/" imaginé
    # au cadrage R1, jamais confirmé avant cette demande).
    traites = tmp_path / "a_traiter" / "BL" / "Traités"
    bl = _bl("BL 123.096.pdf")

    cible = archiver_bl(source, bl, traites)

    assert not source.exists()
    assert cible.exists()
    # Un sous-dossier par commande, demande explicite de l'acheteur (voir
    # _dossier_pour_commande) — pas directement à plat sous Traités/.
    assert cible.parent == traites / "123.096"
    assert "735136" in cible.name
    assert "123.096" in cible.name


def test_archiver_bl_numero_bl_avec_slash_ne_casse_pas_le_chemin(tmp_path):
    """BUG RÉEL CORRIGÉ (nouveau fournisseur YESSS) : un n° de BL imprimé
    avec un "/" (ex. "CAM/040759") faisait échouer archiver_bl() avec une
    erreur "chemin d'accès introuvable" — Windows interprète le "/" comme
    un séparateur de dossier dans le nom de fichier généré. Le Suivi avait
    pourtant déjà été écrit avec succès à ce stade (constaté en
    recette)."""

    source = tmp_path / "a_traiter" / "BL M4.273.pdf"
    source.parent.mkdir(parents=True)
    source.write_bytes(b"%PDF-1.4 factice")

    traites = tmp_path / "a_traiter" / "BL" / "Traités"
    bl = _bl("BL M4.273.pdf", numero_commande="M4.273", numero_bl="CAM/040759")

    cible = archiver_bl(source, bl, traites)

    assert not source.exists()
    assert cible.exists()
    assert cible.parent == traites / "M4.273"
    assert "CAM-040759" in cible.name


def test_archiver_bl_evite_lecrasement(tmp_path):

    traites = tmp_path / "Traités"
    bl = _bl("BL 123.096.pdf")

    source1 = tmp_path / "BL 123.096.pdf"
    source1.write_bytes(b"1")
    cible1 = archiver_bl(source1, bl, traites)

    source2 = tmp_path / "BL 123.096_bis.pdf"
    source2.write_bytes(b"2")
    cible2 = archiver_bl(source2, bl, traites)

    assert cible1 != cible2
    assert cible1.exists() and cible2.exists()


def test_appliquer_et_archiver_une_archive_qui_echoue_nempeche_pas_le_reste(tmp_path, monkeypatch):
    """Cas réel rencontré en recette (session R2 suite) : le déplacement
    d'un BL vers Traités/ a échoué (fichier verrouillé par autre chose,
    réseau/antivirus) — ça ne doit PAS faire perdre le résumé de
    l'écriture Suivi déjà faite, ni empêcher le rapport final, ni bloquer
    l'archivage des AUTRES BL du même lot."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    ws.append(["REF1", 10, 0, None, None, None])
    ws.append(["REF2", 5, 0, None, None, None])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    installer_feuille_pieces(chemin_suivi, tmp_path / "backups")  # P1 : la feuille Pièces doit exister

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "bl_verrouille.pdf").write_bytes(b"1")
    (dossier_a_traiter / "bl_ok.pdf").write_bytes(b"2")

    bl_verrouille = _bl("bl_verrouille.pdf", numero_commande="C1")
    bl_ok = _bl("bl_ok.pdf", numero_commande="C2")

    c1 = Correspondance(
        LigneBL("REF1", "", 10.0, 1.0, 10.0),
        LigneSuivi(2, "REF1", "", 10, 0, None, None, "", ""),
        Statut.SUR,
    )
    c2 = Correspondance(
        LigneBL("REF2", "", 5.0, 1.0, 5.0),
        LigneSuivi(3, "REF2", "", 5, 0, None, None, "", ""),
        Statut.SUR,
    )

    rapport = RapportRapprochement(surs=[(bl_verrouille, c1), (bl_ok, c2)], fichier_suivi=chemin_suivi)

    appel_original = pipeline_bl.archiver_bl

    def _archiver_bl_simule_verrou(chemin_source, bl, dossier_traites, dossier_commandes_bc=None):
        if bl.fichier == "bl_verrouille.pdf":
            raise OSError("simulé : fichier verrouillé")
        return appel_original(chemin_source, bl, dossier_traites, dossier_commandes_bc)

    monkeypatch.setattr(pipeline_bl, "archiver_bl", _archiver_bl_simule_verrou)

    resume = appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, rapport.surs)

    assert resume["lignes_ecrites"] == 2
    # P1 : une ligne Pièces de type BL par ligne écrite, Fichier = chemin
    # d'archive (même nom que l'archivage qui suit), y compris pour le BL
    # dont l'archivage a ensuite échoué (la ligne de document existe).
    assert resume["pieces_ecrites"] == 2
    pieces = lire_pieces(chemin_suivi)
    assert [p["Type"] for p in pieces] == ["BL", "BL"]
    assert [p["N° pièce"] for p in pieces] == ["735136", "735136"]
    assert [p["N° de commande"] for p in pieces] == ["C1", "C2"]
    assert pieces[0]["Référence Suivi"] == "REF1" and pieces[0]["Qté"] == 10.0 and pieces[0]["Montant HT"] == 10.0
    assert pieces[0]["Fichier"] == "2026-08-06 - 109 DISTRIBUTION - 735136 - BC C1.pdf"
    assert pieces[0]["Mode de rapprochement"] == "Auto"
    assert [f for f, _ in resume["bl_archives"]] == ["bl_ok.pdf"]
    assert [f for f, _ in resume["archivage_echoue"]] == ["bl_verrouille.pdf"]
    assert resume["chemin_rapport"].exists()
    assert (dossier_a_traiter / "bl_verrouille.pdf").exists()  # jamais supprimé si l'archivage a échoué


def test_appliquer_et_archiver_deplace_les_bl_a_confirmer_vers_a_verifier(tmp_path):
    """Demande explicite de l'acheteur (session R2 suite) : un BL avec au
    moins une ligne "inconnue" (ex. référence différente entre BL et
    Suivi, jamais rapprochée automatiquement) doit être déplacé vers
    a_traiter/BL/À vérifier/ — jamais laissé mélangé avec les BL pas
    encore traités, pour repérer facilement ceux qui attendent une
    décision humaine."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    installer_feuille_pieces(chemin_suivi, tmp_path / "backups")  # P1 : la feuille Pièces doit exister

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "bl_litigieux.pdf").write_bytes(b"1")

    bl = _bl("bl_litigieux.pdf", numero_commande="131.157")
    c_inconnu = Correspondance(
        LigneBL("GOUJON8X75", "", 100.0, None, None),
        None,
        Statut.INCONNU,
        ["Aucune ligne du Suivi ne correspond à cette référence pour cette commande"],
    )

    rapport = RapportRapprochement(inconnus=[(bl, c_inconnu)], fichier_suivi=chemin_suivi)

    resume = appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, [])

    assert resume["lignes_ecrites"] == 0
    assert resume["bl_archives"] == []
    assert not (dossier_a_traiter / "bl_litigieux.pdf").exists()

    [(fichier, cible, raisons)] = resume["bl_a_verifier"]
    assert fichier == "bl_litigieux.pdf"
    assert cible == dossier_a_traiter / "À vérifier" / "bl_litigieux.pdf"
    assert cible.exists()
    assert "Aucune ligne du Suivi" in raisons[0]


def test_appliquer_et_archiver_fichier_multi_bl_archive_le_resolu_sans_attendre_ses_freres(tmp_path):
    """BUG RÉEL CRITIQUE trouvé en recette (session R2 suite, Cominter
    Ouest) : un fichier contenant PLUSIEURS BL (voir moteur.ocr.
    pages_par_identifiant) restait bloqué EN BLOC dès qu'UN de ses BL
    n'était pas résolu (ex. n° de commande introuvable) — sur 8 BL scannés
    dans un seul fichier lors de la recette réelle, un seul avait fini
    correctement traité et renommé dans Traités/. Chaque BL doit désormais
    être extrait (découpage PDF par pages, voir bl.pages) et
    archivé/déplacé INDIVIDUELLEMENT, sans attendre ses frères du même
    fichier."""

    import fitz

    from conftest import FIXTURES

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    ws.append(["REF1", 10, 10, 1.0, "2026-08-13", None])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    installer_feuille_pieces(chemin_suivi, tmp_path / "backups")  # P1 : la feuille Pièces doit exister

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    source = dossier_a_traiter / "bl_multi.pdf"
    # Vrai PDF de 8 pages (voir tests/test_parsers_bl_cominter.py) — un
    # fichier factice ne suffit pas ici, le découpage par pages a besoin
    # d'un PDF réellement ouvrable.
    source.write_bytes((FIXTURES / "bl_cominter_6_multi_bl_8pages.pdf").read_bytes())

    # BL n°1 du fichier (page 0) : entièrement à jour (résolu).
    bl_resolu = BonLivraison(
        fournisseur="COMINTER", fichier="bl_multi.pdf", numero_bl="OBL1",
        date_bl="06/08/2026", numero_commande="C1", pages=[0],
    )
    c_deja_a_jour = Correspondance(
        LigneBL("REF1", "", 10.0, 1.0, 10.0),
        LigneSuivi(2, "REF1", "", 10, 10, 1.0, "2026-08-13", "", ""),
        Statut.DEJA_A_JOUR,
    )

    # BL n°2 du MÊME fichier (page 1) : commande introuvable -> anomalie
    # rattachée à CE BL précisément (rapport.anomalies_bl), jamais de
    # correspondance produite pour lui.
    bl_anomalie = BonLivraison(
        fournisseur="COMINTER", fichier="bl_multi.pdf", numero_bl="OBL2",
        date_bl="07/08/2026", numero_commande="", pages=[1],
    )

    rapport = RapportRapprochement(
        deja_a_jour=[(bl_resolu, c_deja_a_jour)],
        anomalies_bl=[(bl_anomalie, "N° de commande introuvable sur le BL")],
        fichier_suivi=chemin_suivi,
    )

    resume = appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, [])

    [(fichier_archive, cible_archivee)] = resume["bl_archives"]
    assert fichier_archive == "bl_multi.pdf"
    assert cible_archivee.exists()
    # Sous-dossier par commande (C1, voir bl_resolu) sous Traités/.
    assert cible_archivee.parent.name == "C1"
    assert cible_archivee.parent.parent.name == "Traités"

    [(fichier_verif, cible_verif, raisons)] = resume["bl_a_verifier"]
    assert fichier_verif == "bl_multi.pdf"
    assert cible_verif.exists()
    assert cible_verif.parent.name == "À vérifier"
    assert "N° de commande introuvable" in raisons[0]

    # Les 2 BL traités ne couvrent que 2 des 8 pages d'origine : le
    # fichier source doit avoir été réduit aux 6 pages restantes, jamais
    # supprimé ni laissé intact.
    assert source.exists()
    with fitz.open(source) as doc:
        assert doc.page_count == 6


def test_appliquer_et_archiver_fichier_multi_bl_supprime_la_source_une_fois_tout_redistribue(tmp_path):
    """Une fois que TOUS les BL d'un fichier multi-BL ont été extraits
    (archivés ou déplacés vers "à vérifier"), le fichier source ne doit
    plus rester dans a_traiter/BL/ avec la totalité de ses pages — sinon
    il serait relu en entier (OCR + matching) à chaque exécution suivante
    pour rien."""

    from conftest import FIXTURES

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    ws.append(["REF1", 10, 10, 1.0, "2026-08-13", None])
    ws.append(["REF2", 5, 0, None, None, None])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    installer_feuille_pieces(chemin_suivi, tmp_path / "backups")  # P1 : la feuille Pièces doit exister

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    source = dossier_a_traiter / "bl_multi.pdf"
    source.write_bytes((FIXTURES / "bl_cominter_6_multi_bl_8pages.pdf").read_bytes())

    bl1 = BonLivraison(
        fournisseur="COMINTER", fichier="bl_multi.pdf", numero_bl="OBL1",
        date_bl="06/08/2026", numero_commande="C1", pages=[0],
    )
    c1 = Correspondance(
        LigneBL("REF1", "", 10.0, 1.0, 10.0),
        LigneSuivi(2, "REF1", "", 10, 10, 1.0, "2026-08-13", "", ""),
        Statut.DEJA_A_JOUR,
    )

    # Couvre le RESTE des 8 pages du fixture (pas juste la page 1) : les
    # deux BL doivent consommer la TOTALITÉ des pages pour que le fichier
    # source soit supprimé (sinon il serait seulement réduit, voir le test
    # précédent).
    bl2 = BonLivraison(
        fournisseur="COMINTER", fichier="bl_multi.pdf", numero_bl="OBL2",
        date_bl="06/08/2026", numero_commande="C2", pages=[1, 2, 3, 4, 5, 6, 7],
    )
    c2 = Correspondance(
        LigneBL("REF2", "", 5.0, 1.0, 5.0),
        LigneSuivi(3, "REF2", "", 5, 0, None, None, "", ""),
        Statut.SUR,
    )

    rapport = RapportRapprochement(deja_a_jour=[(bl1, c1)], surs=[(bl2, c2)], fichier_suivi=chemin_suivi)

    resume = appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, rapport.surs)

    assert len(resume["bl_archives"]) == 2
    assert resume["bl_a_verifier"] == []
    assert not source.exists()


def test_appliquer_et_archiver_un_seul_bl_mais_pages_partielles_est_decoupe_pas_deplace_en_bloc(tmp_path):
    """Détection de fournisseur par page (moteur.rapprochement.lecture_bl) :
    un fichier peut désormais produire un SEUL BonLivraison dont bl.pages
    ne couvre qu'UNE PARTIE des pages du fichier (les autres appartiennent
    à un fournisseur différent, ou n'ont pas pu être lues). Ce cas ne doit
    JAMAIS emprunter le chemin historique "fichier entier déplacé tel
    quel" (_traiter_bl_unique_du_fichier) — sinon les pages qui n'ont rien
    à voir avec ce BL partiraient avec lui. Il doit être extrait comme un
    fichier multi-BL ordinaire (une seule page ici), laissant le reste du
    fichier source intact pour un futur passage."""

    from conftest import FIXTURES

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    ws.append(["REF1", 10, 10, 1.0, "2026-08-13", None])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    installer_feuille_pieces(chemin_suivi, tmp_path / "backups")  # P1 : la feuille Pièces doit exister

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    source = dossier_a_traiter / "bl_multi.pdf"
    # Réutilise le fixture 8 pages (peu importe son contenu réel ici, seul
    # le nombre de pages compte) — un seul BL, page 0 seulement.
    source.write_bytes((FIXTURES / "bl_cominter_6_multi_bl_8pages.pdf").read_bytes())

    bl = BonLivraison(
        fournisseur="COMINTER", fichier="bl_multi.pdf", numero_bl="OBL1",
        date_bl="06/08/2026", numero_commande="C1", pages=[0],
    )
    c = Correspondance(
        LigneBL("REF1", "", 10.0, 1.0, 10.0),
        LigneSuivi(2, "REF1", "", 10, 10, 1.0, "2026-08-13", "", ""),
        Statut.DEJA_A_JOUR,
    )

    rapport = RapportRapprochement(deja_a_jour=[(bl, c)], fichier_suivi=chemin_suivi)

    resume = appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, [])

    [(fichier_archive, cible_archivee)] = resume["bl_archives"]
    assert fichier_archive == "bl_multi.pdf"
    assert cible_archivee.exists()
    assert cible_archivee.parent.name == "C1"

    # Le fichier source doit avoir été RÉDUIT (page 0 extraite), pas
    # déplacé/renommé en bloc, et pas non plus laissé intact.
    import fitz
    assert source.exists()
    with fitz.open(source) as doc:
        assert doc.page_count == 7


def test_appliquer_et_archiver_anomalie_de_lecture_necoupe_pas_les_bl_resolus_du_meme_fichier(tmp_path):
    """Détection de fournisseur par page : un fichier peut avoir À LA FOIS
    une anomalie de lecture (ex. une page d'un fournisseur reconnu mais
    sans parser BL) ET un BonLivraison résolu depuis une AUTRE page. Avant
    correction, la présence de CE fichier dans rapport.anomalies_lecture
    suffisait à le faire déplacer ENTIER vers "à vérifier" avant même que
    son BL résolu ait pu être archivé individuellement — le BL résolu ne
    doit jamais être perdu/bloqué par une anomalie concernant une AUTRE
    page du même fichier."""

    from conftest import FIXTURES

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    ws.append(["REF1", 10, 10, 1.0, "2026-08-13", None])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    installer_feuille_pieces(chemin_suivi, tmp_path / "backups")  # P1 : la feuille Pièces doit exister

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    source = dossier_a_traiter / "bl_multi.pdf"
    source.write_bytes((FIXTURES / "bl_cominter_6_multi_bl_8pages.pdf").read_bytes())

    bl = BonLivraison(
        fournisseur="COMINTER", fichier="bl_multi.pdf", numero_bl="OBL1",
        date_bl="06/08/2026", numero_commande="C1", pages=[0],
    )
    c = Correspondance(
        LigneBL("REF1", "", 10.0, 1.0, 10.0),
        LigneSuivi(2, "REF1", "", 10, 10, 1.0, "2026-08-13", "", ""),
        Statut.DEJA_A_JOUR,
    )

    rapport = RapportRapprochement(
        deja_a_jour=[(bl, c)],
        anomalies_lecture=[("bl_multi.pdf", "Page(s) 2 : fournisseur RAVATE PRO reconnu mais pas encore de parser BL")],
        fichier_suivi=chemin_suivi,
    )

    resume = appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, [])

    # Le BL résolu (page 0) a bien été archivé — PAS avalé par l'anomalie
    # de lecture, qui ne concerne qu'une autre page du même fichier.
    [(fichier_archive, cible_archivee)] = resume["bl_archives"]
    assert fichier_archive == "bl_multi.pdf"
    assert cible_archivee.exists()

    # L'anomalie de lecture n'a PAS déplacé le fichier entier vers "à
    # vérifier" (rien d'autre à y déplacer : elle ne correspond à aucun
    # BonLivraison, juste une page non reconnue).
    assert resume["bl_a_verifier"] == []


# --- Dossier par commande (BC + BL + retours), demande explicite de
# l'acheteur : "dans traités, il faudra créer un dossier pour chaque
# commande [...] dedans on y met ce bon de commande, tous les BL et bons de
# retours associés" ---------------------------------------------------------


def test_trouver_bon_de_commande_un_seul_candidat(tmp_path):

    chantier = tmp_path / "Lacouture"
    chantier.mkdir()
    (chantier / "Lacouture - BC 108.271 - COMINTER.pdf").write_bytes(b"x")

    resultat = trouver_bon_de_commande("108.271", tmp_path)

    assert resultat == chantier / "Lacouture - BC 108.271 - COMINTER.pdf"


def test_trouver_bon_de_commande_fichier_a_plat_ou_en_sous_dossier(tmp_path):
    """L'archive réelle mélange fichiers en vrac à la racine ET rangés dans
    des sous-dossiers par chantier — recherche récursive dans les deux cas."""

    (tmp_path / "Raccordement Eclairage DEAL - BC 162.001 - COMINTER.pdf").write_bytes(b"x")
    sous_dossier = tmp_path / "Maintenance"
    sous_dossier.mkdir()
    (sous_dossier / "Bornes de Recharge - BC M2.16.011 - COREDIME.pdf").write_bytes(b"x")

    assert trouver_bon_de_commande("162.001", tmp_path) is not None
    assert trouver_bon_de_commande("M2.16.011", tmp_path) == (
        sous_dossier / "Bornes de Recharge - BC M2.16.011 - COREDIME.pdf"
    )


def test_trouver_bon_de_commande_aucun_candidat_retourne_none(tmp_path):

    (tmp_path / "Lacouture - BC 108.271 - COMINTER.pdf").write_bytes(b"x")

    assert trouver_bon_de_commande("142.033", tmp_path) is None


def test_trouver_bon_de_commande_plusieurs_candidats_ambigus_retourne_none(tmp_path):
    """Jamais un choix au hasard entre plusieurs BC qui matcheraient le même
    numéro (règle d'or du projet)."""

    (tmp_path / "Chantier A - BC 108.271 - COMINTER.pdf").write_bytes(b"x")
    (tmp_path / "Chantier A - BC 108.271 - COMINTER.xlsx").write_bytes(b"x")

    assert trouver_bon_de_commande("108.271", tmp_path) is None


def test_trouver_bon_de_commande_dossier_ou_numero_absent_retourne_none(tmp_path):

    assert trouver_bon_de_commande("108.271", tmp_path / "n_existe_pas") is None


def test_trouver_bon_de_commande_deux_copies_identiques_pas_ambigu(tmp_path):
    """BUG RÉEL CORRIGÉ (signalé par l'acheteur, commande M3.23.033) : le
    même BC se retrouve couramment archivé à la fois dans "Commandes/
    <année>/" ET "Commandes/BdCPDF/" — deux fichiers de MÊME NOM ET MÊME
    TAILLE, donc pas une vraie ambiguïté (contrairement au test précédent,
    où les 2 fichiers ont des noms différents). Doit renvoyer l'un des
    deux, jamais None."""

    contenu = b"contenu identique du meme BC"
    sous_dossier_1 = tmp_path / "2026" / "Maintenance"
    sous_dossier_1.mkdir(parents=True)
    sous_dossier_2 = tmp_path / "BdCPDF"
    sous_dossier_2.mkdir()

    (sous_dossier_1 / "Mabooc SODEGIS - BC M3.23.033 - RAVATE ELEC.pdf").write_bytes(contenu)
    (sous_dossier_2 / "Mabooc SODEGIS - BC M3.23.033 - RAVATE ELEC.pdf").write_bytes(contenu)

    resultat = trouver_bon_de_commande("M3.23.033", tmp_path)

    assert resultat is not None
    assert resultat.name == "Mabooc SODEGIS - BC M3.23.033 - RAVATE ELEC.pdf"
    assert trouver_bon_de_commande("", tmp_path) is None
    assert trouver_bon_de_commande("108.271", None) is None


def test_archiver_bl_copie_le_bon_de_commande_trouve(tmp_path):

    source = tmp_path / "BL 123.096.pdf"
    source.write_bytes(b"%PDF-1.4 factice")

    dossier_bc = tmp_path / "Commandes" / "2026"
    dossier_bc.mkdir(parents=True)
    (dossier_bc / "Chantier X - BC 123.096 - 109 DISTRIBUTION.pdf").write_bytes(b"%PDF-1.4 bc")

    traites = tmp_path / "Traités"
    bl = _bl("BL 123.096.pdf")

    cible = archiver_bl(source, bl, traites, dossier_bc)

    fichiers_bc = list(cible.parent.glob("BC - *"))
    assert len(fichiers_bc) == 1
    assert fichiers_bc[0].name == "BC - Chantier X - BC 123.096 - 109 DISTRIBUTION.pdf"


def test_archiver_bl_ne_recopie_pas_le_bc_deja_present(tmp_path):
    """Idempotent : un 2e BL de la même commande archivé plus tard ne
    recherche/copie rien de plus si le BC est déjà là."""

    dossier_bc = tmp_path / "Commandes" / "2026"
    dossier_bc.mkdir(parents=True)
    (dossier_bc / "Chantier X - BC 123.096 - 109 DISTRIBUTION.pdf").write_bytes(b"%PDF-1.4 bc")

    traites = tmp_path / "Traités"

    source1 = tmp_path / "BL 123.096.pdf"
    source1.write_bytes(b"1")
    cible1 = archiver_bl(source1, _bl("BL 123.096.pdf", numero_bl="735136"), traites, dossier_bc)

    # Le BC "disparaît" ensuite de son dossier d'origine (déjà copié) —
    # s'il était recherché une 2e fois, trouver_bon_de_commande ne le
    # retrouverait plus, prouvant que la copie n'a bien lieu qu'une fois.
    (dossier_bc / "Chantier X - BC 123.096 - 109 DISTRIBUTION.pdf").unlink()

    source2 = tmp_path / "BL 123.096_bis.pdf"
    source2.write_bytes(b"2")
    archiver_bl(source2, _bl("BL 123.096_bis.pdf", numero_bl="735137"), traites, dossier_bc)

    fichiers_bc = list(cible1.parent.glob("BC - *"))
    assert len(fichiers_bc) == 1


def test_archiver_bl_sans_bc_trouve_narchive_pas_moins_bien(tmp_path):
    """L'absence de BC (poste sans accès au dossier réseau, ou BC pas
    encore filé) ne doit jamais empêcher l'archivage du BL lui-même."""

    source = tmp_path / "BL 999.pdf"
    source.write_bytes(b"1")
    traites = tmp_path / "Traités"

    cible = archiver_bl(source, _bl("BL 999.pdf", numero_commande="999.999"), traites, None)

    assert cible.exists()
    assert list(cible.parent.glob("BC - *")) == []


def test_est_resolu_bon_de_retour_toujours_resolu():
    """Un bon de retour n'a par nature rien à écrire (voir
    rapprocher_dossier) — il ne doit jamais rester coincé dans "à vérifier"
    à cause de l'anomalie informative que rapprocher_dossier lui attache
    systématiquement."""

    bl_retour = BonLivraison(
        fournisseur="109 DISTRIBUTION", fichier="retour.pdf", numero_bl="25894",
        date_bl="14/08/2026", numero_commande="M3.10.175", type_document="RETOUR",
        numero_bl_origine="737760",
    )
    g = {
        "bl": bl_retour, "sur": [], "a_confirmer": [], "deja_a_jour": [], "inconnu": [],
        "anomalies": ["Bon de retour — annule R9PRA263 du BL 737760 : rien à écrire depuis ce document"],
    }

    assert _est_resolu(g, set()) is True


def test_appliquer_et_archiver_bon_de_retour_rejoint_traites_avec_sa_commande(tmp_path):

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    installer_feuille_pieces(chemin_suivi, tmp_path / "backups")  # P1 : la feuille Pièces doit exister

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "retour.pdf").write_bytes(b"1")

    bl_retour = BonLivraison(
        fournisseur="109 DISTRIBUTION", fichier="retour.pdf", numero_bl="25894",
        date_bl="14/08/2026", numero_commande="M3.10.175", type_document="RETOUR",
        numero_bl_origine="737760",
    )

    rapport = RapportRapprochement(
        anomalies_bl=[(bl_retour, "Bon de retour — annule R9PRA263 du BL 737760 : rien à écrire")],
        fichier_suivi=chemin_suivi,
    )

    resume = appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, [])

    [(fichier, cible)] = resume["bl_archives"]
    assert fichier == "retour.pdf"
    assert cible.exists()
    assert cible.parent.name == "M3.10.175"
    assert cible.parent.parent.name == "Traités"
    assert "RETOUR" in cible.name
    assert resume["bl_a_verifier"] == []


# --- P1 : lignes Pièces de type BL -------------------------------------------


def test_pieces_pour_bl_une_ligne_par_ligne_ecrite(tmp_path):
    bl = _bl("bl.pdf", numero_commande="123.096", numero_bl="735136", date_bl="06/08/2026")
    ls = LigneSuivi(5, "81000298", "", 2, 0, None, None, "", "", chantier="Chantier A", sous_chantier="SC1")
    c_sur = Correspondance(_ligne_bl(), ls, Statut.SUR)
    c_confirme = Correspondance(LigneBL("X1", "Désig", 3.0, None, None), ls, Statut.A_CONFIRMER, ["tarif"])
    c_deja = Correspondance(_ligne_bl(), ls, Statut.DEJA_A_JOUR)

    pieces = pieces_pour_bl([(bl, c_sur), (bl, c_confirme), (bl, c_deja)], tmp_path / "Traités")

    assert len(pieces) == 2  # « déjà à jour » ne produit rien
    p = pieces[0]
    assert p["Type"] == "BL" and p["Fournisseur"] == "109 Distribution"  # nom Suivi via MAPPING_FOURNISSEURS
    assert p["N° pièce"] == "735136" and p["Date pièce"] == date(2026, 8, 6)
    assert p["N° de commande"] == "123.096" and p["Chantier"] == "Chantier A" and p["Sous-Chantier"] == "SC1"
    assert p["Référence Suivi"] == "81000298" and p["Référence fournisseur"] == "81000298"
    assert p["Qté"] == 2.0 and p["PU HT"] == 32.0 and p["Montant HT"] == 64.0
    assert p["Mode de rapprochement"] == "Auto" and p["Commentaire"] is None
    assert p["Fichier"].texte.endswith('123.096\\2026-08-06 - 109 DISTRIBUTION - 735136 - BC 123.096.pdf","2026-08-06 - 109 DISTRIBUTION - 735136 - BC 123.096.pdf")')
    assert p["ID pièce"] == "109 Distribution|BL|735136|123.096|81000298|"
    q = pieces[1]
    assert q["Mode de rapprochement"] == "Confirmé"
    assert q["Montant HT"] is None and q["PU HT"] is None  # aucun prix sur le BL : rien d'inventé


def test_pieces_pour_bl_montant_recalcule_est_commente():
    bl = _bl("bl.pdf")
    c = Correspondance(LigneBL("R", "", 4.0, 2.5, None), _ligne_suivi(5), Statut.SUR)
    [p] = pieces_pour_bl([(bl, c)], "Traités")
    assert p["Montant HT"] == 10.0 and p["Commentaire"] == "montant recalculé"


def test_appliquer_et_archiver_refuse_sans_feuille_pieces(tmp_path):
    """P1 : sans feuille Pièces, rien n'est écrit (ni Commandes ni Pièces),
    rien n'est déplacé — jamais une Qté livrée sans sa ligne de document."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(["Référence", "Qté commandée", "Qté livrée", "Tarif BL", "Date de livraison", "Note"])
    ws.append(["REF1", 10, 0, None, None, None])
    chemin_suivi = tmp_path / "suivi.xlsx"
    wb.save(chemin_suivi)
    avant = chemin_suivi.read_bytes()

    dossier_a_traiter = tmp_path / "a_traiter" / "BL"
    dossier_a_traiter.mkdir(parents=True)
    (dossier_a_traiter / "bl.pdf").write_bytes(b"1")
    c = Correspondance(_ligne_bl(), _ligne_suivi(2), Statut.SUR)
    rapport = RapportRapprochement(surs=[(_bl("bl.pdf"), c)], fichier_suivi=chemin_suivi)

    with pytest.raises(FeuillePiecesAbsente):
        appliquer_et_archiver(tmp_path, dossier_a_traiter, rapport, rapport.surs)
    assert chemin_suivi.read_bytes() == avant
    assert (dossier_a_traiter / "bl.pdf").exists()
