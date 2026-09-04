"""
P1 — Feuille « Pièces » : installation sur le Suivi commandes VIVANT
(étape 7 du plan, voir CLAUDE.md « Feuille Pièces — modèle, socles,
migration ») et passage d'un lot de factures via les pipelines réécrits.

    py -3 installer_pieces.py vivant     # 0d -> 1 -> 3 -> 4 -> 5 + contrôles
    py -3 installer_pieces.py lot        # lot de factures (lignes SÛRES seulement)
    py -3 installer_pieces.py copie <chemin.xlsx>   # même chose sur une copie

Garde-fous du mode `vivant`, dans l'ordre : verrou Excel absent ; hash MD5
du vivant identique au dernier backup permanent (backups/permanents/…AVANT
PIECES…) — sinon STOP (le classeur a changé depuis le relevé, refaire
l'étape 0) ; sauvegarde horodatée (rotation) ; relevé des 5 colonnes
facture avant bascule ; chaque étape suivie de verifier_structure() ;
puis vérification par Excel lui-même sur une COPIE du vivant écrit
(aucune réparation, 17 tableaux, KPI Dashboard, recalcul complet) et
contrôle des 5 colonnes (0 écart attendu hors PU moyen pondéré). STOP à la
première anomalie, exception propagée.
"""

import glob
import hashlib
import json
import shutil
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from moteur.rapprochement.ecriture import est_verrouille, sauvegarder
from moteur.rapprochement.installation_pieces import COLONNES_LIGNE2_CONSTATEES, executer, verifier_structure
from moteur.rapprochement.migration_pieces import lire_lignes_facturees
from moteur.rapprochement.pieces import COLONNES_FACTURE_CALCULEES, lire_pieces
from moteur.rapprochement.pipeline_bl import trouver_fichier_suivi_vivant
from moteur.rapprochement.pipeline_facture import (
    DOSSIER_A_TRAITER_FACTURES,
    appliquer_et_archiver_factures,
    rapprocher_dossier_factures,
)
from moteur.rapprochement.verification_excel import resume_verification, verifier_dans_excel

PROJET = Path(__file__).resolve().parent
BACKUPS = PROJET / "backups"
DOSSIER_VERIF = PROJET / "rapports" / "verif_excel"

KPI_CELLULES = ("B6", "E6", "H6", "K6", "N6", "Q6", "B9", "E9", "H9", "K9", "N9", "Q9")
# Original intact RECALCULÉ par Excel (2026-09-04) ; la réparation de la
# ligne 2 (étape 0d) ajoute légitimement 218,54 € à B6/N6/H9.
KPI_ORIGINAL_RECALCULE = {
    "B6": 2103024.8774794196, "E6": 1732.0, "H6": 6554.0, "K6": 119.0, "N6": 886419.7314460782,
    "Q6": -0.5, "B9": 61.0, "E9": 1316.80362, "H9": 856458.0341660788, "K9": 174.0,
    "N9": 1882.5600000000002, "Q9": 0.6899603295697284,
}
KPI_ATTENDU = {k: v + (218.54 if k in ("B6", "N6", "H9") else 0.0) for k, v in KPI_ORIGINAL_RECALCULE.items()}


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def md5(chemin) -> str:
    h = hashlib.md5()
    with open(chemin, "rb") as f:
        for bloc in iter(lambda: f.read(1 << 20), b""):
            h.update(bloc)
    return h.hexdigest()


def _controle_5_colonnes(recalcule, avant) -> tuple:
    wb = load_workbook(recalcule, read_only=True, data_only=True)
    try:
        ws = wb["Commandes"]
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c is not None else None for c in next(rows)]
        idx = {h: i for i, h in enumerate(hdr) if h}
        par_ligne, statuts = {}, Counter()
        for r_, row in enumerate(rows, start=2):
            par_ligne[r_] = {c: (row[idx[c]] if idx[c] < len(row) else None)
                             for c in COLONNES_FACTURE_CALCULEES + ("Statut facture",)}
            statuts[par_ligne[r_]["Statut facture"]] += 1
    finally:
        wb.close()
    compteur, pu_max = Counter(), 0.0
    for l in avant:
        n = par_ligne[l.ligne_excel]
        if str(n["N° facture"] or "") != str(l.numero_facture):
            compteur["N° facture"] += 1
        if abs(float(n["Qté facturée"] or 0) - float(l.qte or 0)) > 0.001:
            compteur["Qté"] += 1
        if abs(float(n["Montant facturé HT"] or 0) - float(l.montant or 0)) > 0.005:
            compteur["Montant"] += 1
        d = l.date_facture.date() if hasattr(l.date_facture, "date") else l.date_facture
        nd = n["Date facture"].date() if hasattr(n["Date facture"], "date") else n["Date facture"]
        if d is not None and d != nd:
            compteur["Date"] += 1
        if l.pu is not None and n["PU facturé"] not in (None, "") and abs(float(n["PU facturé"]) - float(l.pu)) > 0.0001:
            compteur["PU (moyen pondéré)"] += 1
            pu_max = max(pu_max, abs(float(n["PU facturé"]) - float(l.pu)))
    return compteur, pu_max, statuts


def installer(fichier, est_vivant: bool) -> None:
    fichier = Path(fichier)
    DOSSIER_VERIF.mkdir(parents=True, exist_ok=True)
    log(f"classeur : {fichier}")
    assert not est_verrouille(fichier), "STOP : verrou Excel présent"

    if est_vivant:
        permanents = sorted(glob.glob(str(BACKUPS / "permanents" / "*AVANT PIECES*.xlsx")))
        assert permanents, "STOP : aucun backup permanent (étape 0b)"
        h_vivant, h_perm = md5(fichier), md5(permanents[-1])
        log(f"md5 vivant {h_vivant} / backup permanent {h_perm}")
        assert h_vivant == h_perm, "STOP : le vivant a changé depuis le backup permanent — refaire l'étape 0"
        assert not est_verrouille(fichier), "STOP : verrou Excel présent (2e contrôle)"
        log(f"sauvegarde rotation : {sauvegarder(fichier, BACKUPS)}")

    avant = lire_lignes_facturees(fichier)
    json.dump([l.__dict__ for l in avant], open(DOSSIER_VERIF / "valeurs_5_colonnes_avant.json", "w", encoding="utf-8"),
              default=str, ensure_ascii=False)
    log(f"{len(avant)} lignes facturées relevées avant bascule")

    t0 = time.time()
    res = executer(fichier, PROJET, BACKUPS, etapes=("0d", "1", "3", "4", "5"), journal=log,
                   colonnes_ligne2=COLONNES_LIGNE2_CONSTATEES)
    log(f"étapes 0d→5 terminées en {round(time.time() - t0)} s")
    r4 = res["4"]
    log(f"migration : {r4['statistiques']} ; PDF lus {r4['pdf_lus']} ; au centime avant {r4['au_centime_avant_ecriture']} "
        f"/ après {r4['au_centime']} ; rapport {r4['chemin_rapport']}")
    for f, a, b, e in r4["sommes"]:
        log(f"  {f}: Commandes {a:.2f} / Pièces {b:.2f} / écart {e:+.2f}")
    assert r4["au_centime"], "STOP : contrôle au centime KO"

    copie = DOSSIER_VERIF / f"{fichier.stem} - verif Excel.xlsx"
    recalcule = DOSSIER_VERIF / f"{fichier.stem} - recalcule.xlsx"
    shutil.copy2(fichier, copie)
    cellules = [("Dashboard", a) for a in KPI_CELLULES] + [("Commandes", a) for a in ("Y2", "Z2", "AA2", "BD1", "BG1")]
    r = verifier_dans_excel(copie, recalculer=True, cellules=cellules, enregistrer_sous=recalcule)
    print(resume_verification(r), flush=True)
    assert r["ouvert"] and not r["reparation"] and not r["erreur"], "STOP : Excel a réparé ou n'a pas ouvert la copie"
    assert "Pièces" in r["feuilles"] and "Pieces" in r["tableaux"], "STOP : feuille/tableau Pièces absents après ouverture Excel"
    if est_vivant:
        for (feuille, adr), v in r["cellules"].items():
            if feuille == "Dashboard":
                attendu = KPI_ATTENDU[adr]
                log(f"  KPI {adr}: {v} (attendu {attendu:.2f}) {'OK' if abs(float(v) - attendu) < 0.01 else 'ÉCART — à expliquer'}")

    compteur, pu_max, statuts = _controle_5_colonnes(recalcule, avant)
    log(f"contrôle 5 colonnes sur {len(avant)} lignes : écarts {dict(compteur)} (PU : écart max {pu_max:.4f})")
    log(f"Statut facture : {dict(statuts)}")
    assert not (set(compteur) - {"PU (moyen pondéré)"}), "STOP : écart hors PU"
    copie.unlink(missing_ok=True)
    log("Installation P1 vérifiée.")


def lot_factures() -> None:
    vivant = trouver_fichier_suivi_vivant(PROJET)
    assert not est_verrouille(vivant), "STOP : verrou Excel"
    pieces_avant = lire_pieces(vivant)
    assert pieces_avant is not None, "STOP : feuille Pièces absente — lancer `installer_pieces.py vivant` d'abord"
    log(f"Pièces avant le lot : {len(pieces_avant)} lignes")

    rapport = rapprocher_dossier_factures(PROJET / DOSSIER_A_TRAITER_FACTURES, PROJET)
    log(f"lot : sûres {len(rapport.surs)}, à confirmer {len(rapport.a_confirmer)}, déjà à jour {len(rapport.deja_a_jour)}, "
        f"inconnues {len(rapport.inconnus)}, frais {len(rapport.frais)}, anomalies lecture {len(rapport.anomalies_lecture)}, "
        f"anomalies facture {len(rapport.anomalies_facture)}")
    for f, c in rapport.surs:
        log(f"  SÛRE {f.fichier} facture {f.numero_facture} cde {c.ligne_facture.numero_commande} "
            f"réf {c.ligne_facture.reference_fournisseur} qté {c.ligne_facture.quantite_facturee:g} -> L{c.ligne_suivi.ligne_excel}")
    for f, c in rapport.a_confirmer:
        log(f"  À CONFIRMER {f.fichier} {f.numero_facture} réf {c.ligne_facture.reference_fournisseur} : "
            f"{c.cause.value if c.cause else '?'} — {'; '.join(c.raisons)[:150]}")
    for f, r in rapport.anomalies_facture:
        log(f"  ANOMALIE {f.fichier} : {r[:150]}")

    assert not est_verrouille(vivant), "STOP : verrou Excel (avant écriture)"
    resume = appliquer_et_archiver_factures(PROJET, PROJET / DOSSIER_A_TRAITER_FACTURES, rapport, rapport.surs)
    log(f"écriture : {resume['lignes_ecrites']} ligne(s) rapprochée(s), {resume['pieces_ecrites']} pièce(s) écrite(s), "
        f"{len(resume['pieces_ignorees'])} ignorée(s), sauvegarde {resume['sauvegarde']}")
    log(f"archivées {[f for f, _ in resume['factures_archivees']]} ; à vérifier {[f for f, _, _ in resume['factures_a_verifier']]} ; "
        f"sans parser {[f for f, _ in resume['factures_sans_parser']]} ; archivage échoué {resume['archivage_echoue']}")
    log(f"résorption : {resume['resorption']} ; rapport : {resume['chemin_rapport']}")

    nouvelles = lire_pieces(vivant)[len(pieces_avant):]
    log(f"Pièces après le lot : +{len(nouvelles)} ; types {dict(Counter(p['Type'] for p in nouvelles))}")
    log(f"structure : { {k: v for k, v in verifier_structure(vivant, attendre_nouvelles_colonnes=True).items() if k != 'feuilles_openpyxl'} }")
    DOSSIER_VERIF.mkdir(parents=True, exist_ok=True)
    copie = DOSSIER_VERIF / "Suivi apres lot - verif Excel.xlsx"
    shutil.copy2(vivant, copie)
    r = verifier_dans_excel(copie, recalculer=True, cellules=[("Dashboard", a) for a in KPI_CELLULES])
    print(resume_verification(r), flush=True)
    assert r["ouvert"] and not r["reparation"] and not r["erreur"]
    copie.unlink(missing_ok=True)
    log("Lot de factures passé et vérifié.")


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else ""
    if mode == "vivant":
        installer(trouver_fichier_suivi_vivant(PROJET), est_vivant=True)
    elif mode == "copie" and len(sys.argv) > 2:
        installer(sys.argv[2], est_vivant=False)
    elif mode == "lot":
        lot_factures()
    else:
        print(__doc__)
        sys.exit(2)
