"""
Génération du Panier de commande (étape 4, après décision de l'acheteur).

Le Comparatif.xlsx (étape 3) est le document de décision : sa colonne
"Fournisseur retenu" est pré-remplie au mieux-disant mais reste un texte
libre, modifiable à la main dans Excel. Une fois la décision prise (ou
validée telle quelle), ce module relit CE texte — jamais les formules
Excel des colonnes voisines (Réf./N° devis/Disponibilité), qui ne se sont
peut-être jamais recalculées si le fichier n'a pas été rouvert dans Excel —
et reconstruit, en repartant du besoin et des devis, un bloc de lignes par
fournisseur retenu, prêt à coller dans la feuille "Commandes" du Suivi
commandes.

Les colonnes du Panier CALQUENT celles de cette feuille "Commandes",
lues directement dans un export du Suivi déposé à la racine du projet
(fichier "*Suivi commandes*.xlsx") : même ordre, mêmes en-têtes. Seules
les colonnes que Consultation AI peut remplir avec certitude (Référence,
Désignation, Qté commandée, Tarif convenu, Devis associé, Chantier,
Fournisseur, Type de commande) sont renseignées ; le reste (mode de
livraison, statut, dates, facturation...) appartient au suivi de la
commande APRÈS passation et reste vide, à compléter par l'acheteur comme
aujourd'hui.

Si aucun export du Suivi n'est trouvé, un jeu de colonnes minimal de repli
est utilisé (les mêmes 8 colonnes connues, dans l'ordre) — le Panier reste
utilisable, juste pas calé sur l'ordre exact du Suivi.
"""

import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from moteur.besoin import lire_fichier_besoin, nom_chantier
from moteur.lecture_pdf import analyser_devis
from moteur.comparateur import comparer
from moteur.base import BaseArticles
from moteur.referentiel import Referentiel
from moteur.excel import _entete, _largeurs, GRIS
from moteur.journal import capturer_journal


# Fournisseur interne (moteur/detecteur.py, moteur/parsers.py) -> nom EXACT
# de la liste "Fournisseurs" du Suivi commandes (feuille "Listes
# Paramètres"). Sans entrée ici, le nom interne est recopié tel quel, et
# signalé à vérifier : jamais de nom inventé.
MAPPING_FOURNISSEURS = {
    "RAVATE": "RAVATE ELEC",
    "RAVATE PRO": "RAVATE PRO",
    "IMPORELEC": "Imporelec",
    "IMPORTER": "Importer",
    "STAND 64": "STAND 64",
    "CLAREO": "CLAREO",
    "COREDIME": "COREDIME",
    "COMINTER MAYOTTE": "COMINTER Mayotte",
    "COMINTER": "COMINTER",
    "ELECTRIC PLUS": "GMR",   # GMR = marque publique du canal Electric Plus
    "GMR": "GMR",
    "109 DISTRIBUTION": "109 Distribution",
    "EDOI": "EDOI",
    "COMPTOIR DU CABLING": "Comptoir du Cabling",
    "TELENCO": "Telenco",
    "SAGEES": "Sagees",
    "DEM": "DEM",
    "YESSS": "YESSS",
    "YESSS MAYOTTE": "YESSS MAYOTTE",
    "PROTECTHOMS": "Protecthoms",
    # Vérifié dans la liste "Fournisseurs" du vrai Suivi commandes (feuille
    # "Listes Paramètres") : "Artdeco.re" y figure telle quelle (à côté
    # d'une entrée distincte "LEDS'Run" non retenue ici, le domaine
    # artdeco.re étant l'ancre de détection de ce fournisseur — voir
    # moteur/detecteur.py). À confirmer auprès de l'acheteuse si un Panier
    # réel révèle que c'est en fait "LEDS'Run" qui est utilisé.
    "ART DECO": "Artdeco.re",
}

TYPE_COMMANDE_DEFAUT = "Chantier"

# En-têtes -> champ connu, une fois normalisés (majuscules, sans accents).
_CHAMPS_CONNUS = {
    "REFERENCE": "reference",
    "DESIGNATION": "designation",
    "QTE COMMANDEE": "qte",
    "TARIF CONVENU": "tarif",
    "DEVIS ASSOCIE": "devis",
    "CHANTIER": "chantier",
    "FOURNISSEUR": "fournisseur",
    "TYPE DE COMMANDE": "type_commande",
}

# Repli si aucun export du Suivi n'est trouvé à la racine du projet.
_ENTETES_PAR_DEFAUT = [
    "Référence", "Désignation", "Qté commandée", "Tarif convenu",
    "Devis associé", "Chantier", "Fournisseur", "Type de commande",
]


def _sans_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )


def _normaliser(s) -> str:
    return _sans_accents(str(s or "")).strip().upper()


def trouver_fichier_suivi(dossier_projet):
    """Le dernier export du Suivi commandes déposé à la racine du projet
    (ex. "1.3.0.1. Suivi commandes - 2026.xlsx"), ou None si absent."""

    dossier_projet = Path(dossier_projet)
    candidats = [
        f for f in dossier_projet.glob("*Suivi commandes*.xlsx")
        if not f.name.startswith("~$")
    ]
    if not candidats:
        return None
    return max(candidats, key=lambda f: f.stat().st_mtime)


def _entetes_commandes(fichier_suivi):
    """Ordre EXACT des colonnes de la feuille "Commandes" du Suivi, pour
    que le Panier s'y colle sans décalage."""

    wb = load_workbook(fichier_suivi, read_only=True, data_only=True)
    try:
        ws = wb["Commandes"] if "Commandes" in wb.sheetnames else wb.worksheets[0]
        premiere = next(ws.iter_rows(min_row=1, max_row=1))
        return [str(c.value).strip() if c.value is not None else "" for c in premiere]
    finally:
        wb.close()


def _chantiers_suivi(fichier_suivi):
    """Chantiers connus du Suivi (feuille "Listes Paramètres", colonne
    "Chantier"), pour retrouver le code exact ("137 ZAC Doujani")."""

    wb = load_workbook(fichier_suivi, read_only=True, data_only=True)
    try:
        if "Listes Paramètres" not in wb.sheetnames:
            return []
        ws = wb["Listes Paramètres"]
        return [
            str(row[1]).strip()
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True)
            if row[1]
        ]
    finally:
        wb.close()


def _resoudre_chantier(nom, chantiers_connus):
    """Code exact du chantier ("137 ZAC Doujani") à partir du nom déduit du
    fichier besoin ("Doujani"). Ambigu ou introuvable -> nom tel quel,
    laissé à vérifier (jamais de code deviné)."""

    if not nom or not chantiers_connus:
        return nom, False

    cible = _normaliser(nom)
    correspondances = [c for c in chantiers_connus if cible in _normaliser(c)]

    if len(correspondances) == 1:
        return correspondances[0], True

    return nom, False


def nom_fournisseur_canonique(nom):
    """Nom exact attendu par le Suivi commandes, ou (nom interne, False) si
    absent du mapping connu — à vérifier avant collage."""

    canonique = MAPPING_FOURNISSEURS.get(nom)
    return (canonique, True) if canonique else (nom, False)


def _lire_decisions(fichier_comparatif):
    """Relit la colonne "Fournisseur retenu" (texte libre) du Comparatif,
    dans l'ordre des lignes. Ignore la ligne de totaux et les lignes vides."""

    wb = load_workbook(fichier_comparatif, data_only=True)
    try:
        ws = wb["Comparatif"] if "Comparatif" in wb.sheetnames else wb.worksheets[0]
        entetes = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_retenu = entetes.index("Fournisseur retenu") if "Fournisseur retenu" in entetes else None

        decisions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            if row[0] == "TOTAL (si tout pris chez ce fournisseur)":
                continue
            retenu = row[col_retenu] if col_retenu is not None and col_retenu < len(row) else None
            decisions.append(str(retenu).strip() if retenu else "")

        return decisions
    finally:
        wb.close()


def _preparer_comparaison(dossier_projet, fichier_besoin, dossier_devis):
    """Recalcule le rapprochement besoin/devis à l'identique de l'étape 3
    (moteur/pipeline.py generer_comparatif), pour retrouver les offres
    RÉELLES (référence, n° devis, disponibilité, prix) de chaque ligne."""

    dossier_projet = Path(dossier_projet)

    base = BaseArticles(dossier_projet / "base")
    base.importer_bdd(dossier_projet / "base" / "BDD_articles.csv")
    base.importer_familles(dossier_projet / "base" / "familles.csv")

    dossier_referentiel = dossier_projet / "referentiel"
    referentiel = Referentiel(dossier_projet / "moteur")
    referentiel.importer_bdd(dossier_projet / "base" / "BDD_articles.csv")
    referentiel.importer_composes(dossier_referentiel / "composes.csv")
    referentiel.appliquer_confirmations(dossier_referentiel / "A_confirmer.xlsx")

    fichier_besoin = Path(fichier_besoin)
    chantier = nom_chantier(fichier_besoin)
    besoin = lire_fichier_besoin(fichier_besoin)

    articles = analyser_devis(dossier_devis)
    if articles:
        base.enrichir_depuis_devis(articles)

    resultat = comparer(articles, besoin, base, referentiel)

    referentiel.fermer()
    base.fermer()

    return resultat, chantier


def generer_panier(
    dossier_projet, fichier_besoin, dossier_devis, fichier_comparatif,
    dossier_resultats=None,
):
    """
    Construit resultats/Panier_<chantier>_<date>.xlsx à partir des
    décisions de l'acheteur dans fichier_comparatif (Comparatif.xlsx généré
    puis, le cas échéant, corrigé et enregistré depuis Excel).

    fichier_besoin et dossier_devis doivent être les MÊMES qu'à la
    génération de ce comparatif : ils servent à retrouver, pour chaque
    ligne et chaque fournisseur retenu, l'offre réelle (référence, n° de
    devis, prix). dossier_resultats est le dossier où écrire le Panier ;
    par défaut dossier_projet/"resultats" (pour une consultation, passer
    son sous-dossier resultats/, voir moteur/consultation.py). Retourne le
    chemin du Panier généré, ou None si aucune ligne n'a pu être commandée.
    """

    dossier_projet = Path(dossier_projet)
    dossier_resultats = Path(dossier_resultats) if dossier_resultats else dossier_projet / "resultats"

    with capturer_journal(dossier_resultats, prefixe="panier"):

        resultat, chantier = _preparer_comparaison(dossier_projet, fichier_besoin, dossier_devis)
        decisions = _lire_decisions(fichier_comparatif)

        if len(decisions) != len(resultat["lignes"]):
            print(
                f"/!\\ Le Comparatif a {len(decisions)} ligne(s), le besoin "
                f"recalculé en a {len(resultat['lignes'])} : des lignes ont "
                "peut-être été ajoutées/supprimées à la main dans le Comparatif. "
                "Rapprochement par position, au mieux."
            )

        fichier_suivi = trouver_fichier_suivi(dossier_projet)
        if fichier_suivi:
            entetes = _entetes_commandes(fichier_suivi)
            chantiers_connus = _chantiers_suivi(fichier_suivi)
            print(f"Colonnes calées sur : {fichier_suivi.name}")
        else:
            entetes = list(_ENTETES_PAR_DEFAUT)
            chantiers_connus = []
            print(
                "/!\\ Aucun export du Suivi commandes trouvé à la racine du "
                "projet (fichier attendu : *Suivi commandes*.xlsx). "
                "Colonnes de repli utilisées."
            )

        chantier_resolu, chantier_ok = _resoudre_chantier(chantier, chantiers_connus)
        if chantiers_connus and not chantier_ok and chantier:
            print(
                f"/!\\ Chantier « {chantier} » non retrouvé sans ambiguïté dans "
                "le Suivi : nom recopié tel quel, à corriger avant collage."
            )

        lignes_panier = []
        non_commandees = []
        fournisseurs_a_verifier = set()

        for i, ligne in enumerate(resultat["lignes"]):

            demande = ligne.get("demande") or ligne["reference"]
            fournisseur_retenu = decisions[i] if i < len(decisions) else ""

            if not ligne["offres"]:
                non_commandees.append({
                    "demande": demande, "reference": ligne["reference"],
                    "qte": ligne["qte_besoin"], "motif": "Aucune offre reçue dans les devis",
                })
                continue

            if not fournisseur_retenu:
                non_commandees.append({
                    "demande": demande, "reference": ligne["reference"],
                    "qte": ligne["qte_besoin"],
                    "motif": "Aucun fournisseur retenu dans le Comparatif",
                })
                continue

            offre = ligne["offres"].get(fournisseur_retenu)
            if offre is None:
                cible = _normaliser(fournisseur_retenu)
                offre = next(
                    (o for f, o in ligne["offres"].items() if _normaliser(f) == cible),
                    None,
                )
            if offre is None:
                non_commandees.append({
                    "demande": demande, "reference": ligne["reference"],
                    "qte": ligne["qte_besoin"],
                    "motif": (
                        f"Fournisseur retenu « {fournisseur_retenu} » sans offre "
                        "correspondante sur cette ligne (faute de frappe ?)"
                    ),
                })
                continue

            nom_canon, connu = nom_fournisseur_canonique(fournisseur_retenu)
            if not connu:
                fournisseurs_a_verifier.add(fournisseur_retenu)

            lignes_panier.append({
                "reference": offre["reference"],
                "designation": ligne["designation"] or demande,
                "qte": ligne["qte_besoin"],
                "tarif": offre["prix_norm"],
                "devis": offre["devis"],
                "chantier": chantier_resolu,
                "fournisseur": nom_canon,
                "type_commande": TYPE_COMMANDE_DEFAUT,
            })

        if fournisseurs_a_verifier:
            print(
                "/!\\ Fournisseur(s) absent(s) de la liste du Suivi commandes, "
                "recopié(s) tel quel (à vérifier avant collage) : "
                + ", ".join(sorted(fournisseurs_a_verifier))
            )

        if not lignes_panier:
            print("\nAucune ligne à commander (pas de décision prise, ou aucune offre).")
            return None

        lignes_panier.sort(key=lambda l: (l["fournisseur"], str(l["reference"])))

        fichier = _ecrire_panier(
            dossier_projet, entetes, lignes_panier, non_commandees, chantier,
            dossier_resultats=dossier_resultats,
        )

        print(f"\nPanier créé : {fichier}")
        print(f"  {len(lignes_panier)} ligne(s) à commander, {len(non_commandees)} en attente.")

        return fichier


def _ecrire_panier(
    dossier_projet, entetes, lignes_panier, non_commandees, chantier,
    dossier_resultats=None,
):

    wb = Workbook()

    ws = wb.active
    ws.title = "Panier"
    ws.append(entetes)
    _entete(ws)

    index_champ = {}
    for i, e in enumerate(entetes):
        champ = _CHAMPS_CONNUS.get(_normaliser(e))
        if champ:
            index_champ[champ] = i

    col_qte = index_champ.get("qte")
    col_tarif = index_champ.get("tarif")

    fournisseur_actuel = None
    bande = False

    for ligne in lignes_panier:

        valeurs = [""] * len(entetes)
        for champ, col in index_champ.items():
            valeurs[col] = ligne.get(champ, "")

        ws.append(valeurs)
        num = ws.max_row

        if ligne["fournisseur"] != fournisseur_actuel:
            fournisseur_actuel = ligne["fournisseur"]
            bande = not bande

        if bande:
            for cel in ws[num]:
                cel.fill = GRIS

        if col_qte is not None:
            ws.cell(row=num, column=col_qte + 1).number_format = "#,##0.###"
        if col_tarif is not None:
            ws.cell(row=num, column=col_tarif + 1).number_format = "#,##0.0000"

    ws.freeze_panes = "A2"
    _largeurs(ws)

    if non_commandees:
        ws2 = wb.create_sheet("Non commandées")
        ws2.append(["Demande d'origine", "Référence", "Qté", "Motif"])
        _entete(ws2)
        for n in non_commandees:
            ws2.append([n["demande"], n["reference"], n["qte"], n["motif"]])
        ws2.freeze_panes = "A2"
        _largeurs(ws2)

    dossier = Path(dossier_resultats) if dossier_resultats else Path(dossier_projet) / "resultats"
    dossier.mkdir(parents=True, exist_ok=True)

    nom_base = f"Panier {chantier}" if chantier else "Panier"
    horodatage = datetime.now().strftime("%Y%m%d")
    fichier = dossier / f"{nom_base}_{horodatage}.xlsx"

    try:
        wb.save(fichier)
    except PermissionError:
        fichier = dossier / f"{nom_base}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        print(f"{nom_base}_{horodatage}.xlsx est ouvert dans Excel.")
        wb.save(fichier)

    return fichier
