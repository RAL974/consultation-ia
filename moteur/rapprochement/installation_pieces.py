"""
Orchestration de P1 (voir CLAUDE.md, « Feuille Pièces — modèle, socles,
migration ») — les mêmes étapes, dans le même ordre, d'abord sur une COPIE
du Suivi commandes, puis sur le vivant (étape 7) :

  0d  reparer_formules_ligne()        ligne 2 de Commandes, colonnes listées
  1   installer_feuille_pieces()      feuille Pièces + tableau Pieces
  3   ajouter_colonnes_commandes()    4 colonnes calculées en fin de Commandes
  4   migrer_factures_vers_pieces()   1 ligne Pièces par ligne facture existante
  5   basculer_colonnes_facture()     les 5 colonnes facture deviennent formules
      (uniquement si l'étape 4 est au centime)

Chaque étape est suivie d'une vérification faite par l'outil
(verifier_structure : zip relisible, XML bien formé, feuilles/tableaux
attendus) ; la vérification par Excel lui-même (ouverture sans réparation,
recalcul) est dans moteur.rapprochement.verification_excel.
"""

import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from moteur.rapprochement.ecriture import lire_entetes, reparer_formules_ligne
from moteur.rapprochement.migration_pieces import controler_sommes, migrer_factures_vers_pieces
from moteur.rapprochement.pieces import (
    COLONNES_COMMANDES_NOUVELLES,
    COLONNES_FACTURE_CALCULEES,
    COLONNES_PIECES,
    FEUILLE_COMMANDES,
    FEUILLE_PIECES,
    TABLE_PIECES,
    ajouter_colonnes_commandes,
    basculer_colonnes_facture,
    installer_feuille_pieces,
)

# Étape 0d : cellules de la ligne 2 de Commandes dont la formule est
# décalée (voir CLAUDE.md, §0.3 du dossier P1) — liste EXPLICITE. Le plan
# cite les colonnes 25 à 30 (Y..AD) ; l'inspection du vivant (2026-09-04)
# montre le même décalage sur 22 à 24 (V..X : Reliquat, RAL, Soldé) —
# réparées aussi, sur décision [HUMAIN-P1-a] (réponse par défaut : oui).
COLONNES_LIGNE2_PLAN = tuple(range(25, 31))
COLONNES_LIGNE2_CONSTATEES = tuple(range(22, 31))


def verifier_structure(fichier, attendre_pieces=True, attendre_nouvelles_colonnes=False) -> dict:
    """Vérification outil, sans Excel : zip relisible (zipfile + openpyxl
    lecture seule), TOUTES les parties XML bien formées, feuilles et
    tableaux attendus, calcChain absent quand des formules ont été
    ajoutées. Lève AssertionError avec le détail à la 1re anomalie."""

    fichier = Path(fichier)
    resultat = {}
    with zipfile.ZipFile(fichier) as z:
        assert z.testzip() is None, "zip corrompu"
        noms = z.namelist()
        for nom in noms:
            if nom.endswith(".xml") or nom.endswith(".rels"):
                ET.fromstring(z.read(nom))  # XML bien formé
        resultat["parties"] = len(noms)
        resultat["tableaux"] = len([n for n in noms if re.match(r"xl/tables/table\d+\.xml$", n)])
        resultat["calc_chain"] = "xl/calcChain.xml" in noms
        wbx = z.read("xl/workbook.xml").decode("utf-8")
        resultat["feuilles"] = re.findall(r'<sheet\b[^>]*\bname="([^"]*)"', wbx)
        resultat["full_calc_on_load"] = 'fullCalcOnLoad="1"' in wbx
        if attendre_pieces:
            chemin_table = next(
                (n for n in noms if re.match(r"xl/tables/table\d+\.xml$", n)
                 and f'name="{TABLE_PIECES}"' in z.read(n).decode("utf-8")), None,
            )
            assert chemin_table, f"tableau {TABLE_PIECES} absent"
            xml_table = z.read(chemin_table).decode("utf-8")
            resultat["ref_pieces"] = re.search(r'<table\b[^>]*\sref="([^"]*)"', xml_table).group(1)
            noms_colonnes = re.findall(r'<tableColumn\b[^>]*\bname="([^"]*)"', xml_table)
            assert [n.replace("&amp;", "&") for n in noms_colonnes] == list(COLONNES_PIECES), "colonnes Pieces inattendues"

    wb = load_workbook(fichier, read_only=True)
    try:
        resultat["feuilles_openpyxl"] = list(wb.sheetnames)
    finally:
        wb.close()
    assert resultat["feuilles"] == resultat["feuilles_openpyxl"]
    if attendre_pieces:
        assert resultat["feuilles"][1] == FEUILLE_PIECES, "Pièces doit être juste après Commandes"
    entetes = lire_entetes(fichier, FEUILLE_COMMANDES)
    resultat["nb_colonnes_commandes"] = len(entetes)
    if attendre_nouvelles_colonnes:
        assert set(COLONNES_COMMANDES_NOUVELLES) <= set(entetes), "colonnes calculées absentes"
        assert not resultat["calc_chain"], "calcChain devrait avoir été supprimé"
        assert resultat["full_calc_on_load"], "fullCalcOnLoad attendu"
    return resultat


def etape_0d(fichier, dossier_backups, colonnes=COLONNES_LIGNE2_CONSTATEES) -> dict:
    return reparer_formules_ligne(fichier, dossier_backups, colonnes=colonnes)


def etape_1(fichier, dossier_backups) -> Path:
    return installer_feuille_pieces(fichier, dossier_backups)


def etape_3(fichier, dossier_backups) -> Path:
    return ajouter_colonnes_commandes(fichier, dossier_backups)


def etape_4(fichier, dossier_projet, dossier_backups, limite=None, journal=None) -> dict:
    return migrer_factures_vers_pieces(fichier, dossier_projet, dossier_backups, limite=limite, journal=journal)


def etape_5(fichier, dossier_backups) -> Path:
    controle = controler_sommes(fichier)
    if not controle["au_centime"]:
        raise ValueError(f"Étape 5 refusée : le contrôle au centime échoue — {controle['detail']}")
    return basculer_colonnes_facture(fichier, dossier_backups)


def executer(fichier, dossier_projet, dossier_backups, etapes=("0d", "1", "3", "4", "5"), journal=print,
             colonnes_ligne2=COLONNES_LIGNE2_CONSTATEES, limite_migration=None) -> dict:
    """Enchaîne les étapes demandées, chacune suivie de verifier_structure().
    S'arrête à la 1re erreur (exception propagée, rien de plus n'est
    écrit). Retourne {étape: résultat}."""

    fichier = Path(fichier)
    resultats = {}

    def _log(msg):
        journal(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

    for etape in etapes:
        _log(f"Étape {etape} — début ({fichier.name})")
        if etape == "0d":
            resultats[etape] = etape_0d(fichier, dossier_backups, colonnes_ligne2)
            _log(f"  formules réécrites en ligne 2 : {sorted(resultats[etape])}")
            resultats[etape + "_verif"] = verifier_structure(fichier, attendre_pieces=False)
        elif etape == "1":
            resultats[etape] = etape_1(fichier, dossier_backups)
            resultats[etape + "_verif"] = verifier_structure(fichier)
            _log(f"  tableaux : {resultats[etape + '_verif']['tableaux']}, feuilles : {resultats[etape + '_verif']['feuilles'][:3]}…")
        elif etape == "3":
            resultats[etape] = etape_3(fichier, dossier_backups)
            resultats[etape + "_verif"] = verifier_structure(fichier, attendre_nouvelles_colonnes=True)
            _log(f"  colonnes Commandes : {resultats[etape + '_verif']['nb_colonnes_commandes']}, calcChain : {resultats[etape + '_verif']['calc_chain']}")
        elif etape == "4":
            resultats[etape] = etape_4(fichier, dossier_projet, dossier_backups, limite_migration, journal=_log)
            resultats[etape + "_verif"] = verifier_structure(fichier)
            r = resultats[etape]
            _log(f"  {r['pieces_construites']} pièces, au centime : {r['au_centime']}, rapport : {r['chemin_rapport']}")
            _log(f"  ref Pieces : {resultats[etape + '_verif']['ref_pieces']}")
        elif etape == "5":
            resultats[etape] = etape_5(fichier, dossier_backups)
            resultats[etape + "_verif"] = verifier_structure(fichier, attendre_nouvelles_colonnes=True)
            entetes = lire_entetes(fichier, FEUILLE_COMMANDES)
            assert set(COLONNES_FACTURE_CALCULEES) <= set(entetes)
            _log("  5 colonnes facture basculées en formules")
        else:
            raise ValueError(f"Étape inconnue : {etape}")
        _log(f"Étape {etape} — OK")
    return resultats
