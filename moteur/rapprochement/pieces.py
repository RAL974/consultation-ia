"""
Feuille « Pièces » du Suivi commandes — modèle, formules et lecture
(voir CLAUDE.md, « Feuille Pièces — modèle, socles, migration »).

Une ligne de Pièces = une ligne de DOCUMENT (BL, Retour, Facture, Avoir,
Frais, Demande d'avoir), écrite par l'outil (moteur.rapprochement.
ecriture.ajouter_lignes_tableau), jamais saisie à la main — seule la
colonne « Commentaire » est humaine. Les 5 anciennes colonnes de saisie
facture de Commandes (« N° facture », « Date facture », « Qté facturée »,
« PU facturé », « Montant facturé HT ») sont devenues des colonnes
CALCULÉES à partir de Pièces (FORMULES_COMMANDES_BASCULE), complétées par
4 nouvelles colonnes calculées (FORMULES_COMMANDES_NOUVELLES). « Statut
commande » n'est jamais modifié.

Idempotence : chaque ligne porte un « ID pièce » unique
(construire_id_piece) — une ligne dont l'ID existe déjà dans la feuille
n'est jamais réécrite (voir ecriture.ajouter_lignes_tableau, colonne_id).

Ce module ne contient AUCUNE règle de parsing : il ne lit que le classeur
et les objets déjà extraits par les parsers/pipelines.
"""

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from moteur.panier import MAPPING_FOURNISSEURS
from moteur.rapprochement.ecriture import (
    FEUILLE_COMMANDES,
    Formule,
    ajouter_colonnes_calculees,
    ajouter_feuille_tableau,
    ajouter_lignes_tableau,
    basculer_colonnes_en_formules,
)

FEUILLE_PIECES = "Pièces"
TABLE_PIECES = "Pieces"

# Les 26 colonnes, dans l'ordre exact du tableau (voir CLAUDE.md).
COLONNES_PIECES = (
    "ID pièce", "Type", "Fournisseur", "N° pièce", "Date pièce", "N° de commande",
    "Chantier", "Sous-Chantier", "Référence Suivi", "Référence fournisseur", "Désignation",
    "Qté", "PU HT", "Montant HT", "N° BL lié", "N° facture liée",
    "Prix de référence", "Source prix", "Écart PU", "Écart ligne €", "Contrôle prix",
    "Mode de rapprochement", "Demande d'avoir", "Fichier", "Date d'écriture", "Commentaire",
)

TYPE_BL = "BL"
TYPE_RETOUR = "Retour"
TYPE_FACTURE = "Facture"
TYPE_AVOIR = "Avoir"
TYPE_FRAIS = "Frais"
TYPE_DEMANDE_AVOIR = "Demande d'avoir"
TYPES_PIECE = (TYPE_BL, TYPE_RETOUR, TYPE_FACTURE, TYPE_AVOIR, TYPE_FRAIS, TYPE_DEMANDE_AVOIR)

MODE_AUTO = "Auto"
MODE_CONFIRME = "Confirmé"
MODE_EQUIVALENCE = "Équivalence"
MODE_MIGRE = "Migré"
MODE_MIGRE_SANS_PDF = "Migré sans PDF"

COMMENTAIRE_MONTANT_RECALCULE = "montant recalculé"


class FeuillePiecesAbsente(ValueError):
    """Le classeur visé n'a pas (encore) de feuille Pièces — aucune ligne de
    document ne peut y être écrite (jamais d'écriture perdue en silence :
    installer_feuille_pieces() d'abord, voir CLAUDE.md, P1 étape 1)."""

# Styles logiques (voir ecriture._STYLES_LOGIQUES) des colonnes typées.
STYLES_PIECES = {
    "Date pièce": "date",
    "PU HT": "monnaie4",
    "Montant HT": "monnaie",
    "Prix de référence": "monnaie4",
    "Écart PU": "monnaie4",
    "Écart ligne €": "monnaie",
    "Date d'écriture": "datetime",
}

LARGEURS_PIECES = {
    "ID pièce": 12, "Type": 9, "Fournisseur": 16, "N° pièce": 12, "Date pièce": 11,
    "N° de commande": 13, "Chantier": 22, "Sous-Chantier": 16, "Référence Suivi": 16,
    "Référence fournisseur": 16, "Désignation": 40, "Qté": 8, "PU HT": 12, "Montant HT": 13,
    "N° BL lié": 12, "N° facture liée": 13, "Prix de référence": 12, "Source prix": 10,
    "Écart PU": 10, "Écart ligne €": 11, "Contrôle prix": 16, "Mode de rapprochement": 16,
    "Demande d'avoir": 13, "Fichier": 45, "Date d'écriture": 16, "Commentaire": 40,
}

# Les 5 colonnes facture de Commandes, désormais CALCULÉES (elles restent
# lues telles quelles par matching_facture/fnp — valeurs en cache du
# dernier recalcul Excel ; la source de vérité est la feuille Pièces).
COLONNES_FACTURE_CALCULEES = (
    "N° facture", "Date facture", "Qté facturée", "PU facturé", "Montant facturé HT",
)

COLONNES_COMMANDES_NOUVELLES = (
    "Reste à facturer HT", "Écart facture €", "Qté retournée", "Statut facture",
)

# Formules ANGLAISES telles que stockées dans le XML (préfixes _xlfn/_xlws
# compris — MAXIFS, TEXTJOIN, UNIQUE, FILTER sont des fonctions
# post-2007 ; SUMIFS/IFERROR n'en ont pas besoin), références structurées
# sous leur forme longue. Même chaîne sur toutes les lignes.
_CRITERES = (
    "Pieces[N° de commande],Commandes[[#This Row],[N° de commande]],"
    "Pieces[Référence Suivi],Commandes[[#This Row],[Référence]]"
)
_MAXIFS_DATE = f'_xlfn.MAXIFS(Pieces[Date pièce],{_CRITERES},Pieces[Type],"Facture")'

FORMULES_COMMANDES_BASCULE = {
    "Qté facturée": Formule(
        f'SUMIFS(Pieces[Qté],{_CRITERES},Pieces[Type],"Facture")'
        f'+SUMIFS(Pieces[Qté],{_CRITERES},Pieces[Type],"Avoir")'
    ),
    "Montant facturé HT": Formule(
        f'SUMIFS(Pieces[Montant HT],{_CRITERES},Pieces[Type],"Facture")'
        f'+SUMIFS(Pieces[Montant HT],{_CRITERES},Pieces[Type],"Avoir")'
    ),
    "N° facture": Formule(
        '_xlfn.TEXTJOIN("; ",TRUE,_xlfn.UNIQUE(_xlfn._xlws.FILTER(Pieces[N° pièce],'
        "(Pieces[N° de commande]=Commandes[[#This Row],[N° de commande]])"
        "*(Pieces[Référence Suivi]=Commandes[[#This Row],[Référence]])"
        '*(Pieces[Type]="Facture"),"")))',
        array=True,
    ),
    # MAXIFS vaut 0 sans facture -> afficherait 00/01/1900 : vide à la place.
    "Date facture": Formule(f'IF({_MAXIFS_DATE}=0,"",{_MAXIFS_DATE})'),
    # Moyen pondéré — les PU unitaires restent dans Pièces.
    "PU facturé": Formule(
        "IFERROR(Commandes[[#This Row],[Montant facturé HT]]/Commandes[[#This Row],[Qté facturée]],\"\")"
    ),
}

FORMULES_COMMANDES_NOUVELLES = {
    "Reste à facturer HT": Formule(
        "IFERROR(N(Commandes[[#This Row],[Facturé BL]])-N(Commandes[[#This Row],[Montant facturé HT]]),\"\")"
    ),
    "Écart facture €": Formule(f'SUMIFS(Pieces[Écart ligne €],{_CRITERES},Pieces[Type],"Facture")'),
    "Qté retournée": Formule(f'-SUMIFS(Pieces[Qté],{_CRITERES},Pieces[Type],"Retour")'),
    # P1 : vide / 🔵 En attente facture / 🟠 Partiellement facturée /
    # ✅ Facturée / ⛔ Sur-facturée (Qté facturée > Qté livrée = garde-fou
    # double facturation). Libellés affinés en P2.
    "Statut facture": Formule(
        'IF(N(Commandes[[#This Row],[Qté facturée]])=0,'
        'IF(N(Commandes[[#This Row],[Qté livrée]])>0,"🔵 En attente facture",""),'
        'IF(N(Commandes[[#This Row],[Qté facturée]])>N(Commandes[[#This Row],[Qté livrée]]),"⛔ Sur-facturée",'
        'IF(N(Commandes[[#This Row],[Qté facturée]])<N(Commandes[[#This Row],[Qté livrée]]),'
        '"🟠 Partiellement facturée","✅ Facturée")))'
    ),
}

STYLES_COMMANDES_NOUVELLES = {"Reste à facturer HT": "monnaie", "Écart facture €": "monnaie"}


# --- Aides de modèle ------------------------------------------------------


def nom_fournisseur_suivi(nom) -> str:
    """Nom Suivi (liste « Fournisseurs ») d'un nom de parser ou d'une
    saisie libre de la colonne Fournisseur de Commandes ('Coredime' ->
    'COREDIME', '109 DISTRIBUTION' -> '109 Distribution') — via
    moteur.panier.MAPPING_FOURNISSEURS ; inchangé si inconnu."""

    nom = str(nom or "").strip()
    return MAPPING_FOURNISSEURS.get(nom.upper(), nom)


def cle_reference(valeur) -> str:
    """Référence Suivi -> clé texte pour les index Python (jamais pour
    Excel, qui compare les valeurs telles quelles) : '5120' et 5120 et
    5120.0 donnent '5120'."""

    if valeur is None:
        return ""
    if isinstance(valeur, float) and valeur.is_integer():
        return str(int(valeur))
    return str(valeur).strip()


def cle_commande(valeur) -> str:
    return cle_reference(valeur)


def construire_id_piece(fournisseur, type_piece, numero_piece, numero_commande, reference_suivi,
                        numero_bl_lie="") -> str:
    """Clé unique d'une ligne Pièces :
    `<Fournisseur>|<Type>|<N° pièce>|<N° de commande>|<Référence Suivi>|<N° BL lié>`
    (Référence Suivi = référence fournisseur préfixée "F:" pour un Frais
    sans ligne Suivi, voir nouvelle_piece). Deux lignes réellement
    identiques sur ces 6 champs (même article deux fois sur le même BL)
    sont suffixées « #2 », « #3 »... par dedoublonner_ids()."""

    return "|".join(
        cle_reference(v) for v in (fournisseur, type_piece, numero_piece, numero_commande, reference_suivi, numero_bl_lie)
    )


def dedoublonner_ids(pieces) -> list:
    """Suffixe « #2 », « #3 »... les ID en double AU SEIN du lot `pieces`
    (dans l'ordre) — déterministe : rejouer le même lot redonne les mêmes
    ID, que l'idempotence de ecrire_pieces() (ID déjà présent dans la
    feuille) laisse alors de côté. Jamais de suffixe par rapport aux ID
    déjà écrits : un ID déjà présent EST la même ligne de document."""

    vus = set()
    for piece in pieces:
        base = piece["ID pièce"]
        candidat, k = base, 1
        while candidat in vus:
            k += 1
            candidat = f"{base}#{k}"
        piece["ID pièce"] = candidat
        vus.add(candidat)
    return pieces


def formule_fichier(chemin) -> Formule:
    """=HYPERLINK("<chemin>","<nom>") — une formule, jamais une relation
    OOXML (aucune partie de relations à maintenir, survit à un tri)."""

    chemin = Path(chemin)
    texte = str(chemin).replace('"', "'")
    nom = chemin.name.replace('"', "'")
    return Formule(f'HYPERLINK("{texte}","{nom}")', cache=nom)


def nouvelle_piece(type_piece, fournisseur, numero_piece, date_piece, numero_commande, chantier, sous_chantier,
                   reference_suivi, reference_fournisseur, designation, qte, pu_ht, montant_ht,
                   numero_bl_lie="", numero_facture_liee="", mode=MODE_AUTO, fichier=None,
                   commentaire=None, date_ecriture=None, id_piece=None) -> dict:
    """Une ligne Pièces complète ({colonne: valeur}, colonnes P2/P4
    laissées vides). Qté signée (Retour et Avoir négatives — à la charge
    de l'appelant). `fichier` : chemin -> formule HYPERLINK."""

    if type_piece not in TYPES_PIECE:
        raise ValueError(f"Type de pièce inconnu : {type_piece!r}")
    fournisseur = nom_fournisseur_suivi(fournisseur)
    cle_id = reference_suivi if reference_suivi not in (None, "") else f"F:{reference_fournisseur}"
    if id_piece is None:
        id_piece = construire_id_piece(fournisseur, type_piece, numero_piece, numero_commande, cle_id, numero_bl_lie)

    if isinstance(date_piece, datetime):
        date_piece = date_piece.date()

    piece = {nom: None for nom in COLONNES_PIECES}
    piece.update({
        "ID pièce": id_piece,
        "Type": type_piece,
        "Fournisseur": fournisseur,
        "N° pièce": cle_reference(numero_piece) or None,
        "Date pièce": date_piece,
        "N° de commande": cle_commande(numero_commande) or None,
        "Chantier": chantier if chantier not in ("", None) else None,
        "Sous-Chantier": sous_chantier if sous_chantier not in ("", None) else None,
        "Référence Suivi": reference_suivi if reference_suivi not in ("", None) else None,
        "Référence fournisseur": str(reference_fournisseur).strip() if reference_fournisseur not in ("", None) else None,
        "Désignation": str(designation).strip() if designation not in ("", None) else None,
        "Qté": qte,
        "PU HT": pu_ht,
        "Montant HT": montant_ht,
        "N° BL lié": cle_reference(numero_bl_lie) or None,
        "N° facture liée": cle_reference(numero_facture_liee) or None,
        "Mode de rapprochement": mode,
        "Fichier": formule_fichier(fichier) if fichier else None,
        "Date d'écriture": date_ecriture or datetime.now().replace(microsecond=0),
        "Commentaire": commentaire or None,
    })
    return piece


# --- Écriture -------------------------------------------------------------


def ecrire_pieces(fichier, pieces, dossier_backups) -> dict:
    """Écrit `pieces` (dicts de nouvelle_piece) en fin du tableau Pieces —
    idempotent par « ID pièce » (voir ajouter_lignes_tableau). Lève
    ValueError si la feuille Pièces n'existe pas encore (jamais d'écriture
    silencieusement perdue)."""

    if not feuille_pieces_presente(fichier):
        raise FeuillePiecesAbsente(
            f"« {Path(fichier).name} » n'a pas de feuille « {FEUILLE_PIECES} » — rien n'a été écrit "
            "(installer_feuille_pieces() d'abord)."
        )
    return ajouter_lignes_tableau(
        fichier, FEUILLE_PIECES, TABLE_PIECES, pieces, dossier_backups,
        colonne_id="ID pièce", styles_colonnes=STYLES_PIECES,
    )


def installer_feuille_pieces(fichier, dossier_backups) -> Path:
    """Étape 1 de P1 : la feuille Pièces + le tableau Pieces (26 colonnes),
    juste après Commandes."""

    return ajouter_feuille_tableau(
        fichier, FEUILLE_PIECES, TABLE_PIECES, COLONNES_PIECES, dossier_backups,
        feuille_modele=FEUILLE_COMMANDES, largeurs=LARGEURS_PIECES,
    )


def basculer_colonnes_facture(fichier, dossier_backups) -> Path:
    """Étape 5 de P1 (1/2) : les 5 colonnes facture de Commandes deviennent
    calculées depuis Pièces. À n'appeler qu'après la migration contrôlée
    au centime (moteur.rapprochement.migration_pieces)."""

    return basculer_colonnes_en_formules(
        fichier, FEUILLE_COMMANDES, FEUILLE_COMMANDES, FORMULES_COMMANDES_BASCULE, dossier_backups,
    )


def ajouter_colonnes_commandes(fichier, dossier_backups) -> Path:
    """Étape 5 de P1 (2/2) — aussi utilisée seule à l'étape 3 pour prouver
    le socle : les 4 nouvelles colonnes calculées en fin de Commandes."""

    return ajouter_colonnes_calculees(
        fichier, FEUILLE_COMMANDES, FEUILLE_COMMANDES, FORMULES_COMMANDES_NOUVELLES, dossier_backups,
        styles_colonnes=STYLES_COMMANDES_NOUVELLES,
    )


# --- Lecture --------------------------------------------------------------


def lire_pieces(fichier) -> list | None:
    """Toutes les lignes de la feuille Pièces ({colonne: valeur}, valeurs
    en cache — la colonne Fichier rend le libellé du lien) ; None si la
    feuille n'existe pas (classeur pas encore migré — les appelants se
    rabattent alors sur les 5 colonnes de Commandes)."""

    wb = load_workbook(fichier, read_only=True, data_only=True)
    try:
        if FEUILLE_PIECES not in wb.sheetnames:
            return None
        ws = wb[FEUILLE_PIECES]
        lignes = ws.iter_rows(values_only=True)
        entetes = [str(c).strip() if c is not None else "" for c in next(lignes, ())]
        resultat = []
        for i, row in enumerate(lignes, start=2):
            if not any(v is not None for v in row):
                continue
            piece = {nom: (row[j] if j < len(row) else None) for j, nom in enumerate(entetes) if nom}
            piece["_ligne_excel"] = i
            resultat.append(piece)
        return resultat
    finally:
        wb.close()


def feuille_pieces_presente(fichier) -> bool:
    """Lecture directe de xl/workbook.xml (pas de chargement openpyxl)."""
    import re
    import zipfile

    with zipfile.ZipFile(fichier) as z:
        noms = re.findall(r'<sheet\b[^>]*\bname="([^"]*)"', z.read("xl/workbook.xml").decode("utf-8"))
    return FEUILLE_PIECES in [n.replace("&amp;", "&") for n in noms]


class IndexPieces:
    """Index en mémoire des lignes Pièces pour les pipelines (idempotence
    par ID, cumul facturé par ligne de commande, doublons de facture).
    `pieces` = lire_pieces() ; None -> index vide, `disponible` False."""

    def __init__(self, pieces):
        self.disponible = pieces is not None
        self.pieces = list(pieces or [])
        self.ids = {str(p.get("ID pièce") or "") for p in self.pieces}
        self._par_ligne = {}
        self._par_facture = {}
        for p in self.pieces:
            cle = (cle_commande(p.get("N° de commande")), cle_reference(p.get("Référence Suivi")))
            self._par_ligne.setdefault(cle, []).append(p)
            if p.get("Type") == TYPE_FACTURE:
                cle_f = (nom_fournisseur_suivi(p.get("Fournisseur")).upper(), cle_reference(p.get("N° pièce")))
                self._par_facture.setdefault(cle_f, []).append(p)

    def pieces_ligne(self, numero_commande, reference_suivi, types=None) -> list:
        lignes = self._par_ligne.get((cle_commande(numero_commande), cle_reference(reference_suivi)), [])
        if types is None:
            return list(lignes)
        return [p for p in lignes if p.get("Type") in types]

    def qte_facturee(self, numero_commande, reference_suivi) -> float:
        """Σ Qté des Factures et Avoirs (signés) déjà écrits pour cette
        ligne de commande."""
        return round(sum(
            float(p.get("Qté") or 0)
            for p in self.pieces_ligne(numero_commande, reference_suivi, (TYPE_FACTURE, TYPE_AVOIR))
        ), 4)

    def montant_facture(self, numero_commande, reference_suivi) -> float:
        return round(sum(
            float(p.get("Montant HT") or 0)
            for p in self.pieces_ligne(numero_commande, reference_suivi, (TYPE_FACTURE, TYPE_AVOIR))
        ), 2)

    def numeros_factures(self, numero_commande, reference_suivi) -> list:
        vus = []
        for p in self.pieces_ligne(numero_commande, reference_suivi, (TYPE_FACTURE,)):
            n = cle_reference(p.get("N° pièce"))
            if n and n not in vus:
                vus.append(n)
        return vus

    def date_facture(self, numero_commande, reference_suivi):
        dates = [p.get("Date pièce") for p in self.pieces_ligne(numero_commande, reference_suivi, (TYPE_FACTURE,))]
        dates = [d for d in dates if d is not None]
        return max(dates) if dates else None

    def fichiers_facture(self, fournisseur, numero_facture) -> set:
        """Libellés « Fichier » des lignes Facture déjà écrites pour ce
        n° chez ce fournisseur (vide si aucune)."""
        return {
            str(p.get("Fichier") or "")
            for p in self._par_facture.get((nom_fournisseur_suivi(fournisseur).upper(), cle_reference(numero_facture)), [])
        }
