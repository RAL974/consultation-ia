"""
Orchestration du rapprochement factures — même principe que
moteur/rapprochement/pipeline_bl.py (BL), en deux temps :

1. `rapprocher_dossier_factures()` — LECTURE SEULE : lit les factures de
   `dossier_a_traiter`, résout leur(s) commande(s) (voir
   `_resoudre_commandes_facture`), rapproche chaque bloc de BL cité contre
   le Suivi (moteur.rapprochement.matching_facture). Ne modifie jamais rien.
2. `appliquer_et_archiver_factures()` — écrit les 5 colonnes facture (voir
   moteur.rapprochement.ecriture.ENTETES_FACTURE, créées pour de vrai dans
   le Suivi vivant le 2026-09-01, réutilisées telles quelles ici — voir
   `ecritures_pour_facture()`), puis range chaque facture lue :
   - entièrement résolue -> COPIÉE (jamais déplacée : une facture peut
     concerner PLUSIEURS commandes, voir CLAUDE.md "Volet 3" — une facture
     multi-BC ne peut pas être découpée par page comme un BL Cominter
     multi-BL, un PDF de facture 109 n'a qu'une seule page utile) dans
     a_traiter/BL/Traités/<n° de commande>/ POUR CHAQUE commande concernée
     (nommage voir `_nom_archive_facture`), puis le fichier source est
     supprimé de a_traiter/Factures/ ;
   - au moins une ligne non résolue (inconnue, "à confirmer" non cochée,
     commande introuvable, avoir) -> a_traiter/Factures/À vérifier/, fichier
     entier, jamais découpé (voir ci-dessus).
   - fournisseur RECONNU mais SANS parser facture (`parser_facture()` de
     moteur.rapprochement.parsers_facture retourne None, voir lecture_facture.
     lire_facture) -> laissé EN PLACE dans a_traiter/Factures/, jamais
     déplacé vers À vérifier/ (ce n'est pas une décision humaine en attente,
     c'est un fournisseur qui n'est simplement pas encore couvert — voir
     `_est_anomalie_sans_parser`) — juste listé dans le rapport et dans
     `resume["factures_sans_parser"]`. Seules les VRAIES anomalies de
     lecture (fournisseur non reconnu, PDF illisible) partent en À vérifier/.
   Puis écrit un rapport dans rapports/, avec le compteur de résorption
   (lignes livrées sans facture, voir `compter_lignes_a_facturer`).

Réutilise autant que possible l'infrastructure déjà éprouvée côté BL
(moteur.rapprochement.pipeline_bl) plutôt que de la dupliquer : recherche du
Suivi vivant, recherche/copie du bon de commande, dossier de commande,
nettoyage de nom de fichier, écriture d'un rapport horodaté.
"""

import shutil
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from moteur.referentiel import Referentiel
from moteur.rapprochement.ecriture import Ecriture, appliquer
from moteur.rapprochement.pieces import (
    COMMENTAIRE_MONTANT_RECALCULE,
    MODE_AUTO,
    MODE_CONFIRME,
    MODE_EQUIVALENCE,
    TYPE_AVOIR,
    TYPE_FACTURE,
    TYPE_FRAIS,
    IndexPieces,
    dedoublonner_ids,
    ecrire_pieces,
    lire_pieces,
    nouvelle_piece,
)
from moteur.rapprochement.lecture_facture import analyser_dossier
from moteur.rapprochement.matching import deduire_commande_par_contenu
from moteur.rapprochement.matching_facture import (
    CauseFacture,
    CorrespondanceFacture,
    StatutFacture,
    apparier_facture,
    charger_frais_fournisseurs,
    est_bdc_manuel_24x,
    index_pieces_du_suivi,
    lire_lignes_commande_facture,
    lire_lignes_fournisseur_facture,
)
from moteur.rapprochement.modele_bl import LigneBL
from moteur.rapprochement.pipeline_bl import (
    DOSSIER_BACKUPS,
    DOSSIER_REFERENTIEL,
    DOSSIER_TRAITES,
    _copier_bon_de_commande_si_absent,
    _dossier_pour_commande,
    _parser_date_bl,
    _sans_caracteres_interdits,
    deplacer_vers_a_verifier,
    trouver_dossier_commandes,
    trouver_fichier_suivi_vivant,
)

DOSSIER_A_TRAITER_FACTURES = "a_traiter/Factures"
DOSSIER_A_VERIFIER_FACTURES = "À vérifier"  # sous-dossier de a_traiter/Factures/
DOSSIER_TRAITES_BL = "a_traiter/BL/" + DOSSIER_TRAITES  # archivage final DANS le
# même arbre que les BL, par commande (demande explicite de l'acheteur, déjà
# actée côté BL : "ainsi nous aurons tout le flux commande-BL-facture
# facilement consultable" — voir CLAUDE.md, "dossier par commande").
DOSSIER_RAPPORTS = "rapports"
NOM_A_CONFIRMER_FACTURE = "A_confirmer_Facture.xlsx"  # fichier À PART de
# A_confirmer_BL.xlsx et A_confirmer.xlsx (devis) — même moteur/articles.db
# partagé (un alias confirmé vaut pour tous les flux), mais chacun régénère
# SA PROPRE file d'attente à chaque exécution (voir moteur/referentiel.py).
NOM_FRAIS_FOURNISSEURS = "frais_fournisseurs.csv"  # referentiel/ — voir
# moteur.rapprochement.matching_facture.charger_frais_fournisseurs, session
# S0 : références connues qui ne sont PAS de vrais articles (frais de port,
# éco-taxe...), jamais rapprochées à une ligne du Suivi, jamais bloquantes.
NOM_FEUILLE_SUBSTITUTIONS = "Substitutions probables"  # 2e feuille de
# A_confirmer_Facture.xlsx (même fichier que le référentiel, feuille à
# part) — voir _ecrire_substitutions_probables/_appliquer_confirmations_
# substitutions : les correspondances CauseFacture.SUBSTITUTION_PROBABLE
# (voir matching_facture._residuel_unique) ne viennent PAS de
# Referentiel.resoudre() (aucune ressemblance structurelle de référence à
# proposer), impossible de les glisser dans Referentiel._propositions sans
# dénaturer ce mécanisme — une feuille séparée, écrite/lue directement ici,
# reste le moyen le plus simple de les faire transiter par le MÊME fichier
# de confirmation qu'utilise déjà l'acheteur pour ce flux.

# Seuil de réconciliation Total HT facture vs somme des lignes extraites —
# même tolérance que le contrôle interne (console) de chaque parser (voir
# moteur.fournisseurs.dist109/coredime) : 0,02€, ici SURFACÉ dans le
# rapport (CauseFacture.TOTAL_ECART) plutôt que seulement imprimé au
# terminal. Sans ce contrôle, une facture dont AUCUNE ligne n'est extraite
# (donc rien à comparer, rien "d'inconnu" non plus) pouvait être
# considérée "entièrement résolue" et archivée SANS AUCUNE trace de son
# montant manquant (cas réel trouvé en session S0,
# facture_coredime_6108846_remise_double_x3.pdf : 196,92€ silencieusement
# perdus avant ce correctif).
SEUIL_TOTAL_ECART_FACTURE = 0.02


@dataclass
class RapportRapprochementFacture:

    surs: list = field(default_factory=list)          # [(Facture, CorrespondanceFacture)]
    a_confirmer: list = field(default_factory=list)
    deja_a_jour: list = field(default_factory=list)
    inconnus: list = field(default_factory=list)
    frais: list = field(default_factory=list)          # [(Facture, CorrespondanceFacture)] statut FRAIS — voir charger_frais_fournisseurs
    anomalies_lecture: list = field(default_factory=list)  # [(nom_fichier, raison)]
    anomalies_facture: list = field(default_factory=list)  # [(Facture, raison)] — avoir, commande introuvable...
    fichier_suivi: Path | None = None


def classifier_cause_anomalie(raison: str) -> CauseFacture | None:
    """Dérive une CauseFacture à partir du texte d'une anomalie EN CLAIR
    (anomalies_lecture/anomalies_facture, restées des 2-tuples — voir
    RapportRapprochementFacture, jamais changées de forme pour ne pas
    casser gui_rapprochement_facture.py) : best-effort, sert uniquement au
    compte rendu chiffré par cause en fin de rapport, jamais à une
    décision de rapprochement elle-même. None si aucun motif reconnu."""

    if "DOUBLON" in raison:
        return CauseFacture.DOUBLON_FACTURE
    if "AVOIR" in raison:
        return CauseFacture.AVOIR
    if "bon manuel" in raison:
        return CauseFacture.BDC_MANUEL_24X
    if "Écart de Total HT" in raison:
        return CauseFacture.TOTAL_ECART
    if "introuvable dans le Suivi" in raison or "n° de commande introuvable" in raison:
        return CauseFacture.COMMANDE_ABSENTE
    if "Fournisseur" in raison and "reconnu" in raison and "non" in raison:
        return CauseFacture.FOURNISSEUR_INCONNU
    if "pas encore de parser facture" in raison:
        return CauseFacture.PARSER_ABSENT
    if "Aucune ligne extraite" in raison:
        return CauseFacture.ZERO_LIGNE
    if "PDF illisible" in raison or "Erreur de lecture" in raison:
        return CauseFacture.ANNEXE_SANS_TEXTE
    return None


def _resoudre_commandes_facture(facture, fichier_suivi) -> dict:
    """{numero_bl: (numero_commande | None, deduit: bool, raison_deduction: str | None)}
    pour chaque bloc de BL cité par la facture.

    En-tête (N°Réf.Client, voir moteur.fournisseurs.dist109) fait autorité
    dès qu'elle donne UN candidat clair (68/79 factures du lot de cadrage) —
    appliqué à TOUS les blocs de BL de la facture, y compris quand elle en
    cite plusieurs (89% des factures multi-BL du lot de cadrage ne couvrent
    qu'UNE SEULE commande, livrée en plusieurs fois). En repli — en-tête
    vide ou format interne 109 non exploitable (ex. "BC 241659") — la
    commande de CE bloc est déduite de son propre contenu (même mécanisme
    que pour un n° de commande illisible sur un BL, voir
    matching.deduire_commande_par_contenu : au moins 2 lignes concordantes,
    score sans ambiguïté) contre TOUTES les commandes de ce fournisseur
    dans le Suivi ; jamais utilisée pour un rapprochement "sûr" automatique
    (voir rapprocher_dossier_factures, comme côté BL)."""

    candidat_entete = facture.numeros_commande[0] if len(facture.numeros_commande) == 1 else None

    par_bl = {}
    for l in facture.lignes:
        par_bl.setdefault(l.numero_bl, []).append(l)

    resolutions = {}
    lignes_fournisseur = None

    for numero_bl, lignes_bloc in par_bl.items():

        if candidat_entete:
            resolutions[numero_bl] = (candidat_entete, False, None)
            continue

        if lignes_fournisseur is None:
            lignes_fournisseur = lire_lignes_fournisseur_facture(fichier_suivi, facture.fournisseur)

        pseudo = [
            LigneBL(reference_fournisseur=l.reference_fournisseur, designation=l.designation,
                    quantite_livree=l.quantite_facturee)
            for l in lignes_bloc
        ]
        deduit, score = deduire_commande_par_contenu(pseudo, lignes_fournisseur)

        if deduit:
            entete_affichee = "; ".join(facture.numeros_commande) or "(vide)"
            resolutions[numero_bl] = (
                deduit, True,
                f"N° de commande « {deduit} » déduit du contenu du BL {numero_bl} "
                f"({score} référence(s)/quantité(s) concordantes) — N°Réf.Client de la facture "
                f"non exploitable tel quel (« {entete_affichee} »)",
            )
        else:
            resolutions[numero_bl] = (None, False, None)

    return resolutions


def _verifier_total_ht_facture(facture) -> str | None:
    """Réconciliation Total HT (voir SEUIL_TOTAL_ECART_FACTURE) sur les
    lignes BRUTES du parser (frais compris, avant tout filtrage/
    agrégation) — retourne la raison en clair si l'écart dépasse le seuil,
    None sinon. Cas réel qui motive ce contrôle (session S0) :
    facture_coredime_6108846_remise_double_x3.pdf, 0 ligne extraite pour un
    Total HT affiché de 196,92€ — sans ce contrôle, une facture SANS AUCUNE
    ligne (donc rien "d'inconnu" non plus) pouvait être considérée
    "entièrement résolue" et archivée en silence, perdant toute trace du
    montant manquant (voir _est_resolu_facture, rapprocher_dossier_factures)."""

    if facture.total_ht_affiche is None:
        return None

    total_extrait = round(sum(l.montant_ht or 0.0 for l in facture.lignes), 2)
    if abs(facture.total_ht_affiche - total_extrait) <= SEUIL_TOTAL_ECART_FACTURE:
        return None

    return (
        f"Écart de Total HT : {facture.total_ht_affiche - total_extrait:+.2f}€ "
        f"(affiché {facture.total_ht_affiche:.2f}€, extrait {total_extrait:.2f}€) "
        "— une ligne a peut-être été oubliée ou mal lue"
    )


def _appliquer_confirmations_substitutions(dossier_referentiel: Path, nom_fichier: str,
                                            referentiel: Referentiel) -> int:
    """Lit la feuille NOM_FEUILLE_SUBSTITUTIONS de A_confirmer_Facture.xlsx
    (si elle existe) : chaque ligne dont Décision == "OUI" est apprise dans
    referentiel/equivalences_bl.csv (voir Referentiel.apprendre_equivalence)
    — JAMAIS directement dans la table alias comme un repli référentiel
    ordinaire (voir bandeau de NOM_FEUILLE_SUBSTITUTIONS). Une ligne dont
    Décision est vide ou "NON" est simplement laissée de côté (ni apprise,
    ni marquée rejetée — elle réapparaîtra tant que le résiduel unique se
    reproduit ; contrairement au workflow référentiel classique, il n'y a
    pas de notion de "rejet définitif" ici, une substitution proposée par
    élimination n'a pas vocation à être mémorisée comme fausse pour
    toujours). Retourne le nombre de paires nouvellement apprises."""

    fichier = Path(dossier_referentiel) / nom_fichier
    if not fichier.exists():
        return 0

    wb = load_workbook(fichier, data_only=True)
    if NOM_FEUILLE_SUBSTITUTIONS not in wb.sheetnames:
        return 0
    ws = wb[NOM_FEUILLE_SUBSTITUTIONS]

    lignes = list(ws.iter_rows(values_only=True))
    if not lignes:
        return 0

    entetes = [str(c).strip() if c else "" for c in lignes[0]]

    def idx(nom):
        return entetes.index(nom) if nom in entetes else None

    i_ref_facturee = idx("Référence facturée")
    i_ref_suivi = idx("Référence Suivi (connue)")
    i_decision = idx("Décision")
    i_facture = idx("Facture")
    i_commande = idx("Commande")

    if i_ref_facturee is None or i_ref_suivi is None or i_decision is None:
        return 0

    n = 0
    for ligne in lignes[1:]:

        if ligne is None or i_decision >= len(ligne):
            continue

        decision = str(ligne[i_decision] or "").strip().upper()
        if decision != "OUI":
            continue

        ref_facturee = str(ligne[i_ref_facturee] or "").strip()
        ref_suivi = str(ligne[i_ref_suivi] or "").strip()
        if not ref_facturee or not ref_suivi:
            continue

        note = (
            f"Substitution probable confirmée (résiduel unique) — facture "
            f"{ligne[i_facture] if i_facture is not None else '?'}, commande "
            f"{ligne[i_commande] if i_commande is not None else '?'}"
        )
        if referentiel.apprendre_equivalence(
            Path(dossier_referentiel) / "equivalences_bl.csv", ref_suivi, ref_facturee, note,
        ):
            n += 1

    if n:
        print(f"{n} équivalence(s) apprise(s) depuis « {NOM_FEUILLE_SUBSTITUTIONS} » (equivalences_bl.csv).")

    return n


def _ecrire_substitutions_probables(dossier_referentiel: Path, nom_fichier: str,
                                     rapport: "RapportRapprochementFacture") -> None:
    """Ajoute (ou retire, si plus aucune) la feuille NOM_FEUILLE_SUBSTITUTIONS
    dans A_confirmer_Facture.xlsx — appelée APRÈS referentiel.
    ecrire_a_confirmer() (qui régénère/supprime la feuille primaire "À
    confirmer" du référentiel), donc le fichier peut ou non déjà exister à
    cet instant. Round-trip openpyxl (load_workbook/save) : sans risque
    pour ce fichier de confirmation jetable (pas le classeur Suivi vivant,
    qui lui passe TOUJOURS par un patch XML chirurgical, jamais
    openpyxl.save() — voir moteur.rapprochement.ecriture)."""

    lignes = [
        (facture, c) for facture, c in rapport.a_confirmer
        if c.cause is CauseFacture.SUBSTITUTION_PROBABLE
    ]

    fichier = Path(dossier_referentiel) / nom_fichier

    if not lignes:
        if not fichier.exists():
            return
        wb = load_workbook(fichier)
        if NOM_FEUILLE_SUBSTITUTIONS not in wb.sheetnames:
            return
        del wb[NOM_FEUILLE_SUBSTITUTIONS]
        if wb.sheetnames:
            wb.save(fichier)
        else:
            fichier.unlink()
        return

    if fichier.exists():
        wb = load_workbook(fichier)
        if NOM_FEUILLE_SUBSTITUTIONS in wb.sheetnames:
            del wb[NOM_FEUILLE_SUBSTITUTIONS]
        ws = wb.create_sheet(NOM_FEUILLE_SUBSTITUTIONS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = NOM_FEUILLE_SUBSTITUTIONS

    entetes = [
        "Fournisseur", "Facture", "Commande", "Référence facturée", "Désignation facturée",
        "Qté facturée", "PU facturé", "Référence Suivi (connue)", "Désignation Suivi",
        "Qté Suivi", "Tarif Suivi", "Décision",
    ]
    ws.append(entetes)

    for facture, c in lignes:
        lf, ls = c.ligne_facture, c.ligne_suivi
        ws.append([
            facture.fournisseur, facture.numero_facture, lf.numero_commande,
            lf.reference_fournisseur, lf.designation, lf.quantite_facturee, lf.prix_unitaire_ht,
            ls.reference, ls.designation, ls.qte_livree, ls.tarif_bl or ls.tarif_convenu,
            "",
        ])

    for cellule in ws[1]:
        cellule.font = Font(bold=True)

    wb.save(fichier)
    print(f"{len(lignes)} substitution(s) probable(s) à confirmer : feuille « {NOM_FEUILLE_SUBSTITUTIONS} » de {fichier}")


def rapprocher_dossier_factures(dossier_a_traiter, dossier_projet) -> RapportRapprochementFacture:
    """Lecture seule : lit toutes les factures du dossier, les rapproche du
    Suivi. Ne modifie ni le Suivi ni les fichiers de `dossier_a_traiter`."""

    dossier_projet = Path(dossier_projet)
    fichier_suivi = trouver_fichier_suivi_vivant(dossier_projet)

    factures, anomalies = analyser_dossier(dossier_a_traiter)

    rapport = RapportRapprochementFacture(anomalies_lecture=list(anomalies), fichier_suivi=fichier_suivi)

    if fichier_suivi is None:
        for f in factures:
            rapport.anomalies_facture.append((f, "Suivi commandes introuvable"))
        return rapport

    dossier_referentiel = dossier_projet / DOSSIER_REFERENTIEL
    referentiel = Referentiel(dossier_projet / "moteur")
    referentiel.importer_bdd(dossier_projet / "base" / "BDD_articles.csv")
    referentiel.importer_equivalences_bl(dossier_referentiel / "equivalences_bl.csv")
    referentiel.appliquer_confirmations(dossier_referentiel / NOM_A_CONFIRMER_FACTURE)
    _appliquer_confirmations_substitutions(dossier_referentiel, NOM_A_CONFIRMER_FACTURE, referentiel)
    frais_connus = charger_frais_fournisseurs(dossier_referentiel / NOM_FRAIS_FOURNISSEURS)
    index_pieces = index_pieces_du_suivi(fichier_suivi)  # feuille Pièces lue UNE fois par lot

    for facture in factures:

        doublon = _doublon_facture(facture, index_pieces)
        if doublon:
            # Même n° de facture déjà écrit dans Pièces pour ce fournisseur
            # depuis un AUTRE document (date/fichier différents) : jamais
            # écrit, jamais rapproché — décision humaine (À vérifier).
            rapport.anomalies_facture.append((facture, doublon))
            continue

        if facture.type_document == "AVOIR":
            # Jamais rapprochée automatiquement (voir CLAUDE.md, Volet 3) —
            # aucun exemple réel à ce jour, aucune règle inventée au-delà de
            # « ne jamais toucher », en attendant un vrai cas.
            rapport.anomalies_facture.append((
                facture, "Facture d'AVOIR — jamais rapprochée automatiquement, à traiter manuellement",
            ))
            continue

        # Non bloquant pour le reste (les lignes lisibles sont quand même
        # rapprochées ci-dessous), mais empêche la facture d'être
        # considérée "résolue" tant que l'écart n'a pas été vérifié à la
        # main (anomalies_facture alimente g["anomalies"], voir
        # _est_resolu_facture).
        raison_ecart = _verifier_total_ht_facture(facture)
        if raison_ecart:
            rapport.anomalies_facture.append((facture, raison_ecart))

        resolutions = _resoudre_commandes_facture(facture, fichier_suivi)

        par_bl = {}
        for l in facture.lignes:
            par_bl.setdefault(l.numero_bl, []).append(l)

        # Regroupe par commande RÉSOLUE (pas par BL brut) : un même article
        # peut être réparti sur PLUSIEURS BL d'une même facture pour la
        # MÊME commande (cas réel Facture_362840.pdf, voir CLAUDE.md
        # session S0, correction 1a) — les comparer à la Qté livrée
        # nécessite de les additionner d'abord (voir
        # matching_facture.agreger_lignes_meme_reference), jamais bloc par
        # bloc indépendamment (chaque bloc, confronté isolément à la Qté
        # livrée TOTALE, ressortirait "à confirmer" à tort sur une
        # livraison fractionnée qui, additionnée, correspond exactement).
        par_commande = {}
        for numero_bl, lignes_bloc in par_bl.items():

            numero_commande, deduit, raison_deduction = resolutions.get(numero_bl, (None, False, None))

            if not numero_commande:
                entete_affichee = "; ".join(facture.numeros_commande) or "(vide)"
                if est_bdc_manuel_24x(facture.numeros_commande_bruts):
                    rapport.anomalies_facture.append((
                        facture,
                        f"BL {numero_bl} : commande passée sur un bon manuel "
                        f"(« {'; '.join(facture.numeros_commande_bruts)} ») — hors format "
                        "Suivi, à rattacher à la main",
                    ))
                else:
                    rapport.anomalies_facture.append((
                        facture,
                        f"BL {numero_bl} : n° de commande introuvable (N°Réf.Client « {entete_affichee} » "
                        "non exploitable, et déduction par contenu non concluante)",
                    ))
                continue

            groupe = par_commande.setdefault(
                numero_commande, {"lignes": [], "deduit": False, "raisons_deduction": []},
            )
            for l in lignes_bloc:
                l.numero_commande = numero_commande
            groupe["lignes"].extend(lignes_bloc)
            if deduit:
                groupe["deduit"] = True
                groupe["raisons_deduction"].append(raison_deduction)

        for numero_commande, groupe in par_commande.items():

            try:
                lignes_suivi = lire_lignes_commande_facture(
                    fichier_suivi, facture.fournisseur, numero_commande, index_pieces=index_pieces,
                )
            except Exception as e:
                rapport.anomalies_facture.append((facture, f"Erreur de lecture du Suivi ({e})"))
                continue

            if not lignes_suivi:
                rapport.anomalies_facture.append((
                    facture,
                    f"Commande {numero_commande} introuvable dans le Suivi pour « {facture.fournisseur} »",
                ))
                continue

            for c in apparier_facture(
                groupe["lignes"], lignes_suivi, facture.numero_facture,
                referentiel=referentiel, fournisseur=facture.fournisseur, devis=facture.numero_facture,
                frais_connus=frais_connus,
            ):
                if groupe["deduit"] and c.statut is StatutFacture.SUR:
                    # Un n° de commande DÉDUIT (pas lu directement) n'est
                    # jamais assez sûr pour une écriture automatique — même
                    # principe que pipeline_bl.py.
                    c = CorrespondanceFacture(
                        c.ligne_facture, c.ligne_suivi, StatutFacture.A_CONFIRMER,
                        list(groupe["raisons_deduction"]) + c.raisons, c.cause,
                    )

                if c.statut is StatutFacture.SUR:
                    rapport.surs.append((facture, c))
                elif c.statut is StatutFacture.A_CONFIRMER:
                    rapport.a_confirmer.append((facture, c))
                elif c.statut is StatutFacture.DEJA_A_JOUR:
                    rapport.deja_a_jour.append((facture, c))
                elif c.statut is StatutFacture.FRAIS:
                    rapport.frais.append((facture, c))
                else:
                    rapport.inconnus.append((facture, c))

    _desamorcer_conflits_meme_ligne_suivi_facture(rapport)

    referentiel.ecrire_a_confirmer(dossier_referentiel, NOM_A_CONFIRMER_FACTURE)
    _ecrire_substitutions_probables(dossier_referentiel, NOM_A_CONFIRMER_FACTURE, rapport)
    referentiel.fermer()

    return rapport


def _desamorcer_conflits_meme_ligne_suivi_facture(rapport: RapportRapprochementFacture) -> None:
    """Même garde-fou que pipeline_bl._desamorcer_conflits_meme_ligne_suivi :
    si deux FICHIERS différents du lot proposent chacun une correspondance
    "sûre" vers la même ligne du Suivi, seule la 1ère reste sûre — l'autre
    bascule "à confirmer" (doublon de dépôt probable, ou vrai litige de
    facturation à trancher à la main)."""

    lignes_vues = {}
    surs_filtres = []

    for facture, c in rapport.surs:

        cle = c.ligne_suivi.ligne_excel
        premier_fichier = lignes_vues.get(cle)

        if premier_fichier is not None and premier_fichier != facture.fichier:
            rapport.a_confirmer.append((facture, CorrespondanceFacture(
                c.ligne_facture, c.ligne_suivi, StatutFacture.A_CONFIRMER,
                [
                    f"Une autre facture de ce même lot ({premier_fichier}) cible aussi cette "
                    "ligne du Suivi — vérifier qu'il ne s'agit pas de la même facture déposée "
                    "deux fois avant de confirmer"
                ],
            )))
            continue

        lignes_vues.setdefault(cle, facture.fichier)
        surs_filtres.append((facture, c))

    rapport.surs = surs_filtres


def regrouper_par_facture(rapport: RapportRapprochementFacture) -> dict:
    """{id(facture): {"facture":..., "sur": [...], "a_confirmer": [...],
    "deja_a_jour": [...], "inconnu": [...], "frais": [...],
    "anomalies": [raison, ...]}} — même principe que
    pipeline_bl.regrouper_par_bl. "frais" ne participe PAS au calcul de
    _est_resolu_facture (ni au numérateur ni au dénominateur) : un frais
    connu (voir charger_frais_fournisseurs) n'est jamais bloquant, jamais
    besoin d'être "confirmé" pour que la facture soit considérée résolue."""

    groupes = {}

    def _groupe(facture):
        return groupes.setdefault(id(facture), {
            "facture": facture, "sur": [], "a_confirmer": [], "deja_a_jour": [], "inconnu": [],
            "frais": [], "anomalies": [],
        })

    for facture, c in rapport.surs:
        _groupe(facture)["sur"].append(c)
    for facture, c in rapport.a_confirmer:
        _groupe(facture)["a_confirmer"].append(c)
    for facture, c in rapport.deja_a_jour:
        _groupe(facture)["deja_a_jour"].append(c)
    for facture, c in rapport.inconnus:
        _groupe(facture)["inconnu"].append(c)
    for facture, c in rapport.frais:
        _groupe(facture)["frais"].append(c)
    for facture, raison in rapport.anomalies_facture:
        _groupe(facture)["anomalies"].append(raison)

    return groupes


# Exception ENCADRÉE, demandée et validée par l'acheteur en une phrase
# (session F4, cadrage avant code Coredime) : ses BL n'affichent JAMAIS de
# prix (voir moteur.fournisseurs.coredime, "GABARIT BL" — réglé à la
# facture) donc "Tarif BL" du Suivi reste éternellement vide pour ce
# fournisseur si on ne le renseigne QUE depuis le flux BL comme pour les
# autres — bloquant le contrôle de prix Excel ("Statut commande",
# ⚠️ Surfacturation, qui LIT Tarif BL directement, indépendamment de ce
# module). Un fournisseur de cette liste blanche voit son PU facturé
# recopié dans Tarif BL, mais UNIQUEMENT si Tarif BL est encore vide
# (jamais un écrasement d'une valeur BL réelle déjà présente) et
# UNIQUEMENT pour une ligne réellement écrite (sûre, ou "à confirmer"
# cochée) — jamais pour une ligne laissée de côté. Tracé dans
# resume["tarif_bl_ecrit_depuis_facture"] (voir appliquer_et_archiver_
# factures), jamais un effet de bord silencieux.
FOURNISSEURS_TARIF_BL_DEPUIS_FACTURE = {"COREDIME"}


def _prefixe_archive_facture(facture) -> str:
    """"<date> - <fournisseur> - Facture <n°>" : identité d'un DOCUMENT dans
    le libellé « Fichier » de Pièces (voir _nom_archive_facture), commune à
    toutes ses copies par commande."""
    return _nom_archive_facture(facture, "x").rsplit(" - BC ", 1)[0]


def _doublon_facture(facture, index_pieces) -> str | None:
    """Garde-fou double facturation au niveau du DOCUMENT : le même n° de
    facture, chez le même fournisseur, est déjà écrit dans Pièces depuis un
    fichier dont l'identité (date - fournisseur - n°) diffère -> raison en
    clair (jamais écrit), sinon None. Un redépôt du MÊME document (même
    identité) n'est pas un doublon : ses lignes ressortent « déjà à jour »
    par leur ID (voir matching_facture._comparer_facture)."""

    if index_pieces is None or not index_pieces.disponible or facture.type_document == "AVOIR":
        return None
    fichiers = index_pieces.fichiers_facture(facture.fournisseur, facture.numero_facture)
    if not fichiers:
        return None
    prefixe = _prefixe_archive_facture(facture)
    if any(f.startswith(prefixe + " - BC") for f in fichiers):
        return None
    return (
        f"DOUBLON : la facture {facture.numero_facture} de {facture.fournisseur} est déjà enregistrée dans Pièces "
        f"depuis un autre document ({', '.join(sorted(fichiers))}) — jamais réécrite, à vérifier à la main"
    )


def _chemin_archive_facture(facture, numero_commande: str, dossier_traites_bl) -> Path:
    """Chemin FUTUR (ou présent) de la copie archivée de cette facture pour
    cette commande — c'est ce chemin qui va dans Pièces[Fichier]
    (=HYPERLINK), écrit AVANT l'archivage (voir appliquer_et_archiver_
    factures : l'écriture précède le rangement) ; identique quel que soit
    le sort du fichier source (archivé tout de suite, ou après passage par
    À vérifier/)."""

    suffixe = Path(facture.fichier or "").suffix or ".pdf"
    return _dossier_pour_commande(Path(dossier_traites_bl), numero_commande) / f"{_nom_archive_facture(facture, numero_commande)}{suffixe}"


def _mode_rapprochement(c) -> str:
    if c.statut is StatutFacture.SUR:
        return MODE_AUTO
    if c.cause in (CauseFacture.CLE_PARTIELLE, CauseFacture.SUBSTITUTION_PROBABLE):
        return MODE_EQUIVALENCE  # ligne obtenue par un repli (référence proche / référentiel / résiduel), confirmée
    return MODE_CONFIRME


def pieces_pour_facture(correspondances, dossier_traites_bl, frais=()) -> tuple:
    """Construit les lignes Pièces (voir moteur.rapprochement.pieces) pour une
    liste de (Facture, CorrespondanceFacture) déjà décidées « à écrire » :
    une ligne de type Facture (Avoir si le document en est un) par
    correspondance, avec Chantier/Sous-Chantier/Référence Suivi copiés de
    la ligne Commandes rapprochée, N° BL lié = BL cité sur la facture pour
    cette ligne, Fichier = chemin d'archive (voir _chemin_archive_facture),
    Mode = Auto (sûre) / Confirmé (cochée) / Équivalence (repli confirmé).
    Les `frais` (StatutFacture.FRAIS, jamais rapprochés à une ligne du
    Suivi) des factures effectivement écrites deviennent des lignes de
    type Frais, sans Référence Suivi.

    « Montant HT » = montant IMPRIMÉ (LigneFacture.montant_ht) ; s'il
    manque, recalculé Qté × PU avec Commentaire « montant recalculé » ET
    une entrée dans `montants_recalcules` — jamais silencieusement.

    Retourne (pieces, ecritures_tarif_bl, montants_recalcules,
    tarif_bl_ecrit) — les Ecriture « Tarif BL » de la liste blanche
    FOURNISSEURS_TARIF_BL_DEPUIS_FACTURE restent écrites dans Commandes
    (seule colonne de Commandes encore écrite par ce flux)."""

    pieces = []
    ecritures = []
    montants_recalcules = []
    tarif_bl_ecrit = []
    factures_ecrites = {}  # id(facture) -> {n° commande: (chantier, sous-chantier)}

    for facture, c in correspondances:

        if c.statut is StatutFacture.DEJA_A_JOUR or c.ligne_suivi is None:
            continue

        ls, lf = c.ligne_suivi, c.ligne_facture
        numero_commande = lf.numero_commande or ls.numero_commande
        commentaire = None

        if lf.montant_ht:
            montant = lf.montant_ht
        else:
            montant = round(lf.quantite_facturee * (lf.prix_unitaire_ht or 0.0), 2)
            commentaire = COMMENTAIRE_MONTANT_RECALCULE
            montants_recalcules.append({
                "fichier": facture.fichier,
                "facture": facture.numero_facture,
                "reference": lf.reference_fournisseur,
                "ligne_excel": ls.ligne_excel,
                "montant": montant,
            })

        qte = lf.quantite_facturee
        type_piece = TYPE_FACTURE
        if facture.type_document == "AVOIR":
            type_piece, qte, montant = TYPE_AVOIR, -abs(qte), -abs(montant)

        pieces.append(nouvelle_piece(
            type_piece, facture.fournisseur, facture.numero_facture, _parser_date_bl(facture.date_facture),
            numero_commande, ls.chantier, ls.sous_chantier, ls.reference,
            lf.reference_fournisseur, lf.designation, qte, lf.prix_unitaire_ht, montant,
            numero_bl_lie=lf.numero_bl, mode=_mode_rapprochement(c),
            fichier=_chemin_archive_facture(facture, numero_commande, dossier_traites_bl),
            commentaire=commentaire,
        ))
        factures_ecrites.setdefault(id(facture), {})[numero_commande] = (ls.chantier, ls.sous_chantier)

        if (
            lf.prix_unitaire_ht
            and facture.fournisseur.upper() in FOURNISSEURS_TARIF_BL_DEPUIS_FACTURE
            and not ls.tarif_bl
        ):
            ecritures.append(Ecriture(ls.ligne_excel, "Tarif BL", lf.prix_unitaire_ht))
            tarif_bl_ecrit.append({
                "fichier": facture.fichier,
                "facture": facture.numero_facture,
                "reference": lf.reference_fournisseur,
                "ligne_excel": ls.ligne_excel,
                "tarif_bl": lf.prix_unitaire_ht,
            })

    for facture, c in frais:
        commandes = factures_ecrites.get(id(facture))
        if not commandes:
            continue  # facture pas écrite dans ce passage : ses frais attendront
        lf = c.ligne_facture
        numero_commande = lf.numero_commande or (next(iter(commandes)) if len(commandes) == 1 else "")
        if not numero_commande:
            continue
        chantier, sous_chantier = commandes.get(numero_commande, (None, None))
        montant = lf.montant_ht if lf.montant_ht is not None else round(lf.quantite_facturee * (lf.prix_unitaire_ht or 0.0), 2)
        pieces.append(nouvelle_piece(
            TYPE_FRAIS, facture.fournisseur, facture.numero_facture, _parser_date_bl(facture.date_facture),
            numero_commande, chantier, sous_chantier, None,
            lf.reference_fournisseur, lf.designation, lf.quantite_facturee, lf.prix_unitaire_ht, montant,
            numero_bl_lie=lf.numero_bl, mode=MODE_AUTO,
            fichier=_chemin_archive_facture(facture, numero_commande, dossier_traites_bl),
            commentaire="; ".join(c.raisons) or None,
        ))

    return dedoublonner_ids(pieces), ecritures, montants_recalcules, tarif_bl_ecrit


def _nom_archive_facture(facture, numero_commande: str) -> str:
    """"<date> - <fournisseur> - Facture <n° facture> - BC <n° commande>"
    (voir CLAUDE.md, Volet 3) — même convention que _nom_archive_bl côté BL."""

    jour = _parser_date_bl(facture.date_facture) or date.today()
    numero_facture = _sans_caracteres_interdits(facture.numero_facture or "sans-numero", "-")
    numero_commande = _sans_caracteres_interdits(numero_commande or "inconnue", "-")
    prefixe = "AVOIR" if facture.type_document == "AVOIR" else facture.fournisseur

    return f"{jour.isoformat()} - {prefixe} - Facture {numero_facture} - BC {numero_commande}"


def archiver_facture(chemin_source: Path, facture, commandes_concernees: list,
                      dossier_traites_bl: Path, dossier_commandes_bc=None) -> list:
    """Copie (jamais déplace, voir bandeau du module) `chemin_source` vers
    a_traiter/BL/Traités/<n° de commande>/ POUR CHAQUE commande de
    `commandes_concernees` (une facture peut en couvrir plusieurs — voir
    CLAUDE.md, Volet 2 : cas réel Facture_365533.pdf). Le fichier source
    n'est supprimé qu'une fois TOUTES les copies faites avec succès."""

    chemin_source = Path(chemin_source)
    cibles = []

    for numero_commande in commandes_concernees:

        dossier = _dossier_pour_commande(dossier_traites_bl, numero_commande)
        dossier.mkdir(parents=True, exist_ok=True)

        base = _nom_archive_facture(facture, numero_commande)
        cible = dossier / f"{base}{chemin_source.suffix}"
        compteur = 1
        while cible.exists():
            compteur += 1
            cible = dossier / f"{base} ({compteur}){chemin_source.suffix}"

        shutil.copy2(chemin_source, cible)
        _copier_bon_de_commande_si_absent(dossier, numero_commande, dossier_commandes_bc)
        cibles.append(cible)

    chemin_source.unlink()
    return cibles


def _est_resolu_facture(g: dict, cles_ecrites: set) -> bool:
    """Même principe que pipeline_bl._est_resolu (sans le cas particulier
    RETOUR, propre à 109 Distribution côté BL, sans équivalent facture)."""

    if g["inconnu"] or g["anomalies"]:
        return False

    resolues = len(g["deja_a_jour"]) + sum(1 for c in g["sur"] + g["a_confirmer"] if id(c) in cles_ecrites)
    total = len(g["sur"]) + len(g["a_confirmer"]) + len(g["deja_a_jour"])

    return resolues >= total


def _raisons_non_resolu_facture(g: dict, cles_ecrites: set) -> list:
    return [r for c in g["inconnu"] for r in c.raisons] + [
        r for c in g["a_confirmer"] if id(c) not in cles_ecrites for r in c.raisons
    ] + list(g["anomalies"])


_MOTIF_SANS_PARSER = "pas encore de parser facture"  # voir lecture_facture.lire_facture, texte exact de la raison


def _est_anomalie_sans_parser(raison: str) -> bool:
    """True si `raison` (une entrée de RapportRapprochementFacture.
    anomalies_lecture) signale un fournisseur RECONNU mais sans parser
    facture (voir lecture_facture.lire_facture) — à distinguer d'une VRAIE
    anomalie de lecture (fournisseur non reconnu, PDF illisible), qui,
    elle, part vers À vérifier/ (voir appliquer_et_archiver_factures)."""

    return _MOTIF_SANS_PARSER in raison


def compter_lignes_a_facturer(chemin_suivi, fournisseur: str) -> dict:
    """Diagnostic de résorption — l'indicateur du chantier F2 (voir
    CLAUDE.md) : lignes du Suivi, pour `fournisseur`, avec une Qté livrée
    non nulle mais sans N° facture renseigné. Fonctionne même si les
    colonnes facture n'existent pas encore dans ce Suivi (voir
    matching_facture, colonnes_facture_disponibles) : toutes les lignes
    livrées ressortent alors "à facturer" faute de pouvoir vérifier — signal
    honnête (pas un 0 trompeur qui masquerait l'absence des colonnes)."""

    lignes = lire_lignes_fournisseur_facture(chemin_suivi, fournisseur)
    livrees = [l for l in lignes if l.qte_livree > 0]
    a_facturer = [l for l in livrees if not l.numero_facture]

    return {
        "livrees": len(livrees),
        "a_facturer": len(a_facturer),
        "deja_facturees": len(livrees) - len(a_facturer),
    }


def ecrire_rapport_facture(dossier_rapports: Path, texte: str) -> Path:
    dossier_rapports = Path(dossier_rapports)
    dossier_rapports.mkdir(parents=True, exist_ok=True)
    horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
    chemin = dossier_rapports / f"rapprochement_facture_{horodatage}.txt"
    chemin.write_text(texte, encoding="utf-8")
    return chemin


def appliquer_et_archiver_factures(dossier_projet, dossier_a_traiter, rapport: RapportRapprochementFacture,
                                    correspondances_a_ecrire: list) -> dict:
    """Écrit `correspondances_a_ecrire` dans le Suivi, puis archive/déplace
    chaque facture — voir bandeau du module. Retourne un résumé
    {sauvegarde, lignes_ecrites, factures_archivees, factures_a_verifier,
    archivage_echoue, chemin_rapport, resorption (dict {fournisseur: ...},
    voir compter_lignes_a_facturer), montants_recalcules,
    tarif_bl_ecrit_depuis_facture (voir FOURNISSEURS_TARIF_BL_DEPUIS_FACTURE),
    factures_sans_parser}.

    **Piège pour tout futur appel direct hors GUI, même piège que côté BL
    (voir pipeline_bl.appliquer_et_archiver, "Piège pour tout futur appel
    direct hors GUI") — RENCONTRÉ POUR DE VRAI ici en session F4** :
    `dossier_a_traiter` DOIT TOUJOURS être `a_traiter/Factures/` lui-même
    (parent direct de `DOSSIER_A_VERIFIER_FACTURES`), jamais un de ses
    sous-dossiers. Un appel pointé par erreur directement sur
    `a_traiter/Factures/À vérifier/` (pour retraiter seulement les factures
    qui y étaient déjà, après un correctif de matching) a créé un
    sous-dossier imbriqué `À vérifier/À vérifier/` au lieu d'y laisser les
    78 factures encore non résolues à plat — repéré et corrigé à la main
    dans la foulée (déplacement des fichiers, aucune perte : ni le Suivi ni
    aucun PDF n'a été affecté, seul le RANGEMENT était faux).
    `rapprocher_dossier_factures()` (lecture seule), lui, peut être pointé
    sur n'importe quel dossier sans risque (c'est ce qui permet de tester
    "À vérifier" isolément) — seul `appliquer_et_archiver_factures()` a
    cette contrainte."""

    dossier_projet = Path(dossier_projet)
    dossier_commandes_bc = trouver_dossier_commandes(dossier_projet)
    dossier_traites_bl = dossier_projet / DOSSIER_TRAITES_BL

    resume = {
        "sauvegarde": None, "lignes_ecrites": 0, "pieces_ecrites": 0, "pieces_ignorees": [],
        "factures_archivees": [], "factures_a_verifier": [],
        "archivage_echoue": [], "chemin_rapport": None, "resorption": None,
        "montants_recalcules": [], "factures_sans_parser": [],
        "tarif_bl_ecrit_depuis_facture": [], "causes": {},
    }

    if correspondances_a_ecrire:
        # Depuis P1 : les lignes de document vont dans la feuille Pièces
        # (une par ligne de facture + les frais des factures écrites) ;
        # dans Commandes ne reste écrit que « Tarif BL » (liste blanche).
        # Pièces d'abord (l'essentiel, refusé net si la feuille manque —
        # FeuillePiecesAbsente), Tarif BL ensuite.
        pieces, ecritures_tarif_bl, montants_recalcules, tarif_bl_ecrit = pieces_pour_facture(
            correspondances_a_ecrire, dossier_traites_bl, frais=rapport.frais,
        )
        resultat_pieces = ecrire_pieces(rapport.fichier_suivi, pieces, dossier_projet / DOSSIER_BACKUPS)
        resume["sauvegarde"] = resultat_pieces["sauvegarde"]
        resume["pieces_ecrites"] = resultat_pieces["ajoutees"]
        resume["pieces_ignorees"] = resultat_pieces["ignorees"]
        if ecritures_tarif_bl:
            sauvegarde_tarif = appliquer(rapport.fichier_suivi, ecritures_tarif_bl, dossier_projet / DOSSIER_BACKUPS)
            resume["sauvegarde"] = resume["sauvegarde"] or sauvegarde_tarif
        resume["lignes_ecrites"] = len(correspondances_a_ecrire)
        resume["montants_recalcules"] = montants_recalcules
        resume["tarif_bl_ecrit_depuis_facture"] = tarif_bl_ecrit

    cles_ecrites = {id(c) for _, c in correspondances_a_ecrire}
    groupes = regrouper_par_facture(rapport)

    dossier_a_traiter = Path(dossier_a_traiter)

    fichiers_avec_facture = {g["facture"].fichier for g in groupes.values()}

    # Un fournisseur RECONNU mais sans parser facture n'est pas une
    # "anomalie" au sens où une décision humaine serait attendue — c'est
    # simplement un fournisseur pas encore couvert. Laissé EN PLACE dans
    # a_traiter/Factures/ (jamais déplacé vers À vérifier/), juste
    # rapporté — voir bandeau du module et `_est_anomalie_sans_parser`.
    fichiers_en_echec_total = set()
    for fichier, raison in rapport.anomalies_lecture:
        if fichier in fichiers_avec_facture:
            continue
        if _est_anomalie_sans_parser(raison):
            resume["factures_sans_parser"].append((fichier, raison))
        else:
            fichiers_en_echec_total.add(fichier)

    for fichier in fichiers_en_echec_total:

        chemin_source = dossier_a_traiter / fichier
        if not chemin_source.exists():
            continue

        raisons = [r for f, r in rapport.anomalies_lecture if f == fichier]
        try:
            cible = deplacer_vers_a_verifier(chemin_source, dossier_a_traiter / DOSSIER_A_VERIFIER_FACTURES)
            resume["factures_a_verifier"].append((fichier, cible, raisons))
        except OSError as e:
            resume["archivage_echoue"].append((fichier, str(e)))

    for g in groupes.values():

        facture = g["facture"]
        chemin_source = dossier_a_traiter / facture.fichier
        if not chemin_source.exists():
            continue

        # Un déplacement/archivage qui échoue (fichier verrouillé...) ne
        # doit jamais faire perdre le résumé de l'écriture déjà faite dans
        # le Suivi ni empêcher le rapport final (leçon retenue de
        # pipeline_bl.py, "bug de robustesse").
        if _est_resolu_facture(g, cles_ecrites):
            commandes = sorted({
                c.ligne_facture.numero_commande
                for c in g["sur"] + g["deja_a_jour"]
                if c.ligne_facture.numero_commande
            })
            try:
                cibles = archiver_facture(chemin_source, facture, commandes, dossier_traites_bl, dossier_commandes_bc)
                resume["factures_archivees"].append((facture.fichier, cibles))
            except OSError as e:
                resume["archivage_echoue"].append((facture.fichier, str(e)))
        else:
            raisons = _raisons_non_resolu_facture(g, cles_ecrites)
            try:
                cible = deplacer_vers_a_verifier(chemin_source, dossier_a_traiter / DOSSIER_A_VERIFIER_FACTURES)
                resume["factures_a_verifier"].append((facture.fichier, cible, raisons))
            except OSError as e:
                resume["archivage_echoue"].append((facture.fichier, str(e)))

    # Résorption PAR FOURNISSEUR (demande explicite, session F4) — pour
    # chaque fournisseur réellement présent dans CE lot, jamais figé sur
    # "109 DISTRIBUTION" comme avant (ne disait plus rien dès qu'un 2e
    # fournisseur était traité).
    resume["resorption"] = {}
    for fournisseur in sorted({g["facture"].fournisseur for g in groupes.values()}):
        try:
            resume["resorption"][fournisseur] = compter_lignes_a_facturer(rapport.fichier_suivi, fournisseur)
        except Exception:
            resume["resorption"][fournisseur] = None

    lignes_rapport = [
        f"Rapprochement factures — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"{resume['lignes_ecrites']} ligne(s) rapprochée(s) écrite(s) — {resume['pieces_ecrites']} ligne(s) "
        f"dans la feuille Pièces (Facture + Frais), {len(resume['pieces_ignorees'])} déjà présente(s) (ID pièce).",
        f"{len(resume['factures_archivees'])} facture(s) archivée(s) : "
        + ", ".join(f for f, _ in resume["factures_archivees"]),
        f"{len(resume['factures_a_verifier'])} facture(s) déplacée(s) vers "
        f"{DOSSIER_A_TRAITER_FACTURES}/{DOSSIER_A_VERIFIER_FACTURES}/ (décision humaine nécessaire) :",
    ] + [
        f"  - {fichier} : {' ; '.join(raisons) if raisons else '(voir détail)'}"
        for fichier, _, raisons in resume["factures_a_verifier"]
    ] + [
        f"{len(resume['archivage_echoue'])} facture(s) écrite(s) mais PAS déplacée(s) "
        "(fichier verrouillé, à ranger à la main) : "
        + ", ".join(f for f, _ in resume["archivage_echoue"]),
        f"{len(rapport.deja_a_jour)} ligne(s) déjà à jour (rien écrit, doublon évité).",
        f"{len(rapport.anomalies_lecture)} anomalie(s) de lecture : "
        + "; ".join(f"{f} ({r})" for f, r in rapport.anomalies_lecture),
        f"{len(rapport.anomalies_facture)} bloc(s) non rapproché(s) (avoir, commande introuvable...) : "
        + "; ".join(f"{f.fichier} ({r})" for f, r in rapport.anomalies_facture),
        f"{len(resume['factures_sans_parser'])} fichier(s) non traité(s) — pas de parser (laissé(s) en place "
        f"dans {DOSSIER_A_TRAITER_FACTURES}/, PAS déplacé(s)) : "
        + ", ".join(f for f, _ in resume["factures_sans_parser"]),
    ]

    for fournisseur, r in (resume["resorption"] or {}).items():
        if r:
            lignes_rapport.append(
                f"Résorption {fournisseur} : {r['a_facturer']} ligne(s) livrée(s) encore sans "
                f"facture sur {r['livrees']} livrée(s) au total ({r['deja_facturees']} déjà facturée(s))."
            )

    if resume["tarif_bl_ecrit_depuis_facture"]:
        lignes_rapport.append(
            f"{len(resume['tarif_bl_ecrit_depuis_facture'])} « Tarif BL » renseigné(s) depuis la facture "
            f"(fournisseur en liste blanche {sorted(FOURNISSEURS_TARIF_BL_DEPUIS_FACTURE)}, "
            "Tarif BL était vide — voir bandeau ecritures_pour_facture) :"
        )
        lignes_rapport += [
            f"  - {t['fichier']} (facture {t['facture']}, réf. {t['reference']}, ligne Excel {t['ligne_excel']}) : "
            f"{t['tarif_bl']:.4f}€"
            for t in resume["tarif_bl_ecrit_depuis_facture"]
        ]

    if resume["montants_recalcules"]:
        lignes_rapport.append(
            f"{len(resume['montants_recalcules'])} « Montant facturé HT » RECALCULÉ (Qté × PU, "
            "aucun montant imprimé pour cette ligne sur la facture) — JAMAIS silencieux, à vérifier :"
        )
        lignes_rapport += [
            f"  - {m['fichier']} (facture {m['facture']}, réf. {m['reference']}, ligne Excel {m['ligne_excel']}) : "
            f"{m['montant']:.2f}€"
            for m in resume["montants_recalcules"]
        ]

    if rapport.frais:
        lignes_rapport.append(
            f"{len(rapport.frais)} ligne(s) de frais connus (jamais rapprochées à une ligne du Suivi, "
            "jamais bloquantes — voir referentiel/frais_fournisseurs.csv) :"
        )
        lignes_rapport += [
            f"  - {f.fichier} (facture {f.numero_facture}, réf. {c.ligne_facture.reference_fournisseur}) : "
            f"{c.ligne_facture.montant_ht or 0.0:.2f}€"
            for f, c in rapport.frais
        ]

    # Compte rendu chiffré PAR CAUSE (session S0, correction 1e) : combine
    # les causes déjà posées directement (a_confirmer/inconnus/frais, voir
    # apparier_facture) et celles dérivées en best-effort du texte des
    # anomalies "fichier entier" (voir classifier_cause_anomalie) — jamais
    # un résidu sans cause identifiée qui resterait invisible du compte
    # rendu ("Pas de résiduel unique", cadrage de session).
    compteur_causes = {}
    for _, c in rapport.a_confirmer + rapport.inconnus + rapport.frais:
        if c.cause is not None:
            compteur_causes[c.cause] = compteur_causes.get(c.cause, 0) + 1
    for _, raison in rapport.anomalies_facture:
        cause = classifier_cause_anomalie(raison)
        if cause is not None:
            compteur_causes[cause] = compteur_causes.get(cause, 0) + 1
    for _, raison in rapport.anomalies_lecture:
        cause = classifier_cause_anomalie(raison)
        if cause is not None:
            compteur_causes[cause] = compteur_causes.get(cause, 0) + 1

    resume["causes"] = {cause.value: n for cause, n in compteur_causes.items()}

    if compteur_causes:
        lignes_rapport.append("Répartition par cause :")
        lignes_rapport += [
            f"  - {cause.value} : {compteur_causes[cause]}"
            for cause in CauseFacture
            if cause in compteur_causes
        ]

    resume["chemin_rapport"] = ecrire_rapport_facture(dossier_projet / DOSSIER_RAPPORTS, "\n".join(lignes_rapport))

    return resume
