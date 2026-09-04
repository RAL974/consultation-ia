"""
Écriture sécurisée dans le classeur Suivi commandes (feuille "Commandes") —
socle de Rapprochement AI (voir CLAUDE.md, branche "Rapprochement AI").

Le classeur est VIVANT : ouvert dans Excel une bonne partie de la journée,
potentiellement par plusieurs personnes (Xavier compris). Trois garde-fous,
dans cet ordre :

1. Verrou Excel (fichier "~$..." à côté du classeur) : jamais d'écriture en
   force, `appliquer()` lève ClasseurVerrouille et laisse réessayer plus
   tard (voir est_verrouille()).
2. Sauvegarde horodatée AVANT toute écriture, avec rotation (voir
   sauvegarder()).
3. Écriture CHIRURGICALE : seule la valeur des cellules demandées est
   modifiée, en patchant directement la partie XML de la feuille visée dans
   le zip OOXML (un .xlsx est un zip) — jamais de désérialisation/
   réécriture complète du classeur.

   openpyxl (load_workbook() + save()) a été essayé en premier sur une
   COPIE du vrai Suivi commandes et écarté pour cet usage : même sans
   toucher à rien d'autre qu'une seule cellule, la réécriture complète fait
   disparaître xl/calcChain.xml, xl/metadata.xml et customXml/* (item de
   propriétés du document), et dégrade des validations de données
   ("Data Validation extension is not supported and will be removed",
   avertissement obtenu sur le classeur réel à l'ouverture) — inacceptable
   sur un fichier vivant riche en formules, tableaux structurés (16 Excel
   Tables) et validations. Le patch XML ciblé ne touche STRICTEMENT rien
   d'autre que le texte des cellules demandées ; chaque autre partie du zip
   (styles, tableaux, validations, calcChain, sharedStrings, customXml,
   printerSettings...) est recopiée octet pour octet.

Seules 4 colonnes de la feuille "Commandes" sont de VRAIES données saisies
au clavier (vérifié cellule par cellule sur le classeur réel) :
"Date de livraison", "Qté livrée", "Tarif BL", et "Note" (texte libre déjà
utilisé comme valeur "magique" par plusieurs formules : "Rupture
fournisseur", "Reliquat soldé", "Commande annulée"). TOUT le reste de la
feuille — Statut commande, Reliquat, RAL, Soldé, Reste à facturer, Facturé
BL, Potentiel factu... — est calculé par une formule Excel à partir de ces
4 colonnes (XLOOKUP/IF en cascade côté "Statut commande" notamment) :
écrire une valeur figée dedans romprait le calcul à la prochaine ouverture
Excel, silencieusement. `appliquer()` refuse toute colonne hors de
COLONNES_MODIFIABLES.

Les 5 colonnes facture de "Commandes" ("N° facture", "Date facture", "Qté
facturée", "PU facturé", "Montant facturé HT"), créées par l'outil le
2026-09-01 comme colonnes de saisie (`ajouter_entetes_saisie()`, patch de
la feuille ET de la définition du tableau), sont depuis P1 (2026-09-04,
voir CLAUDE.md « Feuille Pièces — modèle, socles, migration ») des colonnes
CALCULÉES par formule à partir de la feuille « Pièces » (une ligne par
ligne de document : BL, Retour, Facture, Avoir, Frais, Demande d'avoir —
écrite par l'outil, jamais saisie ; seule sa colonne « Commentaire » est
humaine). L'outil n'écrit donc plus jamais une valeur dans ces 5 colonnes
(retirées de COLONNES_MODIFIABLES) : il ajoute des lignes à Pièces
(ajouter_lignes_tableau(), idempotent par « ID pièce »). Les socles
génériques de P1 (fin de ce fichier : ajouter_feuille_tableau(),
ajouter_lignes_tableau(), ajouter_colonnes_calculees()/
basculer_colonnes_en_formules(), reparer_formules_ligne()) prennent le
chemin du classeur en paramètre et resservent tels quels pour M1 et T1.
"Statut commande" (la formule existante, dont dépend le Dashboard) n'est
et ne sera jamais modifié par ce module.
"""

import re
import shutil
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.datetime import to_excel as _date_vers_excel

FEUILLE_COMMANDES = "Commandes"

# Les seules colonnes de "Commandes" que Rapprochement AI a le droit
# d'écrire (voir bandeau ci-dessus) — toute autre colonne de "Commandes"
# est une formule. Depuis P1 (feuille Pièces, voir CLAUDE.md), les 5
# colonnes facture ("N° facture", "Date facture", "Qté facturée",
# "PU facturé", "Montant facturé HT") sont devenues des colonnes CALCULÉES
# (formules écrites par l'outil, voir moteur.rapprochement.pieces.
# FORMULES_COMMANDES_BASCULE) : l'outil n'y écrit plus jamais une valeur —
# tout passe par des lignes de la feuille Pièces (ajouter_lignes_tableau).
COLONNES_MODIFIABLES = (
    "Date de livraison", "Qté livrée", "Tarif BL", "Note",
)

RETENTION_BACKUPS_JOURS = 30

_NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_NS_PKGREL = "{http://schemas.openxmlformats.org/package/2006/relationships}"


class ClasseurVerrouille(Exception):
    """Le Suivi commandes est actuellement ouvert dans Excel (fichier ~$ présent)."""


class ColonneNonModifiable(ValueError):
    """Tentative d'écriture dans une colonne calculée par formule (ou
    absente des en-têtes) — jamais autorisée, voir COLONNES_MODIFIABLES."""


@dataclass(frozen=True)
class Ecriture:
    """Une écriture demandée : ligne Excel 1-based (la ligne 1 est
    l'en-tête, donc la 1ère commande est en ligne 2), nom d'en-tête (doit
    être dans COLONNES_MODIFIABLES), valeur Python (str / int / float /
    date/datetime)."""

    ligne: int
    colonne: str
    valeur: object


def est_verrouille(fichier) -> bool:
    """True si `fichier` est actuellement ouvert dans Excel (présence du
    fichier de verrou "~$<nom>" dans le même dossier)."""

    fichier = Path(fichier)
    return (fichier.parent / f"~${fichier.name}").exists()


def sauvegarder(fichier, dossier_backups) -> Path:
    """Copie horodatée de `fichier` dans `dossier_backups`, puis purge les
    sauvegardes de plus de RETENTION_BACKUPS_JOURS jours. Retourne le
    chemin de la copie créée."""

    fichier = Path(fichier)
    dossier_backups = Path(dossier_backups)
    dossier_backups.mkdir(parents=True, exist_ok=True)

    horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
    cible = dossier_backups / f"{fichier.stem}_{horodatage}{fichier.suffix}"
    compteur = 1
    while cible.exists():  # deux sauvegardes dans la même seconde (tests, double-clic...)
        compteur += 1
        cible = dossier_backups / f"{fichier.stem}_{horodatage}_{compteur}{fichier.suffix}"
    shutil.copy2(fichier, cible)

    _purger_backups_anciens(dossier_backups)
    return cible


def _purger_backups_anciens(dossier_backups, jours=RETENTION_BACKUPS_JOURS):
    seuil = datetime.now().timestamp() - jours * 86400
    for f in Path(dossier_backups).glob("*.xlsx"):
        if f.stat().st_mtime < seuil:
            f.unlink()


def lire_entetes(fichier, feuille=FEUILLE_COMMANDES) -> dict:
    """{nom d'en-tête: index de colonne 1-based} de la 1ère ligne de
    `feuille`."""

    wb = load_workbook(fichier, read_only=True, data_only=True)
    try:
        ws = wb[feuille]
        premiere = next(ws.iter_rows(min_row=1, max_row=1))
        return {
            str(c.value).strip(): i + 1
            for i, c in enumerate(premiere)
            if c.value is not None
        }
    finally:
        wb.close()


def _verifier_colonnes(ecritures, entetes):
    for e in ecritures:
        if e.colonne not in COLONNES_MODIFIABLES:
            raise ColonneNonModifiable(
                f"« {e.colonne} » n'est pas une colonne saisissable (c'est une "
                f"formule dans le Suivi commandes) — colonnes autorisées : "
                f"{', '.join(COLONNES_MODIFIABLES)}"
            )
        if e.colonne not in entetes:
            raise ColonneNonModifiable(
                f"« {e.colonne} » introuvable dans les en-têtes de la feuille "
                f"« {FEUILLE_COMMANDES} » — export du Suivi commandes différent "
                f"de celui attendu ?"
            )


def simuler(fichier, ecritures, feuille=FEUILLE_COMMANDES) -> list:
    """Rapport « ce qui sera écrit », sans rien modifier dans `fichier`.
    Une entrée par Ecriture : ligne, colonne, ancienne_valeur,
    nouvelle_valeur. C'est ce rapport que l'acheteur doit voir et confirmer
    avant tout appel à appliquer() — mode simulation par défaut."""

    entetes = lire_entetes(fichier, feuille)
    _verifier_colonnes(ecritures, entetes)

    # Une seule passe séquentielle sur la feuille : en mode read_only,
    # ws.cell(row=, column=) ré-analyse le flux XML depuis le début à
    # CHAQUE appel (accès aléatoire non prévu par openpyxl en read_only) —
    # sur ~5 900 lignes, des dizaines d'écritures rendaient ça minutes,
    # voire jamais fini. On collecte donc d'abord {ligne -> {colonne: idx}}
    # demandées, puis on ne lit qu'une fois chaque ligne concernée.
    lignes_demandees = {}
    for e in ecritures:
        lignes_demandees.setdefault(e.ligne, set()).add(entetes[e.colonne])

    derniere_ligne = max(lignes_demandees)
    valeurs_lues = {}  # (ligne, idx_colonne) -> valeur actuelle
    wb = load_workbook(fichier, read_only=True, data_only=True)
    try:
        ws = wb[feuille]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i in lignes_demandees:
                for idx in lignes_demandees[i]:
                    valeurs_lues[(i, idx)] = row[idx - 1] if idx - 1 < len(row) else None
            if i >= derniere_ligne:
                break
    finally:
        wb.close()

    rapport = []
    for e in ecritures:
        idx = entetes[e.colonne]
        rapport.append(
            {
                "ligne": e.ligne,
                "colonne": e.colonne,
                "ancienne_valeur": valeurs_lues.get((e.ligne, idx)),
                "nouvelle_valeur": e.valeur,
            }
        )
    return rapport


def appliquer(fichier, ecritures, dossier_backups, feuille=FEUILLE_COMMANDES) -> Path:
    """Écrit réellement `ecritures` dans `fichier` (feuille `feuille`) :
    verrou -> sauvegarde -> patch XML chirurgical. Lève ClasseurVerrouille
    si le fichier est ouvert dans Excel (jamais d'écriture en force) et
    ColonneNonModifiable si une écriture vise une colonne hors de
    COLONNES_MODIFIABLES (jamais de compromis là-dessus, voir bandeau).
    Retourne le chemin de la sauvegarde créée avant écriture."""

    fichier = Path(fichier)
    if est_verrouille(fichier):
        raise ClasseurVerrouille(
            f"« {fichier.name} » est actuellement ouvert dans Excel — "
            "ferme-le puis réessaie (rien n'a été écrit)."
        )

    entetes = lire_entetes(fichier, feuille)
    _verifier_colonnes(ecritures, entetes)

    sauvegarde = sauvegarder(fichier, dossier_backups)

    cellules = {}
    for e in ecritures:
        idx = entetes[e.colonne]
        ref = f"{get_column_letter(idx)}{e.ligne}"
        cellules[ref] = e.valeur

    _patcher_cellules_xlsx(fichier, feuille, cellules)
    return sauvegarde


def ajouter_entetes_saisie(fichier, noms, dossier_backups, feuille=FEUILLE_COMMANDES) -> Path:
    """Ajoute `noms` comme nouvelles colonnes DANS le tableau structuré de
    `feuille` (même nom que la feuille, ex. "Commandes"), à sa suite : une
    cellule d'en-tête (inlineStr) en ligne 1 pour chaque nom, ET les
    <tableColumn> correspondants ajoutés à la définition du tableau (avec
    extension de sa `ref` et de son `<autoFilter>`). Sciemment DANS le
    tableau, pas à côté : une colonne hors tableau ne suit pas un tri du
    tableau, les données se retrouveraient sur les mauvaises lignes (voir
    CLAUDE.md).

    Patch exactement DEUX parties du zip, rien d'autre : la feuille (ligne 1
    + <dimension>) et la définition XML du tableau (ref, autoFilter,
    tableColumns — AUCUN calculatedColumnFormula, AUCUNE totalsRow ajoutée).
    Même esprit chirurgical que appliquer() (voir bandeau) : tout le reste
    du classeur (styles, calcChain, sharedStrings, customXml, les 15 autres
    tableaux, les autres feuilles...) est recopié octet pour octet.

    Mêmes garde-fous que appliquer() : ClasseurVerrouille si le fichier est
    ouvert dans Excel (jamais d'écriture en force), sauvegarde horodatée
    AVANT toute écriture. Lève ValueError si un `nom` existe déjà (en-tête
    de la feuille OU colonne du tableau), ou si une cellule cible n'est pas
    vide (jamais d'écrasement d'une donnée existante). Retourne le chemin
    de la sauvegarde créée avant écriture."""

    fichier = Path(fichier)
    if not noms:
        raise ValueError("Aucun nom d'en-tête fourni.")
    if len(set(noms)) != len(noms):
        raise ValueError(f"Noms d'en-tête en double dans {noms!r}.")

    if est_verrouille(fichier):
        raise ClasseurVerrouille(
            f"« {fichier.name} » est actuellement ouvert dans Excel — "
            "ferme-le puis réessaie (rien n'a été écrit)."
        )

    entetes = lire_entetes(fichier, feuille)
    for nom in noms:
        if nom in entetes:
            raise ValueError(
                f"« {nom} » existe déjà dans les en-têtes de « {feuille} » "
                f"(colonne {get_column_letter(entetes[nom])}) — rien n'a été écrit."
            )

    derniere_colonne = max(entetes.values()) if entetes else 0
    refs = [
        f"{get_column_letter(derniere_colonne + 1 + i)}1" for i in range(len(noms))
    ]

    with zipfile.ZipFile(fichier, "r") as zin:
        chemin_feuille = _chemin_feuille(zin, feuille)
        xml_feuille = zin.read(chemin_feuille).decode("utf-8")
        chemin_table = _chemin_table_par_nom(zin, feuille)
        xml_table = zin.read(chemin_table).decode("utf-8")

    _verifier_cellules_vides(xml_feuille, 1, refs)
    _verifier_colonnes_table_absentes(xml_table, noms)

    colonne_min_requise = derniere_colonne + len(noms)
    nouvelle_dimension = _dimension_etendue(xml_feuille, colonne_min_requise)
    xml_table = _etendre_tableau(xml_table, noms, colonne_min_requise)

    sauvegarde = sauvegarder(fichier, dossier_backups)

    xml_feuille = _remplacer_dans_ligne(xml_feuille, 1, dict(zip(refs, noms)))
    xml_feuille = re.sub(
        r'<dimension ref="[^"]*"\s*/>',
        f'<dimension ref="{nouvelle_dimension}"/>',
        xml_feuille,
        count=1,
    )

    _patcher_parties_xlsx(fichier, {chemin_feuille: xml_feuille, chemin_table: xml_table})
    return sauvegarde


# --- Patch XML chirurgical -------------------------------------------------


def _chemin_feuille(zin: zipfile.ZipFile, nom_feuille: str) -> str:
    """xl/worksheets/sheetN.xml correspondant au nom d'onglet `nom_feuille`,
    en suivant xl/workbook.xml -> xl/_rels/workbook.xml.rels (lecture seule,
    ces deux fichiers de contrôle ne sont jamais réécrits)."""

    workbook_xml = ET.fromstring(zin.read("xl/workbook.xml"))
    rid = None
    for sheet in workbook_xml.iter(f"{_NS_MAIN}sheet"):
        if sheet.get("name") == nom_feuille:
            rid = sheet.get(f"{_NS_REL}id")
            break
    if rid is None:
        raise ValueError(f"Feuille « {nom_feuille} » introuvable dans xl/workbook.xml")

    rels_xml = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
    for rel in rels_xml.iter(f"{_NS_PKGREL}Relationship"):
        if rel.get("Id") == rid:
            cible = rel.get("Target")
            # Target est soit relatif à xl/ ("worksheets/sheet1.xml", forme
            # écrite par Excel), soit relatif à la racine du package
            # ("/xl/worksheets/sheet1.xml", forme écrite par openpyxl).
            return cible.lstrip("/") if cible.startswith("/") else "xl/" + cible
    raise ValueError(f"Relation « {rid} » introuvable dans xl/_rels/workbook.xml.rels")


def _verifier_cellules_vides(xml_feuille: str, num_ligne: int, refs) -> None:
    """Lève ValueError si l'une de `refs` (ex. "AY1") existe déjà comme
    cellule dans la ligne `num_ligne` du XML — jamais d'écrasement d'une
    cellule déjà présente, même vide de valeur."""

    motif_ligne = re.compile(rf'<row r="{num_ligne}"[^>]*>(.*?)</row>', re.S)
    m = motif_ligne.search(xml_feuille)
    contenu = m.group(1) if m else ""
    for ref in refs:
        if re.search(rf'<c r="{ref}"', contenu):
            raise ValueError(
                f"La cellule {ref} existe déjà dans la feuille — écriture "
                "refusée (rien n'a été écrit)."
            )


def _etendre_plage(ref: str, colonne_min_requise: int) -> str:
    """"A1:AX6420" -> "A1:BC6420" si `colonne_min_requise` (1-based) dépasse
    la colonne de fin actuelle — ne réduit jamais, ne touche jamais les
    lignes ni la colonne de début. Partagé par <dimension>, <table ref=...>
    et <autoFilter ref=...> (même forme de plage partout)."""

    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)$", ref)
    if not m:
        raise ValueError(f"Plage de cellules de forme inattendue : {ref!r}")
    debut_col, debut_ligne, fin_col, fin_ligne = m.groups()
    nouvelle_fin_col = get_column_letter(
        max(column_index_from_string(fin_col), colonne_min_requise)
    )
    return f"{debut_col}{debut_ligne}:{nouvelle_fin_col}{fin_ligne}"


def _dimension_etendue(xml_feuille: str, colonne_min_requise: int) -> str:
    """Nouvelle valeur de <dimension ref="..."/> couvrant au moins la
    colonne `colonne_min_requise` (1-based), même ligne de fin qu'avant."""

    m = re.search(r'<dimension ref="([^"]*)"\s*/>', xml_feuille)
    if not m:
        raise ValueError("<dimension> introuvable ou de forme inattendue dans la feuille.")
    return _etendre_plage(m.group(1), colonne_min_requise)


def _chemin_table_par_nom(zin: zipfile.ZipFile, nom_table: str) -> str:
    """xl/tables/tableN.xml dont l'attribut du <table> racine name="..."
    vaut `nom_table` — recherche directe dans xl/tables/ (pas besoin de
    suivre feuille -> _rels -> table, un seul tableau porte ce nom dans
    tout le classeur)."""

    for nom in zin.namelist():
        if re.match(r"xl/tables/table\d+\.xml$", nom):
            data = zin.read(nom).decode("utf-8")
            if re.search(rf'<table\b[^>]*\bname="{re.escape(nom_table)}"', data):
                return nom
    raise ValueError(f"Tableau structuré « {nom_table} » introuvable dans xl/tables/")


def _verifier_colonnes_table_absentes(xml_table: str, noms) -> None:
    """Lève ValueError si l'un de `noms` est déjà le name= d'un
    <tableColumn> existant — vérification indépendante de lire_entetes()
    (qui lit la ligne 1 de la FEUILLE) pour ne jamais dépendre d'une
    hypothèse de synchronisation entre la feuille et le tableau."""

    m = re.search(r"<tableColumns\b[^>]*>(.*)</tableColumns>", xml_table, re.S)
    contenu = m.group(1) if m else ""
    existants = set(re.findall(r'<tableColumn\b[^>]*\bname="([^"]*)"', contenu))
    for nom in noms:
        if nom in existants:
            raise ValueError(
                f"« {nom} » existe déjà comme colonne du tableau structuré — "
                "rien n'a été écrit."
            )


def _etendre_tableau(xml_table: str, noms, colonne_min_requise: int) -> str:
    """Étend la définition XML d'un tableau structuré (xl/tables/tableN.xml)
    pour y ajouter `noms` comme nouvelles colonnes, à la fin (l'ORDRE des
    <tableColumn> dans le XML est l'ordre gauche->droite affiché par Excel
    — vérifié sur le vrai classeur, où l'attribut id="" n'est PAS
    séquentiel par position) :

    - `ref` du <table> et, s'il existe, du <autoFilter> : étendus pour
      couvrir la nouvelle dernière colonne (jamais les lignes) ;
    - <tableColumns count="..."> : compte mis à jour ;
    - un <tableColumn id="…" name="…"/> par nom — SEULEMENT id et name
      (jamais de calculatedColumnFormula : ces colonnes sont de la saisie
      pure, pas un calcul hérité par une future ligne du tableau) — id =
      max(id existants) + 1, +2, ... pour rester unique."""

    m_table = re.search(r'(<table\b[^>]*\sref=")([^"]*)(")', xml_table)
    if not m_table:
        raise ValueError('<table ref="..."> introuvable ou de forme inattendue.')
    nouvelle_ref = _etendre_plage(m_table.group(2), colonne_min_requise)
    xml_table = (
        xml_table[: m_table.start()]
        + m_table.group(1) + nouvelle_ref + m_table.group(3)
        + xml_table[m_table.end() :]
    )

    m_filtre = re.search(r'(<autoFilter\b[^>]*\sref=")([^"]*)(")', xml_table)
    if m_filtre:
        nouvelle_ref_filtre = _etendre_plage(m_filtre.group(2), colonne_min_requise)
        xml_table = (
            xml_table[: m_filtre.start()]
            + m_filtre.group(1) + nouvelle_ref_filtre + m_filtre.group(3)
            + xml_table[m_filtre.end() :]
        )

    m_cols = re.search(r'<tableColumns\b[^>]*\bcount="(\d+)"', xml_table)
    if not m_cols:
        raise ValueError('<tableColumns count="..."> introuvable ou de forme inattendue.')
    nouveau_compte = int(m_cols.group(1)) + len(noms)
    xml_table = xml_table[: m_cols.start(1)] + str(nouveau_compte) + xml_table[m_cols.end(1) :]

    ids_existants = [int(i) for i in re.findall(r'<tableColumn\b[^>]*\bid="(\d+)"', xml_table)]
    id_suivant = max(ids_existants, default=0) + 1

    fragments = []
    for i, nom in enumerate(noms):
        nom_xml = nom
        for a, b in (("&", "&amp;"), ("<", "&lt;"), (">", "&gt;"), ('"', "&quot;")):
            nom_xml = nom_xml.replace(a, b)
        fragments.append(f'<tableColumn id="{id_suivant + i}" name="{nom_xml}"/>')

    position = xml_table.rindex("</tableColumns>")
    return xml_table[:position] + "".join(fragments) + xml_table[position:]


def _contenu_cellule(valeur):
    """(xml_interne, attribut_t) pour la valeur Python donnée. Jamais de
    sharedStrings.xml pour du texte : chaîne "inline" (t="inlineStr"), pour
    ne pas toucher une table partagée par des milliers d'autres cellules."""

    if isinstance(valeur, bool):
        return (f"<v>{1 if valeur else 0}</v>", ' t="b"')
    if isinstance(valeur, (int, float)):
        texte = repr(float(valeur)) if isinstance(valeur, float) else str(valeur)
        return (f"<v>{texte}</v>", "")
    if isinstance(valeur, (datetime, date)):
        return (f"<v>{repr(float(_date_vers_excel(valeur)))}</v>", "")

    texte = str(valeur)
    for a, b in (("&", "&amp;"), ("<", "&lt;"), (">", "&gt;")):
        texte = texte.replace(a, b)
    return (f'<is><t xml:space="preserve">{texte}</t></is>', ' t="inlineStr"')


def _ecrire_cellule(contenu_ligne: str, ref: str, valeur) -> str:
    """Remplace (ou insère, triée par colonne) la cellule `ref` dans le
    contenu XML d'une <row>...</row> (sans les balises <row> elles-mêmes)."""

    contenu_valeur, attr_t = _contenu_cellule(valeur)
    motif_cellule = re.compile(rf'<c r="{ref}"([^>]*?)(?:/>|>.*?</c>)', re.S)
    m = motif_cellule.search(contenu_ligne)
    if m:
        attrs = re.sub(r'\s+t="[^"]*"', "", m.group(1))
        nouvelle_cellule = f'<c r="{ref}"{attrs}{attr_t}>{contenu_valeur}</c>'
        return contenu_ligne[: m.start()] + nouvelle_cellule + contenu_ligne[m.end() :]

    # Cellule absente du XML (jamais écrite depuis la création de la ligne) :
    # insertion triée par colonne — un ordre non croissant fait "réparer" le
    # fichier par Excel à l'ouverture.
    nouvelle_cellule = f'<c r="{ref}"{attr_t}>{contenu_valeur}</c>'
    col_cible = column_index_from_string(re.match(r"[A-Z]+", ref).group())
    position = len(contenu_ligne)
    for m2 in re.finditer(r'<c r="([A-Z]+)\d+"', contenu_ligne):
        if column_index_from_string(m2.group(1)) > col_cible:
            position = m2.start()
            break
    return contenu_ligne[:position] + nouvelle_cellule + contenu_ligne[position:]


def _remplacer_dans_ligne(xml_feuille: str, num_ligne: int, remplacements: dict) -> str:
    motif_ligne = re.compile(rf'(<row r="{num_ligne}"[^>]*>)(.*?)(</row>)', re.S)
    m = motif_ligne.search(xml_feuille)
    if not m:
        raise ValueError(
            f"Ligne {num_ligne} introuvable dans la feuille (ligne vide ou "
            "hors de la plage de données ?)"
        )
    ouverture, contenu, fermeture = m.group(1), m.group(2), m.group(3)

    for ref, valeur in remplacements.items():
        contenu = _ecrire_cellule(contenu, ref, valeur)

    return xml_feuille[: m.start()] + ouverture + contenu + fermeture + xml_feuille[m.end() :]


def _patcher_parties_xlsx(fichier, parties: dict) -> None:
    """Réécrit `fichier` en place : chaque partie nommée dans `parties`
    (chemin dans le zip -> nouveau contenu XML str) est remplacée en une
    seule passe atomique ; TOUTES les autres parties sont recopiées octet
    pour octet depuis l'original. Utilisé par ajouter_entetes_saisie(), qui
    doit patcher DEUX parties (la feuille et la définition du tableau) sans
    laisser le zip dans un état intermédiaire — _patcher_cellules_xlsx()
    reste inchangée, dédiée à appliquer() (une seule partie, la feuille).

    Depuis P1 (socles feuille Pièces, voir plus bas) : une valeur None
    SUPPRIME la partie du zip (cas de xl/calcChain.xml, voir
    _supprimer_calc_chain) ; un chemin absent de l'original est AJOUTÉ
    comme nouvelle partie (nouvelle feuille, nouveau tableau, nouvelle
    relation), en fin de zip."""

    fichier = Path(fichier)
    fichier_tmp = fichier.with_suffix(fichier.suffix + ".tmp")

    with zipfile.ZipFile(fichier, "r") as zin:
        try:
            with zipfile.ZipFile(fichier_tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                existants = set()
                for item in zin.infolist():
                    existants.add(item.filename)
                    if item.filename in parties:
                        if parties[item.filename] is None:
                            continue  # partie supprimée
                        donnees = parties[item.filename].encode("utf-8")
                    else:
                        donnees = zin.read(item.filename)
                    zout.writestr(item, donnees)
                for chemin, contenu in parties.items():
                    if chemin in existants or contenu is None:
                        continue
                    info = zipfile.ZipInfo(chemin, date_time=datetime.now().timetuple()[:6])
                    info.compress_type = zipfile.ZIP_DEFLATED
                    zout.writestr(info, contenu.encode("utf-8"))
        except Exception:
            fichier_tmp.unlink(missing_ok=True)
            raise

    fichier_tmp.replace(fichier)


def _patcher_cellules_xlsx(fichier, feuille: str, cellules_par_ref: dict):
    """Réécrit `fichier` en place : seule la partie XML de `feuille` est
    modifiée dans le zip, toutes les autres parties (styles, tableaux,
    validations, calcChain, sharedStrings, customXml, printerSettings...)
    sont recopiées octet pour octet depuis l'original."""

    fichier = Path(fichier)
    fichier_tmp = fichier.with_suffix(fichier.suffix + ".tmp")

    with zipfile.ZipFile(fichier, "r") as zin:
        chemin_feuille = _chemin_feuille(zin, feuille)
        xml_feuille = zin.read(chemin_feuille).decode("utf-8")

        cellules_par_ligne = {}
        for ref, valeur in cellules_par_ref.items():
            num_ligne = int(re.match(r"[A-Z]+(\d+)", ref).group(1))
            cellules_par_ligne.setdefault(num_ligne, {})[ref] = valeur

        for num_ligne, remplacements in cellules_par_ligne.items():
            xml_feuille = _remplacer_dans_ligne(xml_feuille, num_ligne, remplacements)

        try:
            with zipfile.ZipFile(fichier_tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    donnees = (
                        xml_feuille.encode("utf-8")
                        if item.filename == chemin_feuille
                        else zin.read(item.filename)
                    )
                    zout.writestr(item, donnees)
        except Exception:
            fichier_tmp.unlink(missing_ok=True)
            raise

    fichier_tmp.replace(fichier)


# ===========================================================================
# Socles génériques P1 — feuille + tableau, lignes, colonnes calculées
# (voir CLAUDE.md, « Feuille Pièces — modèle, socles, migration »)
#
# Trois socles, tous par patch XML du zip (jamais openpyxl.save(), voir
# bandeau), tous paramétrés par le chemin du classeur pour resservir tels
# quels à M1 (registre BdC manuels) et T1 (Tableau1 du .xlsm) :
#   1. ajouter_feuille_tableau()   : nouvelle feuille + nouveau tableau
#      structuré (en-têtes seuls) ;
#   2. ajouter_lignes_tableau()    : append de lignes en fin de tableau,
#      idempotent par colonne-clé ;
#   3. ajouter_colonnes_calculees() / basculer_colonnes_en_formules() :
#      colonnes calculées (calculatedColumnFormula dans la définition du
#      tableau + <f> sur chaque ligne du tableau).
# Plus reparer_formules_ligne() (cellules explicitement listées d'une
# ligne, recopie de la formule d'une ligne modèle).
#
# Formules : toujours la chaîne ANGLAISE telle qu'Excel la stocke dans le
# XML (préfixes _xlfn./_xlws. compris, ex. "_xlfn.MAXIFS(", "_xlfn._xlws.
# FILTER("), sans "=" de tête, références structurées sous la forme
# longue "Table[[#This Row],[Colonne]]" (la forme "[@Colonne]" n'existe
# qu'à l'écran). La même chaîne est écrite sur TOUTES les lignes (jamais la
# ligne 2 comme modèle, voir CLAUDE.md).
#
# calcChain : jamais complété à la main — supprimé proprement (partie +
# relation + Override de [Content_Types]) dès qu'une formule est ajoutée,
# et fullCalcOnLoad="1" posé sur <calcPr> : Excel recalcule tout et
# régénère la chaîne à l'ouverture (l'attribut disparaît à sa prochaine
# sauvegarde).
# ===========================================================================


@dataclass(frozen=True)
class Formule:
    """Une formule à écrire dans une cellule : `texte` = formule anglaise
    telle que stockée par Excel (sans "="), `cache` = valeur affichée avant
    le 1er recalcul (facultatif — ex. le libellé d'un HYPERLINK), `array` =
    formule matricielle (t="array", nécessaire dès qu'une fonction à
    résultat tableau — FILTER/UNIQUE — est imbriquée)."""

    texte: str
    cache: object = None
    array: bool = False


# Styles « logiques » -> numFmtId Excel (formats intégrés : 14 = date
# courte, 22 = date+heure, 44 = comptabilité € ; 164 = format personnalisé
# du Suivi "#,##0.0000 €" s'il est déclaré dans styles.xml, sinon repli
# sur 44). Résolus en index de <cellXfs> par _resoudre_styles(), qui
# ajoute l'<xf> manquant en fin de <cellXfs> (jamais de réindexation des
# styles existants).
_STYLES_LOGIQUES = {"date": 14, "datetime": 22, "monnaie": 44, "monnaie4": 164}

_CT_FEUILLE = "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
_CT_TABLE = "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"
_REL_FEUILLE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
_REL_TABLE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/table"
_REL_CALCCHAIN = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain"


def _echapper_xml(texte, attribut=False) -> str:
    texte = str(texte)
    for a, b in (("&", "&amp;"), ("<", "&lt;"), (">", "&gt;")):
        texte = texte.replace(a, b)
    if attribut:
        texte = texte.replace('"', "&quot;")
    return texte


def _desechapper_xml(texte) -> str:
    for a, b in (("&lt;", "<"), ("&gt;", ">"), ("&quot;", '"'), ("&apos;", "'"), ("&amp;", "&")):
        texte = texte.replace(a, b)
    return texte


def _verifier_verrou(fichier):
    if est_verrouille(fichier):
        raise ClasseurVerrouille(
            f"« {Path(fichier).name} » est actuellement ouvert dans Excel — "
            "ferme-le puis réessaie (rien n'a été écrit)."
        )


# --- Styles --------------------------------------------------------------


def _resoudre_styles(xml_styles: str, noms_logiques) -> tuple:
    """(xml_styles éventuellement complété, {nom logique: index cellXfs}).
    Réutilise le 1er <xf> « nu » (font/fill/border 0) portant le numFmtId
    voulu ; sinon en ajoute un en fin de <cellXfs>."""

    m = re.search(r'<cellXfs count="(\d+)">(.*?)</cellXfs>', xml_styles, re.S)
    if not m:
        raise ValueError("<cellXfs> introuvable ou de forme inattendue dans styles.xml.")
    xfs = re.findall(r"<xf\b[^>]*?(?:/>|>.*?</xf>)", m.group(2), re.S)
    numfmts_declares = set(re.findall(r'<numFmt\b[^>]*\bnumFmtId="(\d+)"', xml_styles))

    resultat = {}
    ajouts = []
    for nom in noms_logiques:
        numfmt = _STYLES_LOGIQUES[nom]
        if numfmt >= 164 and str(numfmt) not in numfmts_declares:
            numfmt = _STYLES_LOGIQUES["monnaie"]
        trouve = None
        for i, xf in enumerate(xfs):
            if (
                f'numFmtId="{numfmt}"' in xf
                and 'fontId="0"' in xf and 'fillId="0"' in xf and 'borderId="0"' in xf
            ):
                trouve = i
                break
        if trouve is None:
            ajouts.append(f'<xf numFmtId="{numfmt}" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>')
            trouve = len(xfs) + len(ajouts) - 1
        resultat[nom] = trouve

    if ajouts:
        nouveau = int(m.group(1)) + len(ajouts)
        xml_styles = (
            xml_styles[: m.start()]
            + f'<cellXfs count="{nouveau}">' + m.group(2) + "".join(ajouts) + "</cellXfs>"
            + xml_styles[m.end():]
        )
    return xml_styles, resultat


# --- Cellules / lignes -----------------------------------------------------


def _cellule_xml(ref: str, valeur, style=None, cm: bool = False) -> str:
    """XML complet d'une cellule <c> pour `valeur` (None -> cellule vide
    stylée, Formule -> <f> [+ <v> cache], sinon voir _contenu_cellule)."""

    attr_s = f' s="{style}"' if style is not None else ""
    if valeur is None:
        return f'<c r="{ref}"{attr_s}/>'
    if isinstance(valeur, Formule):
        f = _echapper_xml(valeur.texte)
        attr_f = f' t="array" ref="{ref}"' if valeur.array else ""
        attr_cm = ' cm="1"' if (valeur.array and cm) else ""
        if valeur.cache is None:
            return f'<c r="{ref}"{attr_s}{attr_cm}><f{attr_f}>{f}</f></c>'
        if isinstance(valeur.cache, bool):
            return f'<c r="{ref}"{attr_s}{attr_cm} t="b"><f{attr_f}>{f}</f><v>{1 if valeur.cache else 0}</v></c>'
        if isinstance(valeur.cache, (int, float)):
            return f'<c r="{ref}"{attr_s}{attr_cm}><f{attr_f}>{f}</f><v>{valeur.cache!r}</v></c>'
        return (
            f'<c r="{ref}"{attr_s}{attr_cm} t="str"><f{attr_f}>{f}</f>'
            f"<v>{_echapper_xml(str(valeur.cache))}</v></c>"
        )
    contenu, attr_t = _contenu_cellule(valeur)
    return f'<c r="{ref}"{attr_s}{attr_t}>{contenu}</c>'


def _poser_cellule(contenu_ligne: str, ref: str, fabrique) -> str:
    """Remplace la cellule `ref` (en lui passant son style existant, ou
    None) ou l'insère triée par colonne — `fabrique(style)` retourne le XML
    de la nouvelle cellule."""

    motif = re.compile(rf'<c r="{ref}"([^>]*?)(?:/>|>.*?</c>)', re.S)
    m = motif.search(contenu_ligne)
    if m:
        m_s = re.search(r'\ss="(\d+)"', m.group(1))
        style = int(m_s.group(1)) if m_s else None
        return contenu_ligne[: m.start()] + fabrique(style) + contenu_ligne[m.end():]

    col_cible = column_index_from_string(re.match(r"[A-Z]+", ref).group())
    position = len(contenu_ligne)
    for m2 in re.finditer(r'<c r="([A-Z]+)\d+"', contenu_ligne):
        if column_index_from_string(m2.group(1)) > col_cible:
            position = m2.start()
            break
    return contenu_ligne[:position] + fabrique(None) + contenu_ligne[position:]


def _ajuster_spans(ouverture: str, colonne_max: int) -> str:
    m = re.search(r'spans="(\d+):(\d+)"', ouverture)
    if m and int(m.group(2)) < colonne_max:
        return ouverture[: m.start()] + f'spans="{m.group(1)}:{colonne_max}"' + ouverture[m.end():]
    return ouverture


_MOTIF_ROW = re.compile(r'<row r="(\d+)"([^>]*?)(?:/>|>(.*?)</row>)', re.S)


def _reecrire_lignes(xml_feuille: str, transformer, lignes_a_creer=(), attributs_ligne: str = "") -> str:
    """UNE SEULE passe sur <sheetData> (la feuille Commandes pèse ~58 Mo,
    une recherche par ligne serait quadratique) : chaque <row> existante
    passe par `transformer(num, ouverture, contenu) -> (ouverture,
    contenu)` ; les numéros de `lignes_a_creer` absents du XML sont créés
    vides à leur place (ordre croissant) puis transformés eux aussi."""

    m_sd = re.search(r"<sheetData\b[^>]*?(?:/>|>(.*?)</sheetData>)", xml_feuille, re.S)
    if not m_sd:
        raise ValueError("<sheetData> introuvable dans la feuille.")
    interieur = m_sd.group(1) or ""

    a_creer = sorted(set(lignes_a_creer))
    sortie = []
    position = 0

    def _creer(num):
        ouverture = f'<row r="{num}"{attributs_ligne}>'
        ouverture, contenu = transformer(num, ouverture, "")
        sortie.append(ouverture + contenu + "</row>")

    for m in _MOTIF_ROW.finditer(interieur):
        num = int(m.group(1))
        while a_creer and a_creer[0] < num:
            _creer(a_creer.pop(0))
        if a_creer and a_creer[0] == num:
            a_creer.pop(0)
        sortie.append(interieur[position:m.start()])
        ouverture = f'<row r="{num}"{m.group(2)}>'
        contenu = m.group(3) or ""
        ouverture, contenu = transformer(num, ouverture, contenu)
        sortie.append(ouverture + contenu + "</row>")
        position = m.end()
    sortie.append(interieur[position:])
    while a_creer:
        _creer(a_creer.pop(0))

    return xml_feuille[: m_sd.start()] + "<sheetData>" + "".join(sortie) + "</sheetData>" + xml_feuille[m_sd.end():]


def _plage_avec_derniere_ligne(ref: str, derniere_ligne: int) -> str:
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)$", ref)
    if not m:
        raise ValueError(f"Plage de cellules de forme inattendue : {ref!r}")
    debut_col, debut_ligne, fin_col, fin_ligne = m.groups()
    return f"{debut_col}{debut_ligne}:{fin_col}{max(int(fin_ligne), derniere_ligne)}"


def _remplacer_attribut_ref(xml: str, balise: str, nouvelle_ref: str) -> str:
    m = re.search(rf'(<{balise}\b[^>]*\sref=")([^"]*)(")', xml)
    if not m:
        raise ValueError(f'<{balise} ref="..."> introuvable.')
    return xml[: m.start()] + m.group(1) + nouvelle_ref + m.group(3) + xml[m.end():]


def _remplacer_dimension(xml_feuille: str, nouvelle: str) -> str:
    if re.search(r'<dimension ref="[^"]*"\s*/>', xml_feuille):
        return re.sub(r'<dimension ref="[^"]*"\s*/>', f'<dimension ref="{nouvelle}"/>', xml_feuille, count=1)
    return xml_feuille


# --- Lecture de la structure --------------------------------------------


def _colonnes_table(xml_table: str) -> list:
    """Noms des <tableColumn> dans l'ordre du XML (= ordre affiché)."""
    m = re.search(r"<tableColumns\b[^>]*>(.*)</tableColumns>", xml_table, re.S)
    contenu = m.group(1) if m else ""
    return [_desechapper_xml(n) for n in re.findall(r'<tableColumn\b[^>]*\bname="([^"]*)"', contenu)]


def _ref_table(xml_table: str) -> str:
    m = re.search(r'<table\b[^>]*\sref="([^"]*)"', xml_table)
    if not m:
        raise ValueError('<table ref="..."> introuvable.')
    return m.group(1)


def _lettres_colonnes_table(xml_table: str) -> dict:
    """{nom de colonne: lettre de colonne} d'après la ref du tableau et
    l'ordre des <tableColumn>."""
    ref = _ref_table(xml_table)
    debut = column_index_from_string(re.match(r"([A-Z]+)", ref).group(1))
    return {nom: get_column_letter(debut + i) for i, nom in enumerate(_colonnes_table(xml_table))}


def _chemin_table_de_feuille(zin: zipfile.ZipFile, chemin_feuille: str, nom_table: str) -> str:
    """xl/tables/tableN.xml portant name=`nom_table`, en vérifiant qu'il est
    bien relié à `chemin_feuille` (sinon ValueError)."""
    chemin = _chemin_table_par_nom(zin, nom_table)
    rels = f"xl/worksheets/_rels/{Path(chemin_feuille).name}.rels"
    if rels in zin.namelist():
        contenu = zin.read(rels).decode("utf-8")
        if Path(chemin).name not in contenu:
            raise ValueError(f"Le tableau « {nom_table} » n'est pas relié à la feuille demandée.")
    return chemin


def _valeurs_colonne(zin: zipfile.ZipFile, xml_feuille: str, lettre: str, a_partir_de: int = 2) -> dict:
    """{numéro de ligne: valeur texte} des cellules de la colonne `lettre`
    — chaînes inline (écrites par l'outil) ET chaînes partagées (forme
    réécrite par Excel à sa prochaine sauvegarde), nombres bruts sinon."""

    motif = re.compile(
        rf'<c r="{lettre}(\d+)"([^>]*?)(?:/>|>(.*?)</c>)', re.S
    )
    partagees = None
    resultat = {}
    for m in motif.finditer(xml_feuille):
        num = int(m.group(1))
        if num < a_partir_de:
            continue
        attrs, corps = m.group(2), m.group(3) or ""
        m_t = re.search(r'\st="([^"]*)"', attrs)
        t = m_t.group(1) if m_t else None
        if t == "inlineStr":
            m_v = re.search(r"<t[^>]*>(.*?)</t>", corps, re.S)
            valeur = _desechapper_xml(m_v.group(1)) if m_v else ""
        elif t == "s":
            if partagees is None:
                partagees = _chaines_partagees(zin)
            m_v = re.search(r"<v>(\d+)</v>", corps)
            valeur = partagees[int(m_v.group(1))] if m_v else ""
        else:
            m_v = re.search(r"<v>(.*?)</v>", corps, re.S)
            valeur = _desechapper_xml(m_v.group(1)) if m_v else ""
        if valeur != "":
            resultat[num] = valeur
    return resultat


def _chaines_partagees(zin: zipfile.ZipFile) -> list:
    if "xl/sharedStrings.xml" not in zin.namelist():
        return []
    racine = ET.fromstring(zin.read("xl/sharedStrings.xml"))
    resultat = []
    for si in racine.iter(f"{_NS_MAIN}si"):
        resultat.append("".join(t.text or "" for t in si.iter(f"{_NS_MAIN}t")))
    return resultat


def _derniere_ligne_feuille(xml_feuille: str) -> int:
    nums = [int(n) for n in re.findall(r'<row r="(\d+)"', xml_feuille)]
    return max(nums, default=0)


def _a_metadata_dynamique(zin: zipfile.ZipFile) -> bool:
    """True si xl/metadata.xml déclare le type XLDAPR (tableaux dynamiques)
    — permet d'écrire cm="1" sur une cellule matricielle, comme Excel le
    fait lui-même dans le Suivi (ex. R6557)."""
    if "xl/metadata.xml" not in zin.namelist():
        return False
    return 'name="XLDAPR"' in zin.read("xl/metadata.xml").decode("utf-8")


# --- calcChain / recalcul --------------------------------------------------


def _supprimer_calc_chain(zin: zipfile.ZipFile, parties: dict) -> None:
    """Complète `parties` pour retirer proprement xl/calcChain.xml : la
    partie elle-même (None), sa relation dans xl/_rels/workbook.xml.rels et
    son Override dans [Content_Types].xml. Sans effet si absente."""

    if "xl/calcChain.xml" not in zin.namelist():
        return
    parties["xl/calcChain.xml"] = None

    rels = parties.get("xl/_rels/workbook.xml.rels") or zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rels = re.sub(r'<Relationship\b[^>]*Type="' + re.escape(_REL_CALCCHAIN) + r'"[^>]*/>', "", rels)
    parties["xl/_rels/workbook.xml.rels"] = rels

    ct = parties.get("[Content_Types].xml") or zin.read("[Content_Types].xml").decode("utf-8")
    ct = re.sub(r'<Override\b[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', "", ct)
    parties["[Content_Types].xml"] = ct


def _forcer_recalcul_complet(xml_workbook: str) -> str:
    """fullCalcOnLoad="1" sur <calcPr> (créé s'il manque) : Excel recalcule
    tout à l'ouverture et régénère calcChain."""

    m = re.search(r"<calcPr\b([^>]*?)/>", xml_workbook)
    if m:
        if "fullCalcOnLoad" in m.group(1):
            return xml_workbook
        return xml_workbook[: m.start()] + f'<calcPr{m.group(1)} fullCalcOnLoad="1"/>' + xml_workbook[m.end():]
    # Position légale : après <definedNames> (ou <sheets>), avant extLst.
    for balise in ("</definedNames>", "</sheets>"):
        pos = xml_workbook.find(balise)
        if pos >= 0:
            pos += len(balise)
            return xml_workbook[:pos] + '<calcPr fullCalcOnLoad="1"/>' + xml_workbook[pos:]
    raise ValueError("<sheets> introuvable dans xl/workbook.xml.")


# --- Socle 1 : feuille + tableau ------------------------------------------


def ajouter_feuille_tableau(fichier, nom_feuille: str, nom_table: str, colonnes, dossier_backups,
                            feuille_modele=FEUILLE_COMMANDES, table_modele=None, apres_feuille=None,
                            largeurs=None) -> Path:
    """Crée la feuille `nom_feuille` avec un tableau structuré `nom_table`
    dont les en-têtes sont `colonnes` (ligne 1, inlineStr, style d'en-tête
    copié de A1 de `feuille_modele`), ref A1:<fin>2 (une ligne de données
    vide : Excel exige au moins une ligne de données sous l'en-tête),
    autoFilter, tableStyleInfo copié du tableau `table_modele` (par défaut
    le tableau homonyme de `feuille_modele`). Onglet placé juste après
    `apres_feuille` (par défaut `feuille_modele`).

    Parties du zip touchées, en une seule passe atomique : nouvelle
    xl/worksheets/sheetN.xml, nouvelle xl/worksheets/_rels/sheetN.xml.rels,
    nouvelle xl/tables/tableN.xml, xl/workbook.xml (<sheet>, et
    localSheetId/activeTab décalés si besoin), xl/_rels/workbook.xml.rels,
    [Content_Types].xml. Tout le reste est recopié octet pour octet.
    Verrou Excel et sauvegarde comme appliquer(). ValueError si la feuille
    ou le tableau existe déjà. Retourne le chemin de la sauvegarde."""

    fichier = Path(fichier)
    colonnes = list(colonnes)
    if not colonnes or len(set(colonnes)) != len(colonnes):
        raise ValueError("Colonnes vides ou en double.")
    if not re.match(r"^[A-Za-z_][A-Za-z0-9_.]*$", nom_table):
        raise ValueError(f"Nom de tableau invalide pour Excel : {nom_table!r}")
    _verifier_verrou(fichier)
    table_modele = table_modele or feuille_modele
    apres_feuille = apres_feuille or feuille_modele

    with zipfile.ZipFile(fichier, "r") as zin:
        noms = zin.namelist()
        xml_workbook = zin.read("xl/workbook.xml").decode("utf-8")
        xml_rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        xml_ct = zin.read("[Content_Types].xml").decode("utf-8")

        feuilles = re.findall(r'<sheet\b[^>]*\bname="([^"]*)"', xml_workbook)
        feuilles = [_desechapper_xml(f) for f in feuilles]
        if nom_feuille in feuilles:
            raise ValueError(f"La feuille « {nom_feuille} » existe déjà — rien n'a été écrit.")
        if apres_feuille not in feuilles:
            raise ValueError(f"Feuille « {apres_feuille} » introuvable.")
        for nom in noms:
            if re.match(r"xl/tables/table\d+\.xml$", nom):
                data = zin.read(nom).decode("utf-8")
                if re.search(rf'<table\b[^>]*\b(?:name|displayName)="{re.escape(nom_table)}"', data):
                    raise ValueError(f"Le tableau « {nom_table} » existe déjà — rien n'a été écrit.")

        # Style d'en-tête (A1 de la feuille modèle) et style du tableau modèle.
        chemin_modele = _chemin_feuille(zin, feuille_modele)
        xml_modele = zin.read(chemin_modele).decode("utf-8")
        m_a1 = re.search(r'<c r="A1"\s+s="(\d+)"', xml_modele)
        style_entete = int(m_a1.group(1)) if m_a1 else None
        table_style = '<tableStyleInfo name="TableStyleMedium2" showFirstColumn="0" showLastColumn="0" showRowStripes="1" showColumnStripes="0"/>'
        try:
            xml_table_modele = zin.read(_chemin_table_par_nom(zin, table_modele)).decode("utf-8")
        except ValueError:
            xml_table_modele = ""  # pas de tableau modèle (classeur minimal) : style par défaut
        m_style = re.search(r"<tableStyleInfo[^>]*/>", xml_table_modele)
        if m_style:
            table_style = m_style.group(0)

        # Numéros libres.
        num_feuilles = [int(n) for n in re.findall(r"xl/worksheets/sheet(\d+)\.xml", " ".join(noms))]
        num_tables = [int(n) for n in re.findall(r"xl/tables/table(\d+)\.xml", " ".join(noms))]
        ids_tables = []
        for nom in noms:
            if re.match(r"xl/tables/table\d+\.xml$", nom):
                m_id = re.search(r'<table\b[^>]*\sid="(\d+)"', zin.read(nom).decode("utf-8"))
                if m_id:
                    ids_tables.append(int(m_id.group(1)))
        n_feuille = max(num_feuilles, default=0) + 1
        n_table = max(num_tables, default=0) + 1
        id_table = max(ids_tables, default=0) + 1
        sheet_ids = [int(n) for n in re.findall(r'<sheet\b[^>]*\bsheetId="(\d+)"', xml_workbook)]
        sheet_id = max(sheet_ids, default=0) + 1
        rids = [int(n) for n in re.findall(r'Id="rId(\d+)"', xml_rels)]
        rid = f"rId{max(rids, default=0) + 1}"

    chemin_feuille = f"xl/worksheets/sheet{n_feuille}.xml"
    chemin_rels_feuille = f"xl/worksheets/_rels/sheet{n_feuille}.xml.rels"
    chemin_table = f"xl/tables/table{n_table}.xml"
    fin = get_column_letter(len(colonnes))
    ref = f"A1:{fin}2"

    # --- feuille ---
    cellules = "".join(
        _cellule_xml(f"{get_column_letter(i + 1)}1", nom, style_entete) for i, nom in enumerate(colonnes)
    )
    cols = ""
    if largeurs:
        fragments = []
        for i, nom in enumerate(colonnes):
            if nom in largeurs:
                fragments.append(f'<col min="{i + 1}" max="{i + 1}" width="{largeurs[nom]}" customWidth="1"/>')
        if fragments:
            cols = "<cols>" + "".join(fragments) + "</cols>"
    xml_feuille = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<dimension ref="{ref}"/>'
        '<sheetViews><sheetView workbookViewId="0">'
        '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        '<selection pane="bottomLeft" activeCell="A2" sqref="A2"/>'
        "</sheetView></sheetViews>"
        '<sheetFormatPr baseColWidth="10" defaultRowHeight="15"/>'
        f"{cols}"
        f'<sheetData><row r="1" spans="1:{len(colonnes)}">{cellules}</row></sheetData>'
        f'<tableParts count="1"><tablePart r:id="rId1"/></tableParts>'
        "</worksheet>"
    )
    xml_rels_feuille = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'<Relationship Id="rId1" Type="{_REL_TABLE}" Target="../tables/table{n_table}.xml"/>'
        "</Relationships>"
    )
    # --- tableau ---
    colonnes_xml = "".join(
        f'<tableColumn id="{i + 1}" name="{_echapper_xml(nom, attribut=True)}"/>' for i, nom in enumerate(colonnes)
    )
    xml_table = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<table xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'id="{id_table}" name="{nom_table}" displayName="{nom_table}" ref="{ref}" totalsRowShown="0">'
        f'<autoFilter ref="{ref}"/>'
        f'<tableColumns count="{len(colonnes)}">{colonnes_xml}</tableColumns>'
        f"{table_style}</table>"
    )
    # --- workbook.xml : <sheet> juste après `apres_feuille`, indices décalés ---
    m_apres = re.search(
        rf'<sheet\b[^>]*\bname="{re.escape(_echapper_xml(apres_feuille, attribut=True))}"[^>]*/>', xml_workbook
    )
    if not m_apres:
        raise ValueError(f"Feuille « {apres_feuille} » introuvable dans xl/workbook.xml.")
    position_onglet = feuilles.index(apres_feuille) + 1  # index 0-based du nouvel onglet
    nouvelle_balise = f'<sheet name="{_echapper_xml(nom_feuille, attribut=True)}" sheetId="{sheet_id}" r:id="{rid}"/>'
    xml_workbook = xml_workbook[: m_apres.end()] + nouvelle_balise + xml_workbook[m_apres.end():]

    def _decaler(m):
        idx = int(m.group(2))
        return f'{m.group(1)}="{idx + 1 if idx >= position_onglet else idx}"'

    xml_workbook = re.sub(r'\b(localSheetId|activeTab|firstSheet)="(\d+)"', _decaler, xml_workbook)

    xml_rels = xml_rels.replace(
        "</Relationships>",
        f'<Relationship Id="{rid}" Type="{_REL_FEUILLE}" Target="worksheets/sheet{n_feuille}.xml"/></Relationships>',
    )
    xml_ct = xml_ct.replace(
        "</Types>",
        f'<Override PartName="/{chemin_feuille}" ContentType="{_CT_FEUILLE}"/>'
        f'<Override PartName="/{chemin_table}" ContentType="{_CT_TABLE}"/></Types>',
    )

    sauvegarde = sauvegarder(fichier, dossier_backups)
    _patcher_parties_xlsx(fichier, {
        chemin_feuille: xml_feuille,
        chemin_rels_feuille: xml_rels_feuille,
        chemin_table: xml_table,
        "xl/workbook.xml": xml_workbook,
        "xl/_rels/workbook.xml.rels": xml_rels,
        "[Content_Types].xml": xml_ct,
    })
    return sauvegarde


# --- Socle 2 : lignes en fin de tableau ------------------------------------


def ajouter_lignes_tableau(fichier, nom_feuille: str, nom_table: str, lignes, dossier_backups,
                           colonne_id=None, styles_colonnes=None) -> dict:
    """Ajoute `lignes` (liste de dict {nom de colonne: valeur}) en fin du
    tableau `nom_table` de `nom_feuille` : <row> appendues après la
    dernière ligne de <sheetData>, textes en inlineStr (sharedStrings
    jamais touché), nombres en <v>, dates en nombre de série (style date
    ou date+heure, voir _STYLES_LOGIQUES), Formule en <f> (+ <v> cache).
    <dimension>, ref du tableau et de son autoFilter étendus.

    `styles_colonnes` : {nom de colonne: style logique} ("date",
    "datetime", "monnaie", "monnaie4") — les dates/datetimes reçoivent leur
    style automatiquement même sans entrée. Idempotence par `colonne_id` :
    une ligne dont la valeur de cette colonne existe déjà dans la feuille
    (ou en double dans le lot) n'est JAMAIS réécrite — retournée dans
    "ignorees". Retourne {"ajoutees", "ignorees", "premiere_ligne",
    "derniere_ligne", "sauvegarde"} ; sans rien écrire si aucune ligne à
    ajouter (sauvegarde None)."""

    fichier = Path(fichier)
    _verifier_verrou(fichier)
    styles_colonnes = dict(styles_colonnes or {})
    lignes = list(lignes)

    with zipfile.ZipFile(fichier, "r") as zin:
        chemin_feuille = _chemin_feuille(zin, nom_feuille)
        xml_feuille = zin.read(chemin_feuille).decode("utf-8")
        chemin_table = _chemin_table_de_feuille(zin, chemin_feuille, nom_table)
        xml_table = zin.read(chemin_table).decode("utf-8")
        xml_styles = zin.read("xl/styles.xml").decode("utf-8")
        cm = _a_metadata_dynamique(zin)

        lettres = _lettres_colonnes_table(xml_table)
        for i, ligne in enumerate(lignes):
            inconnues = set(ligne) - set(lettres)
            if inconnues:
                raise ValueError(f"Ligne {i}: colonne(s) inconnue(s) du tableau « {nom_table} » : {sorted(inconnues)}")
        if colonne_id is not None and colonne_id not in lettres:
            raise ValueError(f"Colonne-clé « {colonne_id} » absente du tableau « {nom_table} ».")

        ignorees = []
        if colonne_id is not None:
            existants = set(_valeurs_colonne(zin, xml_feuille, lettres[colonne_id]).values())
            retenues = []
            for ligne in lignes:
                cle = str(ligne.get(colonne_id, "") or "")
                if not cle:
                    raise ValueError(f"Ligne sans valeur pour la colonne-clé « {colonne_id} » : {ligne!r}")
                if cle in existants:
                    ignorees.append(cle)
                    continue
                existants.add(cle)
                retenues.append(ligne)
            lignes = retenues

    resultat = {"ajoutees": 0, "ignorees": ignorees, "premiere_ligne": None, "derniere_ligne": None, "sauvegarde": None}
    if not lignes:
        return resultat

    # Styles logiques nécessaires (déclarés + dates détectées).
    for ligne in lignes:
        for nom, valeur in ligne.items():
            if nom not in styles_colonnes:
                if isinstance(valeur, datetime):
                    styles_colonnes[nom] = "datetime"
                elif isinstance(valeur, date):
                    styles_colonnes[nom] = "date"
    for nom, style in styles_colonnes.items():
        if style not in _STYLES_LOGIQUES:
            raise ValueError(f"Style logique inconnu : {style!r} (colonne « {nom} »)")
    xml_styles_patch, index_styles = _resoudre_styles(xml_styles, sorted(set(styles_colonnes.values())))

    premiere = max(_derniere_ligne_feuille(xml_feuille), 1) + 1
    nb_colonnes = len(lettres)
    ordre = list(lettres)  # noms dans l'ordre du tableau

    rows = []
    for k, ligne in enumerate(lignes):
        num = premiere + k
        cellules = []
        for nom in ordre:
            if nom not in ligne or ligne[nom] is None:
                continue
            style = index_styles.get(styles_colonnes.get(nom)) if nom in styles_colonnes else None
            cellules.append(_cellule_xml(f"{lettres[nom]}{num}", ligne[nom], style, cm))
        rows.append(f'<row r="{num}" spans="1:{nb_colonnes}">{"".join(cellules)}</row>')
    derniere = premiere + len(lignes) - 1

    if "</sheetData>" in xml_feuille:
        xml_feuille = xml_feuille.replace("</sheetData>", "".join(rows) + "</sheetData>", 1)
    else:
        xml_feuille = re.sub(r"<sheetData\s*/>", "<sheetData>" + "".join(rows) + "</sheetData>", xml_feuille, count=1)

    ref_table = _plage_avec_derniere_ligne(_ref_table(xml_table), derniere)
    xml_table = _remplacer_attribut_ref(xml_table, "table", ref_table)
    if re.search(r"<autoFilter\b", xml_table):
        xml_table = _remplacer_attribut_ref(xml_table, "autoFilter", ref_table)
    m_dim = re.search(r'<dimension ref="([^"]*)"\s*/>', xml_feuille)
    if m_dim:
        dim = m_dim.group(1)
        if ":" not in dim:
            dim = f"{dim}:{dim}"
        xml_feuille = _remplacer_dimension(xml_feuille, _plage_avec_derniere_ligne(dim, derniere))

    parties = {chemin_feuille: xml_feuille, chemin_table: xml_table}
    if xml_styles_patch != xml_styles:
        parties["xl/styles.xml"] = xml_styles_patch

    resultat["sauvegarde"] = sauvegarder(fichier, dossier_backups)
    _patcher_parties_xlsx(fichier, parties)
    resultat.update(ajoutees=len(lignes), premiere_ligne=premiere, derniere_ligne=derniere)
    return resultat


# --- Socle 3 : colonnes calculées -------------------------------------------


def _en_formule(valeur) -> Formule:
    return valeur if isinstance(valeur, Formule) else Formule(str(valeur))


def _poser_calculated_column(xml_table: str, nom: str, formule: Formule) -> str:
    """Ajoute (ou remplace) <calculatedColumnFormula> dans le <tableColumn>
    nommé `nom`."""

    nom_attr = re.escape(_echapper_xml(nom, attribut=True))
    m = re.search(rf'<tableColumn\b([^>]*\bname="{nom_attr}"[^>]*?)(/>|>(.*?)</tableColumn>)', xml_table, re.S)
    if not m:
        raise ValueError(f"Colonne « {nom} » introuvable dans la définition du tableau.")
    attr_array = ' array="1"' if formule.array else ""
    ccf = f"<calculatedColumnFormula{attr_array}>{_echapper_xml(formule.texte)}</calculatedColumnFormula>"
    interieur = m.group(3) or ""
    interieur = re.sub(r"<calculatedColumnFormula\b[^>]*>.*?</calculatedColumnFormula>", "", interieur, flags=re.S)
    nouveau = f"<tableColumn{m.group(1)}>{ccf}{interieur}</tableColumn>"
    return xml_table[: m.start()] + nouveau + xml_table[m.end():]


def _appliquer_formules_colonnes(fichier, feuille, table, formules: dict, dossier_backups,
                                 creer: bool, styles_colonnes=None) -> Path:
    fichier = Path(fichier)
    if not formules:
        raise ValueError("Aucune formule fournie.")
    _verifier_verrou(fichier)
    formules = {nom: _en_formule(f) for nom, f in formules.items()}
    styles_colonnes = dict(styles_colonnes or {})

    with zipfile.ZipFile(fichier, "r") as zin:
        chemin_feuille = _chemin_feuille(zin, feuille)
        xml_feuille = zin.read(chemin_feuille).decode("utf-8")
        chemin_table = _chemin_table_de_feuille(zin, chemin_feuille, table)
        xml_table = zin.read(chemin_table).decode("utf-8")
        xml_styles = zin.read("xl/styles.xml").decode("utf-8")
        xml_workbook = zin.read("xl/workbook.xml").decode("utf-8")
        cm = _a_metadata_dynamique(zin)
        parties = {}
        _supprimer_calc_chain(zin, parties)

    colonnes_existantes = _colonnes_table(xml_table)
    entetes_feuille = lire_entetes(fichier, feuille)
    noms = list(formules)

    if creer:
        for nom in noms:
            if nom in colonnes_existantes or nom in entetes_feuille:
                raise ValueError(f"« {nom} » existe déjà (feuille ou tableau) — rien n'a été écrit.")
        derniere_colonne = max(entetes_feuille.values()) if entetes_feuille else 0
        refs_entetes = [f"{get_column_letter(derniere_colonne + 1 + i)}1" for i in range(len(noms))]
        _verifier_cellules_vides(xml_feuille, 1, refs_entetes)
        colonne_min_requise = derniere_colonne + len(noms)
        xml_table = _etendre_tableau(xml_table, noms, colonne_min_requise)
        xml_feuille = _remplacer_dimension(xml_feuille, _dimension_etendue(xml_feuille, colonne_min_requise))
        xml_feuille = _remplacer_dans_ligne(xml_feuille, 1, dict(zip(refs_entetes, noms)))
    else:
        for nom in noms:
            if nom not in colonnes_existantes:
                raise ValueError(f"« {nom} » n'est pas une colonne du tableau « {table} » — rien n'a été écrit.")

    lettres = _lettres_colonnes_table(xml_table)
    for nom, formule in formules.items():
        xml_table = _poser_calculated_column(xml_table, nom, formule)

    for nom, style in styles_colonnes.items():
        if style not in _STYLES_LOGIQUES:
            raise ValueError(f"Style logique inconnu : {style!r} (colonne « {nom} »)")
    xml_styles_patch, index_styles = _resoudre_styles(xml_styles, sorted(set(styles_colonnes.values())))

    m_ref = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)$", _ref_table(xml_table))
    premiere_ligne, derniere_ligne = int(m_ref.group(2)) + 1, int(m_ref.group(4))
    colonne_max = column_index_from_string(m_ref.group(3))
    cibles = [(lettres[nom], formules[nom], index_styles.get(styles_colonnes.get(nom))) for nom in noms]

    def _transformer(num, ouverture, contenu):
        if num < premiere_ligne or num > derniere_ligne:
            return ouverture, contenu
        for lettre, formule, style_defaut in cibles:
            ref = f"{lettre}{num}"
            contenu = _poser_cellule(
                contenu, ref,
                lambda s, ref=ref, formule=formule, style_defaut=style_defaut: _cellule_xml(
                    ref, formule, s if s is not None else style_defaut, cm,
                ),
            )
        return _ajuster_spans(ouverture, colonne_max), contenu

    xml_feuille = _reecrire_lignes(xml_feuille, _transformer, range(premiere_ligne, derniere_ligne + 1))

    parties[chemin_feuille] = xml_feuille
    parties[chemin_table] = xml_table
    parties["xl/workbook.xml"] = _forcer_recalcul_complet(xml_workbook)
    if xml_styles_patch != xml_styles:
        parties["xl/styles.xml"] = xml_styles_patch

    sauvegarde = sauvegarder(fichier, dossier_backups)
    _patcher_parties_xlsx(fichier, parties)
    return sauvegarde


def ajouter_colonnes_calculees(fichier, feuille, table, formules: dict, dossier_backups,
                               styles_colonnes=None) -> Path:
    """Ajoute, en fin du tableau `table` de `feuille`, une colonne calculée
    par entrée de `formules` ({nom: formule anglaise | Formule}) : en-tête
    en ligne 1, <tableColumn> avec <calculatedColumnFormula>, et la MÊME
    formule (<f>) sur chaque ligne du tableau (jamais la ligne 2 comme
    modèle). xl/calcChain.xml supprimé proprement, fullCalcOnLoad posé.
    Verrou/sauvegarde comme appliquer(). Retourne la sauvegarde."""

    return _appliquer_formules_colonnes(fichier, feuille, table, formules, dossier_backups, True, styles_colonnes)


def basculer_colonnes_en_formules(fichier, feuille, table, formules: dict, dossier_backups,
                                  styles_colonnes=None) -> Path:
    """Même chose qu'ajouter_colonnes_calculees() pour des colonnes qui
    EXISTENT déjà dans le tableau (ex. les 5 colonnes facture de
    Commandes) : leurs valeurs saisies sont REMPLACÉES par la formule sur
    chaque ligne (style de cellule conservé). À n'appeler qu'une fois la
    source de vérité migrée (voir CLAUDE.md, étape 5 : contrôle au centime
    AVANT la bascule)."""

    return _appliquer_formules_colonnes(fichier, feuille, table, formules, dossier_backups, False, styles_colonnes)


# --- Réparation d'une ligne (formules décalées) ------------------------------


def _decaler_references_ligne(formule: str, ligne_modele: int, ligne: int) -> str:
    """Références A1 RELATIVES en ligne `ligne_modele` -> `ligne` (H3 -> H2),
    comme Excel le ferait en recopiant la cellule d'une ligne à l'autre.
    Les références structurées (Commandes[[#This Row],[X]]) et les lignes
    absolues ($5) ne sont pas touchées."""

    motif = re.compile(r"(?<![A-Za-z0-9_\[\]'\"])(\$?[A-Z]{1,3})(\d+)(?![\d\(A-Za-z_])")

    def _rempl(m):
        if int(m.group(2)) == ligne_modele:
            return f"{m.group(1)}{ligne}"
        return m.group(0)

    return motif.sub(_rempl, formule)


def reparer_formules_ligne(fichier, dossier_backups, colonnes, feuille=FEUILLE_COMMANDES,
                           ligne: int = 2, ligne_modele: int = 3) -> dict:
    """Réécrit dans les cellules (`colonnes`, index 1-based, liste
    EXPLICITE et limitative) de la ligne `ligne` la formule exacte lue en
    `ligne_modele` (références relatives décalées, voir
    _decaler_references_ligne ; formule matricielle conservée). Hors
    COLONNES_MODIFIABLES par construction : fonction dédiée à la réparation
    d'une ligne de formules décalées (voir CLAUDE.md, étape 0d de P1),
    jamais à une écriture de valeur. Une cellule modèle sans formule, ou une
    formule partagée (t="shared") non expansée, arrête tout (ValueError)
    avant la moindre écriture. Retourne {colonne: formule écrite} ; ne
    réécrit pas une cellule déjà identique."""

    fichier = Path(fichier)
    colonnes = [int(c) for c in colonnes]
    if not colonnes:
        raise ValueError("Aucune colonne à réparer.")
    _verifier_verrou(fichier)

    with zipfile.ZipFile(fichier, "r") as zin:
        chemin_feuille = _chemin_feuille(zin, feuille)
        xml_feuille = zin.read(chemin_feuille).decode("utf-8")
        xml_workbook = zin.read("xl/workbook.xml").decode("utf-8")
        cm = _a_metadata_dynamique(zin)

    def _ligne(num):
        m = re.search(rf'<row r="{num}"[^>]*>(.*?)</row>', xml_feuille, re.S)
        if not m:
            raise ValueError(f"Ligne {num} introuvable dans la feuille « {feuille} ».")
        return m.group(1)

    contenu_modele = _ligne(ligne_modele)
    contenu_cible = _ligne(ligne)

    a_ecrire = {}
    for col in colonnes:
        lettre = get_column_letter(col)
        m = re.search(rf'<c r="{lettre}{ligne_modele}"[^>]*>(.*?)</c>', contenu_modele, re.S)
        m_f = re.search(r"<f\b([^>]*)>(.*?)</f>", m.group(1), re.S) if m else None
        if not m_f:
            raise ValueError(f"Cellule modèle {lettre}{ligne_modele} sans formule — réparation refusée.")
        if 't="shared"' in m_f.group(1):
            raise ValueError(f"Cellule modèle {lettre}{ligne_modele} : formule partagée non expansée — réparation refusée.")
        texte = _decaler_references_ligne(_desechapper_xml(m_f.group(2)), ligne_modele, ligne)
        est_array = 't="array"' in m_f.group(1)
        m_c = re.search(rf'<c r="{lettre}{ligne}"[^>]*>(.*?)</c>', contenu_cible, re.S)
        m_fc = re.search(r"<f\b[^>]*>(.*?)</f>", m_c.group(1), re.S) if m_c else None
        if m_fc and _desechapper_xml(m_fc.group(1)) == texte:
            continue
        a_ecrire[col] = Formule(texte, array=est_array)

    if not a_ecrire:
        return {}

    def _transformer(num, ouverture, contenu):
        if num != ligne:
            return ouverture, contenu
        for col, formule in a_ecrire.items():
            ref = f"{get_column_letter(col)}{num}"
            contenu = _poser_cellule(
                contenu, ref, lambda s, ref=ref, formule=formule: _cellule_xml(ref, formule, s, cm),
            )
        return ouverture, contenu

    xml_feuille = _reecrire_lignes(xml_feuille, _transformer)

    sauvegarder(fichier, dossier_backups)
    _patcher_parties_xlsx(fichier, {
        chemin_feuille: xml_feuille,
        "xl/workbook.xml": _forcer_recalcul_complet(xml_workbook),
    })
    return {col: f.texte for col, f in a_ecrire.items()}
