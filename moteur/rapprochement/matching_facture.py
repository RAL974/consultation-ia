"""
Rapprochement d'une facture déjà lue (voir lecture_facture.py) avec les
lignes de la commande correspondante dans le Suivi commandes — étape 2 du
flux factures (session F2, voir CLAUDE.md "Rapprochement factures").

Réutilise le moteur de correspondance de référence déjà éprouvé côté BL
(moteur.rapprochement.matching : `_memes_references`/`_repli_reference_proche`/
`_repli_referentiel`, référentiel articles compris — importés tels quels,
jamais dupliqués) plutôt que d'en écrire un second. Fonctionne par duck
typing : `LigneFacture` porte les mêmes champs `reference_fournisseur`/
`designation` que `LigneBL`, donc ces fonctions BL s'appliquent sans
adaptation. Seule la SÉMANTIQUE DE COMPARAISON change (voir
`_comparer_facture`) : une facture ne "cumule" jamais une quantité comme un
BL (`qte_livree_cumulee`) — elle est confrontée à ce qui est DÉJÀ enregistré
comme livré (Qté livrée) et facturé (N° facture, si déjà renseigné).

Colonnes facture (voir moteur.rapprochement.ecriture.ENTETES_FACTURE,
réutilisé tel quel ici — jamais une 2e liste de noms redéfinie en double,
voir CLAUDE.md "colonnes créées dans le Suivi vivant") : créées pour de
vrai dans le vrai Suivi commandes le 2026-09-01 (colonnes 51 à 55, table
structurée "Commandes" étendue). "Montant facturé HT" est une colonne de
SAISIE (pas une formule Excel, contrairement au 1er plan F2/Volet 1) —
c'est donc `pipeline_facture.ecritures_pour_facture()` qui doit l'écrire
lui-même (Qté facturée × PU facturé), jamais laissé au calcul Excel.
`lire_lignes_commande_facture()` lit les 5 colonnes si présentes dans LE
Suivi passé en argument, sinon leurs champs restent `None` sur chaque
`LigneSuiviFacture` — le rapprochement en LECTURE SEULE reste alors
utilisable même sur un classeur qui ne les aurait pas encore (diagnostic de
résorption sur les colonnes déjà existantes : Qté livrée / Tarif BL / Tarif
convenu, voir `colonnes_facture_disponibles()`) ; seule l'ÉCRITURE réelle
exige ces colonnes, et `moteur.rapprochement.ecriture.appliquer()` refuse
déjà proprement si elles sont absentes des en-têtes du fichier réel
(`ColonneNonModifiable`, voir son bandeau) — rien à dupliquer ici non plus.
"""

import csv
import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook

from moteur.outils import to_float
from moteur.panier import MAPPING_FOURNISSEURS
from moteur.rapprochement.ecriture import ENTETES_FACTURE
from moteur.rapprochement.matching import (
    _cle,
    _memes_references,
    _repli_reference_proche,
    _repli_referentiel,
)
from moteur.rapprochement.modele_facture import LigneFacture

FEUILLE_COMMANDES = "Commandes"

# Colonnes qui existent déjà dans le Suivi de tout poste (indépendamment des
# colonnes facture, voir plus bas) — sans elles, aucun rapprochement facture
# n'est possible du tout (export du Suivi différent de celui attendu).
_COLONNES_REQUISES_FACTURE = (
    "Référence", "Désignation", "Qté commandée", "N° de commande", "Fournisseur",
    "Qté livrée", "Tarif BL", "Tarif convenu",
)

# Colonnes facture proprement dites — réutilise ENTETES_FACTURE
# (moteur.rapprochement.ecriture, source de vérité unique, voir bandeau) —
# OPTIONNELLES à la LECTURE (leur absence ne bloque pas le diagnostic en
# lecture seule) mais dans moteur.rapprochement.ecriture.COLONNES_MODIFIABLES
# pour l'écriture réelle (qui, elle, échoue proprement si elles manquent).
COLONNES_FACTURE_OPTIONNELLES = ENTETES_FACTURE


@dataclass
class LigneSuiviFacture:

    ligne_excel: int  # numéro de ligne réel dans la feuille (1-based, pour l'écriture)

    reference: str
    designation: str
    qte_commandee: float
    qte_livree: float
    tarif_bl: float | None
    tarif_convenu: float | None
    # Renseigné uniquement par lire_lignes_fournisseur_facture() (déduction
    # de commande par contenu, même principe que côté BL) — vide sinon.
    numero_commande: str = ""

    # None si les colonnes facture n'existent pas encore dans CE Suivi (voir
    # bandeau du module) — jamais une valeur devinée à leur place.
    numero_facture: str | None = None
    date_facture: object = None
    qte_facturee: float | None = None
    pu_facture: float | None = None


class StatutFacture(Enum):
    SUR = "sûr"                  # à écrire après OK global
    A_CONFIRMER = "à confirmer"    # écart, question par question
    DEJA_A_JOUR = "déjà à jour"    # idempotence : rien à écrire
    INCONNU = "inconnu"          # aucune ligne Suivi correspondante sûre
    FRAIS = "frais"              # référence connue comme un frais (voir charger_frais_fournisseurs) — jamais rapprochée à une ligne du Suivi, jamais bloquante


class CauseFacture(Enum):
    """Classification de chaque ligne non "sûre"/anomalie — permet un
    compte rendu CHIFFRÉ par cause (session S0, étape 1e), sans changer les
    messages "raisons" déjà en clair (lus par la GUI, voir
    gui_rapprochement_facture.py) ni la forme des tuples déjà consommés
    ailleurs (RapportRapprochementFacture.anomalies_facture/anomalies_lecture
    restent des 2-tuples — les changer casserait cette même GUI, qui fait
    `for facture, raison in rapport.anomalies_facture`). `cause` est donc :
    - directement renseignée par apparier_facture()/_comparer_facture() pour
      tout ce qui transite par une CorrespondanceFacture (FRAIS, QTE_PARTIELLE,
      QTE_SUPERIEURE, DOUBLON_FACTURE, CLE_PARTIELLE, REF_INCONNUE,
      PRIX_DIFF_MEME_REF) ;
    - dérivée du texte de la raison, en best-effort, pour les anomalies
      "fichier entier" qui restent des 2-tuples (voir
      moteur.rapprochement.pipeline_facture.classifier_cause_anomalie) :
      ANNEXE_SANS_TEXTE, FOURNISSEUR_INCONNU, PARSER_ABSENT, ZERO_LIGNE,
      TOTAL_ECART, COMMANDE_ABSENTE, BDC_MANUEL_24X, AVOIR."""

    ANNEXE_SANS_TEXTE = "annexe_sans_texte"
    FOURNISSEUR_INCONNU = "fournisseur_inconnu"
    PARSER_ABSENT = "parser_absent"
    ZERO_LIGNE = "zero_ligne"
    TOTAL_ECART = "total_ecart"
    COMMANDE_ABSENTE = "commande_absente"
    BDC_MANUEL_24X = "bdc_manuel_24x"
    REF_INCONNUE = "ref_inconnue"
    CLE_PARTIELLE = "cle_partielle"
    QTE_PARTIELLE = "qte_partielle"
    QTE_SUPERIEURE = "qte_superieure"
    DOUBLON_FACTURE = "doublon_facture"
    FRAIS = "frais"
    AVOIR = "avoir"
    PRIX_DIFF_MEME_REF = "prix_diff_meme_ref"
    SUBSTITUTION_PROBABLE = "substitution_probable"


@dataclass
class CorrespondanceFacture:

    ligne_facture: LigneFacture
    ligne_suivi: LigneSuiviFacture | None
    statut: StatutFacture
    raisons: list = field(default_factory=list)
    cause: CauseFacture | None = None


# "N°Réf.Client"/"Réf.:" contenant un bon manuel type "BC 241766"/"BCN
# 241461" (préfixe BC/BCN + "24" + 4 chiffres) — carnet papier d'un chargé
# de travaux, même famille que les "carnets manuels des gars" déjà
# documentés côté BL (CLAUDE.md) : structurellement PAS rattachable à une
# commande du Suivi, jamais un résultat à deviner. Cas réels confirmés :
# Facture_362777.pdf (109, "BC 241766"), 6100226.pdf (Coredime, "Réf.:
# BCN 241461").
MOTIF_BDC_MANUEL_24X = re.compile(r"\bBCN?\s*24\d{4}\b", re.IGNORECASE)


def est_bdc_manuel_24x(candidats_bruts) -> bool:
    """True si l'un des candidats bruts (N°Réf.Client/Réf.: NON parsé en
    n° de commande exploitable, voir Facture.numeros_commande_bruts)
    ressemble à un bon manuel "BC/BCN 24XXXX" — sert à distinguer cette
    cause précise (BDC_MANUEL_24X) d'une commande simplement absente du
    Suivi (COMMANDE_ABSENTE)."""

    return any(MOTIF_BDC_MANUEL_24X.search(str(c)) for c in candidats_bruts if c)


def charger_frais_fournisseurs(chemin_csv) -> dict:
    """Charge referentiel/frais_fournisseurs.csv (colonnes Fournisseur;
    Reference;Libelle) -> {FOURNISSEUR: {REFERENCE: libellé}} (clés
    normalisées en majuscules). Fichier absent -> dict vide, comportement
    inchangé (aucun frais reconnu) — jamais une erreur bloquante.

    Sert à isoler des références qui ne sont PAS de vrais articles (frais
    de port, éco-taxe...) : reconnues, elles ne sont jamais confrontées au
    Suivi (ni "inconnue", ni bloquantes) — voir apparier_facture(). Une
    référence absente de ce fichier suit le comportement EXISTANT sans
    aucun changement, quel que soit son montant (jamais de seuil "petit
    montant = ignoré" inventé, décision explicite de cadrage — voir
    CLAUDE.md session S0)."""

    chemin_csv = Path(chemin_csv)
    resultat = {}
    if not chemin_csv.is_file():
        return resultat

    with open(chemin_csv, encoding="utf-8-sig") as f:
        # Lignes de commentaire ("# ...", en tête de fichier pour
        # documenter le format) ignorées avant l'en-tête — même idiome que
        # moteur.referentiel.Referentiel.importer_equivalences_bl.
        lignes_utiles = (l for l in f if not l.lstrip().startswith("#"))
        for ligne in csv.DictReader(lignes_utiles, delimiter=";"):
            fournisseur = (ligne.get("Fournisseur") or "").strip().upper()
            reference = (ligne.get("Reference") or "").strip().upper()
            libelle = (ligne.get("Libelle") or "").strip()
            if not fournisseur or not reference:
                continue
            resultat.setdefault(fournisseur, {})[reference] = libelle or reference

    return resultat


def colonnes_facture_disponibles(entetes: dict) -> bool:
    return all(c in entetes for c in COLONNES_FACTURE_OPTIONNELLES)


def _ouvrir_feuille_commandes_facture(chemin_suivi):
    wb = load_workbook(chemin_suivi, read_only=True, data_only=True)
    ws = wb[FEUILLE_COMMANDES]
    entetes = {
        c.value: i
        for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
    }

    manquantes = [c for c in _COLONNES_REQUISES_FACTURE if c not in entetes]
    if manquantes:
        wb.close()
        raise KeyError(
            f"Colonne(s) introuvable(s) dans « {FEUILLE_COMMANDES} » : "
            f"{', '.join(manquantes)} — export du Suivi commandes différent de celui attendu ?"
        )

    return wb, ws, entetes


def _valeur_texte(v):
    v = str(v).strip() if v is not None else ""
    return v or None


def _ligne_suivi_facture_depuis_row(i, row, entetes, avec_commande=False) -> LigneSuiviFacture:

    facture_ok = colonnes_facture_disponibles(entetes)

    return LigneSuiviFacture(
        ligne_excel=i,
        reference=row[entetes["Référence"]],
        designation=row[entetes["Désignation"]],
        qte_commandee=to_float(row[entetes["Qté commandée"]]),
        qte_livree=to_float(row[entetes["Qté livrée"]]),
        tarif_bl=row[entetes["Tarif BL"]],
        tarif_convenu=row[entetes["Tarif convenu"]],
        numero_commande=str(row[entetes["N° de commande"]] or "").strip() if avec_commande else "",
        numero_facture=_valeur_texte(row[entetes["N° facture"]]) if facture_ok else None,
        date_facture=row[entetes["Date facture"]] if facture_ok else None,
        qte_facturee=(
            to_float(row[entetes["Qté facturée"]]) if facture_ok and row[entetes["Qté facturée"]] is not None else None
        ),
        pu_facture=(
            to_float(row[entetes["PU facturé"]]) if facture_ok and row[entetes["PU facturé"]] is not None else None
        ),
    )


def lire_lignes_commande_facture(chemin_suivi, fournisseur: str, numero_commande: str) -> list[LigneSuiviFacture]:
    """Même principe que moteur.rapprochement.matching.lire_lignes_commande
    (conversion MAPPING_FOURNISSEURS comprise), pour les colonnes facture."""

    fournisseur = MAPPING_FOURNISSEURS.get(fournisseur.upper(), fournisseur)

    wb, ws, entetes = _ouvrir_feuille_commandes_facture(chemin_suivi)
    try:
        i_cde = entetes["N° de commande"]
        i_fourn = entetes["Fournisseur"]

        resultat = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            cde = row[i_cde]
            fourn = row[i_fourn]

            if cde is None or str(cde).strip() != numero_commande:
                continue
            if fourn is None or _cle(fourn) != _cle(fournisseur):
                continue

            resultat.append(_ligne_suivi_facture_depuis_row(i, row, entetes))

        return resultat
    finally:
        wb.close()


def lire_lignes_fournisseur_facture(chemin_suivi, fournisseur: str) -> list[LigneSuiviFacture]:
    """Même principe que moteur.rapprochement.matching.lire_lignes_fournisseur
    — utilisé pour la déduction de commande par contenu quand l'en-tête de
    la facture (N°Réf.Client) ne suffit pas (voir pipeline_facture.py)."""

    fournisseur = MAPPING_FOURNISSEURS.get(fournisseur.upper(), fournisseur)

    wb, ws, entetes = _ouvrir_feuille_commandes_facture(chemin_suivi)
    try:
        i_fourn = entetes["Fournisseur"]

        resultat = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            fourn = row[i_fourn]
            if fourn is None or _cle(fourn) != _cle(fournisseur):
                continue

            resultat.append(_ligne_suivi_facture_depuis_row(i, row, entetes, avec_commande=True))

        return resultat
    finally:
        wb.close()


def agreger_lignes_meme_reference(lignes_facture: list) -> tuple:
    """Regroupe les lignes d'une même facture qui partagent la même
    référence normalisée (même `_cle()` que la comparaison exacte) —
    quantités SOMMÉES avant comparaison à la Qté livrée, montants sommés,
    n° de BL concaténés ("détail par BL conservé"). Cas réel qui motive ce
    correctif (session S0) : Facture_362840.pdf (109, commande 123.089) —
    P03200 et F2U15RVVOO sont chacun répartis sur 2 "Bon de livraison"
    DIFFÉRENTS de la MÊME facture (livraison fractionnée) ; comparés
    isolément bloc par bloc, chaque bloc ressortait "à confirmer" à tort
    (100 facturés sur ce bloc vs 200 déjà livrés en tout), alors que la
    somme (100+100=200, 100+200=300) correspond exactement.

    PU exigé IDENTIQUE entre les lignes regroupées (arrondi à 4 décimales,
    comme le reste du projet) — sinon AUCUNE agrégation n'est faite pour ce
    groupe : chaque ligne reste séparée (retournée telle quelle dans
    `lignes_facture`), sa référence ajoutée à `refs_prix_differents` pour
    qu'apparier_facture() la fasse ressortir "à confirmer"
    (CauseFacture.PRIX_DIFF_MEME_REF) plutôt que de deviner quel prix est
    le bon. Une référence qui n'apparaît qu'UNE fois traverse inchangée.

    Retourne (lignes_regroupees, refs_prix_differents)."""

    groupes = {}
    ordre = []
    for l in lignes_facture:
        cle = _cle(l.reference_fournisseur)
        groupes.setdefault(cle, [])
        if cle not in ordre:
            ordre.append(cle)
        groupes[cle].append(l)

    resultat = []
    refs_prix_differents = set()

    for cle in ordre:
        lot = groupes[cle]

        if len(lot) == 1:
            resultat.append(lot[0])
            continue

        prix_distincts = {round(l.prix_unitaire_ht, 4) for l in lot if l.prix_unitaire_ht is not None}
        if len(prix_distincts) > 1:
            refs_prix_differents.add(cle)
            resultat.extend(lot)
            continue

        montants = [l.montant_ht for l in lot if l.montant_ht is not None]
        resultat.append(LigneFacture(
            reference_fournisseur=lot[0].reference_fournisseur,
            designation=lot[0].designation,
            quantite_facturee=sum(l.quantite_facturee for l in lot),
            prix_unitaire_ht=lot[0].prix_unitaire_ht,
            montant_ht=round(sum(montants), 2) if montants else None,
            numero_commande=lot[0].numero_commande,
            numero_bl=" + ".join(dict.fromkeys(l.numero_bl for l in lot if l.numero_bl)),
        ))

    return resultat, refs_prix_differents


def _repli_premier_token(ligne_facture: LigneFacture, disponibles: list) -> tuple:
    """Repli : la référence Suivi porte un SUFFIXE LIBRE après le vrai
    code — texte descriptif ajouté à la saisie, jamais présent sur la
    facture. Cas réel (session S0, Facture_6108234.pdf, commande
    M3.14.342) : Suivi "SIXGPCP35 PVC" vs facture "SIXGPCP35" — comparaison
    sur le PREMIER TOKEN (avant le 1er espace) de la référence Suivi.
    N'accepte que s'il existe UNE SEULE ligne candidate parmi les
    `disponibles` (jamais un choix au hasard, même principe que
    _repli_reference_proche) ET que la référence Suivi contient
    RÉELLEMENT un espace (sinon ce serait une simple comparaison exacte,
    déjà couverte par _memes_references, pas la peine de la refaire ici).
    Toujours "à confirmer", jamais "sûr"."""

    cle_facture = str(ligne_facture.reference_fournisseur or "").strip().upper()
    if not cle_facture:
        return None, None

    candidats = [
        ls for ls in disponibles
        if " " in str(ls.reference or "").strip()
        and str(ls.reference).strip().split()[0].upper() == cle_facture
    ]

    if len(candidats) == 1:
        ls = candidats[0]
        return ls, (
            f"Référence Suivi « {ls.reference} » à suffixe libre — le premier terme "
            f"correspond exactement à la référence facturée « {ligne_facture.reference_fournisseur} »"
        )

    return None, None


def _residuel_unique(resultats: list, lignes_a_apparier: list, disponibles: list) -> None:
    """Modifie `resultats` EN PLACE (dernier repli, appelé après tous les
    autres) : si, pour cette commande, il ne reste plus qu'UNE SEULE ligne
    facture INCONNUE et qu'UNE SEULE ligne Suivi jamais réclamée
    (`disponibles`), ET que cette ligne Suivi n'a encore AUCUN n° de
    facture, ET que les deux quantités concordent (et les PU aussi, à
    0,02€ près, quand les deux sont connus) -> "substitution probable",
    TOUJOURS "à confirmer" (jamais "sûr" : aucune preuve textuelle ou
    structurelle, seulement un processus d'élimination — contrairement à
    _repli_reference_proche/_repli_premier_token/_repli_referentiel, qui
    exigent tous une ressemblance).

    Cas réel qui motive ce repli (Facture_6108234.pdf/Coredime, commande
    M3.14.342) : référence facturée "LEG06620" (aucune ligne Suivi ne s'en
    approche, même avec tous les replis existants) alors qu'il ne reste,
    à la fin, QU'UNE ligne Suivi non facturée pour cette commande ("5120",
    même désignation "ICT 20 BLEU TURBO G-ROUL 100M", même quantité 100,
    même tarif 0,37€) — "5120" est manifestement une ancienne référence
    reprise par Coredime sous un autre code, sans AUCUN rapport textuel ni
    numérique. Une correspondance retenue ici est destinée, une fois
    confirmée par l'acheteur, à être apprise dans
    referentiel/equivalences_bl.csv (voir moteur.rapprochement.
    pipeline_facture._ecrire_substitutions_probables/
    _appliquer_confirmations_substitutions) — jamais directement dans la
    table alias comme un repli référentiel ordinaire, ce fichier CSV
    étant le ledger partagé BL+Facture des substitutions pures (voir
    CLAUDE.md)."""

    # Garde-fou : n'a de sens qu'après une VRAIE élimination parmi
    # plusieurs lignes facture (c'est la convergence à 1 seul survivant sur
    # 9, par exemple, qui rend le rapprochement crédible malgré l'absence
    # de ressemblance) — jamais sur une commande à 1 seule ligne facture
    # dès le départ, où "il ne reste qu'une ligne inconnue" est vrai par
    # construction et ne prouve rien (bug réel trouvé en confrontant les
    # tests synthétiques déjà existants, qui utilisent tous une quantité/
    # un prix par défaut identiques des deux côtés pour une raison sans
    # rapport : sans ce garde-fou, n'importe quelle paire 1 vs 1 aux
    # références totalement différentes se faisait proposer comme
    # "substitution probable" au lieu de rester "inconnue").
    if len(lignes_a_apparier) <= 1:
        return

    inconnus = [i for i, c in enumerate(resultats) if c.statut is StatutFacture.INCONNU]
    if len(inconnus) != 1 or len(disponibles) != 1:
        return

    ls = disponibles[0]
    if ls.numero_facture:
        return

    i = inconnus[0]
    lf = lignes_a_apparier[i]

    if abs(lf.quantite_facturee - ls.qte_livree) > 0.001:
        return

    prix_ref = ls.tarif_bl or ls.tarif_convenu
    if lf.prix_unitaire_ht and prix_ref and abs(lf.prix_unitaire_ht - prix_ref) > 0.02:
        return

    disponibles.remove(ls)
    resultats[i] = CorrespondanceFacture(
        lf, ls, StatutFacture.A_CONFIRMER,
        [
            f"Substitution probable : après rapprochement, il ne reste plus, pour cette "
            f"commande, qu'une seule ligne facture non rapprochée (« {lf.reference_fournisseur} ») "
            f"et qu'une seule ligne Suivi sans facture (« {ls.reference} ») — même quantité "
            f"({lf.quantite_facturee:g}). À confirmer avant d'apprendre cette équivalence."
        ],
        CauseFacture.SUBSTITUTION_PROBABLE,
    )


def _comparer_facture(ligne_facture: LigneFacture, ligne_suivi: LigneSuiviFacture, numero_facture: str) -> CorrespondanceFacture:
    """Une facture n'est jamais "cumulée" comme un BL : elle est confrontée
    à ce qui est DÉJÀ enregistré (Qté livrée, Tarif BL/Tarif convenu, et le
    N° de facture déjà présent le cas échéant — idempotence).

    Une facture reçue AVANT que son BL soit rapproché dans le Suivi (Qté
    livrée encore à 0) N'EST PAS un motif de blocage — décision explicite
    de l'acheteur (session F4, Coredime) : "ce ne sont pas des factures non
    parvenues puisqu'on les a reçues ! [...] il y a les BL manquants
    là-dedans, ils sont signés" — la livraison est réellement survenue
    (BL papier signé), seul son rapprochement dans l'outil n'a pas encore
    eu lieu. Le contrôle de cohérence de quantité (facturée vs livrée) est
    alors simplement IGNORÉ (rien à comparer, ce n'est pas signe d'anomalie).

    Le PRIX facturé N'EST PLUS NON PLUS un motif de blocage — décision
    explicite de l'acheteur, qui affine sa position dans la même session :
    d'abord une tolérance de 0,01€ envisagée (bruit d'arrondi type
    GEWDX40020/GEWDX27720), puis, quelques secondes plus tard : "il faut
    écrire tout ce qui apparaît sur les factures rapprochables à des
    commandes, quel que soit le prix". Revient donc entièrement sur le
    "aucune tolérance" du cadrage initial (Volet 3, CLAUDE.md). Le PU
    facturé est écrit tel quel (voir pipeline_facture.ecritures_pour_facture)
    ; tout écart avec Tarif BL/Tarif convenu reste visible directement dans
    le Suivi (colonnes côte à côte), sans jamais bloquer l'écriture. Seul ce
    qui concerne l'IDENTITÉ du rapprochement (bon n° de facture, bonne
    ligne — pas son prix) reste un motif de blocage."""

    raisons = []
    cause = None

    if ligne_suivi.numero_facture:
        if ligne_suivi.numero_facture == numero_facture:
            # Idempotence : ce même n° de facture est déjà enregistré sur
            # cette ligne — un doublon de dépôt/traitement, rien à réécrire
            # (même principe que côté BL, voir moteur.rapprochement.matching._comparer).
            return CorrespondanceFacture(ligne_facture, ligne_suivi, StatutFacture.DEJA_A_JOUR)
        raisons.append(
            f"Cette ligne porte déjà un autre n° de facture ({ligne_suivi.numero_facture}) "
            "— doublon de dépôt ou litige de facturation à vérifier avant d'écraser"
        )
        return CorrespondanceFacture(
            ligne_facture, ligne_suivi, StatutFacture.A_CONFIRMER, raisons, CauseFacture.DOUBLON_FACTURE,
        )

    if ligne_suivi.qte_livree > 0 and abs(ligne_facture.quantite_facturee - ligne_suivi.qte_livree) > 0.001:
        raisons.append(
            f"Qté facturée ({ligne_facture.quantite_facturee:g}) différente de la Qté livrée "
            f"déjà enregistrée ({ligne_suivi.qte_livree:g}) — facturation partielle/multiple ?"
        )
        cause = (
            CauseFacture.QTE_SUPERIEURE
            if ligne_facture.quantite_facturee > ligne_suivi.qte_livree
            else CauseFacture.QTE_PARTIELLE
        )

    statut = StatutFacture.A_CONFIRMER if raisons else StatutFacture.SUR
    return CorrespondanceFacture(ligne_facture, ligne_suivi, statut, raisons, cause)


def apparier_facture(lignes_facture: list[LigneFacture], lignes_suivi: list[LigneSuiviFacture],
                      numero_facture: str, referentiel=None, fournisseur: str = "",
                      devis: str = "", frais_connus: dict | None = None) -> list[CorrespondanceFacture]:
    """Même algorithme deux passes que moteur.rapprochement.matching.apparier()
    (voir son bandeau pour le bug réel qu'il corrige — une correspondance
    EXACTE ne doit jamais être volée par un repli approximatif traité avant
    elle dans l'ordre du document) : 1re passe, uniquement les
    correspondances EXACTES (candidat unique) sur toutes les lignes de la
    facture ; 2e passe, ce qui reste (repli référence proche / premier
    token / référentiel / ambiguïté), sur les lignes Suivi déjà réduites.

    Quatre ajouts (session S0, voir CLAUDE.md) :
    - `frais_connus` (voir charger_frais_fournisseurs) : toute ligne dont la
      référence est un FRAIS connu pour ce fournisseur (ex. COREDIME ECO-23,
      COREDIME 9993 LIVRAISON AVION) est retirée du rapprochement AVANT
      toute comparaison — statut FRAIS direct, jamais "inconnue", jamais
      bloquante. Une référence absente de ce whitelist suit le comportement
      normal, quel que soit son montant (jamais de seuil deviné).
    - Agrégation par référence (voir agreger_lignes_meme_reference) AVANT
      comparaison : un même article facturé sur PLUSIEURS BL de cette
      facture est comparé à la Qté livrée UNE SEULE FOIS, quantités
      sommées — sinon chaque bloc, confronté isolément à la Qté livrée
      TOTALE, ressort "à confirmer" à tort sur une livraison fractionnée
      qui, additionnée, correspond exactement. PU différent selon le bloc
      -> pas d'agrégation, chaque ligne reste "à confirmer"
      (CauseFacture.PRIX_DIFF_MEME_REF).
    - Repli "premier token" (voir _repli_premier_token), entre le repli
      référence-proche et le repli référentiel : la référence Suivi porte
      un suffixe libre (texte descriptif ajouté à la saisie).
    - Résiduel unique (voir _residuel_unique), en tout dernier repli :
      exactement 1 ligne facture encore inconnue + exactement 1 ligne
      Suivi jamais réclamée pour cette commande, mêmes quantité/PU ->
      "substitution probable", toujours à confirmer."""

    frais_du_fournisseur = (frais_connus or {}).get(fournisseur.upper(), {})

    correspondances_frais = []
    lignes_a_apparier = []
    for lf in lignes_facture:
        libelle = frais_du_fournisseur.get(str(lf.reference_fournisseur or "").strip().upper())
        if libelle:
            correspondances_frais.append(CorrespondanceFacture(
                lf, None, StatutFacture.FRAIS,
                [f"Frais connu ({libelle}) — jamais rapproché à une ligne du Suivi"],
                CauseFacture.FRAIS,
            ))
        else:
            lignes_a_apparier.append(lf)

    lignes_a_apparier, refs_prix_differents = agreger_lignes_meme_reference(lignes_a_apparier)

    disponibles = list(lignes_suivi)
    resultat_par_index = {}

    for i, lf in enumerate(lignes_a_apparier):
        candidats = [
            ls for ls in disponibles
            if _memes_references(ls.reference, lf.reference_fournisseur, referentiel,
                                  ls.designation, lf.designation, fournisseur, devis)
        ]
        if len(candidats) == 1:
            ls = candidats[0]
            disponibles.remove(ls)
            resultat_par_index[i] = _comparer_facture(lf, ls, numero_facture)

    for i, lf in enumerate(lignes_a_apparier):

        if i in resultat_par_index:
            continue

        candidats = [
            ls for ls in disponibles
            if _memes_references(ls.reference, lf.reference_fournisseur, referentiel,
                                  ls.designation, lf.designation, fournisseur, devis)
        ]

        if not candidats:
            ls_repli, raison_repli = _repli_reference_proche(lf, disponibles)
            if ls_repli is None:
                ls_repli, raison_repli = _repli_premier_token(lf, disponibles)
            if ls_repli is None:
                ls_repli, raison_repli = _repli_referentiel(lf, disponibles, referentiel, fournisseur, devis)
            if ls_repli is not None:
                disponibles.remove(ls_repli)
                c = _comparer_facture(lf, ls_repli, numero_facture)
                if c.statut is StatutFacture.DEJA_A_JOUR:
                    # Même garde-fou que côté BL (voir matching.apparier) :
                    # un repli approximatif vers une ligne déjà à jour ne
                    # doit jamais être "recomparé" comme un nouvel écart —
                    # il ressort déjà_a_jour, raison du repli conservée.
                    resultat_par_index[i] = CorrespondanceFacture(
                        c.ligne_facture, c.ligne_suivi, StatutFacture.DEJA_A_JOUR, [raison_repli],
                    )
                else:
                    resultat_par_index[i] = CorrespondanceFacture(
                        c.ligne_facture, c.ligne_suivi, StatutFacture.A_CONFIRMER,
                        [raison_repli] + c.raisons, CauseFacture.CLE_PARTIELLE,
                    )
            else:
                resultat_par_index[i] = CorrespondanceFacture(
                    lf, None, StatutFacture.INCONNU,
                    ["Aucune ligne du Suivi ne correspond à cette référence pour cette commande"],
                    CauseFacture.REF_INCONNUE,
                )
        else:
            exacts = [
                ls for ls in candidats
                if str(ls.reference).strip().upper() == lf.reference_fournisseur.strip().upper()
            ]
            if len(exacts) == 1:
                ls = exacts[0]
                disponibles.remove(ls)
                resultat_par_index[i] = _comparer_facture(lf, ls, numero_facture)
            else:
                resultat_par_index[i] = CorrespondanceFacture(
                    lf, None, StatutFacture.INCONNU,
                    [f"{len(candidats)} lignes du Suivi correspondent à cette référence (ambigu)"],
                    CauseFacture.REF_INCONNUE,
                )

    resultats = [resultat_par_index[i] for i in range(len(lignes_a_apparier))]

    for i, lf in enumerate(lignes_a_apparier):
        if _cle(lf.reference_fournisseur) not in refs_prix_differents:
            continue
        c = resultats[i]
        if c.statut in (StatutFacture.DEJA_A_JOUR, StatutFacture.INCONNU):
            # Rien à confirmer sur une ligne déjà à jour ; et forcer
            # "à confirmer" sur une ligne INCONNUE laisserait ligne_suivi à
            # None avec un statut qui suppose pourtant un candidat — jamais
            # cette incohérence (voir ecritures_pour_facture, qui lit
            # c.ligne_suivi.ligne_excel sans garde).
            continue
        resultats[i] = CorrespondanceFacture(
            c.ligne_facture, c.ligne_suivi, StatutFacture.A_CONFIRMER,
            c.raisons + [
                "Cette référence est facturée sur plusieurs BL de cette facture à des PU "
                "différents — pas d'agrégation automatique, prix à vérifier ligne à ligne"
            ],
            c.cause or CauseFacture.PRIX_DIFF_MEME_REF,
        )

    _residuel_unique(resultats, lignes_a_apparier, disponibles)

    return resultats + correspondances_frais
