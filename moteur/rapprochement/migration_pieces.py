"""
Étape 4 de P1 (voir CLAUDE.md, « Feuille Pièces — modèle, socles,
migration ») : migration des lignes facture EXISTANTES de la feuille
Commandes (les 5 colonnes de saisie « N° facture », « Date facture », « Qté
facturée », « PU facturé », « Montant facturé HT ») vers la feuille Pièces.

Source de vérité = les 5 colonnes + les PDF archivés :
- les PDF sont relus en LECTURE SEULE avec les parsers existants
  (moteur.rapprochement.lecture_facture.lire_facture — aucune règle de
  parsing ici) pour retrouver, ligne par ligne, le N° BL lié, la référence
  fournisseur, la désignation ;
- Qté / PU HT / Montant HT viennent des 5 colonnes de Commandes : ce sont
  les valeurs que le pipeline y a écrites (montant imprimé, ou recalculé et
  signalé à l'époque) — c'est ce qui garantit le contrôle AU CENTIME
  (Σ Montant facturé HT de Commandes = Σ Pieces[Montant HT] type Facture,
  par fournisseur) exigé avant la bascule des colonnes en formules
  (étape 5). Un montant imprimé différent sur le PDF est signalé en
  Commentaire, jamais substitué en silence.
- Une ligne Commandes qui agrège PLUSIEURS lignes du PDF (même référence
  sur plusieurs BL de la même facture, voir
  matching_facture.agreger_lignes_meme_reference) redevient une ligne
  Pièces PAR ligne de document (une par BL), si et seulement si leurs
  quantités et montants imprimés se somment exactement aux valeurs de
  Commandes.
- PDF introuvable, n° de facture absent du PDF, ligne non retrouvée ou
  ambiguë -> ligne Pièces créée depuis les 5 colonnes seules,
  Mode « Migré sans PDF », Commentaire explicite. Sinon Mode « Migré ».

Où sont les PDF : a_traiter/BL/Traités/<commande>/ (nom
« <date> - <fournisseur> - Facture <n°> - BC <commande>.pdf », copie par
commande), et, pour une facture écrite partiellement puis laissée à la
décision de l'acheteur, encore dans a_traiter/Factures/À vérifier/ (ou à la
racine de a_traiter/Factures/) sous son nom d'origine — cherchée par son
numéro dans le nom de fichier, puis vérifiée par le numéro LU dans le PDF.
"""

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from moteur.rapprochement.lecture_facture import lire_facture
from moteur.rapprochement.matching import _cle
from moteur.rapprochement.pieces import (
    COLONNES_FACTURE_CALCULEES,
    FEUILLE_COMMANDES,
    MODE_MIGRE,
    MODE_MIGRE_SANS_PDF,
    TYPE_FACTURE,
    IndexPieces,
    cle_reference,
    dedoublonner_ids,
    ecrire_pieces,
    lire_pieces,
    nom_fournisseur_suivi,
    nouvelle_piece,
)
from moteur.rapprochement.pipeline_bl import _nom_dossier_commande, _parser_date_bl, _sans_caracteres_interdits
from moteur.rapprochement.pipeline_facture import (
    DOSSIER_A_TRAITER_FACTURES,
    DOSSIER_A_VERIFIER_FACTURES,
    DOSSIER_RAPPORTS,
    DOSSIER_TRAITES_BL,
)

TOLERANCE_CENTIME = 0.005
TOLERANCE_QTE = 0.001
EXTENSIONS_PDF = (".pdf",)


@dataclass
class LigneFacturee:
    """Une ligne de Commandes portant un N° facture (les 5 colonnes + ce
    qu'il faut pour Pièces)."""

    ligne_excel: int
    fournisseur: str
    numero_commande: str
    reference: object
    chantier: object
    sous_chantier: object
    numero_facture: str
    date_facture: object
    qte: float
    pu: float | None
    montant: float

    @property
    def fournisseur_suivi(self) -> str:
        return nom_fournisseur_suivi(self.fournisseur)


def _nombre(v, defaut=None):
    if v is None or v == "":
        return defaut
    try:
        return float(v)
    except (TypeError, ValueError):
        return defaut


def lire_lignes_facturees(fichier, limite=None) -> list:
    """Toutes les lignes de Commandes avec un N° facture non vide, dans
    l'ordre de la feuille (`limite` : n premières, pour un extrait de
    test)."""

    wb = load_workbook(fichier, read_only=True, data_only=True)
    try:
        ws = wb[FEUILLE_COMMANDES]
        lignes = ws.iter_rows(values_only=True)
        entetes = {str(c).strip(): i for i, c in enumerate(next(lignes)) if c is not None}
        manquantes = [c for c in ("Référence", "N° de commande", "Fournisseur", "Chantier", "Sous-Chantier") + COLONNES_FACTURE_CALCULEES if c not in entetes]
        if manquantes:
            raise KeyError(f"Colonne(s) introuvable(s) dans « {FEUILLE_COMMANDES} » : {', '.join(manquantes)}")

        def _val(row, nom):
            i = entetes[nom]
            return row[i] if i < len(row) else None

        resultat = []
        for i, row in enumerate(lignes, start=2):
            numero = _val(row, "N° facture")
            if numero is None or str(numero).strip() == "":
                continue
            resultat.append(LigneFacturee(
                ligne_excel=i,
                fournisseur=str(_val(row, "Fournisseur") or "").strip(),
                numero_commande=str(_val(row, "N° de commande") or "").strip(),
                reference=_val(row, "Référence"),
                chantier=_val(row, "Chantier"),
                sous_chantier=_val(row, "Sous-Chantier"),
                numero_facture=cle_reference(numero),
                date_facture=_val(row, "Date facture"),
                qte=_nombre(_val(row, "Qté facturée"), 0.0),
                pu=_nombre(_val(row, "PU facturé")),
                montant=_nombre(_val(row, "Montant facturé HT"), 0.0),
            ))
            if limite is not None and len(resultat) >= limite:
                break
        return resultat
    finally:
        wb.close()


def trouver_pdf_facture(dossier_projet, numero_commande: str, numero_facture: str) -> list:
    """Candidats PDF pour (commande, n° facture), du plus sûr au moins sûr :
    archive de la commande (nom normalisé), puis a_traiter/Factures/
    À vérifier/ et a_traiter/Factures/ (nom d'origine contenant le n°)."""

    dossier_projet = Path(dossier_projet)
    numero = cle_reference(numero_facture)
    if not numero:
        return []
    candidats = []

    dossier_commande = dossier_projet / DOSSIER_TRAITES_BL / _nom_dossier_commande(numero_commande)
    if dossier_commande.is_dir():
        motif = f"* - Facture {_sans_caracteres_interdits(numero, '-')} - BC *"
        candidats += sorted(p for p in dossier_commande.glob(motif) if p.suffix.lower() in EXTENSIONS_PDF)

    if len(numero) >= 4:
        for dossier in (
            dossier_projet / DOSSIER_A_TRAITER_FACTURES / DOSSIER_A_VERIFIER_FACTURES,
            dossier_projet / DOSSIER_A_TRAITER_FACTURES,
        ):
            if dossier.is_dir():
                candidats += sorted(
                    p for p in dossier.iterdir()
                    if p.is_file() and p.suffix.lower() in EXTENSIONS_PDF and numero.lower() in p.stem.lower()
                )
    return candidats


class _CacheFactures:
    """Un PDF n'est lu qu'une fois ; les lignes déjà rattachées à une ligne
    Commandes ne sont plus proposées (une ligne de document = une ligne
    Pièces, jamais deux)."""

    def __init__(self):
        self.factures = {}
        self.consommees = defaultdict(set)
        self.erreurs = {}

    def lire(self, chemin: Path) -> list:
        chemin = Path(chemin)
        if chemin not in self.factures:
            try:
                factures, raisons = lire_facture(chemin)
            except Exception as e:  # jamais bloquant : la ligne sera « sans PDF »
                factures, raisons = [], [f"erreur de lecture ({e})"]
            self.factures[chemin] = factures
            if raisons:
                self.erreurs[chemin] = raisons
        return self.factures[chemin]

    def libres(self, chemin: Path, facture) -> list:
        return [l for l in facture.lignes if id(l) not in self.consommees[chemin]]

    def consommer(self, chemin: Path, lignes) -> None:
        self.consommees[chemin].update(id(l) for l in lignes)


def _proche(a, b, tolerance) -> bool:
    return a is not None and b is not None and abs(float(a) - float(b)) <= tolerance


def _date_piece(ligne: LigneFacturee, facture):
    d = ligne.date_facture
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    if facture is not None:
        return _parser_date_bl(facture.date_facture)
    return None


def _piece_base(ligne: LigneFacturee, facture, chemin_pdf, mode, commentaire, **champs) -> dict:
    base = dict(
        type_piece=TYPE_FACTURE,
        fournisseur=ligne.fournisseur_suivi,
        numero_piece=ligne.numero_facture,
        date_piece=_date_piece(ligne, facture),
        numero_commande=ligne.numero_commande,
        chantier=ligne.chantier,
        sous_chantier=ligne.sous_chantier,
        reference_suivi=ligne.reference,
        reference_fournisseur=ligne.reference,
        designation=None,
        qte=ligne.qte,
        pu_ht=ligne.pu,
        montant_ht=ligne.montant,
        numero_bl_lie="",
        mode=mode,
        fichier=chemin_pdf,
        commentaire=commentaire,
    )
    base.update(champs)
    return nouvelle_piece(**base)


def _migrer_ligne(ligne: LigneFacturee, dossier_projet, cache: _CacheFactures, statistiques: dict) -> list:
    """Les lignes Pièces (1..n) pour UNE ligne de Commandes — voir bandeau
    pour les règles de rattachement."""

    chemins = trouver_pdf_facture(dossier_projet, ligne.numero_commande, ligne.numero_facture)
    facture, chemin_pdf = None, None
    for chemin in chemins:
        for f in cache.lire(chemin):
            if cle_reference(f.numero_facture) == ligne.numero_facture:
                facture, chemin_pdf = f, chemin
                break
        if facture is not None:
            break

    if facture is None:
        if not chemins:
            motif = "PDF introuvable (archive de la commande, À vérifier, Factures)"
            statistiques["pdf_introuvable"] += 1
        else:
            motif = "n° de facture non retrouvé dans " + ", ".join(p.name for p in chemins)
            statistiques["numero_absent_du_pdf"] += 1
        return [_piece_base(ligne, None, chemins[0] if chemins else None, MODE_MIGRE_SANS_PDF, motif)]

    commentaires = []
    if nom_fournisseur_suivi(facture.fournisseur).upper() != ligne.fournisseur_suivi.upper():
        commentaires.append(f"fournisseur PDF « {facture.fournisseur} » ≠ Commandes « {ligne.fournisseur} »")

    libres = cache.libres(chemin_pdf, facture)
    cle_ref = _cle(ligne.reference)
    candidats = [l for l in libres if _cle(l.reference_fournisseur) == cle_ref]
    alias = False
    if not candidats:
        par_valeurs = [
            l for l in libres
            if _proche(l.quantite_facturee, ligne.qte, TOLERANCE_QTE)
            and (ligne.pu is None or l.prix_unitaire_ht is None or _proche(l.prix_unitaire_ht, ligne.pu, TOLERANCE_CENTIME))
            and (l.montant_ht is None or _proche(l.montant_ht, ligne.montant, TOLERANCE_CENTIME))
        ]
        if len(par_valeurs) == 1:
            candidats, alias = par_valeurs, True

    if not candidats:
        statistiques["ligne_non_retrouvee"] += 1
        return [_piece_base(
            ligne, facture, chemin_pdf, MODE_MIGRE_SANS_PDF,
            "; ".join([f"ligne non retrouvée dans {chemin_pdf.name}"] + commentaires),
        )]

    if alias:
        commentaires.append(f"réf. facture « {candidats[0].reference_fournisseur} » ≠ réf. Suivi « {ligne.reference} »")

    def _piece_depuis(l, qte, pu, montant, commentaire):
        return _piece_base(
            ligne, facture, chemin_pdf, MODE_MIGRE, commentaire,
            reference_fournisseur=l.reference_fournisseur, designation=l.designation,
            numero_bl_lie=l.numero_bl or "", qte=qte, pu_ht=pu, montant_ht=montant,
        )

    if len(candidats) == 1:
        l = candidats[0]
        cache.consommer(chemin_pdf, [l])
        if l.montant_ht is not None and not _proche(l.montant_ht, ligne.montant, TOLERANCE_CENTIME):
            commentaires.append(f"montant PDF {l.montant_ht:.2f} ≠ Commandes {ligne.montant:.2f} (valeur Commandes conservée)")
            statistiques["montant_pdf_different"] += 1
        statistiques["migre"] += 1
        return [_piece_depuis(l, ligne.qte, ligne.pu, ligne.montant, "; ".join(commentaires) or None)]

    # Plusieurs lignes du PDF pour cette référence : une ligne Pièces par
    # ligne de document si, et seulement si, elles se somment exactement.
    somme_qte = sum(l.quantite_facturee for l in candidats)
    montants = [l.montant_ht for l in candidats]
    somme_montant = sum(m for m in montants if m is not None) if all(m is not None for m in montants) else None
    if _proche(somme_qte, ligne.qte, TOLERANCE_QTE) and somme_montant is not None and _proche(somme_montant, ligne.montant, TOLERANCE_CENTIME):
        cache.consommer(chemin_pdf, candidats)
        statistiques["migre"] += 1
        statistiques["eclatees_multi_bl"] += 1
        pieces = []
        for l in candidats:
            pieces.append(_piece_depuis(
                l, l.quantite_facturee, l.prix_unitaire_ht, round(l.montant_ht, 2),
                "; ".join([f"ligne Commandes éclatée en {len(candidats)} BL"] + commentaires),
            ))
        # Les montants arrondis se somment au centime près à la valeur
        # Commandes (garanti par la tolérance) — rien à ajuster.
        return pieces

    exacts = [l for l in candidats if l.montant_ht is not None and _proche(l.montant_ht, ligne.montant, TOLERANCE_CENTIME)
              and _proche(l.quantite_facturee, ligne.qte, TOLERANCE_QTE)]
    if len(exacts) == 1:
        l = exacts[0]
        cache.consommer(chemin_pdf, [l])
        statistiques["migre"] += 1
        return [_piece_depuis(l, ligne.qte, ligne.pu, ligne.montant, "; ".join(commentaires) or None)]

    statistiques["ligne_ambigue"] += 1
    return [_piece_base(
        ligne, facture, chemin_pdf, MODE_MIGRE_SANS_PDF,
        "; ".join([
            f"{len(candidats)} lignes candidates dans {chemin_pdf.name} (qté {somme_qte:g}, montants non concordants) — champs PDF non repris"
        ] + commentaires),
    )]


def _sommes_par_fournisseur(lignes: list) -> dict:
    sommes = defaultdict(float)
    for l in lignes:
        sommes[l.fournisseur_suivi] += l.montant
    return {f: round(s, 2) for f, s in sommes.items()}


def _sommes_pieces_par_fournisseur(pieces: list) -> dict:
    sommes = defaultdict(float)
    for p in pieces:
        if p.get("Type") == TYPE_FACTURE:
            sommes[nom_fournisseur_suivi(p.get("Fournisseur"))] += float(p.get("Montant HT") or 0)
    return {f: round(s, 2) for f, s in sommes.items()}


def comparer_sommes(commandes: dict, pieces: dict) -> tuple:
    """(au_centime: bool, détail [(fournisseur, Σ Commandes, Σ Pièces, écart)])."""
    detail = []
    ok = True
    for f in sorted(set(commandes) | set(pieces)):
        a, b = commandes.get(f, 0.0), pieces.get(f, 0.0)
        ecart = round(b - a, 2)
        detail.append((f, a, b, ecart))
        if abs(a - b) > TOLERANCE_CENTIME:
            ok = False
    return ok, detail


def controler_sommes(fichier) -> dict:
    """Contrôle indépendant, relu depuis le classeur : Σ « Montant facturé
    HT » de Commandes (valeurs en cache) vs Σ Pieces[Montant HT] type
    Facture, par fournisseur (nom Suivi)."""

    lignes = lire_lignes_facturees(fichier)
    pieces = lire_pieces(fichier) or []
    ok, detail = comparer_sommes(_sommes_par_fournisseur(lignes), _sommes_pieces_par_fournisseur(pieces))
    return {"au_centime": ok, "detail": detail, "lignes_commandes": len(lignes), "pieces_facture": sum(1 for p in pieces if p.get("Type") == TYPE_FACTURE)}


def migrer_factures_vers_pieces(fichier, dossier_projet, dossier_backups, limite=None, ecrire=True,
                                journal=None) -> dict:
    """Voir bandeau. `limite` : n premières lignes facturées (extrait de
    test). `ecrire=False` : tout est construit et contrôlé, rien n'est écrit.
    Retourne le résumé (compteurs, sommes, au_centime, chemin_rapport,
    pieces construites)."""

    fichier = Path(fichier)
    dossier_projet = Path(dossier_projet)
    journal = journal or (lambda *a, **k: None)

    lignes = lire_lignes_facturees(fichier, limite)
    journal(f"{len(lignes)} ligne(s) Commandes avec N° facture à migrer")

    existantes = lire_pieces(fichier)
    if existantes is None:
        raise ValueError("La feuille Pièces n'existe pas encore — étape 1 (installer_feuille_pieces) d'abord.")
    index = IndexPieces(existantes)

    cache = _CacheFactures()
    statistiques = defaultdict(int)
    pieces = []
    for k, ligne in enumerate(lignes, start=1):
        pieces.extend(_migrer_ligne(ligne, dossier_projet, cache, statistiques))
        if k % 100 == 0:
            journal(f"  … {k}/{len(lignes)} lignes traitées, {len(cache.factures)} PDF lus")
    dedoublonner_ids(pieces)

    sommes_commandes = _sommes_par_fournisseur(lignes)
    sommes_pieces = _sommes_pieces_par_fournisseur(pieces)
    au_centime, detail = comparer_sommes(sommes_commandes, sommes_pieces)

    resume = {
        "lignes_commandes": len(lignes),
        "pieces_construites": len(pieces),
        "statistiques": dict(statistiques),
        "pdf_lus": len(cache.factures),
        "pdf_en_erreur": {str(p): r for p, r in cache.erreurs.items()},
        "sommes": detail,
        "au_centime_avant_ecriture": au_centime,
        "au_centime": None,
        "ecriture": None,
        "controle_apres": None,
        "chemin_rapport": None,
    }

    if ecrire:
        if not au_centime:
            raise ValueError("Contrôle au centime en échec AVANT écriture — rien n'a été écrit (voir résumé).")
        resume["ecriture"] = ecrire_pieces(fichier, pieces, dossier_backups)
        journal(f"{resume['ecriture']['ajoutees']} ligne(s) Pièces écrite(s), {len(resume['ecriture']['ignorees'])} déjà présente(s)")
        resume["controle_apres"] = controler_sommes(fichier)
        resume["au_centime"] = resume["controle_apres"]["au_centime"]
    else:
        resume["au_centime"] = au_centime

    resume["chemin_rapport"] = ecrire_rapport_migration(dossier_projet / DOSSIER_RAPPORTS, resume, pieces)
    resume["pieces"] = pieces
    return resume


def ecrire_rapport_migration(dossier_rapports, resume: dict, pieces: list) -> Path:
    dossier_rapports = Path(dossier_rapports)
    dossier_rapports.mkdir(parents=True, exist_ok=True)
    horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
    chemin = dossier_rapports / f"migration_pieces_{horodatage}.txt"

    st = resume["statistiques"]
    lignes = [
        f"Migration Commandes -> Pièces — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"{resume['lignes_commandes']} ligne(s) Commandes avec N° facture lue(s), {resume['pieces_construites']} ligne(s) Pièces construite(s).",
        f"  - Migré (ligne retrouvée dans le PDF) : {st.get('migre', 0)}, dont {st.get('eclatees_multi_bl', 0)} éclatée(s) en plusieurs BL",
        f"  - Migré sans PDF : PDF introuvable {st.get('pdf_introuvable', 0)}, n° absent du PDF {st.get('numero_absent_du_pdf', 0)}, "
        f"ligne non retrouvée {st.get('ligne_non_retrouvee', 0)}, ligne ambiguë {st.get('ligne_ambigue', 0)}",
        f"  - montant PDF ≠ Commandes (valeur Commandes conservée, commentée) : {st.get('montant_pdf_different', 0)}",
        f"{resume['pdf_lus']} PDF lu(s), {len(resume['pdf_en_erreur'])} en erreur de lecture.",
        "Contrôle au centime par fournisseur (Σ Montant facturé HT Commandes vs Σ Pieces[Montant HT] Facture) :",
    ] + [
        f"  - {f} : Commandes {a:.2f} € / Pièces {b:.2f} € / écart {e:+.2f} €"
        for f, a, b, e in resume["sommes"]
    ] + [
        f"Au centime avant écriture : {'OUI' if resume['au_centime_avant_ecriture'] else 'NON'}",
    ]
    if resume["ecriture"] is not None:
        e = resume["ecriture"]
        lignes.append(f"Écriture : {e['ajoutees']} ajoutée(s), {len(e['ignorees'])} ignorée(s) (ID déjà présent), lignes {e['premiere_ligne']}-{e['derniere_ligne']}, sauvegarde {e['sauvegarde']}")
    if resume["controle_apres"] is not None:
        c = resume["controle_apres"]
        lignes.append(f"Contrôle après écriture (relu du classeur) : {'OUI' if c['au_centime'] else 'NON'} — "
                      f"{c['lignes_commandes']} lignes Commandes, {c['pieces_facture']} pièces Facture")
        lignes += [f"  - {f} : Commandes {a:.2f} € / Pièces {b:.2f} € / écart {e:+.2f} €" for f, a, b, e in c["detail"]]
    if resume["pdf_en_erreur"]:
        lignes.append("PDF en erreur de lecture :")
        lignes += [f"  - {p} : {'; '.join(r)}" for p, r in resume["pdf_en_erreur"].items()]
    sans_pdf = [p for p in pieces if p.get("Mode de rapprochement") == MODE_MIGRE_SANS_PDF]
    if sans_pdf:
        lignes.append(f"Lignes « Migré sans PDF » ({len(sans_pdf)}) :")
        lignes += [
            f"  - {p['Fournisseur']} facture {p['N° pièce']} cde {p['N° de commande']} réf {p['Référence Suivi']} : {p['Commentaire']}"
            for p in sans_pdf
        ]
    commentees = [p for p in pieces if p.get("Mode de rapprochement") == MODE_MIGRE and p.get("Commentaire")]
    if commentees:
        lignes.append(f"Lignes « Migré » avec commentaire ({len(commentees)}) :")
        lignes += [
            f"  - {p['Fournisseur']} facture {p['N° pièce']} cde {p['N° de commande']} réf {p['Référence Suivi']} : {p['Commentaire']}"
            for p in commentees
        ]
    chemin.write_text("\n".join(lignes), encoding="utf-8")
    return chemin
