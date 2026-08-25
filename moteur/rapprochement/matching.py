"""
Rapprochement d'un bon de livraison déjà lu (voir lecture_bl.py) avec les
lignes de la commande correspondante dans le Suivi commandes — étape 2 de
Rapprochement AI (session R2, voir CLAUDE.md, "Rapprochement AI").

Ordre de confiance retenu (cadrage R1) :
  a) n° de commande lu sur le BL -> lignes de cette commande dans le Suivi ;
  b) et c) (fournisseur+références+dates, puis bac "inconnu") — pas encore
     nécessaires ici : les BL 109 Distribution testés portent TOUJOURS un
     n° de commande exploitable (voir tableau de flux R1, CLAUDE.md).

Une ligne de BL n'est JAMAIS appariée "au hasard" : si la référence
normalisée (voir moteur.base.coeur_numerique, déjà utilisé pour le même
problème côté devis) ne désigne pas EXACTEMENT une ligne du Suivi pour cette
commande, la ligne part au bac "inconnu" plutôt que de deviner.

Référentiel articles (`referentiel`, optionnel — voir moteur/referentiel.py) :
demande explicite de l'acheteur ("il faut créer une base des équivalences,
ce genre de cas va se présenter très souvent", cas réel 59210/CFF1BIS —
remplacement fournisseur d'un article générique par un autre du même
fabricant/famille, texte et cœur numérique totalement différents, aucun
repli ci-dessus ne peut le deviner). Réutilise TEL QUEL le référentiel déjà
construit et éprouvé côté devis (moteur/comparateur.py, même philosophie :
un alias "confirmé" une fois fait autorité pour toujours, une simple
"proposition" ne fusionne JAMAIS automatiquement) plutôt que d'en bâtir un
second en parallèle. Voir `_memes_references()` (équivalent confirmé -> OR
avec la comparaison par cœur numérique existante, jamais un remplacement —
voir son bandeau, une régression réelle sinon) et `_repli_referentiel()`
(candidat proposé mais pas confirmé -> "à confirmer", enregistré par
`referentiel.resoudre()` pour `referentiel/A_confirmer_BL.xlsx`, voir
moteur/rapprochement/pipeline_bl.py). `referentiel=None` (défaut) désactive
entièrement ce mécanisme — comportement inchangé pour tout appelant
existant.
"""

import difflib
import re
import unicodedata
from dataclasses import dataclass, field
from enum import Enum

from openpyxl import load_workbook

from moteur.base import coeur_numerique
from moteur.outils import to_float
from moteur.panier import MAPPING_FOURNISSEURS
from moteur.rapprochement.modele_bl import LigneBL

FEUILLE_COMMANDES = "Commandes"

_COLONNES_REQUISES = (
    "Référence", "Désignation", "Qté commandée", "N° de commande",
    "Fournisseur", "Date de livraison", "Qté livrée", "Tarif BL",
    "Statut commande", "Note",
)


@dataclass
class LigneSuivi:

    ligne_excel: int  # numéro de ligne réel dans la feuille (1-based, pour l'écriture)

    reference: str
    designation: str
    qte_commandee: float
    qte_livree: float
    tarif_bl: float | None
    date_livraison: object
    statut: str
    note: str
    # Renseigné uniquement par lire_lignes_fournisseur() (déduction de
    # commande par contenu, voir deduire_commande_par_contenu()) — vide
    # sinon, lire_lignes_commande() connaît déjà la commande par son
    # paramètre.
    numero_commande: str = ""


class Statut(Enum):
    SUR = "sûr"                # à écrire après OK global
    A_CONFIRMER = "à confirmer"  # écart, question par question
    DEJA_A_JOUR = "déjà à jour"  # idempotence : rien à écrire
    INCONNU = "inconnu"        # aucune ligne Suivi correspondante sûre


@dataclass
class Correspondance:

    ligne_bl: LigneBL
    ligne_suivi: LigneSuivi | None
    statut: Statut
    raisons: list[str] = field(default_factory=list)
    # True quand la quantité du BL est DÉJÀ comptée dans qte_livree du
    # Suivi (ex. correction de date seule sur une ligne par ailleurs déjà
    # à jour, voir _comparer()) — qte_livree_cumulee ne doit alors JAMAIS
    # recumuler, sous peine de doubler une quantité déjà exacte (bug réel
    # rencontré, voir CLAUDE.md "sur-livraison fantôme").
    qte_deja_incluse: bool = False

    @property
    def qte_livree_cumulee(self) -> float:
        deja = self.ligne_suivi.qte_livree if self.ligne_suivi else 0.0
        if self.qte_deja_incluse:
            return deja
        return round(deja + self.ligne_bl.quantite_livree, 4)


def _cle(ref) -> str:
    """Référence -> clé comparable, tolérante aux préfixes/zéros de tête
    (ex. "086101L" vs "86101L", "EUR52302" vs "52302" — écarts réels
    observés entre le Suivi et l'OCR du BL, session R2). Réutilise
    moteur.base.coeur_numerique (même outil que pour le rapprochement des
    devis) ; repli sur le texte alphanumérique brut si la référence n'a pas
    assez de chiffres pour un cœur numérique fiable (seuil 4 chiffres)."""

    coeur = coeur_numerique(ref)
    if coeur:
        return coeur
    return re.sub(r"[^A-Z0-9]", "", str(ref or "").upper())


def _memes_references(ref_a, ref_b, referentiel, designation_a: str = "", designation_b: str = "",
                       fournisseur: str = "", devis: str = "") -> bool:
    """`_cle(ref_a) == _cle(ref_b)` (comme avant), OU — si un référentiel
    est fourni — les deux références ont un alias CONFIRMÉ (statut "connu")
    vers la MÊME clé. Un OR, jamais un remplacement de `_cle()` : le
    référentiel ajoute des correspondances, il n'en retire jamais.

    BUG RÉEL ÉVITÉ (recette réelle, "LEG069831L" vs "069831L", commande
    139.112) : une 1re version comparait EXCLUSIVEMENT la clé référentiel
    dès qu'elle était disponible ("connu") — mais un alias "connu" peut
    être un simple AUTO-alias d'une référence vers ELLE-MÊME (ex. la BDD
    achats retient parfois la forme préfixée "LEG069831L" comme Clé_Réf
    telle quelle, sans la réduire à "069831L") : les deux références
    ressortaient alors "connu" mais vers des clés DIFFÉRENTES ("LEG069831L"
    et "069831L" chacune vers elle-même), un résultat STRICTEMENT PIRE que
    la comparaison par cœur numérique déjà en place (qui, elle, les
    retombait toutes les deux sur "69831"). D'où le OR : le référentiel ne
    fait jamais perdre une correspondance déjà trouvée par `_cle()`."""

    if _cle(ref_a) == _cle(ref_b):
        return True

    if referentiel is None:
        return False

    cle_a, statut_a = referentiel.resoudre(
        ref_a, designation=designation_a, fournisseur=fournisseur, devis=devis,
    )
    cle_b, statut_b = referentiel.resoudre(
        ref_b, designation=designation_b, fournisseur=fournisseur, devis=devis,
    )
    return statut_a == "connu" and statut_b == "connu" and cle_a == cle_b


def _repli_referentiel(ligne_bl: LigneBL, disponibles: list, referentiel,
                        fournisseur: str = "", devis: str = "") -> tuple:
    """Repli référentiel articles — 3e chance quand ni une correspondance
    exacte (voir `_memes_references()`) ni `_repli_reference_proche()` n'ont abouti :
    cas réel signalé par l'acheteur, un fournisseur remplace un article par
    un autre du même fabricant/de la même famille, avec une référence sans
    AUCUN rapport textuel ni numérique (ex. BL "59210" pour la ligne Suivi
    "CFF1BIS", même désignation "colliers à embase 16/32 boîte de 100",
    même tarif — ou "092897"/"411651", même interrupteur différentiel
    Legrand sous sa référence grand public plutôt que professionnelle).
    Consulte `referentiel.resoudre()` : si la référence du BL a un candidat
    PROPOSÉ (pas encore confirmé) dont la clé correspond à EXACTEMENT une
    ligne du Suivi encore disponible, la retient — toujours "à confirmer"
    (voir apparier()), jamais un choix au hasard. La proposition est
    automatiquement mémorisée par `resoudre()` pour
    `referentiel/A_confirmer_BL.xlsx` : une fois confirmée par l'acheteur,
    la MÊME correspondance deviendra une correspondance exacte via
    `_memes_references()` dès le prochain passage, pour tous les futurs
    BL — plus besoin de la reconfirmer à chaque fois."""

    if referentiel is None:
        return None, ""

    cle_bl, statut_bl = referentiel.resoudre(
        ligne_bl.reference_fournisseur, designation=ligne_bl.designation,
        fournisseur=fournisseur, devis=devis,
    )
    if not cle_bl or statut_bl != "propose":
        return None, ""

    candidats = [
        ls for ls in disponibles
        if referentiel.resoudre(str(ls.reference), designation=ls.designation)[0] == cle_bl
    ]

    if len(candidats) != 1:
        return None, ""

    ls = candidats[0]
    raison = (
        f"Référentiel articles : « {ligne_bl.reference_fournisseur} » proposé comme équivalent "
        f"de « {ls.reference} » (désignation Suivi « {ls.designation} » vs BL "
        f"« {ligne_bl.designation} ») — à confirmer dans referentiel/A_confirmer_BL.xlsx "
        "pour que ce rapprochement soit reconnu automatiquement la prochaine fois"
    )
    return ls, raison


def _ouvrir_feuille_commandes(chemin_suivi):
    """Ouvre le classeur en lecture seule, valide les colonnes requises,
    retourne (wb, ws, entetes) — factorisé entre lire_lignes_commande() et
    lire_lignes_fournisseur() pour ne valider les colonnes qu'à un seul
    endroit."""

    wb = load_workbook(chemin_suivi, read_only=True, data_only=True)
    ws = wb[FEUILLE_COMMANDES]
    entetes = {
        c.value: i
        for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
    }

    manquantes = [c for c in _COLONNES_REQUISES if c not in entetes]
    if manquantes:
        wb.close()
        raise KeyError(
            f"Colonne(s) introuvable(s) dans « {FEUILLE_COMMANDES} » : "
            f"{', '.join(manquantes)} — export du Suivi commandes différent de celui attendu ?"
        )

    return wb, ws, entetes


def _ligne_suivi_depuis_row(i, row, entetes, avec_commande=False) -> LigneSuivi:
    return LigneSuivi(
        ligne_excel=i,
        reference=row[entetes["Référence"]],
        designation=row[entetes["Désignation"]],
        qte_commandee=to_float(row[entetes["Qté commandée"]]),
        qte_livree=to_float(row[entetes["Qté livrée"]]),
        tarif_bl=row[entetes["Tarif BL"]],
        date_livraison=row[entetes["Date de livraison"]],
        statut=row[entetes["Statut commande"]] or "",
        note=row[entetes["Note"]] or "",
        numero_commande=str(row[entetes["N° de commande"]] or "").strip() if avec_commande else "",
    )


def lire_lignes_commande(chemin_suivi, fournisseur: str, numero_commande: str) -> list[LigneSuivi]:
    """Toutes les lignes du Suivi pour ce fournisseur + ce n° de commande,
    quel que soit leur statut actuel (une ligne déjà "Reçue" doit pouvoir
    être signalée en cas de re-livraison inattendue, pas juste les lignes
    ouvertes).

    `fournisseur` est le nom DÉTECTEUR (moteur.detecteur, ex. "ELECTRIC
    PLUS"), pas toujours celui écrit dans la colonne "Fournisseur" du
    Suivi (ex. "GMR" — Electric Plus est l'enseigne grand public, GMR sa
    branche grands comptes pros, confirmé par l'acheteur, session R2
    suite) : converti via moteur.panier.MAPPING_FOURNISSEURS, déjà la
    source de vérité pour cette correspondance côté Panier — pas de
    mapping dupliqué ici."""

    fournisseur = MAPPING_FOURNISSEURS.get(fournisseur.upper(), fournisseur)

    wb, ws, entetes = _ouvrir_feuille_commandes(chemin_suivi)
    try:
        i_cde = entetes["N° de commande"]
        i_fourn = entetes["Fournisseur"]

        resultat = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            cde = row[i_cde]
            fourn = row[i_fourn]

            if cde is None or str(cde).strip() != numero_commande:
                continue
            if fourn is None or _cle(fourn) != _cle(fournisseur):
                continue

            resultat.append(_ligne_suivi_depuis_row(i, row, entetes))

        return resultat
    finally:
        wb.close()


def lire_lignes_fournisseur(chemin_suivi, fournisseur: str) -> list[LigneSuivi]:
    """Toutes les lignes du Suivi pour ce fournisseur, TOUTES commandes
    confondues (numero_commande renseigné sur chaque ligne, contrairement
    à lire_lignes_commande()) — utilisé UNIQUEMENT pour la déduction de
    commande par contenu (voir deduire_commande_par_contenu()) quand le
    n° de commande n'est pas lisible sur le BL (cas réel : trou de
    perforateur sur la zone du BL papier où il est imprimé)."""

    fournisseur = MAPPING_FOURNISSEURS.get(fournisseur.upper(), fournisseur)

    wb, ws, entetes = _ouvrir_feuille_commandes(chemin_suivi)
    try:
        i_fourn = entetes["Fournisseur"]

        resultat = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            fourn = row[i_fourn]
            if fourn is None or _cle(fourn) != _cle(fournisseur):
                continue

            resultat.append(_ligne_suivi_depuis_row(i, row, entetes, avec_commande=True))

        return resultat
    finally:
        wb.close()


def deduire_commande_par_contenu(lignes_bl: list[LigneBL], lignes_suivi_fournisseur: list[LigneSuivi]) -> tuple:
    """Quand le n° de commande n'est pas lisible sur le BL (cas réel signalé
    par l'acheteur : les gars de l'atelier perforent les BL pour les
    classer dans un classeur, ce qui abîme parfois la zone où le n° de
    commande est imprimé), tente de le DÉDUIRE en cherchant, parmi TOUTES
    les lignes du Suivi pour ce fournisseur (lire_lignes_fournisseur), la
    commande dont le plus de lignes correspondent EXACTEMENT (référence ET
    quantité commandée) à celles du BL — une signature de contenu plutôt
    qu'un numéro illisible.

    N'accepte que si :
    - au moins 2 lignes du BL matchent la même commande (jamais une
      coïncidence sur une seule référence, trop répandue pour être fiable
      seule — ex. une référence de câble ou de connecteur générique
      réutilisée dans des dizaines de commandes différentes) ;
    - la meilleure commande a un score STRICTEMENT supérieur à toutes les
      autres (aucune égalité, jamais un choix au hasard entre 2 candidats
      aussi plausibles l'un que l'autre).

    Retourne (numero_commande, nb_lignes_matchées) ou (None, 0) si aucune
    déduction fiable. Un numéro DÉDUIT n'est jamais utilisé pour un
    rapprochement "sûr" — voir rapprocher_dossier(), toute correspondance
    obtenue via ce numéro reste "à confirmer", quel que soit le résultat de
    _comparer()."""

    scores = {}

    for idx, lbl in enumerate(lignes_bl):
        coeur_bl = coeur_numerique(lbl.reference_fournisseur)
        if not coeur_bl:
            continue
        for ls in lignes_suivi_fournisseur:
            if coeur_numerique(ls.reference) != coeur_bl:
                continue
            if ls.qte_commandee is None or abs(ls.qte_commandee - lbl.quantite_livree) > 0.001:
                continue
            scores.setdefault(ls.numero_commande, set()).add(idx)
            break

    if not scores:
        return None, 0

    classement = sorted(scores.items(), key=lambda kv: -len(kv[1]))
    meilleure_commande, meilleures_lignes = classement[0]
    meilleur_score = len(meilleures_lignes)

    if meilleur_score < 2:
        return None, 0

    if len(classement) > 1 and len(classement[1][1]) == meilleur_score:
        return None, 0

    return meilleure_commande, meilleur_score


def _comparer(ligne_bl: LigneBL, ligne_suivi: LigneSuivi, date_bl_reelle=None) -> Correspondance:
    """`date_bl_reelle` (objet date, ou None si non fournie/illisible) est
    la vraie date du BL, déjà extraite par l'appelant (voir apparier()) —
    utilisée pour vérifier que la date ENREGISTRÉE dans le Suivi est la
    bonne, pas seulement qu'une date QUELCONQUE est présente (voir bug
    ci-dessous)."""

    raisons = []

    deja_qte = ligne_suivi.qte_livree
    prix_bl = ligne_bl.prix_net or 0.0
    prix_bl_connu = bool(ligne_bl.prix_net)  # certains fournisseurs (Coredime...)
    # n'affichent JAMAIS de prix sur le BL (réglé à la facture, voir
    # moteur/fournisseurs/coredime.py) — la comparaison de tarif doit alors
    # être ignorée plutôt que bloquer l'idempotence indéfiniment.
    tarif_suivi = to_float(ligne_suivi.tarif_bl) if ligne_suivi.tarif_bl is not None else None

    qte_et_tarif_identiques = (
        abs(deja_qte - ligne_bl.quantite_livree) < 0.001
        and (not prix_bl_connu or (tarif_suivi is not None and abs(tarif_suivi - prix_bl) < 0.01))
    )

    date_suivi_brute = ligne_suivi.date_livraison
    date_suivi = date_suivi_brute.date() if hasattr(date_suivi_brute, "date") else date_suivi_brute
    date_coherente = date_bl_reelle is None or date_suivi is None or date_suivi == date_bl_reelle

    # Idempotence : cette même quantité, un tarif cohérent (ou pas de tarif
    # à comparer pour ce fournisseur) et LA BONNE date de livraison sont
    # DÉJÀ enregistrés -> ce BL (ou un doublon) a déjà été traité, on
    # n'écrit rien deux fois (voir CLAUDE.md, "double traitement du même
    # BL : détecté, refusé").
    deja_a_jour = qte_et_tarif_identiques and date_suivi_brute is not None and date_coherente

    if deja_a_jour:
        return Correspondance(ligne_bl, ligne_suivi, Statut.DEJA_A_JOUR)

    # BUG RÉEL CORRIGÉ (recette réelle, voir CLAUDE.md "déjà à jour ne
    # vérifie jamais que la date enregistrée est la bonne") : des lignes
    # écrites avant que l'extraction de date fonctionne (repli sur
    # date.today()) restaient bloquées pour toujours avec la mauvaise
    # date, "déjà à jour" ne revérifiant que qté/tarif. Ici, qté/tarif
    # sont déjà exacts mais la date ne l'est pas -> signalé "à confirmer"
    # pour correction, `qte_deja_incluse=True` pour ne JAMAIS recumuler
    # une quantité déjà correcte (seule la date doit changer).
    if qte_et_tarif_identiques and date_suivi_brute is not None and not date_coherente:
        raisons.append(
            f"Date de livraison enregistrée ({date_suivi}) différente de la date du BL "
            f"({date_bl_reelle}) — quantité/tarif déjà corrects, seule la date semble fausse"
        )
        return Correspondance(ligne_bl, ligne_suivi, Statut.A_CONFIRMER, raisons, qte_deja_incluse=True)

    nouvelle_qte = round(deja_qte + ligne_bl.quantite_livree, 4)

    if nouvelle_qte > ligne_suivi.qte_commandee + 0.001:
        raisons.append(
            f"Sur-livraison : {nouvelle_qte:g} livrée(s) au total pour "
            f"{ligne_suivi.qte_commandee:g} commandée(s)"
        )

    if tarif_suivi and prix_bl and abs(tarif_suivi - prix_bl) > 0.01:
        raisons.append(f"Tarif différent : Suivi {tarif_suivi:g}€ / BL {prix_bl:g}€")

    statut = Statut.A_CONFIRMER if raisons else Statut.SUR

    return Correspondance(ligne_bl, ligne_suivi, statut, raisons)


def _cle_designation(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def _distance_courte(a: str, b: str) -> int:
    """Nombre de caractères différents entre deux chaînes de même
    longueur ; pénalise aussi un écart de longueur. Suffisant pour repérer
    UN caractère abîmé (trou de perforateur, tache...), pas une vraie
    distance de Levenshtein — pas besoin de plus pour ce cas d'usage."""

    ecart_longueur = abs(len(a) - len(b))
    difference_alignee = sum(1 for x, y in zip(a, b) if x != y)
    return ecart_longueur + difference_alignee


def _chiffres_tete_manquants(a: str, b: str) -> bool:
    """True si l'une des deux chaînes est un SUFFIXE de l'autre, avec un
    écart de 1 ou 2 chiffres de tête — cas réel signalé par l'acheteur : un
    trou de perforateur peut carrément EFFACER un ou deux chiffres du DÉBUT
    de la référence (pas juste l'abîmer/le substituer, déjà couvert par
    _distance_courte) — ex. BL "9894" pour la vraie référence Suivi
    "69894" (chiffre de tête "6" entièrement disparu)."""

    plus_court, plus_long = (a, b) if len(a) <= len(b) else (b, a)
    ecart = len(plus_long) - len(plus_court)
    return 1 <= ecart <= 2 and plus_long.endswith(plus_court)


def _cle_brute(ref) -> str:
    """Référence -> texte alphanumérique brut (garde les LETTRES,
    contrairement à coeur_numerique) — utilisé en repli quand le cœur
    numérique est trop court (< 4 chiffres) pour être comparé de façon
    fiable, cas réel : "H07VK16BL" (BL) vs "H07VK16B" (Suivi), un seul
    caractère d'écart en fin de référence, cœur numérique "716" des deux
    côtés (sous le seuil de 4 chiffres)."""

    return re.sub(r"[^A-Z0-9]", "", str(ref or "").upper())


def _cle_normalisee_i_un(ref) -> str:
    """Texte alphanumérique brut (voir _cle_brute) avec le chiffre "1" et
    la lettre majuscule "I" normalisés vers le même caractère — cas réel
    signalé par l'acheteur : "XVR111STI" (Suivi, "WittyOne 11kW") lu
    "XVR1IISTI" par l'OCR, DEUX des trois "1" confondus avec des "I" dans
    la même référence. Comparaison par ÉGALITÉ après normalisation (pas
    une distance à 1 caractère près comme _distance_courte) : plusieurs
    confusions 1/I peuvent survenir dans la même référence, une seule ne
    suffirait pas à les capturer toutes."""

    return _cle_brute(ref).replace("I", "1")


def _repli_reference_proche(ligne_bl: LigneBL, disponibles: list) -> tuple:
    """2e chance quand AUCUNE référence exacte ne correspond (voir
    apparier()) : cherche une ligne du Suivi dont la référence est proche
    de celle du BL — cas réel signalé par l'acheteur : les gars de
    l'atelier perforent les BL pour les classer, ce qui abîme parfois un
    chiffre de la référence imprimée (ex. OCR "L405205" pour la vraie
    référence "405209", un seul chiffre différent — _distance_courte) ou
    en efface complètement un ou deux au début (ex. "9894" pour la vraie
    référence "69894" — _chiffres_tete_manquants). Ces deux critères
    comparent le CŒUR NUMÉRIQUE ; quand celui-ci est trop court pour être
    fiable des deux côtés (< 4 chiffres, ex. "H07VK16BL"/"H07VK16B" — cœur
    "716" identique mais sous le seuil), repli sur le texte alphanumérique
    brut (_cle_brute, garde les lettres) avec le même critère 1 caractère
    d'écart. Un 4e critère, indépendant du cœur numérique, cherche une
    confusion OCR "1"/"I" (voir _cle_normalisee_i_un — ex. "XVR111STI" lu
    "XVR1IISTI", deux confusions dans la même référence). N'accepte que
    s'il existe EXACTEMENT UNE ligne du Suivi correspondant à l'un de ces
    critères pour cette commande — jamais un choix au hasard. Retourne
    (ligne_suivi, raison) ou (None, "") si aucun repli fiable. Le statut
    reste TOUJOURS "à confirmer" (voir apparier()) : un rapprochement de
    repli n'est jamais écrit automatiquement, la désignation est donnée en
    clair dans la raison pour que l'acheteur puisse trancher d'un coup
    d'œil."""

    coeur_bl = coeur_numerique(ligne_bl.reference_fournisseur)
    brute_bl = _cle_brute(ligne_bl.reference_fournisseur)
    norm_i_un_bl = _cle_normalisee_i_un(ligne_bl.reference_fournisseur)

    candidats = []
    for ls in disponibles:
        coeur_ls = coeur_numerique(ls.reference)

        if coeur_bl and coeur_ls and _distance_courte(coeur_bl, coeur_ls) == 1:
            candidats.append((ls, "1 caractère d'écart — trou de perforateur ?"))
        elif coeur_bl and coeur_ls and _chiffres_tete_manquants(coeur_bl, coeur_ls):
            candidats.append((ls, "chiffre(s) de tête manquant(s) — trou de perforateur sur le début de la référence ?"))
        elif (
            norm_i_un_bl and norm_i_un_bl == _cle_normalisee_i_un(ls.reference)
            and brute_bl != _cle_brute(ls.reference)
        ):
            candidats.append((ls, 'confusion OCR "1"/"I" (chiffre "1" lu comme la lettre "I") — même référence une fois normalisée'))
        elif not coeur_bl or not coeur_ls:
            brute_ls = _cle_brute(ls.reference)
            if brute_bl and brute_ls and _distance_courte(brute_bl, brute_ls) == 1:
                candidats.append((ls, "1 caractère d'écart, référence trop courte pour un cœur numérique fiable — trou de perforateur ?"))

    if len(candidats) != 1:
        return None, ""

    ls, raison_ecart = candidats[0]
    similarite = round(
        difflib.SequenceMatcher(
            None, _cle_designation(ligne_bl.designation), _cle_designation(ls.designation),
        ).ratio() * 100
    )
    raison = (
        f"Référence proche mais pas identique (BL « {ligne_bl.reference_fournisseur} » "
        f"vs Suivi « {ls.reference} », {raison_ecart}) : "
        f"désignation Suivi « {ls.designation} » vs BL « {ligne_bl.designation} » "
        f"(similarité ~{similarite}%) — à vérifier"
    )
    return ls, raison


def apparier(lignes_bl: list[LigneBL], lignes_suivi: list[LigneSuivi], date_bl_reelle=None,
             referentiel=None, fournisseur: str = "", devis: str = "") -> list[Correspondance]:
    """Associe chaque ligne du BL à AU PLUS une ligne du Suivi (une ligne du
    Suivi n'est utilisée qu'une seule fois). Ambigu (0 ou plusieurs
    candidats) -> bac "inconnu", jamais de choix au hasard.

    `date_bl_reelle` (objet date, ou None) est la vraie date du BL, déjà
    extraite par l'appelant — voir _comparer(), utilisée pour détecter une
    "Date de livraison" du Suivi incohérente avec le vrai document.

    `referentiel` (moteur.referentiel.Referentiel, ou None — voir bandeau
    du module) : quand fourni, un alias CONFIRMÉ compte comme une
    correspondance exacte (`_memes_references()`, dès la 1re passe, en
    OR avec la comparaison par cœur numérique déjà en place — jamais un
    remplacement) ; une simple proposition non confirmée est essayée en
    3e repli (`_repli_referentiel()`, toujours "à confirmer").
    `fournisseur`/`devis` n'enrichissent que
    `referentiel/A_confirmer_BL.xlsx` (pour que l'acheteur sache de quel BL
    vient une proposition), aucun effet sur le résultat du rapprochement.

    DEUX PASSES, jamais une seule (BUG RÉEL CORRIGÉ — cas réel, commande
    130.036, lignes "L600001"/"L600002") : une seule passe, ligne de BL
    après ligne de BL dans l'ordre du document, faisait dépendre le
    résultat de l'ORDRE des lignes plutôt que de leur contenu — "L600001"
    (aucun exact, mais à 1 caractère de "600002" côté Suivi) traité AVANT
    "L600002" (exact) consommait la ligne Suivi "600002" par repli, privant
    ensuite "L600002" — pourtant une correspondance EXACTE — de toute
    ligne à apparier (ressorti "inconnu" à tort, alors qu'une correspondance
    exacte existait bel et bien). Un repli approximatif ne doit JAMAIS
    pouvoir voler une ligne Suivi à une autre ligne de BL qui la matche
    EXACTEMENT, quel que soit l'ordre d'apparition sur le document — la
    1re passe résout donc TOUTES les correspondances exactes (candidat
    unique) sur TOUTES les lignes de BL avant qu'aucun repli ne soit
    tenté ; la 2e passe traite ce qui reste (0 ou plusieurs candidats
    exacts) avec les `disponibles` déjà réduits en conséquence."""

    disponibles = list(lignes_suivi)
    resultat_par_index = {}

    # 1re passe : uniquement les correspondances EXACTES (candidat unique),
    # sur l'ensemble des lignes de BL, avant tout repli.
    for i, lbl in enumerate(lignes_bl):
        candidats = [
            ls for ls in disponibles
            if _memes_references(ls.reference, lbl.reference_fournisseur, referentiel,
                                  ls.designation, lbl.designation, fournisseur, devis)
        ]
        if len(candidats) == 1:
            ls = candidats[0]
            disponibles.remove(ls)
            resultat_par_index[i] = _comparer(lbl, ls, date_bl_reelle)

    # 2e passe : ce qui reste (0 ou plusieurs candidats exacts au moment de
    # la 1re passe) — repli / ambiguïté, sur les `disponibles` déjà réduits.
    for i, lbl in enumerate(lignes_bl):

        if i in resultat_par_index:
            continue

        candidats = [
            ls for ls in disponibles
            if _memes_references(ls.reference, lbl.reference_fournisseur, referentiel,
                                  ls.designation, lbl.designation, fournisseur, devis)
        ]

        if not candidats:
            ls_repli, raison_repli = _repli_reference_proche(lbl, disponibles)
            if ls_repli is None:
                ls_repli, raison_repli = _repli_referentiel(lbl, disponibles, referentiel, fournisseur, devis)
            if ls_repli is not None:
                disponibles.remove(ls_repli)
                c = _comparer(lbl, ls_repli, date_bl_reelle)
                if c.statut is Statut.DEJA_A_JOUR:
                    # BUG RÉEL CORRIGÉ (recette réelle, écriture sur le vrai
                    # Suivi) : forcer "à confirmer" ICI aussi aurait fait
                    # calculer qte_livree_cumulee = déjà_qte + quantité_bl
                    # SANS savoir que _comparer() a déjà déterminé que ces
                    # deux quantités sont LES MÊMES (donc rien à cumuler) —
                    # un repli vers une ligne 405209 déjà à 3/3 livrée s'est
                    # ainsi retrouvé écrit à tort 3+3=6/3 (sur-livraison
                    # fantôme). Si le repli pointe vers une ligne DÉJÀ à
                    # jour, aucune écriture n'est proposée, jamais — la
                    # raison du repli est gardée pour que l'acheteur voie
                    # pourquoi ce rapprochement approximatif ne déclenche
                    # rien.
                    resultat_par_index[i] = Correspondance(
                        c.ligne_bl, c.ligne_suivi, Statut.DEJA_A_JOUR, [raison_repli],
                    )
                else:
                    resultat_par_index[i] = Correspondance(
                        c.ligne_bl, c.ligne_suivi, Statut.A_CONFIRMER, [raison_repli] + c.raisons,
                        qte_deja_incluse=c.qte_deja_incluse,
                    )
            else:
                resultat_par_index[i] = Correspondance(
                    lbl, None, Statut.INCONNU,
                    ["Aucune ligne du Suivi ne correspond à cette référence pour cette commande"],
                )
        else:
            # BUG RÉEL CORRIGÉ (cas réel, commande M3.10.175) : deux
            # références DIFFÉRENTES peuvent avoir le MÊME cœur numérique
            # (ex. "R9PRC263" et "R9PRA263" -> "9263" toutes les deux, la
            # lettre médiane C/A n'étant pas un chiffre) — une
            # correspondance de TEXTE EXACT parmi les candidats ambigus
            # est toujours plus fiable qu'une simple coïncidence de cœur
            # numérique, donc toujours préférée avant de renoncer.
            exacts = [
                ls for ls in candidats
                if str(ls.reference).strip().upper() == lbl.reference_fournisseur.strip().upper()
            ]
            if len(exacts) == 1:
                ls = exacts[0]
                disponibles.remove(ls)
                resultat_par_index[i] = _comparer(lbl, ls, date_bl_reelle)
            else:
                resultat_par_index[i] = Correspondance(
                    lbl, None, Statut.INCONNU,
                    [f"{len(candidats)} lignes du Suivi correspondent à cette référence (ambigu)"],
                )

    return [resultat_par_index[i] for i in range(len(lignes_bl))]
