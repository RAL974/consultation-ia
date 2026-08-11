"""
Assistant de contrôle de la base d'équivalences.

Produit resultats/Audit_BDD.xlsx avec :

1. Groupes suspects   : codes internes / groupes regroupant des désignations
                        trop hétérogènes (signe d'un mauvais rapprochement).
2. Couverture         : par catégorie, combien d'articles ont un Code interne
                        et combien n'en ont pas (pour prioriser la saisie).
3. Sans code interne  : la liste des articles non codés, par catégorie.
4. Suggestions        : parmi les articles NON codés d'une même catégorie,
                        des groupes de désignations très proches — candidats
                        à un même Code interne, à VALIDER (jamais appliqués
                        automatiquement).

Aucune modification de la base : l'audit ne fait que lire et signaler.
"""

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from moteur.base import BaseArticles
from moteur.normalisation import code_interne_auto


GRIS = PatternFill(fill_type="solid", fgColor="EFEFEF")
ROUGE = PatternFill(fill_type="solid", fgColor="F4CCCC")
ORANGE = PatternFill(fill_type="solid", fgColor="FCE5CD")

# Mots vides ignorés pour comparer les désignations
_STOP = {
    "DE", "LA", "LE", "DU", "DES", "A", "AVEC", "SANS", "POUR", "EN", "ET",
    "CABLE", "CÂBLE", "T500", "T250", "C100", "C50", "T", "MM", "MM2", "MM²",
    "BLANC", "NOIR", "ROUGE", "BLEU", "GRIS", "IVOIRE", "BEIGE", "VERT",
    "COUPE", "TOURET", "U1000", "1M", "IM",
}


def _tokens(designation: str) -> set:
    mots = re.split(r"[^A-Z0-9.]+", (designation or "").upper())
    return {m for m in mots if m and m not in _STOP and len(m) > 1}


def _jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def _entete(ws):
    for c in ws[1]:
        if c.value:
            c.font = Font(bold=True)
            c.fill = GRIS
            c.alignment = Alignment(horizontal="center", wrap_text=True)


def _largeurs(ws, maxi=60):
    for col in ws.columns:
        lg = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(lg + 2, maxi)


def auditer(dossier_projet):

    dossier = Path(dossier_projet)

    base = BaseArticles(dossier / "base")
    base.importer_bdd(dossier / "base" / "BDD_articles.csv")
    base.importer_familles(dossier / "base" / "familles.csv")

    # Charger tous les articles 'bdd'
    articles = base.cx.execute(
        "SELECT ref_origine, code_interne, fabricant, designation, categorie "
        "FROM articles WHERE source='bdd'"
    ).fetchall()

    wb = Workbook()

    # ------------------------------------------------------------------
    # 1. Groupes suspects : Code interne mélangeant des produits que le
    #    normalisateur reconnaît comme DIFFERENTS (re-dérivation).
    # ------------------------------------------------------------------
    groupes_ci = defaultdict(list)
    for ref, ci, fab, des, cat in articles:
        if ci:
            groupes_ci[ci].append((ref, fab, des, cat))

    ws = wb.active
    ws.title = "Groupes suspects"
    ws.append(["Code interne", "Codes déduits en conflit", "Nb réfs",
               "Réf.", "Désignation", "Code déduit", "Catégorie"])
    _entete(ws)

    suspects = []
    for ci, membres in groupes_ci.items():
        if len(membres) < 2:
            continue
        # Re-dérive le code auto de chaque désignation
        derives = [(m, code_interne_auto(m[2])) for m in membres]
        codes = {d for _, d in derives if d}
        # Conflit RÉEL : les membres se dérivent en au moins deux codes
        # AUTO différents (le groupe mélange des produits distincts).
        # On n'alerte PAS quand le code auto diffère simplement du nom
        # manuel du groupe (ex. couleur incluse à la main).
        if len(codes) > 1:
            suspects.append((ci, derives, codes))

    # Trie par nombre de codes en conflit décroissant
    suspects.sort(key=lambda x: -len(x[2]))
    for ci, derives, codes in suspects:
        first = True
        for (ref, fab, des, cat), dcode in derives:
            ws.append([
                ci if first else "",
                " | ".join(sorted(codes)) if first else "",
                len(derives) if first else "",
                ref, des, dcode, cat,
            ])
            r = ws.max_row
            # Surligne le membre qui contredit le code du groupe
            if dcode and dcode != ci:
                for c in ws[r]:
                    c.fill = ROUGE
            first = False

    ws.freeze_panes = "A2"
    _largeurs(ws)

    # ------------------------------------------------------------------
    # 2. Couverture par catégorie
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Couverture")
    ws2.append(["Catégorie", "Total", "Avec code interne", "Sans", "% couvert"])
    _entete(ws2)

    par_cat = defaultdict(lambda: [0, 0])
    for ref, ci, fab, des, cat in articles:
        cat = cat or "(sans catégorie)"
        par_cat[cat][0] += 1
        if ci:
            par_cat[cat][1] += 1

    for cat in sorted(par_cat, key=lambda c: -(par_cat[c][0] - par_cat[c][1])):
        tot, avec = par_cat[cat]
        sans = tot - avec
        pct = avec / tot if tot else 0
        ws2.append([cat, tot, avec, sans, round(pct, 2)])
        ws2.cell(row=ws2.max_row, column=5).number_format = "0 %"
        if pct < 0.5 and sans >= 10:
            ws2.cell(row=ws2.max_row, column=4).fill = ORANGE
    ws2.freeze_panes = "A2"
    _largeurs(ws2, 30)

    # ------------------------------------------------------------------
    # 3. Sans code interne
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Sans code interne")
    ws3.append(["Catégorie", "Réf.", "Fabricant", "Désignation"])
    _entete(ws3)
    sans_code = [(cat or "", ref, fab, des)
                 for ref, ci, fab, des, cat in articles if not ci]
    for cat, ref, fab, des in sorted(sans_code):
        ws3.append([cat, ref, fab, des])
    ws3.freeze_panes = "A2"
    _largeurs(ws3)

    # ------------------------------------------------------------------
    # 4. Suggestions de regroupement (candidats à valider)
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("Suggestions")
    ws4.append(["Suggestion", "Proximité", "Réf.", "Fabricant", "Désignation", "Catégorie"])
    _entete(ws4)

    # Par catégorie, regrouper les non-codés dont les désignations se ressemblent
    par_cat_sans = defaultdict(list)
    for cat, ref, fab, des in sans_code:
        par_cat_sans[cat].append((ref, fab, des))

    num_sugg = 0
    for cat, items in par_cat_sans.items():
        if len(items) < 2 or len(items) > 400:
            continue
        toks = [(_tokens(d), (r, f, d)) for r, f, d in items]
        utilises = set()
        for i in range(len(toks)):
            if i in utilises:
                continue
            cluster = [i]
            for j in range(i + 1, len(toks)):
                if j in utilises:
                    continue
                if _jaccard(toks[i][0], toks[j][0]) >= 0.6:
                    cluster.append(j)
            if len(cluster) >= 2:
                num_sugg += 1
                for k in cluster:
                    utilises.add(k)
                    r, f, d = toks[k][1]
                    ws4.append([f"S{num_sugg}", "", r, f, d, cat])
    ws4.freeze_panes = "A2"
    _largeurs(ws4)

    base.fermer()

    dossier_res = dossier / "resultats"
    dossier_res.mkdir(exist_ok=True)
    fichier = dossier_res / "Audit_BDD.xlsx"
    wb.save(fichier)

    print(f"Audit créé : {fichier}")
    print(f"  - {len(suspects)} groupe(s) suspect(s)")
    print(f"  - {len(sans_code)} article(s) sans code interne")
    print(f"  - {num_sugg} suggestion(s) de regroupement")
    return fichier
