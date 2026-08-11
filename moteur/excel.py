"""
Génération du Comparatif.xlsx — le document de décision de l'acheteur.

Onglets :
1. Comparatif   : une ligne par référence du besoin (celles SANS AUCUNE
                  offre remontées en tête, jamais perdues en silence).
                  Prix de chaque fournisseur RAMENÉ À UNE BASE COMPARABLE
                  (€/m pour les câbles, €/unité pour le reste), pour
                  comparer sans se faire piéger par les unités de vente
                  (mètre / barre / boîte / kit). Une ligne de totaux en bas
                  indique, par fournisseur, ce que coûterait le besoin s'il
                  lui était pris intégralement.
2. Détail devis : toutes les lignes extraites, avec le PRIX BRUT et l'UNITÉ
                  réels du devis (pour passer commande dans l'unité du
                  fournisseur).
3. Hors besoin  : lignes de devis non rapprochées du besoin.

Colonnes clés du Comparatif :
- Code interne : le code unifiant (ex. R2V3G1.5) partagé entre fournisseurs
  et fabricants.
- Base : l'unité de comparaison retenue (€/m, €/u...). En orange = conversion
  incertaine (conditionnement inconnu) : le prix est laissé tel quel, à
  vérifier.
- Fournisseur retenu : PRÉ-REMPLI au mieux-disant, mais c'est un texte
  libre (pas une formule) — l'acheteur peut le remplacer par un autre nom
  de la liste déroulante associée. Les colonnes "Réf. fournisseur",
  "N° devis", "Disponibilité" et "Montant retenu" sont des formules qui
  suivent CE choix (via un bloc de colonnes masquées, une par fournisseur,
  juste après "Montant retenu") : elles ne restent pas figées sur le moins
  cher. moteur/panier.py (étape 4) relit uniquement le texte de
  "Fournisseur retenu" — jamais ces formules, qui ne se sont peut-être
  jamais recalculées si le fichier n'a pas été rouvert dans Excel — et
  reconstruit lui-même l'offre réelle correspondante depuis le besoin et
  les devis.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from statistics import median

from moteur.comparateur import comparer, meilleure_offre


VERT = PatternFill(fill_type="solid", fgColor="D9EAD3")
JAUNE = PatternFill(fill_type="solid", fgColor="FFF2CC")
ROUGE = PatternFill(fill_type="solid", fgColor="F4CCCC")
ORANGE = PatternFill(fill_type="solid", fgColor="FCE5CD")
GRIS = PatternFill(fill_type="solid", fgColor="EFEFEF")


def _entete(ws, ligne=1):
    for cellule in ws[ligne]:
        if cellule.value:
            cellule.font = Font(bold=True)
            cellule.fill = GRIS
            cellule.alignment = Alignment(horizontal="center", wrap_text=True)


def _largeurs(ws, maxi=42):
    for colonne in ws.columns:
        longueur = max(
            (len(str(c.value)) for c in colonne if c.value is not None),
            default=0,
        )
        ws.column_dimensions[colonne[0].column_letter].width = min(longueur + 3, maxi)


def exporter_comparatif(
    articles, dossier_projet, besoin=None, base=None, besoin_brut=None,
    nom_chantier="", referentiel=None, dossier_resultats=None,
):

    resultat = comparer(articles, besoin, base, referentiel)

    fournisseurs = resultat["fournisseurs"]

    avec_besoin = besoin is not None and len(besoin) > 0

    wb = Workbook()

    # ------------------------------------------------------------------
    # Onglet 1 : Comparatif
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Comparatif"

    if avec_besoin:
        entetes_gauche = [
            "Demande d'origine", "Référence", "Fabricant",
            "Code interne", "Désignation", "Qté", "Unité",
        ]
    else:
        entetes_gauche = [
            "Référence", "Fabricant", "Code interne", "Désignation",
            "Qté", "Unité",
        ]

    col_premier_f = len(entetes_gauche) + 1

    entetes = list(entetes_gauche)
    entetes += fournisseurs
    entetes += [
        "Meilleur prix", "Base", "Fournisseur retenu",
        "Réf. fournisseur", "N° devis", "Disponibilité", "Montant retenu",
    ]
    # Colonnes masquées : détail (réf. / n° devis / disponibilité) de
    # CHAQUE fournisseur, alignées sur les mêmes colonnes que leur prix.
    # Permettent aux formules de "Réf. fournisseur" / "N° devis" /
    # "Disponibilité" de suivre le fournisseur réellement retenu (la
    # colonne "Fournisseur retenu" est modifiable à la main) plutôt que de
    # rester figées sur le moins cher.
    entetes += [f"{f} (réf. — masqué)" for f in fournisseurs]
    entetes += [f"{f} (devis — masqué)" for f in fournisseurs]
    entetes += [f"{f} (dispo — masqué)" for f in fournisseurs]

    ws.append(entetes)
    _entete(ws)

    col_dernier_f = col_premier_f + len(fournisseurs) - 1
    col_meilleur = col_dernier_f + 1
    col_base = col_meilleur + 1
    col_retenu = col_base + 1
    col_ref_retenu = col_retenu + 1
    col_devis_retenu = col_ref_retenu + 1
    col_dispo_retenu = col_devis_retenu + 1
    col_montant = col_dispo_retenu + 1

    col_masque_ref = col_montant + 1
    col_masque_devis = col_masque_ref + len(fournisseurs)
    col_masque_dispo = col_masque_devis + len(fournisseurs)

    l_premier = get_column_letter(col_premier_f)
    l_dernier = get_column_letter(col_dernier_f)
    l_retenu = get_column_letter(col_retenu)
    l_entete_f = f"${l_premier}$1:${l_dernier}$1"

    # Masque les 3 blocs de colonnes de détail par fournisseur
    for depart in (col_masque_ref, col_masque_devis, col_masque_dispo):
        for i in range(len(fournisseurs)):
            ws.column_dimensions[get_column_letter(depart + i)].hidden = True

    # Position des colonnes Qté / Unité
    col_qte = entetes.index("Qté") + 1
    col_unite = entetes.index("Unité") + 1
    l_qte = get_column_letter(col_qte)

    num = 2

    for ligne in resultat["lignes"]:

        offres = ligne["offres"]
        gagnant, meilleur = meilleure_offre(offres)

        info = base.info(ligne["reference"]) if base is not None else {}
        fabricant = info.get("fabricant", "") if info else ""
        code_interne = info.get("code_interne", "") if info else ""

        qte = ligne["qte_besoin"]
        if qte is None and meilleur is not None:
            qte = meilleur["quantite"]

        # Unité(s) de vente réellement rencontrées
        unites = {o["unite"] for o in offres.values() if o["unite"]}
        unite_affichee = ligne["unite"]
        if len(unites) > 1:
            unite_affichee = " / ".join(sorted(unites))

        if avec_besoin:
            valeurs = [
                ligne.get("demande", ""), ligne["reference"], fabricant,
                code_interne, ligne["designation"], qte, unite_affichee,
            ]
        else:
            valeurs = [
                ligne["reference"], fabricant, code_interne,
                ligne["designation"], qte, unite_affichee,
            ]

        # Prix normalisés par fournisseur (base comparable)
        for fournisseur in fournisseurs:
            offre = offres.get(fournisseur)
            valeurs.append(offre["prix_norm"] if offre else None)

        if offres:
            plage = f"{l_premier}{num}:{l_dernier}{num}"
            retenu_cel = f"{l_retenu}{num}"
            valeurs.append(f"=MIN({plage})")               # Meilleur prix
            valeurs.append(meilleur["base"])               # Base
            valeurs.append(gagnant)                         # Fournisseur retenu (éditable)
            valeurs.append(
                f'=IFERROR(INDEX({get_column_letter(col_masque_ref)}{num}:'
                f'{get_column_letter(col_masque_ref + len(fournisseurs) - 1)}{num},'
                f'MATCH({retenu_cel},{l_entete_f},0)),"")'
            )                                                # Réf. fournisseur
            valeurs.append(
                f'=IFERROR(INDEX({get_column_letter(col_masque_devis)}{num}:'
                f'{get_column_letter(col_masque_devis + len(fournisseurs) - 1)}{num},'
                f'MATCH({retenu_cel},{l_entete_f},0)),"")'
            )                                                # N° devis
            valeurs.append(
                f'=IFERROR(INDEX({get_column_letter(col_masque_dispo)}{num}:'
                f'{get_column_letter(col_masque_dispo + len(fournisseurs) - 1)}{num},'
                f'MATCH({retenu_cel},{l_entete_f},0)),"")'
            )                                                # Disponibilité
            valeurs.append(
                f'=IFERROR({l_qte}{num}*INDEX({plage},MATCH({retenu_cel},{l_entete_f},0)),"")'
            )                                                # Montant retenu
        else:
            valeurs += ["AUCUNE OFFRE", "", "", "", "", "", ""]

        # Détail masqué par fournisseur (réf. / devis / disponibilité),
        # dans le MÊME ORDRE que les colonnes de prix ci-dessus.
        for fournisseur in fournisseurs:
            offre = offres.get(fournisseur)
            valeurs.append(offre["reference"] if offre else "")
        for fournisseur in fournisseurs:
            offre = offres.get(fournisseur)
            valeurs.append(offre["devis"] if offre else "")
        for fournisseur in fournisseurs:
            offre = offres.get(fournisseur)
            valeurs.append(offre["disponibilite"] if offre else "")

        ws.append(valeurs)

        if offres:
            for i, fournisseur in enumerate(fournisseurs):
                cel = ws.cell(row=num, column=col_premier_f + i)
                cel.number_format = "#,##0.0000"
                if fournisseur == gagnant:
                    cel.fill = VERT

            ws.cell(row=num, column=col_meilleur).number_format = "#,##0.0000"
            ws.cell(row=num, column=col_montant).number_format = "#,##0.00"

            # Liste déroulante sur "Fournisseur retenu" : seuls les
            # fournisseurs qui ont effectivement une offre sur cette ligne.
            # Évite qu'une faute de frappe casse la formule de rapatriement.
            noms = sorted(offres.keys())
            liste = ",".join(noms)
            if len(liste) <= 250:
                dv = DataValidation(
                    type="list", formula1=f'"{liste}"', allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(ws.cell(row=num, column=col_retenu))

            # Conversion incertaine sur une offre -> base en orange
            if any(not o["fiable"] for o in offres.values()):
                ws.cell(row=num, column=col_base).fill = ORANGE

            # Anomalie de prix : une offre s'écarte fortement de la médiane
            # (souvent un conditionnement mal converti). On surligne la cellule.
            prix_list = [o["prix_norm"] for o in offres.values() if o["prix_norm"]]
            if len(prix_list) >= 3:
                med = median(prix_list)
                if med:
                    for i, fournisseur in enumerate(fournisseurs):
                        o = offres.get(fournisseur)
                        # Écart ENORME (>6x) = quasi certainement une erreur
                        # d'unité/conditionnement, pas une différence de prix.
                        if o and o["prix_norm"] and (
                            o["prix_norm"] < med / 6 or o["prix_norm"] > med * 6
                        ):
                            ws.cell(row=num, column=col_premier_f + i).fill = ROUGE

            # Quantité devis différente du besoin -> jaune
            if (
                avec_besoin and ligne["qte_besoin"] and meilleur["quantite"]
                and meilleur["quantite"] != ligne["qte_besoin"]
            ):
                ws.cell(row=num, column=col_qte).fill = JAUNE
        else:
            ws.cell(row=num, column=col_meilleur).fill = ROUGE

        num += 1

    # Ligne de totaux : ce que coûterait le besoin si tout était pris chez
    # un seul fournisseur (somme des lignes où il a répondu — les lignes
    # sans offre de sa part sont ignorées, pas comptées comme un manque).
    if resultat["totaux_fournisseurs"]:
        ligne_totaux = ["TOTAL (si tout pris chez ce fournisseur)"]
        ligne_totaux += [""] * (col_premier_f - 2)
        for fournisseur in fournisseurs:
            ligne_totaux.append(resultat["totaux_fournisseurs"].get(fournisseur, 0))
        ws.append(ligne_totaux)
        for i in range(len(fournisseurs)):
            cel = ws.cell(row=num, column=col_premier_f + i)
            cel.number_format = "#,##0.00"
        for cel in ws[num]:
            if cel.value not in (None, ""):
                cel.font = Font(bold=True)
                cel.fill = GRIS
        num += 1

    ws.freeze_panes = ws.cell(row=2, column=col_premier_f).coordinate
    _largeurs(ws)

    # ------------------------------------------------------------------
    # Onglet 2 : Détail devis (prix bruts + unités réelles pour commander)
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Détail devis")
    ws2.append([
        "Fournisseur", "N° devis", "Réf. fournisseur", "Réf. distributeur",
        "Désignation", "Qté", "Unité", "Prix brut", "Prix net",
        "Montant", "Disponibilité",
    ])
    _entete(ws2)
    for a in articles:
        ws2.append([
            a.fournisseur, a.devis, a.reference_fournisseur,
            a.reference_distributeur, a.designation, a.quantite, a.unite,
            a.prix_brut, a.prix_net, a.montant, a.disponibilite,
        ])
    ws2.freeze_panes = "A2"
    _largeurs(ws2)

    # ------------------------------------------------------------------
    # Onglet 3 : Hors besoin
    # ------------------------------------------------------------------
    if resultat["hors_besoin"]:
        ws3 = wb.create_sheet("Hors besoin")
        ws3.append([
            "Fournisseur", "Réf. fournisseur", "Désignation", "Qté",
            "Unité", "Prix net", "Montant", "N° devis", "Disponibilité",
        ])
        _entete(ws3)
        for h in resultat["hors_besoin"]:
            ws3.append([
                h["fournisseur"], h["reference"], h["designation"],
                h["quantite"], h["unite"], h["prix"], h["montant"],
                h["devis"], h["disponibilite"],
            ])
        ws3.freeze_panes = "A2"
        _largeurs(ws3)

    # ------------------------------------------------------------------
    # Onglet "Besoin" (demande d'origine) quand elle n'a pas servi au tri
    # ------------------------------------------------------------------
    if besoin_brut:
        wsb = wb.create_sheet("Besoin")
        wsb.append(["Demande d'origine", "Quantité", "Référence (à compléter)"])
        _entete(wsb)
        for b in besoin_brut:
            wsb.append([b.designation, b.quantite, b.reference])
        wsb.freeze_panes = "A2"
        _largeurs(wsb)

    # ------------------------------------------------------------------
    # Enregistrement
    # ------------------------------------------------------------------
    dossier = Path(dossier_resultats) if dossier_resultats else Path(dossier_projet) / "resultats"
    dossier.mkdir(parents=True, exist_ok=True)

    nom_base = f"Comparatif {nom_chantier}" if nom_chantier else "Comparatif"
    fichier = dossier / f"{nom_base}.xlsx"

    try:
        wb.save(fichier)
    except PermissionError:
        fichier = dossier / f"{nom_base}_{datetime.now():%H%M%S}.xlsx"
        print(f"{nom_base}.xlsx est ouvert dans Excel.")
        wb.save(fichier)

    print(f"\nComparatif créé : {fichier}")
    return fichier
