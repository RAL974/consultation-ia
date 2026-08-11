"""
Lecture du besoin chantier.

Objectif : la DEMANDE D'ORIGINE (le texte tel qu'écrit par le service
demandeur) doit apparaître en première colonne du comparatif, pour vérifier
d'un coup d'œil que chaque ligne correspond au bon produit.

------------------------------------------------------------------
FORMAT DU FICHIER BESOIN (dans le dossier besoins/)
------------------------------------------------------------------

Trois colonnes, dans n'importe quel ordre. Les en-têtes sont reconnus
automatiquement :

    Demande d'origine          | Référence | Quantité
    (Demande / Besoin /         (Réf /       (Qté /
     Désignation d'origine)      Référence)   Quantité)

- La colonne "Demande d'origine" est celle qui s'affiche en tête du
  comparatif. Colle ici le texte du demandeur.
- La colonne "Référence" sert au rapprochement avec les devis. Mets la
  référence Legrand (ou fabricant) que tu connais. Si tu la laisses vide,
  la ligne apparaît quand même mais sans offre rapprochée.
- La colonne "Quantité" est la quantité demandée.

EXCEL (.xlsx)
    Ligne 1 = en-têtes, puis une ligne par article.

TEXTE (.txt), une ligne par article, séparateur point-virgule :
    100 douilles pour DCL encastrée ; 710100 ; 100
    10 boîtes ampoules 4000K E27     ; 300007 ; 10
    15 DX3-ID 2P 63A type A 30mA      ; 411651 ; 15

    (Sans point-virgule, l'ancien format "référence quantité" reste accepté.)

Le premier fichier trouvé est utilisé (priorité au .xlsx).
"""

import re
from pathlib import Path

from openpyxl import load_workbook

from moteur.modele import LigneBesoin


def normaliser_ref(ref: str) -> str:
    """Clé brute : majuscules, sans espaces ni tirets ni points."""
    return (
        str(ref)
        .strip()
        .upper()
        .replace(" ", "")
        .replace("-", "")
        .replace(".", "")
    )


def cle_matching(ref: str) -> str:
    """
    Clé de rapprochement inter-fournisseurs.

    Les distributeurs préfixent différemment la même référence fabricant :
        411651 (Ravate FNR) = L411651 (Cominter) = LEG411651 (Coredime, DEM...)

    On extrait le cœur numérique (zéros de tête supprimés). Si la référence
    contient moins de 4 chiffres, on garde la référence complète normalisée.
    """
    ref = normaliser_ref(ref)

    chiffres = "".join(re.findall(r"\d+", ref)).lstrip("0")

    if len(chiffres) >= 4:
        return chiffres

    return ref


def _to_float(valeur) -> float:
    if valeur is None or valeur == "":
        return 0.0
    return float(str(valeur).replace(" ", "").replace(",", "."))


def _identifier_colonnes(entetes):
    """Repère les colonnes demande / référence / quantité par leur en-tête."""

    col_dem = col_ref = col_qte = None

    for i, e in enumerate(entetes):

        e = e.upper()

        if col_dem is None and e.startswith(
            ("DEMAND", "BESOIN", "DÉSIGNATION D", "DESIGNATION D", "LIBELL")
        ):
            col_dem = i

        elif col_ref is None and e.startswith(("REF", "RÉF")):
            col_ref = i

        elif col_qte is None and e.startswith(("QT", "QUANT")):
            col_qte = i

    return col_dem, col_ref, col_qte


def _lire_xlsx(fichier: Path) -> list[LigneBesoin]:

    wb = load_workbook(fichier, data_only=True)
    ws = wb.active

    lignes = list(ws.iter_rows(values_only=True))

    if not lignes:
        return []

    entetes = [str(c).strip() if c is not None else "" for c in lignes[0]]

    col_dem, col_ref, col_qte = _identifier_colonnes(entetes)

    debut = 1

    # Aucun en-tête reconnu : on suppose Demande | Référence | Quantité
    if col_dem is None and col_ref is None:
        col_dem, col_ref, col_qte = 0, 1, 2
        debut = 0

    besoin = []

    for ligne in lignes[debut:]:

        if ligne is None:
            continue

        def cell(idx):
            if idx is None or idx >= len(ligne) or ligne[idx] is None:
                return ""
            return str(ligne[idx]).strip()

        demande = cell(col_dem)
        ref = cell(col_ref)

        if not demande and not ref:
            continue

        qte = _to_float(ligne[col_qte]) if (
            col_qte is not None and col_qte < len(ligne)
        ) else 0.0

        besoin.append(LigneBesoin(reference=ref, quantite=qte, designation=demande))

    return besoin


_MOTS_ENTETE_DEM = (
    "DESIGNATION", "DÉSIGNATION", "DEMANDE", "BESOIN", "LIBELLE", "LIBELLÉ",
)
_MOTS_ENTETE_REF_OU_QTE = (
    "REFERENCE", "RÉFÉRENCE", "REF", "RÉF", "QTE", "QTÉ", "QUANTITE", "QUANTITÉ",
)


def _est_ligne_entete(ligne: str) -> bool:
    """Détecte une ligne d'en-têtes (Désignation/Référence/Qté, collée depuis
    Excel avec des tabulations ou des points-virgules) en tête d'un fichier
    .txt, pour ne pas la lire comme une ligne de besoin fantôme."""

    morceaux = [m.strip().upper() for m in re.split(r"[;\t]", ligne) if m.strip()]

    if len(morceaux) < 2:
        return False

    a_dem = any(m.startswith(_MOTS_ENTETE_DEM) for m in morceaux)
    a_ref_ou_qte = any(m.startswith(_MOTS_ENTETE_REF_OU_QTE) for m in morceaux)

    return a_dem and a_ref_ou_qte


def _lire_txt(fichier: Path) -> list[LigneBesoin]:

    besoin = []
    premiere_ligne = True

    for ligne in fichier.read_text(encoding="utf-8", errors="replace").splitlines():

        ligne = ligne.strip()

        if not ligne or ligne.startswith("#"):
            continue

        if premiere_ligne:
            premiere_ligne = False
            if _est_ligne_entete(ligne):
                continue

        if ";" in ligne:
            # Demande ; Référence ; Quantité
            morceaux = [m.strip() for m in ligne.split(";")]
            demande = morceaux[0]
            ref = morceaux[1] if len(morceaux) > 1 else ""
            qte = _to_float(morceaux[2]) if len(morceaux) > 2 else 0.0

        else:
            morceaux = ligne.split()

            def _est_nombre(x):
                try:
                    float(x.replace(",", "."))
                    return True
                except ValueError:
                    return False

            if len(morceaux) >= 2 and _est_nombre(morceaux[1]):
                # Ancien format : Référence Quantité [Désignation]
                ref = morceaux[0]
                qte = _to_float(morceaux[1])
                demande = " ".join(morceaux[2:])
            elif morceaux and _est_nombre(morceaux[0]):
                # Texte libre du demandeur : "10 interrupteurs double allumage..."
                qte = _to_float(morceaux[0])
                demande = " ".join(morceaux[1:])
                ref = ""
            else:
                # Désignation seule, sans quantité identifiable
                ref = ""
                qte = 0.0
                demande = ligne

        besoin.append(LigneBesoin(reference=ref, quantite=qte, designation=demande))

    return besoin


def trouver_fichier_besoin(dossier) -> Path | None:
    """Retourne le fichier besoin à utiliser (le premier trouvé, priorité au
    .xlsx), ou None si le dossier est vide/absent."""

    dossier = Path(dossier)

    if not dossier.exists():
        return None

    fichiers = sorted(dossier.glob("*.xlsx")) + sorted(dossier.glob("*.txt"))

    fichiers = [f for f in fichiers if not f.name.startswith("~$")]

    return fichiers[0] if fichiers else None


def nom_chantier(fichier) -> str:
    """
    Déduit le nom du chantier à partir du nom du fichier besoin, en retirant
    le mot "besoin(s)" (quelle que soit sa position/casse) et les séparateurs
    superflus.

    "Besoin Doujani.txt"         -> "Doujani"
    "besoin_ecole_dambrevil.txt" -> "ecole dambrevil"
    "Doujani - Besoin.xlsx"      -> "Doujani"

    Retourne "" si fichier est None.
    """

    if fichier is None:
        return ""

    stem = Path(fichier).stem
    sans_mot = re.sub(r"(?i)\bbesoins?\b", " ", stem)
    normalise = re.sub(r"[\s_\-]+", " ", sans_mot).strip(" -_")

    return normalise


def lire_besoin(dossier) -> list[LigneBesoin]:
    """Retourne la liste des lignes du besoin, ou [] si aucun fichier."""

    fichier = trouver_fichier_besoin(dossier)

    if fichier is None:
        return []

    return lire_fichier_besoin(fichier)


def lire_fichier_besoin(fichier) -> list[LigneBesoin]:
    """Lit un fichier besoin précis (déjà localisé)."""

    fichier = Path(fichier)

    print(f"Besoin : {fichier.name}")

    if fichier.suffix.lower() == ".xlsx":
        besoin = _lire_xlsx(fichier)
    else:
        besoin = _lire_txt(fichier)

    print(f"Lignes de besoin : {len(besoin)}")

    return besoin
