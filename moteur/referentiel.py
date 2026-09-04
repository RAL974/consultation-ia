"""
Référentiel articles — normalisation par clé, alias appris, composés.

Rôle : faire tomber sur la même ligne de comparatif des références qui
désignent le même article mais s'écrivent différemment selon le
fournisseur (411651 / LG411651 / LEG411651...), en s'appuyant sur le
référentiel achats RÉEL de l'entreprise (base/BDD_articles.csv, colonnes
CONCAT / Clé_Réf), pas sur une logique de normalisation inventée.

Complète moteur/base.py (BaseArticles), qui reste en place et sert de
repli : moteur/comparateur.py essaie d'abord ce référentiel (alias exact
connu), puis retombe sur base.groupe() si ce module ne sait pas répondre.

Trois tables (base/articles.db) :

1. articles  — amorcée par l'import complet de base/BDD_articles.csv.
2. alias     — toute variante de référence rencontrée (préfixe marque,
               tirets, casse...) -> clé normalisée, avec origine :
                 'import'          : déduite du CSV à l'import (CONCAT/Clé_Réf)
                 'confirme'        : l'acheteur a validé/corrigé une proposition
                 'confirme_nouveau': l'acheteur a explicitement rejeté une
                                     proposition -> la référence devient sa
                                     propre clé, définitivement
                 'devis'           : référence inédite, sans aucun candidat
                                     trouvé -> devient sa propre clé (pas
                                     besoin de confirmation, rien à décider)
3. composes  — un besoin -> plusieurs références (ex. coffret + porte).
               Vide par défaut (aucune donnée de composé dans la BDD
               source) : à peupler à la main dans referentiel/composes.csv,
               ou par apprentissage futur.

Préfixes marque : DÉDUITS à l'import des couples (CONCAT, Clé_Réf,
Fabricant) réellement présents dans base/BDD_articles.csv (voir
`deduire_prefixes`) — jamais codés en dur. Sur les données de ce projet,
seuls 4 fabricants sur ~300 utilisent un préfixe (ex. Legrand -> "LEG") ;
tous les autres n'en ont aucun. Aucun suffixe marque n'a été observé.

Workflow de confirmation (fichier Excel aller-retour, pas de question
console — le GUI Tkinter n'a pas de console) :
  1. Une référence de devis inconnue mais proche d'un article existant est
     PROPOSÉE, jamais rapprochée automatiquement.
  2. En fin d'exécution, referentiel/A_confirmer.xlsx liste les
     propositions en attente.
  3. L'acheteur remplit la colonne Décision (OUI / NON / une clé corrigée)
     à son rythme, PDF ou Excel ouvert à côté.
  4. À l'exécution SUIVANTE, ces décisions sont appliquées (`alias`
     s'enrichit) AVANT de comparer les devis de cette exécution-là.
"""

import csv
import re
import sqlite3
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from moteur.base import normaliser_ref, coeur_numerique
from moteur.outils import to_float


# ----------------------------------------------------------------------
# Déduction des préfixes marque (à partir des données réelles)
# ----------------------------------------------------------------------
def deduire_prefixes(lignes: list[dict]) -> dict:
    """
    Déduit, pour chaque fabricant, le préfixe marque utilisé dans CONCAT
    (ex. Legrand -> "LEG"), à partir des couples (CONCAT, Clé_Réf,
    Fabricant) réellement présents dans `lignes` (une ligne = un dict issu
    du CSV BDD_articles). Règle : le préfixe majoritaire pour ce fabricant
    (strictement plus de la moitié des lignes où CONCAT != Clé_Réf) ; les
    lignes minoritaires (incohérences de saisie dans la source) sont
    ignorées, pas utilisées pour la règle.

    Retourne {fabricant: prefixe}. Un fabricant sans préfixe détecté
    n'apparaît pas dans le résultat.
    """
    candidats = {}  # fabricant -> Counter(prefixe)

    for r in lignes:
        concat = (r.get("CONCAT") or "").strip()
        cle = (r.get("Clé_Réf") or "").strip()
        fab = (r.get("Fabricant") or "").strip()

        if not fab or not concat or not cle or concat == cle:
            continue

        if not concat.endswith(cle):
            continue

        prefixe = concat[: -len(cle)]

        if not prefixe:
            continue

        candidats.setdefault(fab, Counter())[prefixe] += 1

    prefixes = {}

    for fab, compteur in candidats.items():
        prefixe, n = compteur.most_common(1)[0]
        total = sum(compteur.values())
        if n > total / 2:
            prefixes[fab] = prefixe

    return prefixes


def _plier_accents(texte: str) -> str:
    """Majuscules sans accents (même pliage que cle_designation). Sans ce
    pliage, É/Â/Ô agissent comme séparateurs dans _tokens et MUTILENT les
    mots : "dérivations" -> "RIVATIONS", "câblage" -> "BLAGE" (cas réel,
    consultation Rico Carpaye : 9 lignes de besoin, 0 rapprochement)."""
    return "".join(
        c for c in unicodedata.normalize("NFKD", (texte or "").upper())
        if not unicodedata.combining(c)
    )


_DECIMALE_FR = re.compile(r"(\d)\s*,\s*(\d)")

# Mots sans pouvoir discriminant : ils diluent le Jaccard (cas réel :
# "BOITE DERIVATION 155 PAR 110" — "PAR" et "DE" comptaient autant que
# "DERIVATION").
_STOPWORDS = {
    "DE", "DU", "DES", "LE", "LA", "LES", "UN", "UNE", "ET", "EN",
    "AU", "AUX", "PAR", "POUR", "AVEC", "SUR", "NF",
}

# Abréviations/synonymes de conditionnement observés dans les devis réels
# (RAVATE "BTE IP55...", COMINTER "Cartouche silicone..." face à un besoin
# "tube silicone").
_EQUIVALENTS = {"CARTOUCHE": "TUBE", "CART": "TUBE", "BTE": "BOITE"}

_NOMBRE_UNITE = re.compile(r"^(\d+(?:\.\d+)?)([A-Z]+\d*)$")


def _normaliser_texte(designation: str) -> str:
    d = _plier_accents(designation)
    d = _DECIMALE_FR.sub(r"\1.\2", d)   # 1,5mm² / "2, 5MM2" -> 1.5 / 2.5
    return d.replace("²", "2")


def _tokens(designation: str) -> set:
    out = set()
    for m in re.split(r"[^A-Z0-9.]+", _normaliser_texte(designation)):
        m = m.strip(".")
        if not m:
            continue
        # "1.5MM", "40A", "290ML" : le nombre et l'unité comptent chacun
        # pour un token, sinon "1.5mm" (besoin) ne recoupe jamais
        # "1.5 mm²" (devis).
        nu = _NOMBRE_UNITE.match(m)
        parts = [nu.group(1), nu.group(2)] if nu else [m]
        for p in parts:
            if len(p) < 2 and not p.isdigit():
                continue
            # Singulier naïf : EMBOUTS ~ EMBOUT, DERIVATIONS ~ DERIVATION.
            if len(p) > 3 and p.endswith("S") and not p.endswith("SS"):
                p = p[:-1]
            p = _EQUIVALENTS.get(p, p)
            if p in _STOPWORDS:
                continue
            out.add(p)
    return out


# ----------------------------------------------------------------------
# Attributs techniques (section, simple/double, calibre, dimensions...)
# ----------------------------------------------------------------------
# La ressemblance de texte seule ne peut PAS rapprocher
# "inter sectionneur 2x40A Schneider" et "ACTI9 ISW 2P 40A 415VAC"
# (Jaccard = 0). Comme l'étiquette "Type <code>" des bordereaux
# architecte, un attribut technique identique des deux côtés est un
# signal plus fiable qu'un score de mots-clés : il fait remonter le
# score. À l'inverse, deux attributs DÉCLARÉS contradictoires (1.5 mm²
# vs 6 mm², simple vs double) excluent le candidat — jamais de
# proposition sciemment fausse. L'absence d'un attribut d'un côté ne
# pénalise rien.

# Sections de conducteur normalisées (mm²) : borne la détection pour ne
# pas prendre un Ø700MM de luminaire ou un "- 50" de conditionnement
# pour une section.
_SECTIONS = {"0.5", "0.75", "1", "1.5", "2.5", "4", "6", "10", "16",
             "25", "35", "50", "70", "95", "120", "150", "185", "240", "300"}

_RE_DIM3 = re.compile(r"\b(\d{2,4})\s*X\s*(\d{2,4})\s*X\s*(\d{2,4})\b")
_RE_POLES_X_CAL = re.compile(r"\b([1-4])\s*X\s*(\d{1,3})\s*A\b")
_RE_DOUBLE_SEC = re.compile(r"\b2\s*X\s*(\d+(?:\.\d+)?)")
_RE_SEC_MM = re.compile(r"\b(\d+(?:\.\d+)?)\s*MM2?\b")
_RE_CAL = re.compile(r"\b(\d{1,3})\s*A\b")
_RE_POLES = re.compile(r"\b([1-4])\s*P\b")
_RE_VOL = re.compile(r"\b(\d{2,4})\s*ML\b")


def _num(v: str) -> str:
    return v.rstrip("0").rstrip(".") if "." in v else v


def _analyser(designation: str):
    """(tokens, attributs) d'une désignation. Les portions consommées par
    un attribut sont retirées du texte avant tokenisation, pour que les
    dimensions/volumes ne diluent pas le Jaccard."""
    d = _normaliser_texte(designation)
    attrs = {}

    def _consommer(regex, action):
        def _sub(m):
            action(m)
            return " "
        return regex.sub(_sub, d)

    def _dim(m):
        attrs["DIM"] = tuple(int(m.group(i)) for i in (1, 2, 3))
    d = _consommer(_RE_DIM3, _dim)

    def _pxc(m):
        attrs["POLES"], attrs["CAL"] = m.group(1), m.group(2)
    d = _consommer(_RE_POLES_X_CAL, _pxc)

    embout = "EMBOUT" in d

    if embout:
        def _dbl(m):
            v = _num(m.group(1))
            if v not in _SECTIONS:
                return m.group(0)      # pas une section : ne rien consommer
            attrs["NBR"] = "DOUBLE"
            attrs["SEC"] = v
            return " "
        d = _RE_DOUBLE_SEC.sub(_dbl, d)

    def _sec(m):
        v = _num(m.group(1))
        if v not in _SECTIONS:
            return m.group(0)          # ex. Ø700MM d'un luminaire
        attrs.setdefault("SEC", v)
        return " "
    d = _RE_SEC_MM.sub(_sec, d)

    def _cal(m):
        attrs.setdefault("CAL", m.group(1))
    d = _consommer(_RE_CAL, _cal)

    def _pol(m):
        attrs.setdefault("POLES", m.group(1))
    d = _consommer(_RE_POLES, _pol)

    def _vol(m):
        attrs.setdefault("VOL", m.group(1))
    d = _consommer(_RE_VOL, _vol)

    toks = _tokens(d)

    if embout:
        if "DOUBLE" in toks:
            attrs.setdefault("NBR", "DOUBLE")
        elif "SIMPLE" in toks:
            attrs.setdefault("NBR", "SIMPLE")
        if "SEC" not in attrs:
            # "EMBOUT 1.5 NOIR - 50" (RAVATE) : section sans unité. Premier
            # nombre appartenant aux sections normalisées.
            for p in re.findall(r"\d+(?:\.\d+)?", d):
                if _num(p) in _SECTIONS:
                    attrs["SEC"] = _num(p)
                    break

    return toks, attrs


_ATTRS_STRICTS = ("SEC", "NBR", "CAL", "POLES", "VOL")
_TOLERANCE_DIM = 10  # mm par dimension


def _dims_compatibles(a, b):
    return all(abs(x - y) <= _TOLERANCE_DIM for x, y in zip(a, b))


def _attributs_contradictoires(a: dict, b: dict) -> bool:
    for k in _ATTRS_STRICTS:
        if k in a and k in b and a[k] != b[k]:
            return True
    if "DIM" in a and "DIM" in b and not _dims_compatibles(a["DIM"], b["DIM"]):
        return True
    return False


def _attributs_communs(a: dict, b: dict) -> int:
    n = sum(1 for k in _ATTRS_STRICTS if k in a and k in b and a[k] == b[k])
    if "DIM" in a and "DIM" in b and _dims_compatibles(a["DIM"], b["DIM"]):
        n += 1
    return n


def _jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


SEUIL_SIMILARITE = 0.3


def cle_designation(designation: str) -> str:
    """
    Clé canonique dérivée d'une désignation de besoin SANS référence
    (ex. bordereau architecte "Type A1 - Suspension circulaire Open
    space") : "DESC:<désignation normalisée>". Sert à proposer un
    rapprochement par ressemblance de texte plutôt que par référence —
    voir Referentiel.proposer_correspondances_designation().
    """
    d = "".join(
        c for c in unicodedata.normalize("NFKD", (designation or "").strip().upper())
        if not unicodedata.combining(c)
    )
    d = re.sub(r"[^A-Z0-9]+", "_", d).strip("_")
    return f"DESC:{d}" if d else ""


_ETIQUETTE_TYPE = re.compile(r"\bTYPE\s+([A-Z]?\d+[A-Z]?)\b")


def _etiquette_type(designation: str):
    """Étiquette "Type <code>" d'un bordereau architecte (ex. "Type A1",
    "Type C1"), si la désignation en contient une — sinon None."""
    m = _ETIQUETTE_TYPE.search((designation or "").upper())
    return m.group(1) if m else None


class Referentiel:
    """Accès au référentiel articles (moteur/articles.db)."""

    def __init__(self, dossier_moteur):
        self.dossier = Path(dossier_moteur)
        self.dossier.mkdir(exist_ok=True)
        self.chemin_db = self.dossier / "articles.db"
        self.cx = sqlite3.connect(self.chemin_db)
        self.prefixes = {}          # fabricant -> préfixe déduit (rempli par importer_bdd)
        self._propositions = {}     # reference_brute -> (cle_proposee, designation_proposee, score)
        self._creer_tables()

    def _creer_tables(self):
        self.cx.executescript(
            """
            CREATE TABLE IF NOT EXISTS articles (
                cle_normalisee TEXT,
                reference      TEXT,
                fournisseur    TEXT,
                fabricant      TEXT,
                designation    TEXT,
                categorie      TEXT,
                tarif          REAL,
                source         TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_art_cle ON articles(cle_normalisee);

            CREATE TABLE IF NOT EXISTS alias (
                reference_brute TEXT PRIMARY KEY,
                cle_normalisee  TEXT NOT NULL,
                origine         TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS composes (
                cle_besoin      TEXT,
                membre          TEXT,
                quantite_membre REAL DEFAULT 1,
                ordre           INTEGER,
                origine         TEXT DEFAULT 'manuel'
            );
            CREATE INDEX IF NOT EXISTS idx_comp_besoin ON composes(cle_besoin);
            """
        )
        self.cx.commit()

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------
    def importer_bdd(self, csv_path) -> int:
        """
        (Ré)importe base/BDD_articles.csv : amorce `articles` (source=
        'import') et `alias` (origine='import'). Idempotent (recharge
        complète des lignes 'import' à chaque appel, sans jamais toucher
        les lignes apprises en consultation) et rapide (une seule
        transaction pour tout le fichier).
        """
        csv_path = Path(csv_path)

        if not csv_path.exists():
            print(f"Référentiel absent : {csv_path.name} (rapprochement par clé limité)")
            return 0

        with open(csv_path, encoding="utf-8-sig") as f:
            lignes = list(csv.DictReader(f, delimiter=";"))

        self.prefixes = deduire_prefixes(lignes)
        if self.prefixes:
            resume = ", ".join(f"{f} -> {p}" for f, p in sorted(self.prefixes.items()))
            print(f"Préfixes marque déduits : {resume}")

        self.cx.execute("DELETE FROM articles WHERE source='import'")
        self.cx.execute("DELETE FROM alias WHERE origine='import'")

        lignes_articles = []
        lignes_alias = {}  # reference_brute_norm -> cle_norm (dédoublonnage en mémoire)

        for r in lignes:

            ref = (r.get("Référence") or "").strip()
            cle = (r.get("Clé_Réf") or "").strip()

            if not ref or not cle:
                continue

            cle_norm = normaliser_ref(cle)

            lignes_articles.append((
                cle_norm,
                ref,
                (r.get("Fournisseur") or "").strip(),
                (r.get("Fabricant") or "").strip(),
                (r.get("Désignation") or "").strip(),
                (r.get("Catégorie") or "").strip(),
                to_float(r.get("Tarif approximatif", "")),
                "import",
            ))

            # Alias : la référence telle qu'écrite (= CONCAT) -> clé nue,
            # + auto-alias de la clé nue vers elle-même (résolution triviale
            # d'une référence déjà "propre").
            lignes_alias[normaliser_ref(ref)] = cle_norm
            lignes_alias.setdefault(cle_norm, cle_norm)

        self.cx.executemany(
            "INSERT INTO articles VALUES (?,?,?,?,?,?,?,?)", lignes_articles
        )
        # INSERT OR REPLACE (pas un simple INSERT) : une référence de la
        # BDD peut avoir été redirigée par une équivalence BL manuelle
        # (voir importer_equivalences_bl(), origine='manuel', appelée
        # APRÈS importer_bdd() dans moteur.rapprochement.pipeline_bl —
        # l'ordre remet toujours la BDD "propre" en premier, l'équivalence
        # manuelle la re-surcharge ensuite) — sans ce remplacement, un
        # ré-import de la BDD plantait (UNIQUE constraint) sur CETTE seule
        # référence au lieu de simplement reprendre sa valeur d'origine.
        self.cx.executemany(
            "INSERT OR REPLACE INTO alias VALUES (?,?,?)",
            [(ref_brute, cle, "import") for ref_brute, cle in lignes_alias.items()],
        )
        self.cx.commit()

        print(f"Référentiel articles importé : {len(lignes_articles)} références")
        return len(lignes_articles)

    def importer_composes(self, csv_path) -> int:
        """
        (Ré)importe referentiel/composes.csv (Cle_besoin;Membre;Quantite),
        optionnel. Remplace les composés d'origine 'manuel' ; les composés
        appris en consultation (origine='appris') ne sont pas touchés.
        """
        csv_path = Path(csv_path)

        self.cx.execute("DELETE FROM composes WHERE origine='manuel'")

        if not csv_path.exists():
            self.cx.commit()
            return 0

        n = 0

        with open(csv_path, encoding="utf-8-sig") as f:
            for i, r in enumerate(csv.DictReader(f, delimiter=";")):

                cle_besoin = (r.get("Cle_besoin") or r.get("Clé_besoin") or "").strip()
                membre = (r.get("Membre") or "").strip()

                if not cle_besoin or not membre:
                    continue

                qte = to_float(r.get("Quantite") or r.get("Quantité") or "1") or 1.0

                self.cx.execute(
                    "INSERT INTO composes VALUES (?,?,?,?,?)",
                    (normaliser_ref(cle_besoin), normaliser_ref(membre), qte, i, "manuel"),
                )
                n += 1

        self.cx.commit()

        if n:
            print(f"Composés manuels importés : {n} membre(s)")

        return n

    def importer_equivalences_bl(self, csv_path) -> int:
        """
        (Ré)importe referentiel/equivalences_bl.csv (Reference_1;
        Reference_2;Note), optionnel — demande explicite de l'acheteur
        pour le rapprochement des BL (moteur.rapprochement.matching) :
        "il faut créer une base des équivalences, ce genre de cas va se
        présenter très souvent" (cas réel : le fournisseur livre "59210"
        pour la ligne Suivi "CFF1BIS", même article "colliers à embase
        16/32 boîte de 100" — RIEN de commun textuellement ni
        numériquement entre les deux références, donc `resoudre()` ne
        peut PAS le proposer tout seul via préfixe marque/cœur numérique,
        contrairement à un simple écart de préfixe déjà couvert par
        ailleurs — la seule façon fiable de les rapprocher est qu'un
        humain le dise une fois).

        Contrairement à A_confirmer_BL.xlsx (propositions AUTOMATIQUES
        d'un candidat plausible, qu'il faut valider), ce fichier est
        rempli DIRECTEMENT par l'acheteur (ou par une session Claude Code
        en son nom, quand elle explique une substitution en cours de
        travail) — aucune proposition préalable n'est nécessaire, chaque
        ligne est une équivalence déjà tranchée. Réutilise `alias` : les
        DEUX références sont aliasées l'une vers l'autre (`Reference_1`
        fait foi comme clé canonique — à choisir de préférence comme la
        référence déjà connue de la BDD achats quand l'une des deux
        l'est, pour ne pas faire dériver le groupement des lignes de
        DEVIS existantes vers une clé synthétique) avec origine='manuel' — même statut
        "connu" que les alias confirmés dans A_confirmer.xlsx, reconnu
        automatiquement pour toujours dès le prochain rapprochement, sans
        jamais redemander confirmation. Remplace les équivalences
        d'origine 'manuel' à chaque import (idempotent), comme
        importer_composes().
        """
        csv_path = Path(csv_path)

        self.cx.execute("DELETE FROM alias WHERE origine='manuel'")

        if not csv_path.exists():
            self.cx.commit()
            return 0

        n = 0

        with open(csv_path, encoding="utf-8-sig") as f:
            # Lignes de commentaire ("# ...", en tête de fichier pour
            # documenter le format) ignorées avant de passer la 1re ligne
            # utile à csv.DictReader comme en-tête — sans ce filtre, un
            # commentaire pris à tort pour l'en-tête aurait fait échouer
            # la lecture de TOUTES les lignes de données (aucun champ
            # "Reference_1"/"Reference_2" reconnu).
            lignes_utiles = (l for l in f if not l.lstrip().startswith("#"))
            for r in csv.DictReader(lignes_utiles, delimiter=";"):

                ref1 = (r.get("Reference_1") or "").strip()
                ref2 = (r.get("Reference_2") or "").strip()

                if not ref1 or not ref2:
                    continue

                cle = normaliser_ref(ref1)

                for ref in (ref1, ref2):
                    self.cx.execute(
                        "INSERT INTO alias VALUES (?,?,?) "
                        "ON CONFLICT(reference_brute) DO UPDATE SET "
                        "cle_normalisee=excluded.cle_normalisee, origine=excluded.origine",
                        (normaliser_ref(ref), cle, "manuel"),
                    )
                n += 1

        self.cx.commit()

        if n:
            print(f"Équivalences BL importées : {n} paire(s)")

        return n

    def apprendre_equivalence(self, csv_path, reference_connue: str, reference_facturee: str, note: str) -> bool:
        """
        Ajoute une ligne à referentiel/equivalences_bl.csv (Reference_1;
        Reference_2;Note) — utilisé quand une substitution SANS AUCUN
        rapport textuel/numérique (voir importer_equivalences_bl) est
        confirmée par l'acheteur via un mécanisme EXTERNE à ce module (ex.
        le "résiduel unique" du rapprochement factures, voir
        moteur.rapprochement.matching_facture._residuel_unique et
        moteur.rapprochement.pipeline_facture._appliquer_confirmations_
        substitutions) : ce module reste la SEULE façon d'écrire ce
        fichier, mais n'a pas besoin d'être celui qui a PROPOSÉ la paire
        au départ (contrairement à `_propositions`, alimenté uniquement
        par resoudre()/proposer_correspondances_designation()).

        Écrit dans le fichier CSV (portable, suivi par git) plutôt que
        directement dans la table `alias` (SQLite, gitignorée, propre à ce
        poste) : une substitution pure, découverte par élimination et donc
        impossible à re-proposer automatiquement, mérite une trace
        durable et partageable, pas seulement un cache local — voir
        CLAUDE.md, "Enrichissement continu du référentiel partagé".

        Idempotent : ne réécrit rien si cette paire (dans un sens ou
        l'autre) existe déjà. Retourne True si une ligne a été ajoutée.
        """
        csv_path = Path(csv_path)
        r1, r2 = normaliser_ref(reference_connue), normaliser_ref(reference_facturee)

        if not r1 or not r2:
            return False

        paires_existantes = set()
        if csv_path.exists():
            with open(csv_path, encoding="utf-8-sig") as f:
                lignes_utiles = (l for l in f if not l.lstrip().startswith("#"))
                for row in csv.DictReader(lignes_utiles, delimiter=";"):
                    a = normaliser_ref((row.get("Reference_1") or "").strip())
                    b = normaliser_ref((row.get("Reference_2") or "").strip())
                    if a and b:
                        paires_existantes.add(frozenset((a, b)))

        if frozenset((r1, r2)) in paires_existantes:
            return False

        nouveau_fichier = not csv_path.exists()
        with open(csv_path, "a", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, delimiter=";")
            if nouveau_fichier:
                w.writerow(["Reference_1", "Reference_2", "Note"])
            w.writerow([reference_connue, reference_facturee, note])

        print(f"Équivalence apprise : {reference_connue} <-> {reference_facturee} ({csv_path.name})")
        return True

    # ------------------------------------------------------------------
    # Confirmation (fichier Excel aller-retour)
    # ------------------------------------------------------------------
    def appliquer_confirmations(self, xlsx_path) -> int:
        """
        Lit referentiel/A_confirmer.xlsx (s'il existe) et applique chaque
        ligne dont la colonne Décision n'est pas vide, AVANT la résolution
        des devis de cette exécution :
          - "OUI"           -> confirme la clé proposée
          - "NON"           -> rejette : la référence devient sa propre clé,
                               ne sera plus reproposée
          - autre texte     -> clé corrigée à la main par l'acheteur
        """
        xlsx_path = Path(xlsx_path)

        if not xlsx_path.exists():
            return 0

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        lignes = list(ws.iter_rows(values_only=True))
        if not lignes:
            return 0

        entetes = [str(c).strip() if c else "" for c in lignes[0]]

        def idx(nom):
            return entetes.index(nom) if nom in entetes else None

        i_ref = idx("Référence détectée")
        i_prop = idx("Référence déjà connue (proposée)")
        i_decision = idx("Décision")

        if i_ref is None or i_decision is None:
            return 0

        n = 0

        for ligne in lignes[1:]:

            if ligne is None or i_ref >= len(ligne):
                continue

            ref_brute = str(ligne[i_ref] or "").strip()
            decision = str(ligne[i_decision] or "").strip() if i_decision < len(ligne) else ""

            if not ref_brute or not decision:
                continue

            ref_norm = normaliser_ref(ref_brute)
            decision_maj = decision.upper()

            if decision_maj == "OUI":
                cle = normaliser_ref(str(ligne[i_prop] or "")) if i_prop is not None else ""
                if not cle:
                    continue
                origine = "confirme"
            elif decision_maj == "NON":
                cle = ref_norm
                origine = "confirme_nouveau"
            else:
                cle = normaliser_ref(decision)
                origine = "confirme"

            self.cx.execute(
                "INSERT INTO alias VALUES (?,?,?) "
                "ON CONFLICT(reference_brute) DO UPDATE SET "
                "cle_normalisee=excluded.cle_normalisee, origine=excluded.origine",
                (ref_norm, cle, origine),
            )
            n += 1

        self.cx.commit()

        if n:
            print(f"Confirmations appliquées : {n} référence(s)")

        return n

    # ------------------------------------------------------------------
    # Résolution
    # ------------------------------------------------------------------
    def _candidat_prefixe(self, ref_norm: str, fabricant: str = ""):
        for fab, prefixe in self.prefixes.items():
            if fabricant and fabricant.strip().upper() != fab.strip().upper():
                continue
            if ref_norm.startswith(prefixe.upper()) and len(ref_norm) > len(prefixe):
                yield ref_norm[len(prefixe):]

    def resoudre(self, reference: str, fabricant: str = "", designation: str = "",
                 fournisseur: str = "", devis: str = ""):
        """
        Résout une référence de devis en clé normalisée.

        `fournisseur`/`devis` ne servent qu'à enrichir A_confirmer.xlsx
        (voir ecrire_a_confirmer) si cette référence finit en "propose" —
        sans ça, l'acheteur n'a aucun moyen de savoir de quel devis vient
        une ligne à confirmer.

        Retourne (cle, statut) :
          "connu"   -> correspondance certaine (alias exact) ; à utiliser
                       directement comme clé de regroupement.
          "propose" -> candidat plausible trouvé, PAS encore confirmé :
                       ne pas fusionner automatiquement, mais mémorisé pour
                       ecrire_a_confirmer().
          "nouveau" -> aucune correspondance ; la référence devient sa
                       propre clé (ajoutée à `articles`, source='devis').
        """
        ref_norm = normaliser_ref(reference)

        if not ref_norm:
            return "", "nouveau"

        row = self.cx.execute(
            "SELECT cle_normalisee, origine FROM alias WHERE reference_brute=?", (ref_norm,)
        ).fetchone()

        if row:
            cle, origine = row
            if origine == "devis":
                # Auto-alias créé plus bas (référence "nouvelle" vue une
                # 1ère fois dans CETTE exécution ou une précédente) : ce
                # n'est PAS une correspondance confirmée, juste un mémo pour
                # ne pas relancer la recherche de candidats. Doit rester
                # équivalent à "nouveau" (même repli sur base.groupe()),
                # sinon la MÊME référence obtient une clé différente selon
                # qu'elle est résolue pour la 1ère ou la 2e fois dans la
                # même exécution (ex. l'article d'un devis, puis la ligne de
                # besoin correspondante) -> désynchronisation silencieuse.
                return cle, "nouveau"
            return cle, "connu"

        # Candidats : préfixe marque déduit retiré, puis repli sur le cœur
        # numérique (même heuristique, déjà éprouvée, que moteur/base.py).
        candidats = list(self._candidat_prefixe(ref_norm, fabricant))
        coeur = coeur_numerique(ref_norm)
        if coeur:
            candidats.append(coeur)

        toks_devis = _tokens(designation)

        for candidat in candidats:

            art = self.cx.execute(
                "SELECT designation FROM articles WHERE cle_normalisee=? LIMIT 1",
                (candidat,),
            ).fetchone()

            if not art:
                continue

            if toks_devis:
                score = round(_jaccard(toks_devis, _tokens(art[0])), 2)
                if score < SEUIL_SIMILARITE:
                    continue
            else:
                # Pas de désignation à comparer (ex. résolution d'une
                # référence de besoin) : le candidat structurel (préfixe/
                # cœur numérique) est gardé, mais rien ne garantit qu'il
                # corresponde vraiment -> score non calculé, PAS un 1.0
                # trompeur.
                score = None

            # Garde le texte ORIGINAL (pas la forme normalisée) pour que
            # l'acheteur reconnaisse la référence et sa désignation dans
            # A_confirmer.xlsx en les comparant au PDF.
            self._propositions[ref_norm] = {
                "fournisseur": fournisseur, "devis": devis,
                "reference": reference, "designation_devis": designation, "qte_devis": None,
                "cle": candidat, "designation_candidat": art[0], "qte_besoin": None,
                "score": score,
            }
            return candidat, "propose"

        # Rien trouvé : nouvelle référence, devient sa propre clé. Ne
        # l'ajoute que si elle n'y est pas déjà (une même référence peut
        # être vue plusieurs fois dans la même exécution, ou reste d'une
        # exécution précédente) : évite d'accumuler des doublons au fil
        # des exécutions dans `articles` (source='devis') et dans l'export
        # articles_nouveaux.csv.
        deja = self.cx.execute(
            "SELECT 1 FROM articles WHERE cle_normalisee=? AND source='devis' LIMIT 1",
            (ref_norm,),
        ).fetchone()
        if not deja:
            self.cx.execute(
                "INSERT INTO articles VALUES (?,?,?,?,?,?,?,?)",
                (ref_norm, reference, "", fabricant, designation, "", None, "devis"),
            )

        # Auto-alias vers elle-même : une prochaine rencontre de cette
        # référence (même exécution ou une suivante) doit retomber
        # directement en "connu", pas repartir dans la recherche de
        # candidats (qui pourrait, par coïncidence, proposer autre chose).
        self.cx.execute(
            "INSERT OR IGNORE INTO alias VALUES (?,?,?)",
            (ref_norm, ref_norm, "devis"),
        )
        self.cx.commit()

        return ref_norm, "nouveau"

    def proposer_correspondances_designation(self, designation_besoin, quantite_besoin, candidats):
        """
        Rapprochement par DÉSIGNATION pour une ligne de besoin SANS
        référence (ex. bordereau architecte "Type A1 - Suspension
        circulaire..."), impossible à résoudre par resoudre() (qui exige
        une référence). Cas réel : certaines consultations n'ont que des
        désignations côté besoin, face à des devis de fournisseurs qui
        proposent chacun LEUR fabricant — les références ne se recoupent
        jamais, seule la ressemblance de texte (et la quantité, en indice)
        permet de rapprocher.

        `candidats` : liste de dicts {"fournisseur", "reference",
        "designation", "quantite"} — typiquement chaque groupe encore
        disponible de l'index de comparer() (pas déjà rattaché à une autre
        ligne de besoin), un par fournisseur.

        Ne fusionne JAMAIS automatiquement : le(s) meilleur(s) candidat(s)
        (un par fournisseur, score >= SEUIL_SIMILARITE) sont mémorisés
        pour ecrire_a_confirmer(), exactement comme resoudre() le fait pour
        les candidats par référence — l'acheteur les confirme dans
        A_confirmer.xlsx, et à la PROCHAINE exécution ils deviennent des
        alias "connu" comme n'importe quelle référence : la ligne de
        besoin est alors matchée normalement, sans traitement spécial.
        La quantité ne sert qu'à départager deux candidats déjà proches en
        texte pour le MÊME fournisseur (jamais un filtre : le besoin peut
        légitimement différer de la quantité d'un devis général) et est
        montrée à l'acheteur pour l'aider à trancher.

        Étiquette "Type <code>" (ex. "Type A1", "Type C1") : quand un
        bordereau architecte numérote ses postes, certains fournisseurs
        reprennent CETTE étiquette telle quelle dans leur propre
        désignation (cas réel, chantier Cosinus : STAND 64 écrit "TYPE A1
        - SUSPENSION CIRCULAIRE : SD-WOOD RING..."). La ressemblance de
        texte brute (Jaccard) est alors diluée par tout le vocabulaire
        technique autour et repasse SOUS le seuil (score observé ~0.2-0.3
        pour un vrai match) : une étiquette identique des deux côtés est un
        signal bien plus fiable qu'un score de mots-clés, et fait
        remonter le score — jamais en dessous de ce qu'aurait donné le
        texte seul, jamais un filtre non plus (l'absence d'étiquette d'un
        côté ne pénalise rien).

        Retourne la liste des candidats retenus (informatif).
        """
        toks_besoin, attrs_besoin = _analyser(designation_besoin)
        cle_cible = cle_designation(designation_besoin)

        if not toks_besoin or not cle_cible:
            return []

        etiquette_besoin = _etiquette_type(designation_besoin)

        par_fournisseur = {}

        for c in candidats:

            toks_c, attrs_c = _analyser(c["designation"])

            # Attributs techniques DÉCLARÉS contradictoires (1.5 vs 6 mm²,
            # simple vs double, dimensions incompatibles...) : jamais
            # proposé, quel que soit le score de texte.
            if _attributs_contradictoires(attrs_besoin, attrs_c):
                continue

            score = round(_jaccard(toks_besoin, toks_c), 2)

            # Accord d'attributs : même mécanique que l'étiquette "Type".
            # >= 2 attributs identiques (ex. calibre 40A + 2 pôles, ou
            # section + double) valent une quasi-certitude ; 1 attribut
            # (ex. la section seule) suffit à franchir le seuil, le texte
            # départage ensuite entre fournisseurs.
            communs = _attributs_communs(attrs_besoin, attrs_c)
            if communs >= 2:
                score = max(score, 0.9)
            elif communs == 1:
                score = max(score, 0.55)

            if etiquette_besoin and etiquette_besoin == _etiquette_type(c["designation"]):
                score = max(score, 0.9)

            if score < SEUIL_SIMILARITE:
                continue

            # Indice de quantité : ne départage qu'entre deux candidats du
            # MÊME fournisseur déjà retenus sur le texte, jamais un filtre.
            meilleur = par_fournisseur.get(c["fournisseur"])
            if meilleur is not None and score <= meilleur["score"]:
                continue

            par_fournisseur[c["fournisseur"]] = {**c, "score": score}

        retenus = sorted(par_fournisseur.values(), key=lambda x: -x["score"])

        for c in retenus:
            ref_norm = normaliser_ref(c["reference"])
            if not ref_norm:
                continue
            # N'écrase pas une proposition déjà là pour cette même référence
            # (ex. déjà proposée par référence ailleurs dans ce devis).
            self._propositions.setdefault(ref_norm, {
                "fournisseur": c.get("fournisseur", ""), "devis": c.get("devis", ""),
                "reference": c["reference"], "designation_devis": c["designation"],
                "qte_devis": c.get("quantite"),
                "cle": cle_cible, "designation_candidat": designation_besoin,
                "qte_besoin": quantite_besoin,
                "score": c["score"],
            })

        return retenus

    def composants(self, cle: str):
        """[(membre, quantite_membre), ...] déclarés pour cette clé (vide si aucun)."""
        cle_norm = normaliser_ref(cle)
        return [
            (m, q)
            for m, q in self.cx.execute(
                "SELECT membre, quantite_membre FROM composes "
                "WHERE cle_besoin=? ORDER BY ordre", (cle_norm,)
            ).fetchall()
        ]

    # ------------------------------------------------------------------
    # Sorties
    # ------------------------------------------------------------------
    def ecrire_a_confirmer(self, dossier_referentiel, nom_fichier: str = "A_confirmer.xlsx") -> Path | None:
        """
        Régénère referentiel/A_confirmer.xlsx (ou `nom_fichier`, ex.
        "A_confirmer_BL.xlsx" pour le rapprochement des BL — voir
        moteur.rapprochement.pipeline_bl — un fichier À PART pour ne pas
        écraser les propositions de l'autre flux, chacun régénérant le
        sien à chaque exécution avec SES SEULES propositions en attente)
        avec les propositions de cette exécution qui ne sont toujours pas
        tranchées dans `alias` (une proposition déjà confirmée à une
        exécution précédente n'y réapparaît pas : elle est passée en
        "connu").

        Fournisseur/N° devis en tête de ligne : sans ça, l'acheteur n'a
        aucun moyen de savoir dans quel PDF chercher pour vérifier une
        proposition (retour terrain : "ce tableau est incompréhensible").
        """
        dossier = Path(dossier_referentiel)
        dossier.mkdir(exist_ok=True)
        fichier = dossier / nom_fichier

        en_attente = []
        for ref_norm, p in self._propositions.items():
            # Un auto-alias 'devis' (voir resoudre(), branche "nouveau")
            # n'est PAS une décision de l'acheteur — juste un mémo créé
            # par la boucle articles de comparer() pour cette même
            # référence. Sans ce filtre, une proposition par désignation
            # (moteur/comparateur.py, besoin sans référence) n'apparaissait
            # JAMAIS dans A_confirmer.xlsx : la référence de l'article
            # était déjà auto-aliasée avant que la boucle besoin ne
            # tente son rapprochement.
            deja = self.cx.execute(
                "SELECT 1 FROM alias WHERE reference_brute=? AND origine != 'devis'",
                (ref_norm,),
            ).fetchone()
            if not deja:
                en_attente.append(p)

        if not en_attente:
            if fichier.exists():
                fichier.unlink()
            return None

        en_attente.sort(key=lambda p: (p["fournisseur"], p["reference"]))

        wb = Workbook()
        ws = wb.active
        ws.title = "À confirmer"

        entetes = [
            "Fournisseur", "N° devis", "Référence détectée", "Désignation détectée",
            "Qté détectée", "Référence déjà connue (proposée)", "Désignation de cette référence",
            "Qté demandée (besoin)", "Ressemblance", "Décision",
        ]
        ws.append(entetes)

        ws.cell(row=1, column=entetes.index("Référence déjà connue (proposée)") + 1).comment = Comment(
            "Référence déjà présente dans le référentiel (base achats, ou "
            "confirmée lors d'une consultation précédente) qui ressemble le "
            "plus à la référence détectée dans le devis.",
            "Consultation AI", height=120, width=260,
        )
        ws.cell(row=1, column=entetes.index("Décision") + 1).comment = Comment(
            "OUI  = confirme le rapprochement proposé\n"
            "NON  = rejette (ne sera plus reproposé)\n"
            "Autre texte = colle la bonne référence si le rapprochement "
            "proposé est faux\n\n"
            "Ouvre le PDF indiqué en colonne A/B pour vérifier chaque ligne. "
            "Enregistre ce fichier : les décisions sont appliquées à la "
            "PROCHAINE génération du comparatif (pas besoin de relancer "
            "tout de suite).",
            "Consultation AI", height=200, width=280,
        )

        for p in en_attente:
            ws.append([
                p["fournisseur"], p["devis"], p["reference"], p["designation_devis"],
                p["qte_devis"], p["cle"], p["designation_candidat"], p["qte_besoin"],
                round(p["score"] * 100) if p["score"] is not None else None, "",
            ])

        for cellule in ws[1]:
            cellule.font = Font(bold=True)
            cellule.fill = PatternFill(fill_type="solid", fgColor="EFEFEF")
            cellule.alignment = Alignment(horizontal="center", wrap_text=True)

        col_ressemblance = entetes.index("Ressemblance") + 1
        for ligne in ws.iter_rows(min_row=2, min_col=col_ressemblance, max_col=col_ressemblance):
            for cellule in ligne:
                if cellule.value is not None:
                    cellule.number_format = '0"%"'

        for colonne in ws.columns:
            longueur = max(
                (len(str(c.value)) for c in colonne if c.value is not None), default=0,
            )
            ws.column_dimensions[colonne[0].column_letter].width = min(longueur + 3, 45)

        ws.freeze_panes = "A2"

        wb.save(fichier)
        print(f"{len(en_attente)} rapprochement(s) à confirmer : {fichier}")
        return fichier

    def exporter_apprentissage(self, dossier_referentiel) -> None:
        """
        Régénère (à chaque exécution) les 3 CSV de referentiel/exports/ —
        graine du futur référentiel Appro-Tracker :
          aliases_appris.csv    : alias origine != 'import'
          composes.csv          : tous les composés (manuels + appris)
          articles_nouveaux.csv : articles source='devis'
        """
        dossier = Path(dossier_referentiel) / "exports"
        dossier.mkdir(parents=True, exist_ok=True)

        alias_appris = self.cx.execute(
            "SELECT reference_brute, cle_normalisee, origine FROM alias "
            "WHERE origine <> 'import' ORDER BY reference_brute"
        ).fetchall()
        self._ecrire_csv(
            dossier / "aliases_appris.csv",
            "# Alias appris en consultation (hors import initial de la BDD).\n"
            "# Reference_brute;Cle_normalisee;Origine "
            "(confirme = validé/corrigé par l'acheteur, "
            "confirme_nouveau = rejet explicite, devient sa propre clé)\n",
            ["Reference_brute", "Cle_normalisee", "Origine"],
            alias_appris,
        )

        composes = self.cx.execute(
            "SELECT cle_besoin, membre, quantite_membre, origine FROM composes "
            "ORDER BY cle_besoin, ordre"
        ).fetchall()
        self._ecrire_csv(
            dossier / "composes.csv",
            "# Composés (un besoin -> plusieurs références), manuels + appris.\n"
            "# Cle_besoin;Membre;Quantite_membre;Origine\n",
            ["Cle_besoin", "Membre", "Quantite_membre", "Origine"],
            composes,
        )

        nouveaux = self.cx.execute(
            "SELECT cle_normalisee, reference, fournisseur, fabricant, designation "
            "FROM articles WHERE source='devis' ORDER BY cle_normalisee"
        ).fetchall()
        self._ecrire_csv(
            dossier / "articles_nouveaux.csv",
            "# Références rencontrées en devis, absentes de la BDD achats.\n"
            "# Cle_normalisee;Reference;Fournisseur;Fabricant;Designation\n",
            ["Cle_normalisee", "Reference", "Fournisseur", "Fabricant", "Designation"],
            nouveaux,
        )

    @staticmethod
    def _ecrire_csv(chemin, entete_doc, colonnes, lignes):
        with open(chemin, "w", encoding="utf-8-sig", newline="") as f:
            f.write(entete_doc)
            w = csv.writer(f, delimiter=";")
            w.writerow(colonnes)
            w.writerows(lignes)

    def fermer(self):
        self.cx.close()
