# -*- coding: utf-8 -*-
"""
État mensuel des Factures Non Parvenues (FNP) — clôture comptable, demande
directe de la DAF le 31/08/2026 (direction en copie). Deux volets, dans le
même classeur de sortie, mais SANS AUCUN RAPPORT entre eux (confirmé par
l'acheteur, 2026-09-01, cadrage avant code) :

(a) BL non facturés : lignes du Suivi commandes livrées au plus tard le
    dernier jour du mois M mais pas encore facturées (voir
    lire_lignes_bl_non_facturees).
(b) Transitaires en cours de facturation : dossiers de
    "1.3.0. Suivi commandes spéciales.xlsm" arrivés au plus tard fin de mois M
    mais dont la facture transitaire n'a pas encore été reçue/traitée (voir
    lire_dossiers_transitaires_non_factures).

    Le lien entre les deux volets n'existe PAS : la facture du transitaire
    (transport/douane) est indépendante de la facture du fournisseur de la
    marchandise elle-même. Mot de l'acheteur, cadrage : "les n° de commande
    sont liés aux fournisseurs habituels, pas aux transitaires [...] ce
    tableau n'a d'intérêt que pour la partie FNP des transitaires" — un
    dossier de Commandes spéciales sans N° de commande lié (8 des 31 vus au
    cadrage) est donc inclus au même titre que les autres, identifié par
    Désignation/Chantier/Fournisseur/N° dossier revient plutôt que par un
    N° de commande.

LECTURE SEULE partout : ce module n'écrit JAMAIS dans le Suivi commandes ni
dans Commandes spéciales — seulement dans rapports/FNP_<AAAA-MM>.xlsx. Toute
lecture passe par une COPIE temporaire (jamais le fichier vivant ouvert
directement), même prudence que moteur.rapprochement.ecriture pour la
lecture (voir lire_entetes) même si ici aucun verrou n'est nécessaire
puisqu'on n'écrit rien dedans.

Colonnes du Suivi volontairement PAS réutilisées, vérifié au cadrage (formules
relues sur le vrai classeur avant d'écrire une ligne de code) : "Reste à
facturer" (raisonne sur le RELIQUAT non livré, pas sur le livré non facturé),
"Potentiel factu" (projection sur la Qté COMMANDÉE, pas livrée), "Facturé et
livré OK" / "Problème" / "Att réc" (contrôle prix/complétude de livraison,
jamais la présence d'une facture) — même leçon que "Statut commande" (voir
CLAUDE.md, refondu autour du contrôle de prix mi-2026) : aucune colonne
existante ne répond à la question posée par la DAF, ce module reconstruit le
périmètre depuis les colonnes brutes (Date de livraison, Qté livrée, Tarif
BL, Tarif convenu, N° facture, Date facture, Note).
"""

import calendar
import contextlib
import csv
import os
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from moteur.excel import GRIS, JAUNE, ROUGE, _entete, _largeurs
from moteur.outils import to_float
from moteur.rapprochement.pipeline_bl import _parser_date_bl, trouver_fichier_suivi_vivant

FEUILLE_COMMANDES = "Commandes"
NOTE_COMMANDE_ANNULEE = "Commande annulée"

DOSSIER_A_TRAITER_FACTURES = "a_traiter/Factures"
DOSSIER_A_VERIFIER_FACTURES = "À vérifier"  # sous-dossier, voir moteur.rapprochement.pipeline_facture
DOSSIER_REFERENTIEL = "referentiel"

# Date de création des 5 colonnes facture dans le Suivi vivant (voir
# CLAUDE.md, "colonnes créées dans le Suivi vivant", écriture réelle du
# 2026-09-01) : toute ligne livrée AVANT cette date n'a simplement jamais eu
# l'occasion d'être marquée facturée par l'outil, quel que soit son statut
# réel en compta (peut-être déjà réglée par ailleurs) — ne pas la présenter
# comme une anomalie au même titre qu'une ligne récente.
DATE_CREATION_COLONNES_FACTURE = date(2026, 9, 1)

NOM_FICHIER_COMMANDES_SPECIALES = "1.3.0. Suivi commandes spéciales.xlsm"
FEUILLE_SPECIALES = "Suivi"

DOSSIER_RAPPORTS = "rapports"


class SuiviIntrouvable(Exception):
    """Le Suivi commandes vivant est introuvable — voir
    moteur.rapprochement.pipeline_bl.trouver_fichier_suivi_vivant."""


def dernier_jour_mois(annee_mois: str) -> date:
    """"2026-08" -> date(2026, 8, 31)."""
    annee, mois = (int(x) for x in annee_mois.split("-"))
    dernier = calendar.monthrange(annee, mois)[1]
    return date(annee, mois, dernier)


def mois_precedent_complet(reference: date | None = None) -> str:
    """"2026-08" si `reference` (par défaut aujourd'hui) est en septembre
    2026, peu importe le jour — le mois calendaire précédent est TOUJOURS
    complet, jamais le mois en cours. Sert de valeur pré-remplie dans le GUI
    (voir gui_fnp.py), jamais utilisé pour un calcul de périmètre lui-même."""

    ref = reference or date.today()
    if ref.month == 1:
        return f"{ref.year - 1}-12"
    return f"{ref.year}-{ref.month - 1:02d}"


@contextlib.contextmanager
def _copie_temporaire(fichier):
    """Copie temporaire à USAGE UNIQUE de `fichier` (jamais un nom fixe basé
    sur fichier.name — deux lectures concurrentes/rapprochées d'un même nom
    de fichier source, ex. plusieurs tests utilisant tous "suivi.xlsx",
    se marchaient dessus : bug réel trouvé en écrivant les tests, la 2e
    copie écrasait la 1re pendant qu'un handle Windows la tenait encore
    ouverte). Nettoyage tolérant : un verrou Windows résiduel sur un fichier
    TEMPORAIRE ET JETABLE (le nettoyage périodique de l'OS s'en charge) ne
    doit jamais faire échouer la lecture elle-même."""

    fichier = Path(fichier)
    descripteur, chemin_tmp = tempfile.mkstemp(suffix=fichier.suffix, prefix="fnp_lecture_")
    os.close(descripteur)
    tmp = Path(chemin_tmp)
    shutil.copy2(fichier, tmp)
    try:
        yield tmp
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass


def _vers_date(valeur):
    """Cellule Excel (datetime natif via data_only, texte, ou vide) -> date
    Python, ou None. Les dates de Commandes spéciales/Suivi commandes sont
    toujours des datetime natifs en pratique (vérifié au cadrage) ; le repli
    texte est défensif, même prudence que
    moteur.rapprochement.pipeline_bl._parser_date_bl."""

    if valeur is None or valeur == "":
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    texte = str(valeur).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texte, fmt).date()
        except ValueError:
            continue
    return None


@dataclass
class LigneFNP:

    ligne_excel: int
    fournisseur: str
    numero_commande: str
    chantier: str
    reference: str
    designation: str
    qte_livree: float
    montant_ht: float
    source_prix: str  # "Tarif BL" / "Tarif convenu" / "Aucune"
    date_livraison: date
    anciennete_jours: int
    numero_facture: str
    date_facture: object
    note: str


@dataclass
class DossierTransitaire:

    numero_dossier: str
    designation: str
    numero_commande: str
    chantier: str
    fournisseur: str
    transitaire: str
    ref_transport: str
    date_depart: object
    date_arrivee: date
    montant_marchandise: float
    cout_estime: float | None  # None si non calculable (voir repli)
    anciennete_jours: int


@dataclass
class FactureRecueNonRapprochee:
    """Une ligne du Suivi livrée (donc candidate au volet a) mais pour
    laquelle une VRAIE facture PDF a été trouvée dans a_traiter/Factures/
    (racine ou À vérifier/) — datée <= fin de mois — sans avoir encore été
    écrite dans le Suivi (voir _identifier_lignes_excel_facturees). Sortie
    du volet (a), listée séparément : ce n'est PLUS une facture non
    parvenue, seulement pas encore rapprochée dans l'outil (session S0,
    étape 4a)."""

    ligne_excel: int
    fournisseur: str
    numero_commande: str
    reference: str
    numero_facture: str
    date_facture: date | None
    montant_facture_bl: float  # "Facturé BL" déjà calculé (même valeur que LigneFNP.montant_ht)


@dataclass
class AjustementFNP:
    """Une ligne déclarée À LA MAIN par l'acheteur dans
    referentiel/fnp_ajustements_<mois>.csv (jamais par l'outil lui-même,
    voir lire_ajustements_fnp) — un cas connu que le calcul automatique ne
    peut structurellement pas couvrir (bon manuel, dossier transitaire hors
    Commandes spéciales...). Listée à part, JAMAIS fusionnée avec les
    totaux calculés des volets (a)/(b) — décision explicite de cadrage
    (session S0, étape 4b)."""

    type: str  # "BDC_MANUEL" / "TRANSIT" / "AUTRE"
    libelle: str
    fournisseur_ou_transitaire: str
    chantier: str
    piece: str
    date_livraison: date | None
    montant_ht: float
    source: str
    commentaire: str


@dataclass
class ReservesFNP:
    """Réserves de périmètre AUTOMATIQUES (session S0, étape 4c) — ce que
    le calcul ne couvre structurellement pas, jamais un chiffre deviné."""

    n_bdc_manuel_24x: int  # factures dont au moins un bloc est un bon manuel "BC/BCN 24XXXX" (voir CauseFacture.BDC_MANUEL_24X) — matériel livré, commande absente du Suivi, hors périmètre du volet (a)
    n_transitaires_sans_estimation: int  # dossiers de Commandes spéciales non facturés avec ETA <= fin de mois mais Coût estimé absent (déjà compté côté volet b, répété ici pour la section Réserves)
    n_dossiers_speciales_total: int  # nombre TOTAL de dossiers dans Commandes spéciales (tout statut confondu) — rappel que le volet (b) ne couvre que ce qui y est saisi à la main


@dataclass
class RapportFNP:

    mois: str
    fin_de_mois: date
    date_generation: datetime
    chemin_suivi: Path
    suivi_modifie_le: datetime
    depuis: date | None

    lignes_bl: list = field(default_factory=list)          # LigneFNP valorisées
    lignes_sans_prix: list = field(default_factory=list)    # LigneFNP, montant_ht == 0, source_prix == "Aucune"

    dossiers_transitaires: list = field(default_factory=list)   # DossierTransitaire
    chemin_commandes_speciales: Path | None = None
    transitaire_repli_utilise: bool = False
    transitaire_avertissement: str = ""
    commandes_transitaire_non_couvertes: list = field(default_factory=list)  # [(numero_dossier, numero_commande, fournisseur, chantier)]

    # --- v1.1 (session S0) ---
    exclusion_appliquee: bool = True  # False si --sans-exclusion (repli 16h, voir CLAUDE.md)
    factures_recues_non_rapprochees: list = field(default_factory=list)  # FactureRecueNonRapprochee
    ajustements: list = field(default_factory=list)  # AjustementFNP
    chemin_ajustements: Path | None = None
    reserves: ReservesFNP | None = None


def _valeur_texte(v) -> str:
    return str(v).strip() if v is not None else ""


def lire_lignes_bl_non_facturees(chemin_suivi, fin_de_mois: date, depuis: date | None = None):
    """(lignes valorisées, lignes sans prix connu, transitaires_vus) — voir
    RapportFNP pour les deux premières. `transitaires_vus` (3e élément) :
    [(numero_dossier_revient, numero_commande, fournisseur, chantier), ...]
    pour CHAQUE ligne du Suivi (toute date/statut confondu) où "Transitaire"
    est renseigné — collecté dans la même passe pour le contrôle de
    couverture du volet (b) (voir _controler_couverture_transitaires),
    plutôt que relire tout le classeur une 2e fois pour ça.

    Périmètre (voir bandeau du module) : Qté livrée > 0, Date de livraison
    <= fin_de_mois (et >= depuis si fourni), Note != "Commande annulée", et
    (N° facture vide OU Date facture > fin_de_mois). Les autres valeurs
    "magiques" de Note ("Rupture fournisseur", "Reliquat soldé") ne sont PAS
    des motifs d'exclusion ici — seule "Commande annulée" l'est (périmètre
    donné par la DAF) ; elles sont juste reportées telles quelles pour info."""

    with _copie_temporaire(chemin_suivi) as tmp:
        wb = load_workbook(tmp, read_only=True, data_only=True)
        try:
            ws = wb[FEUILLE_COMMANDES]
            lignes = ws.iter_rows(values_only=True)
            entetes = next(lignes)
            idx = {e: i for i, e in enumerate(entetes) if e is not None}

            requis = (
                "Fournisseur", "N° de commande", "Chantier", "Référence", "Désignation",
                "Qté livrée", "Tarif BL", "Tarif convenu", "Date de livraison", "Note",
                "N° facture", "Date facture", "Facturé BL", "Transitaire",
            )
            manquantes = [c for c in requis if c not in idx]
            if manquantes:
                raise KeyError(
                    f"Colonne(s) introuvable(s) dans « {FEUILLE_COMMANDES} » : {', '.join(manquantes)} "
                    "— export du Suivi commandes différent de celui attendu ?"
                )
            # "N° dossier revient" n'est utile qu'au contrôle de couverture du
            # volet (b) (voir _controler_couverture_transitaires) — optionnelle
            # ici, jamais un motif de faire échouer tout le volet (a).
            i_dossier = idx.get("N° dossier revient")

            valorisees, sans_prix, transitaires_vus = [], [], []

            for i, row in enumerate(lignes, start=2):

                transitaire = _valeur_texte(row[idx["Transitaire"]])
                if transitaire:
                    transitaires_vus.append((
                        _valeur_texte(row[i_dossier]) if i_dossier is not None else "",
                        _valeur_texte(row[idx["N° de commande"]]),
                        _valeur_texte(row[idx["Fournisseur"]]),
                        _valeur_texte(row[idx["Chantier"]]),
                    ))

                note = _valeur_texte(row[idx["Note"]])
                if note == NOTE_COMMANDE_ANNULEE:
                    continue

                qte_livree = to_float(row[idx["Qté livrée"]])
                if qte_livree <= 0:
                    continue

                date_livraison = _vers_date(row[idx["Date de livraison"]])
                if date_livraison is None or date_livraison > fin_de_mois:
                    continue
                if depuis is not None and date_livraison < depuis:
                    continue

                numero_facture = _valeur_texte(row[idx["N° facture"]])
                date_facture = _vers_date(row[idx["Date facture"]])

                if numero_facture and (date_facture is None or date_facture <= fin_de_mois):
                    # Un N° facture présent vaut exclusion par défaut — y
                    # compris si la date est illisible/absente (cas réel :
                    # pipeline_facture.ecritures_pour_facture n'écrit la
                    # date QUE si elle a pu être parsée, le N° facture lui
                    # est toujours écrit) : mieux vaut la traiter en
                    # "déjà facturée" que risquer un doublon. Seule une date
                    # facture CONFIRMÉE après la clôture fait rester la
                    # ligne une FNP malgré un N° facture déjà présent.
                    continue

                tarif_bl = to_float(row[idx["Tarif BL"]])
                tarif_convenu = to_float(row[idx["Tarif convenu"]])

                if tarif_bl:
                    source = "Tarif BL"
                elif tarif_convenu:
                    source = "Tarif convenu"
                else:
                    source = "Aucune"

                l = LigneFNP(
                    ligne_excel=i,
                    fournisseur=_valeur_texte(row[idx["Fournisseur"]]),
                    numero_commande=_valeur_texte(row[idx["N° de commande"]]),
                    chantier=_valeur_texte(row[idx["Chantier"]]),
                    reference=_valeur_texte(row[idx["Référence"]]),
                    designation=_valeur_texte(row[idx["Désignation"]]),
                    qte_livree=qte_livree,
                    # Réutilise la valeur DÉJÀ CALCULÉE de "Facturé BL" (Tarif BL
                    # x Qté livrée, replié sur Tarif convenu) plutôt que de la
                    # recalculer soi-même (demande explicite de la DAF/du
                    # cadrage) — "source_prix" ci-dessus n'est qu'une étiquette
                    # de provenance, jamais un recalcul du montant.
                    montant_ht=to_float(row[idx["Facturé BL"]]),
                    source_prix=source,
                    date_livraison=date_livraison,
                    anciennete_jours=(fin_de_mois - date_livraison).days,
                    numero_facture=numero_facture,
                    date_facture=date_facture,
                    note=note,
                )

                (sans_prix if source == "Aucune" else valorisees).append(l)

            return valorisees, sans_prix, transitaires_vus
        finally:
            wb.close()


def lire_dossiers_transitaires_non_factures(chemin_commandes_speciales, fin_de_mois: date):
    """(dossiers, avertissement, dossiers_connus) — dossiers de la feuille
    "Suivi" de Commandes spéciales dont "ETA ou arrivée réelle" <=
    fin_de_mois et "Expédition facturée" == 0 (confirmé par l'acheteur,
    cadrage : 1 = facture transitaire déjà reçue/traitée, 0 = en attente ;
    "ETA ou arrivée réelle" utilisable telle quelle pour ce cut-off, malgré
    son nom qui mélange estimé/réel). `dossiers_connus` (3e élément) :
    ensemble des N° dossier revient présents dans ce classeur (toute date/
    statut confondu), pour le contrôle de couverture (voir
    _controler_couverture_transitaires). `avertissement` non vide si le
    classeur/la feuille/une colonne requise est introuvable — dans ce cas
    `dossiers` est toujours vide, à l'appelant de décider d'un repli."""

    chemin = Path(chemin_commandes_speciales)
    if not chemin.exists():
        return [], f"Classeur « {chemin.name} » introuvable ({chemin}) — volet transitaires non évalué.", set()

    with _copie_temporaire(chemin) as tmp:
        wb = load_workbook(tmp, read_only=True, data_only=True, keep_vba=False)
        try:
            if FEUILLE_SPECIALES not in wb.sheetnames:
                return [], f"Feuille « {FEUILLE_SPECIALES} » introuvable dans « {chemin.name} ».", set()

            ws = wb[FEUILLE_SPECIALES]
            lignes = ws.iter_rows(values_only=True)
            entetes = next(lignes)
            idx = {e: i for i, e in enumerate(entetes) if e is not None}

            requis = (
                "Désignation", "Commande", "Chantier", "Fournisseur", "N° dossier revient",
                "Montant commande", "Transitaire", "ETA ou arrivée réelle", "Expédition facturée",
            )
            manquantes = [c for c in requis if c not in idx]
            if manquantes:
                return [], f"Colonne(s) introuvable(s) dans « {chemin.name} » : {', '.join(manquantes)}.", set()

            dossiers = []
            dossiers_connus = set()

            for row in lignes:

                numero_dossier = _valeur_texte(row[idx["N° dossier revient"]])
                if numero_dossier:
                    dossiers_connus.add(numero_dossier)

                date_arrivee = _vers_date(row[idx["ETA ou arrivée réelle"]])
                if date_arrivee is None or date_arrivee > fin_de_mois:
                    continue

                facture = row[idx["Expédition facturée"]]
                if to_float(facture) != 0:
                    continue

                dossiers.append(DossierTransitaire(
                    numero_dossier=numero_dossier,
                    designation=_valeur_texte(row[idx["Désignation"]]),
                    numero_commande=_valeur_texte(row[idx["Commande"]]),
                    chantier=_valeur_texte(row[idx["Chantier"]]),
                    fournisseur=_valeur_texte(row[idx["Fournisseur"]]),
                    transitaire=_valeur_texte(row[idx["Transitaire"]]),
                    ref_transport=_valeur_texte(row[idx["Réf trsprt"]]) if "Réf trsprt" in idx else "",
                    date_depart=_vers_date(row[idx["Date de départ"]]) if "Date de départ" in idx else None,
                    date_arrivee=date_arrivee,
                    montant_marchandise=to_float(row[idx["Montant commande"]]),
                    cout_estime=(
                        to_float(row[idx["Coût estimé"]])
                        if "Coût estimé" in idx and row[idx["Coût estimé"]] is not None else None
                    ),
                    anciennete_jours=(fin_de_mois - date_arrivee).days,
                ))

            return dossiers, "", dossiers_connus
        finally:
            wb.close()


def _repli_transitaires_suivi_principal(chemin_suivi, fin_de_mois: date) -> list:
    """Repli SI Commandes spéciales est introuvable/inexploitable (voir
    lire_dossiers_transitaires_non_factures) : lignes du Suivi principal
    livrées <= fin_de_mois avec un Transitaire renseigné. Aucun taux
    d'approche moyen par trajet disponible sans le classeur dédié (feuille
    Analyse) — jamais une estimation fabriquée sans base réelle (règle d'or
    du projet) : cout_estime reste None, signalé comme tel dans le rapport."""

    with _copie_temporaire(chemin_suivi) as tmp:
        wb = load_workbook(tmp, read_only=True, data_only=True)
        try:
            ws = wb[FEUILLE_COMMANDES]
            lignes = ws.iter_rows(values_only=True)
            entetes = next(lignes)
            idx = {e: i for i, e in enumerate(entetes) if e is not None}

            requis = ("Transitaire", "N° de commande", "Fournisseur", "Chantier", "Qté livrée", "Date de livraison", "Montant total commande")
            if any(c not in idx for c in requis):
                return []

            dossiers = []
            for row in lignes:
                transitaire = _valeur_texte(row[idx["Transitaire"]])
                if not transitaire:
                    continue
                if to_float(row[idx["Qté livrée"]]) <= 0:
                    continue
                date_livraison = _vers_date(row[idx["Date de livraison"]])
                if date_livraison is None or date_livraison > fin_de_mois:
                    continue

                dossiers.append(DossierTransitaire(
                    numero_dossier="",
                    designation="",
                    numero_commande=_valeur_texte(row[idx["N° de commande"]]),
                    chantier=_valeur_texte(row[idx["Chantier"]]),
                    fournisseur=_valeur_texte(row[idx["Fournisseur"]]),
                    transitaire=transitaire,
                    ref_transport="",
                    date_depart=None,
                    date_arrivee=date_livraison,
                    montant_marchandise=to_float(row[idx["Montant total commande"]]),
                    cout_estime=None,
                    anciennete_jours=(fin_de_mois - date_livraison).days,
                ))

            return dossiers
        finally:
            wb.close()


def _controler_couverture_transitaires(transitaires_vus_suivi, dossiers_connus_speciales) -> list:
    """Commandes du Suivi principal qui portent un Transitaire ET un N°
    dossier revient renseigné, mais dont ce numéro n'apparaît dans AUCUN
    dossier de Commandes spéciales : un vrai trou de couverture (le dossier
    existe côté Suivi principal, jamais saisi côté Commandes spéciales) —
    jamais un rapprochement inventé sans clé fiable commune (voir cadrage :
    le N° de commande n'est pas cette clé, "les n° de commande sont liés aux
    fournisseurs habituels, pas aux transitaires"), donc uniquement les
    lignes où le N° dossier revient LUI-MÊME est renseigné mais absent de
    Commandes spéciales. Les lignes du Suivi avec Transitaire renseigné mais
    SANS N° dossier revient ne sont ni confirmées ni infirmées ici (aucune
    clé pour vérifier) — le message général sur Commandes spéciales "peu
    alimenté" couvre déjà ce risque plus large, pas la peine de fabriquer un
    faux signal ligne à ligne dessus."""

    return [
        t for t in transitaires_vus_suivi
        if t[0] and t[0] not in dossiers_connus_speciales
    ]


# --- v1.1 (session S0) : exclusion "facture reçue non rapprochée" (4a) -----


def _identifier_lignes_excel_facturees(dossier_projet, fin_de_mois: date) -> tuple:
    """Scanne a_traiter/Factures/ (racine ET À vérifier/, LECTURE SEULE via
    moteur.rapprochement.pipeline_facture.rapprocher_dossier_factures —
    jamais de modification, même mécanisme déjà éprouvé en F2/F4) et
    retourne ({ligne_excel: info}, n_bdc_manuel_24x) :

    - `info` (dict) pour CHAQUE correspondance sûre/à confirmer/déjà à jour
      dont la facture est datée <= fin_de_mois (une facture non datée reste
      incluse — mieux vaut la signaler que la perdre, même logique que
      _comparer_facture côté matching) : numero_facture, date_facture,
      reference, fournisseur, numero_commande. Une même ligne_excel visée
      par PLUSIEURS factures garde la PREMIÈRE trouvée (cas rare, pas de
      règle d'arbitrage inventée au-delà de "la première").
    - `n_bdc_manuel_24x` : nombre de blocs anomalies classés
      CauseFacture.BDC_MANUEL_24X sur les DEUX dossiers scannés (voir
      ReservesFNP, étape 4c) — compté dans la même passe pour ne jamais
      rescanner un dossier potentiellement lourd (OCR) deux fois.

    Ne lève JAMAIS d'exception si a_traiter/Factures/ n'existe pas encore
    (aucune facture jamais déposée) — retourne ({}, 0), le volet (a) reste
    alors inchangé (comportement identique à --sans-exclusion)."""

    from moteur.rapprochement.matching_facture import CauseFacture
    from moteur.rapprochement.pipeline_facture import (
        StatutFacture,
        classifier_cause_anomalie,
        rapprocher_dossier_factures,
    )

    dossier_a_traiter = Path(dossier_projet) / DOSSIER_A_TRAITER_FACTURES
    if not dossier_a_traiter.is_dir():
        return {}, 0

    dossiers_a_scanner = [dossier_a_traiter]
    dossier_a_verifier = dossier_a_traiter / DOSSIER_A_VERIFIER_FACTURES
    if dossier_a_verifier.is_dir():
        dossiers_a_scanner.append(dossier_a_verifier)

    lignes_excel_facturees = {}
    n_bdc_manuel_24x = 0

    for dossier in dossiers_a_scanner:

        rapport_facture = rapprocher_dossier_factures(dossier, dossier_projet)

        for facture, c in rapport_facture.surs + rapport_facture.a_confirmer + rapport_facture.deja_a_jour:
            if c.ligne_suivi is None:
                continue
            date_f = _parser_date_bl(facture.date_facture)
            if date_f is not None and date_f > fin_de_mois:
                continue
            lignes_excel_facturees.setdefault(c.ligne_suivi.ligne_excel, {
                "numero_facture": facture.numero_facture,
                "date_facture": date_f,
                "reference": c.ligne_facture.reference_fournisseur,
                "fournisseur": facture.fournisseur,
                "numero_commande": c.ligne_facture.numero_commande,
            })

        for _, raison in rapport_facture.anomalies_facture:
            if classifier_cause_anomalie(raison) is CauseFacture.BDC_MANUEL_24X:
                n_bdc_manuel_24x += 1

    return lignes_excel_facturees, n_bdc_manuel_24x


def _appliquer_exclusion_factures_recues(lignes_bl: list, lignes_sans_prix: list,
                                          lignes_excel_facturees: dict) -> tuple:
    """Fonction PURE (aucune I/O — voir _identifier_lignes_excel_facturees
    pour le scan lui-même) : retire de `lignes_bl`/`lignes_sans_prix` toute
    ligne dont `ligne_excel` figure dans `lignes_excel_facturees`, retourne
    (lignes_bl_restantes, lignes_sans_prix_restantes, factures_recues)."""

    factures_recues = []
    lignes_bl_restantes = []
    lignes_sans_prix_restantes = []

    for l in lignes_bl:
        info = lignes_excel_facturees.get(l.ligne_excel)
        if info is None:
            lignes_bl_restantes.append(l)
            continue
        factures_recues.append(FactureRecueNonRapprochee(
            ligne_excel=l.ligne_excel, fournisseur=l.fournisseur,
            numero_commande=l.numero_commande, reference=l.reference,
            numero_facture=info["numero_facture"], date_facture=info["date_facture"],
            montant_facture_bl=l.montant_ht,
        ))

    for l in lignes_sans_prix:
        info = lignes_excel_facturees.get(l.ligne_excel)
        if info is None:
            lignes_sans_prix_restantes.append(l)
            continue
        factures_recues.append(FactureRecueNonRapprochee(
            ligne_excel=l.ligne_excel, fournisseur=l.fournisseur,
            numero_commande=l.numero_commande, reference=l.reference,
            numero_facture=info["numero_facture"], date_facture=info["date_facture"],
            montant_facture_bl=l.montant_ht,
        ))

    return lignes_bl_restantes, lignes_sans_prix_restantes, factures_recues


# --- v1.1 (session S0) : ajustements déclarés par l'acheteur (4b) ----------


def nom_fichier_ajustements(mois: str) -> str:
    return f"fnp_ajustements_{mois}.csv"


def lire_ajustements_fnp(chemin_csv) -> list:
    """Lit referentiel/fnp_ajustements_<mois>.csv — REMPLI PAR L'ACHETEUR,
    jamais par l'outil (voir AjustementFNP). Fichier absent -> liste vide
    (personne n'a encore rien déclaré ce mois-ci, pas une erreur). Colonnes
    (séparateur ; comme les autres fichiers du projet) : type
    (BDC_MANUEL/TRANSIT/AUTRE) ; libelle ; fournisseur_ou_transitaire ;
    chantier ; piece ; date_livraison (JJ/MM/AAAA) ; montant_ht ; source ;
    commentaire. Une ligne sans "type" est ignorée (ligne d'exemple/vide)."""

    chemin_csv = Path(chemin_csv)
    if not chemin_csv.is_file():
        return []

    resultat = []
    with open(chemin_csv, encoding="utf-8-sig") as f:
        lignes_utiles = (l for l in f if not l.lstrip().startswith("#"))
        for ligne in csv.DictReader(lignes_utiles, delimiter=";"):
            type_ = (ligne.get("type") or "").strip().upper()
            if not type_:
                continue
            resultat.append(AjustementFNP(
                type=type_,
                libelle=(ligne.get("libelle") or "").strip(),
                fournisseur_ou_transitaire=(ligne.get("fournisseur_ou_transitaire") or "").strip(),
                chantier=(ligne.get("chantier") or "").strip(),
                piece=(ligne.get("piece") or "").strip(),
                date_livraison=_vers_date((ligne.get("date_livraison") or "").strip() or None),
                montant_ht=to_float(ligne.get("montant_ht")),
                source=(ligne.get("source") or "").strip(),
                commentaire=(ligne.get("commentaire") or "").strip(),
            ))
    return resultat


# --- v1.1 (session S0) : réserves de périmètre (4c) -------------------------


def compter_dossiers_speciales(chemin_commandes_speciales) -> int:
    """Nombre TOTAL de dossiers dans Commandes spéciales (tout statut
    confondu, pas seulement les non-facturés du volet b) — rappel de
    l'ampleur réelle de ce classeur peu alimenté (voir ReservesFNP).
    0 si le classeur/la feuille est introuvable (jamais bloquant)."""

    chemin = Path(chemin_commandes_speciales)
    if not chemin.exists():
        return 0

    with _copie_temporaire(chemin) as tmp:
        wb = load_workbook(tmp, read_only=True, data_only=True, keep_vba=False)
        try:
            if FEUILLE_SPECIALES not in wb.sheetnames:
                return 0
            ws = wb[FEUILLE_SPECIALES]
            lignes = ws.iter_rows(values_only=True)
            next(lignes, None)  # en-têtes
            return sum(1 for row in lignes if any(v not in (None, "") for v in row))
        finally:
            wb.close()


def calculer_rapport_fnp(dossier_projet, mois: str, depuis: date | None = None,
                          appliquer_exclusion: bool = True) -> RapportFNP:
    """Lecture seule, calcule le RapportFNP sans rien écrire — séparé de
    generer_etat_fnp() pour que le GUI (voir gui_fnp.py) puisse afficher un
    résumé à l'écran à partir des mêmes données que celles écrites dans le
    classeur, sans avoir à rouvrir le fichier généré pour les relire.

    `appliquer_exclusion=False` (option --sans-exclusion, voir fnp.py CLI) :
    désactive l'étape 4a (exclusion "facture reçue non rapprochée") — repli
    prévu si a_traiter/Factures/ est trop volumineux/lent à scanner (OCR) ;
    les étapes 4b/4c restent actives dans les deux cas."""

    dossier_projet = Path(dossier_projet)
    fin_de_mois = dernier_jour_mois(mois)

    chemin_suivi = trouver_fichier_suivi_vivant(dossier_projet)
    if chemin_suivi is None:
        raise SuiviIntrouvable(
            "Le Suivi commandes vivant est introuvable dans "
            "« 1.3.0.1. Commandes courantes/ », à côté du dossier projet."
        )
    suivi_modifie_le = datetime.fromtimestamp(chemin_suivi.stat().st_mtime)

    lignes_valorisees, lignes_sans_prix, transitaires_vus = lire_lignes_bl_non_facturees(
        chemin_suivi, fin_de_mois, depuis,
    )

    chemin_speciales = dossier_projet.parent / NOM_FICHIER_COMMANDES_SPECIALES
    dossiers, avertissement, dossiers_connus = lire_dossiers_transitaires_non_factures(
        chemin_speciales, fin_de_mois,
    )

    repli_utilise = False
    if avertissement:
        repli_utilise = True
        dossiers = _repli_transitaires_suivi_principal(chemin_suivi, fin_de_mois)

    non_couvertes = (
        _controler_couverture_transitaires(transitaires_vus, dossiers_connus)
        if not repli_utilise else []
    )

    # --- 4a : exclusion "facture reçue non rapprochée" ---------------------
    factures_recues = []
    n_bdc_manuel_24x = 0
    if appliquer_exclusion:
        lignes_excel_facturees, n_bdc_manuel_24x = _identifier_lignes_excel_facturees(
            dossier_projet, fin_de_mois,
        )
        lignes_valorisees, lignes_sans_prix, factures_recues = _appliquer_exclusion_factures_recues(
            lignes_valorisees, lignes_sans_prix, lignes_excel_facturees,
        )

    # --- 4b : ajustements déclarés par l'acheteur ---------------------------
    dossier_referentiel = dossier_projet / DOSSIER_REFERENTIEL
    chemin_ajustements = dossier_referentiel / nom_fichier_ajustements(mois)
    ajustements = lire_ajustements_fnp(chemin_ajustements)

    # --- 4c : réserves de périmètre -----------------------------------------
    n_transit_sans_estimation = sum(1 for d in dossiers if d.cout_estime is None) if not repli_utilise else len(dossiers)
    reserves = ReservesFNP(
        n_bdc_manuel_24x=n_bdc_manuel_24x,
        n_transitaires_sans_estimation=n_transit_sans_estimation,
        n_dossiers_speciales_total=compter_dossiers_speciales(chemin_speciales),
    )

    return RapportFNP(
        mois=mois, fin_de_mois=fin_de_mois, date_generation=datetime.now(),
        chemin_suivi=chemin_suivi, suivi_modifie_le=suivi_modifie_le, depuis=depuis,
        lignes_bl=lignes_valorisees, lignes_sans_prix=lignes_sans_prix,
        dossiers_transitaires=dossiers,
        chemin_commandes_speciales=chemin_speciales if chemin_speciales.exists() else None,
        transitaire_repli_utilise=repli_utilise, transitaire_avertissement=avertissement,
        commandes_transitaire_non_couvertes=non_couvertes,
        exclusion_appliquee=appliquer_exclusion,
        factures_recues_non_rapprochees=factures_recues,
        ajustements=ajustements, chemin_ajustements=chemin_ajustements if chemin_ajustements.exists() else None,
        reserves=reserves,
    )


def generer_etat_fnp(dossier_projet, mois: str, depuis: date | None = None,
                      appliquer_exclusion: bool = True) -> Path:
    """Point d'entrée : orchestration complète, écrit
    rapports/FNP_<mois>.xlsx et retourne son chemin. LECTURE SEULE partout —
    ce module n'écrit jamais dans un classeur vivant (Suivi commandes ou
    Commandes spéciales)."""

    dossier_projet = Path(dossier_projet)
    rapport = calculer_rapport_fnp(dossier_projet, mois, depuis, appliquer_exclusion)
    return ecrire_classeur_fnp(dossier_projet, rapport)


# --- Écriture du classeur de sortie -----------------------------------------

_NOMS_MOIS = [
    "", "janvier", "février", "mars", "avril", "mai", "juin",
    "juillet", "août", "septembre", "octobre", "novembre", "décembre",
]


def mois_en_lettres(mois: str) -> str:
    annee, m = mois.split("-")
    return f"{_NOMS_MOIS[int(m)]} {annee}"


def _euros(cellule):
    cellule.number_format = "#,##0.00 €"


def _date_fmt(cellule):
    cellule.number_format = "dd/mm/yyyy"


def _sous_total_par(lignes, champ) -> list:
    """[(valeur_champ, nb_lignes, montant_total)], trié par montant décroissant."""
    totaux = {}
    for l in lignes:
        v = getattr(l, champ) or "(vide)"
        nb, montant = totaux.get(v, (0, 0.0))
        totaux[v] = (nb + 1, montant + l.montant_ht)
    return sorted(((v, nb, m) for v, (nb, m) in totaux.items()), key=lambda x: -x[2])


def ecrire_classeur_fnp(dossier_projet: Path, rapport: RapportFNP) -> Path:

    total_bl = sum(l.montant_ht for l in rapport.lignes_bl)
    total_marchandise_transit = sum(d.montant_marchandise for d in rapport.dossiers_transitaires)
    total_estime_transit = sum(d.cout_estime for d in rapport.dossiers_transitaires if d.cout_estime is not None)
    n_transit_sans_estimation = sum(1 for d in rapport.dossiers_transitaires if d.cout_estime is None)

    anterieures_f1 = [l for l in rapport.lignes_bl if l.date_livraison < DATE_CREATION_COLONNES_FACTURE]
    total_anterieures_f1 = sum(l.montant_ht for l in anterieures_f1)

    buckets = {"< 30 j": (0, 0.0), "30-90 j": (0, 0.0), "> 90 j": (0, 0.0)}
    for l in rapport.lignes_bl:
        cle = "< 30 j" if l.anciennete_jours < 30 else ("30-90 j" if l.anciennete_jours <= 90 else "> 90 j")
        nb, montant = buckets[cle]
        buckets[cle] = (nb + 1, montant + l.montant_ht)

    wb = Workbook()

    # --- Synthèse ------------------------------------------------------
    ws = wb.active
    ws.title = "Synthèse"

    ws["A1"] = f"État des Factures Non Parvenues — {mois_en_lettres(rapport.mois)}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Généré le {rapport.date_generation.strftime('%d/%m/%Y %H:%M')}"
    ws["A3"] = f"Périmètre : livré/arrivé au plus tard le {rapport.fin_de_mois.strftime('%d/%m/%Y')} (dernier jour du mois), facture non reçue à cette date"
    if rapport.depuis:
        ws["A4"] = f"Filtre appliqué : uniquement les livraisons à partir du {rapport.depuis.strftime('%d/%m/%Y')}"

    jours_suivi = (rapport.date_generation - rapport.suivi_modifie_le).days
    ligne = 6
    ws.cell(ligne, 1, "Suivi commandes lu (copie) :").font = Font(bold=True)
    ws.cell(ligne, 2, str(rapport.chemin_suivi))
    ligne += 1
    ws.cell(ligne, 1, "Dernière sauvegarde du Suivi :")
    ws.cell(ligne, 2, rapport.suivi_modifie_le.strftime("%d/%m/%Y %H:%M"))
    ws.cell(ligne, 3, f"(il y a {jours_suivi} jour(s))")
    if jours_suivi > 1:
        ws.cell(ligne, 3).fill = JAUNE
        ws.cell(ligne, 4, "⚠ le Suivi n'a peut-être pas été (r)ouvert/enregistré récemment dans Excel — les valeurs affichées sont celles mises en cache lors de la dernière sauvegarde")
    ligne += 2

    ws.cell(ligne, 1, "VOLET (a) — BL non facturés").font = Font(bold=True, size=12)
    ligne += 1
    ws.cell(ligne, 1, "Total HT (lignes valorisées) :")
    c = ws.cell(ligne, 2, round(total_bl, 2)); _euros(c)
    ws.cell(ligne, 3, f"{len(rapport.lignes_bl)} ligne(s)")
    ligne += 1
    ws.cell(ligne, 1, "Dont livrées sans AUCUN prix connu (non valorisées, voir onglet détail) :")
    ws.cell(ligne, 3, f"{len(rapport.lignes_sans_prix)} ligne(s)")
    ligne += 1
    ws.cell(ligne, 1, f"Dont livraisons antérieures au {DATE_CREATION_COLONNES_FACTURE.strftime('%d/%m/%Y')} (avant la création des colonnes facture — jamais eu l'occasion d'être pointées par l'outil, peut-être déjà réglées par ailleurs) :")
    c = ws.cell(ligne, 2, round(total_anterieures_f1, 2)); _euros(c)
    ws.cell(ligne, 3, f"{len(anterieures_f1)} ligne(s)")
    ligne += 2

    ws.cell(ligne, 1, "Répartition par ancienneté").font = Font(bold=True)
    ligne += 1
    for cle in ("< 30 j", "30-90 j", "> 90 j"):
        nb, montant = buckets[cle]
        ws.cell(ligne, 1, cle)
        ws.cell(ligne, 2, nb)
        c = ws.cell(ligne, 3, round(montant, 2)); _euros(c)
        ligne += 1
    ligne += 1

    ws.cell(ligne, 1, "Répartition par fournisseur").font = Font(bold=True)
    ligne += 1
    debut = ligne
    ws.cell(ligne, 1, "Fournisseur"); ws.cell(ligne, 2, "Nb lignes"); ws.cell(ligne, 3, "Montant HT")
    ligne += 1
    for nom, nb, montant in _sous_total_par(rapport.lignes_bl, "fournisseur"):
        ws.cell(ligne, 1, nom)
        ws.cell(ligne, 2, nb)
        c = ws.cell(ligne, 3, round(montant, 2)); _euros(c)
        ligne += 1
    _entete(ws, debut)
    ligne += 1

    ws.cell(ligne, 1, "Répartition par chantier").font = Font(bold=True)
    ligne += 1
    debut = ligne
    ws.cell(ligne, 1, "Chantier"); ws.cell(ligne, 2, "Nb lignes"); ws.cell(ligne, 3, "Montant HT")
    ligne += 1
    for nom, nb, montant in _sous_total_par(rapport.lignes_bl, "chantier"):
        ws.cell(ligne, 1, nom)
        ws.cell(ligne, 2, nb)
        c = ws.cell(ligne, 3, round(montant, 2)); _euros(c)
        ligne += 1
    _entete(ws, debut)
    ligne += 2

    ws.cell(ligne, 1, "VOLET (b) — Transitaires en cours de facturation").font = Font(bold=True, size=12)
    ligne += 1
    if rapport.transitaire_repli_utilise:
        ws.cell(ligne, 1, "⚠ " + rapport.transitaire_avertissement).fill = ROUGE
        ligne += 1
        ws.cell(ligne, 1, "Repli utilisé : lignes du Suivi principal avec Transitaire renseigné — AUCUNE estimation de coût transitaire disponible sans le classeur dédié (pas de taux d'approche moyen par trajet).")
        ligne += 1
    else:
        ws.cell(ligne, 1, "Source :")
        ws.cell(ligne, 2, str(rapport.chemin_commandes_speciales))
        ligne += 1
    ws.cell(ligne, 1, "Dossiers non facturés par le transitaire :")
    ws.cell(ligne, 2, f"{len(rapport.dossiers_transitaires)} dossier(s)")
    ligne += 1
    ws.cell(ligne, 1, "Montant marchandise concerné (HT) :")
    c = ws.cell(ligne, 2, round(total_marchandise_transit, 2)); _euros(c)
    ligne += 1
    ws.cell(ligne, 1, "Coût transitaire ESTIMÉ (taux d'approche moyen du trajet) :").font = Font(bold=True)
    c = ws.cell(ligne, 2, round(total_estime_transit, 2)); _euros(c)
    ws.cell(ligne, 3, "ESTIMATION — pas un montant facturé").fill = JAUNE
    ligne += 1
    if n_transit_sans_estimation:
        ws.cell(ligne, 1, f"Dont {n_transit_sans_estimation} dossier(s) sans estimation disponible (voir onglet Transitaires).")
        ligne += 1
    ligne += 1

    ws.cell(ligne, 1, "Fiabilité du volet transitaires").font = Font(bold=True)
    ligne += 1
    ws.cell(ligne, 1, "« Commandes spéciales » est un classeur peu alimenté et maintenu à la main (voir CLAUDE.md) — un dossier réel peut très bien ne jamais y avoir été saisi ; les montants ci-dessus ne portent donc que sur les dossiers effectivement enregistrés.")
    ligne += 1
    if rapport.commandes_transitaire_non_couvertes:
        n = len(rapport.commandes_transitaire_non_couvertes)
        ws.cell(ligne, 1, f"⚠ {n} N° de dossier revient cité(s) dans le Suivi principal (colonne Transitaire) mais introuvable(s) dans Commandes spéciales — NON comptés dans le total ci-dessus, à vérifier au cas par cas :").fill = JAUNE
        ligne += 1
        debut = ligne
        ws.cell(ligne, 1, "N° dossier revient"); ws.cell(ligne, 2, "N° de commande"); ws.cell(ligne, 3, "Fournisseur"); ws.cell(ligne, 4, "Chantier")
        ligne += 1
        for num_dossier, num_cde, fournisseur, chantier in rapport.commandes_transitaire_non_couvertes:
            ws.cell(ligne, 1, num_dossier)
            ws.cell(ligne, 2, num_cde)
            ws.cell(ligne, 3, fournisseur)
            ws.cell(ligne, 4, chantier)
            ligne += 1
        _entete(ws, debut)
    ligne += 2

    # --- v1.1 (session S0) : factures reçues non rapprochées (4a) ----------
    ws.cell(ligne, 1, "FACTURES REÇUES EN COURS DE RAPPROCHEMENT (hors volet a)").font = Font(bold=True, size=12)
    ligne += 1
    if not rapport.exclusion_appliquee:
        ws.cell(ligne, 1, "⚠ Exclusion désactivée (--sans-exclusion) — ces lignes restent comptées dans le volet (a) ci-dessus.").fill = JAUNE
        ligne += 1
    else:
        total_recues = sum(f.montant_facture_bl for f in rapport.factures_recues_non_rapprochees)
        ws.cell(ligne, 1, "Une vraie facture PDF a été trouvée dans a_traiter/Factures/ pour ces lignes (datée au plus tard fin de mois) mais n'a pas encore été écrite dans le Suivi — ce ne sont PLUS des factures non parvenues, juste pas encore rapprochées dans l'outil. Sorties du total du volet (a) ci-dessus.")
        ligne += 1
        ws.cell(ligne, 1, "Nombre de lignes :")
        ws.cell(ligne, 2, len(rapport.factures_recues_non_rapprochees))
        ligne += 1
        ws.cell(ligne, 1, "Montant HT (Facturé BL) :").font = Font(bold=True)
        c = ws.cell(ligne, 2, round(total_recues, 2)); _euros(c)
        ws.cell(ligne, 3, "voir onglet « Factures reçues »")
        ligne += 1
    ligne += 1

    # --- v1.1 (session S0) : ajustements déclarés par l'acheteur (4b) ------
    ws.cell(ligne, 1, "DÉCLARÉ PAR L'ACHETEUR (hors outil)").font = Font(bold=True, size=12)
    ligne += 1
    if rapport.chemin_ajustements is None:
        ws.cell(ligne, 1, f"Aucun fichier « {nom_fichier_ajustements(rapport.mois)} » trouvé dans referentiel/ — rien de déclaré ce mois-ci.")
        ligne += 1
    else:
        total_ajustements = sum(a.montant_ht for a in rapport.ajustements)
        ws.cell(ligne, 1, f"Source : referentiel/{nom_fichier_ajustements(rapport.mois)} — jamais fusionné avec les calculs des volets (a)/(b) ci-dessus.")
        ligne += 1
        ws.cell(ligne, 1, "Nombre de lignes déclarées :")
        ws.cell(ligne, 2, len(rapport.ajustements))
        ligne += 1
        ws.cell(ligne, 1, "Montant HT déclaré :").font = Font(bold=True)
        c = ws.cell(ligne, 2, round(total_ajustements, 2)); _euros(c)
        ws.cell(ligne, 3, "voir onglet « Déclaré (hors outil) »")
        ligne += 1
    ligne += 1

    # --- v1.1 (session S0) : réserves de périmètre (4c) ---------------------
    ws.cell(ligne, 1, "RÉSERVES DE PÉRIMÈTRE").font = Font(bold=True, size=12)
    ligne += 1
    r = rapport.reserves
    if r is not None:
        ws.cell(ligne, 1, f"{r.n_bdc_manuel_24x} facture(s) avec au moins une ligne sur bon manuel (\"BC/BCN 24XXXX\", carnet papier) — matériel livré, commande absente du Suivi, hors périmètre de ce calcul.")
        ligne += 1
        ws.cell(ligne, 1, f"{r.n_transitaires_sans_estimation} dossier(s) transitaire non facturé(s) sans estimation de coût disponible (déjà compté dans le volet (b) ci-dessus).")
        ligne += 1
        ws.cell(ligne, 1, "0 ligne(s) migrée(s) sans pièce jointe (fonctionnalité « Pièces » pas encore implémentée dans ce projet).")
        ligne += 1
        ws.cell(ligne, 1, f"Le volet (b) ne couvre QUE les {r.n_dossiers_speciales_total} dossier(s) réellement saisi(s) à la main dans Commandes spéciales — un dossier réel non saisi n'apparaît nulle part dans cet état.")
        ligne += 1
    ligne += 1

    _largeurs(ws, maxi=60)
    for col in ("A",):
        ws.column_dimensions[col].width = 70

    # --- BL non facturés -------------------------------------------------
    ws2 = wb.create_sheet("BL non facturés")
    entetes2 = [
        "Fournisseur", "N° de commande", "Chantier", "Référence", "Désignation",
        "Qté livrée", "Montant HT", "Source du prix", "Date de livraison",
        "Ancienneté (jours)", "N° facture", "Date facture", "Note",
    ]
    ws2.append(entetes2)
    for l in sorted(rapport.lignes_bl, key=lambda l: (l.fournisseur, -l.montant_ht)):
        ws2.append([
            l.fournisseur, l.numero_commande, l.chantier, l.reference, l.designation,
            l.qte_livree, round(l.montant_ht, 2), l.source_prix, l.date_livraison,
            l.anciennete_jours, l.numero_facture, l.date_facture, l.note,
        ])
    for row in ws2.iter_rows(min_row=2, min_col=7, max_col=7):
        _euros(row[0])
    for row in ws2.iter_rows(min_row=2, min_col=9, max_col=9):
        _date_fmt(row[0])
    for row in ws2.iter_rows(min_row=2, min_col=12, max_col=12):
        _date_fmt(row[0])
    _entete(ws2)

    if rapport.lignes_sans_prix:
        # Tout en .append() séquentiel, jamais mélangé à .cell() pour CRÉER
        # une ligne : .append() avance son propre compteur interne
        # (_current_row) indépendamment de max_row — un .cell() intercalé
        # pour "sauter" des lignes désynchronise les deux et peut faire
        # écrire la ligne suivante par-dessus une ligne déjà posée à la
        # main (bug réel trouvé en écrivant les tests). Le style (gras,
        # fond) s'applique APRÈS coup, sur une cellule déjà écrite — jamais
        # pour la créer.
        ws2.append([""])
        ws2.append(["Lignes livrées SANS AUCUN PRIX connu — non valorisées, à traiter à part"])
        ligne_titre = ws2.max_row
        ws2.cell(ligne_titre, 1).font = Font(bold=True)
        ws2.cell(ligne_titre, 1).fill = ROUGE

        ws2.append(entetes2)
        debut_bloc = ws2.max_row
        for l in sorted(rapport.lignes_sans_prix, key=lambda l: l.fournisseur):
            ws2.append([
                l.fournisseur, l.numero_commande, l.chantier, l.reference, l.designation,
                l.qte_livree, 0, l.source_prix, l.date_livraison,
                l.anciennete_jours, l.numero_facture, l.date_facture, l.note,
            ])
        _entete(ws2, debut_bloc)

    _largeurs(ws2, maxi=45)

    # --- Transitaires ------------------------------------------------------
    ws3 = wb.create_sheet("Transitaires")
    if rapport.transitaire_repli_utilise:
        entetes3 = ["N° de commande", "Chantier", "Fournisseur", "Transitaire", "Date de livraison", "Ancienneté (jours)", "Montant marchandise HT", "Coût estimé"]
        ws3.append(entetes3)
        for d in sorted(rapport.dossiers_transitaires, key=lambda d: -d.montant_marchandise):
            ws3.append([
                d.numero_commande, d.chantier, d.fournisseur, d.transitaire,
                d.date_arrivee, d.anciennete_jours, round(d.montant_marchandise, 2),
                "estimation indisponible (classeur Commandes spéciales absent)",
            ])
        for row in ws3.iter_rows(min_row=2, min_col=7, max_col=7):
            _euros(row[0])
        for row in ws3.iter_rows(min_row=2, min_col=5, max_col=5):
            _date_fmt(row[0])
    else:
        entetes3 = [
            "N° dossier revient", "Désignation", "N° de commande", "Chantier", "Fournisseur",
            "Transitaire", "Réf. transport", "Date de départ", "Date d'arrivée",
            "Ancienneté (jours)", "Montant marchandise HT", "Coût transitaire ESTIMÉ",
        ]
        ws3.append(entetes3)
        for d in sorted(rapport.dossiers_transitaires, key=lambda d: -(d.cout_estime or 0)):
            ws3.append([
                d.numero_dossier, d.designation, d.numero_commande or "—", d.chantier, d.fournisseur,
                d.transitaire, d.ref_transport, d.date_depart, d.date_arrivee,
                d.anciennete_jours, round(d.montant_marchandise, 2),
                round(d.cout_estime, 2) if d.cout_estime is not None else "non calculable",
            ])
        for row in ws3.iter_rows(min_row=2, min_col=11, max_col=12):
            if isinstance(row[0].value, (int, float)):
                _euros(row[0])
            if isinstance(row[1].value, (int, float)):
                _euros(row[1])
        for col_lettre in ("H", "I"):
            for row in ws3[f"{col_lettre}2:{col_lettre}{ws3.max_row}"]:
                _date_fmt(row[0])
    _entete(ws3)
    _largeurs(ws3, maxi=45)

    # --- v1.1 (session S0) : Factures reçues en cours de rapprochement (4a) -
    if rapport.exclusion_appliquee:
        ws4 = wb.create_sheet("Factures reçues")
        ws4.append([
            "Fournisseur", "N° de commande", "Référence", "N° facture",
            "Date facture", "Montant HT (Facturé BL)",
        ])
        for f in sorted(rapport.factures_recues_non_rapprochees, key=lambda f: (f.fournisseur, -f.montant_facture_bl)):
            ws4.append([
                f.fournisseur, f.numero_commande, f.reference, f.numero_facture,
                f.date_facture, round(f.montant_facture_bl, 2),
            ])
        for row in ws4.iter_rows(min_row=2, min_col=6, max_col=6):
            _euros(row[0])
        for row in ws4.iter_rows(min_row=2, min_col=5, max_col=5):
            _date_fmt(row[0])
        _entete(ws4)
        _largeurs(ws4, maxi=45)

    # --- v1.1 (session S0) : Déclaré par l'acheteur (hors outil) (4b) -------
    ws5 = wb.create_sheet("Déclaré (hors outil)")
    ws5.append([
        "Type", "Libellé", "Fournisseur/Transitaire", "Chantier", "Pièce",
        "Date de livraison", "Montant HT", "Source", "Commentaire",
    ])
    for a in rapport.ajustements:
        ws5.append([
            a.type, a.libelle, a.fournisseur_ou_transitaire, a.chantier, a.piece,
            a.date_livraison, round(a.montant_ht, 2), a.source, a.commentaire,
        ])
    for row in ws5.iter_rows(min_row=2, min_col=7, max_col=7):
        _euros(row[0])
    for row in ws5.iter_rows(min_row=2, min_col=6, max_col=6):
        _date_fmt(row[0])
    _entete(ws5)
    _largeurs(ws5, maxi=45)

    dossier_rapports = dossier_projet / DOSSIER_RAPPORTS
    dossier_rapports.mkdir(parents=True, exist_ok=True)
    chemin_sortie = dossier_rapports / f"FNP_{rapport.mois}.xlsx"
    wb.save(chemin_sortie)
    return chemin_sortie
